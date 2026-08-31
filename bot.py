# ============================================================
# bot.py - نسخه ۹.۴.۳ (اتصال اتمیک و قابل‌اعتماد حساب بله هنگام ورود)
# ============================================================
import os
import time
import logging
from logging.handlers import RotatingFileHandler
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import psycopg2
from psycopg2 import pool
from datetime import datetime, timedelta, timezone
import threading
from concurrent.futures import ThreadPoolExecutor
from flask import Flask, jsonify
import jdatetime
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
import json
import re
import io
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
import numpy as np
import arabic_reshaper
from bidi.algorithm import get_display
import os.path
import traceback
import html
from collections import deque, OrderedDict
import time as time_module
import pickle
import atexit
import gzip
import hashlib
import binascii
import hmac
import uuid
import base64
import tempfile
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from psycopg2 import sql
from psycopg2.extras import execute_values
from zoneinfo import ZoneInfo  # Python 3.9+

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except Exception:
    load_workbook = None
    OPENPYXL_AVAILABLE = False

try:
    import plotly.graph_objects as go
    PLOTLY_AVAILABLE = True
except Exception:
    go = None
    PLOTLY_AVAILABLE = False

_chart_engine_error = None
ALLOW_LEGACY_CHART_FALLBACK = os.getenv('ALLOW_LEGACY_CHART_FALLBACK', 'false').lower() == 'true'

# ============================================================
# تنظیمات ثابت
# ============================================================
IRAN_TZ = ZoneInfo("Asia/Tehran")
VALID_ROLES = ('admin', 'deputy', 'super_admin')
ALLOWED_UPDATE_FIELDS = {'employee_number', 'full_name', 'title', 'branch_id'}
EDIT_DEADLINE_HOUR = 23   # تا پایان همان روز
EDIT_DEADLINE_MINUTE = 59
SCORE_DEADLINE_HOUR = 16   # ۱۶:۳۰
SCORE_DEADLINE_MINUTE = 30
MAX_AMOUNT_MILLIONS = 4_000_000_000_000
BACKUP_FORMAT_VERSION = 5
BACKUP_SECRET = os.getenv("SUPER_ADMIN_PASSWORD", "")
MAX_BACKUP_COMPRESSED_BYTES = int(os.getenv("MAX_BACKUP_BYTES", 25 * 1024 * 1024))
MAX_BACKUP_UNCOMPRESSED_BYTES = int(os.getenv("MAX_BACKUP_UNCOMPRESSED_BYTES", 200 * 1024 * 1024))
ALLOW_UNSIGNED_LEGACY_BACKUP = os.getenv("ALLOW_UNSIGNED_LEGACY_BACKUP", "false").lower() == "true"
BACKUP_TABLES = (
    'branches', 'users', 'collections', 'notes', 'user_activity_log',
    'settings', 'holidays', 'problems', 'scores', 'feature_settings',
    'actual_stats', 'actual_stats_import_operations', 'actual_stats_audit_log',
    'branch_targets', 'deputy_transfer_operations',
    'deputy_replacement_operations', 'deputy_branch_assignments',
    'message_campaigns', 'message_deliveries'
)
RESTORE_IDENTITY_FIELDS = {
    'branches': ('name',),
    'users': ('employee_number',),
    'collections': ('branch_id', 'shamsi_date'),
    'notes': ('collection_id', 'user_id'),
    'user_activity_log': ('user_id', 'action', 'created_at'),
    'holidays': ('shamsi_date',),
    'problems': ('user_id', 'created_at'),
    'scores': ('collection_id',),
    'actual_stats': ('branch_id', 'shamsi_date'),
    'actual_stats_import_operations': ('operation_uuid',),
    'actual_stats_audit_log': ('operation_id', 'branch_id'),
    'branch_targets': ('branch_id', 'target_date', 'created_at'),
    'deputy_transfer_operations': ('operation_uuid',),
    'deputy_replacement_operations': ('operation_uuid',),
    'deputy_branch_assignments': ('deputy_id', 'branch_id', 'effective_from'),
    'message_campaigns': ('campaign_uuid',),
    'message_deliveries': ('campaign_id', 'user_id'),
}
MAX_CAMPAIGN_TEXT_LENGTH = 3500
MAX_CAMPAIGN_MEDIA_BYTES = 5 * 1024 * 1024
ACTUAL_EXCEL_TEMP_PREFIX = 'zanjan_actual_stats_'
ACTUAL_EXCEL_CONFIRM_TTL_SECONDS = 30 * 60
# تمام تغییرات اتصال حساب بله به‌صورت کوتاه سریالی می‌شوند تا ورودهای هم‌زمان
# نتوانند محدودیت یکتایی telegram_id را دور بزنند یا نشست کاذب بسازند.
AUTH_BIND_ADVISORY_LOCK_KEY = -9_453_100_001

# کدهای رسمی شعب مطابق ستون «کد شعبه» در فایل گزارش سازمانی.
# مقداردهی دیتابیس فقط برای خانه‌های خالی انجام می‌شود و کد موجود را بازنویسی نمی‌کند.
OFFICIAL_BRANCH_CODES = {
    'مرکزی زنجان': '20578',
    'جمهوری': '20784',
    'میدان استقلال': '20602',
    'میدان پایین': '20610',
    'سعدی جنوبی': '20628',
    'سبزه میدان': '20594',
    'اسلام آباد': '20651',
    'خرمشهر': '20644',
    'بعثت': '20636',
    'سعدی شمالی': '20685',
    'سلطانیه': '20834',
    'آببر': '20776',
    'کوچمشکی': '20442',
    'دندی': '20227',
    'مرکزی ابهر': '22053',
    'هیدج': '22103',
    'صايين قلعه': '22111',
    'میدان معلم ابهر': '22061',
    'خدابنده': '22071',
    'خرمدره': '20250',
}

# این نام‌های مستعار فقط هنگام خواندن فایل Excel آمار واقعی استفاده می‌شوند.
# کد اصلی شعبه در دیتابیس، گزارش‌ها و سایر بخش‌های ربات تغییر نمی‌کند.
ACTUAL_EXCEL_CODE_ALIASES = {
    '20842': '20442',  # کوچمشکی
    '20727': '20227',  # دندی
    '20701': '22071',  # خدابنده
    '20750': '20250',  # خرمدره
}

# ============================================================
# تنظیمات لاگین
# ============================================================
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
if logger.hasHandlers():
    logger.handlers.clear()
file_handler = RotatingFileHandler(
    "bot.log", maxBytes=5*1024*1024, backupCount=3, encoding='utf-8'
)
file_handler.setLevel(logging.INFO)
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)
console_handler.setFormatter(formatter)
logger.addHandler(file_handler)
logger.addHandler(console_handler)

# ============================================================
# متغیرهای محیطی
# ============================================================
BOT_TOKEN = os.getenv("BOT_TOKEN")
DB_URL = os.getenv("DATABASE_URL")
PORT = int(os.getenv("PORT", 10000))
SUPER_ADMIN_PASSWORD = os.getenv("SUPER_ADMIN_PASSWORD")
if not BOT_TOKEN or not DB_URL:
    logger.error("❌ BOT_TOKEN and DATABASE_URL are required!")
    exit(1)
if not SUPER_ADMIN_PASSWORD:
    logger.error("❌ SUPER_ADMIN_PASSWORD environment variable is required!")
    exit(1)

# هش کردن رمز عبور برای ذخیره در حافظه (امنیت بیشتر)
PASSWORD_HASH = hashlib.sha256(SUPER_ADMIN_PASSWORD.encode()).hexdigest()

# مخفی کردن توکن در لاگ
safe_base_url = f"https://tapi.bale.ai/bot{'-' * 10}"
BASE_URL = f"https://tapi.bale.ai/bot{BOT_TOKEN}"
logger.info(f"✅ Bale API URL: {safe_base_url}")

# ============================================================
# Flask
# ============================================================
flask_app = Flask(__name__)
_scheduler = None

@flask_app.route('/health')
def health():
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("SELECT 1")
            cur.fetchone()
        scheduler_ok = bool(_scheduler and _scheduler.running)
        status = "healthy" if scheduler_ok else "degraded"
        return jsonify({"status": status, "database": "ok", "scheduler": scheduler_ok,
                        "timestamp": time.time()}), (200 if scheduler_ok else 503)
    except Exception:
        logger.exception("Health check failed")
        return jsonify({"status": "unhealthy", "database": "error", "scheduler": False,
                        "timestamp": time.time()}), 503
    finally:
        if conn:
            return_db_connection(conn)

@flask_app.route('/')
def root():
    return jsonify({"message": "Bot is running", "status": "active"})

def run_flask():
    flask_app.run(host='0.0.0.0', port=PORT)

# ============================================================
# Session
# ============================================================
def create_session():
    session = requests.Session()
    session.headers.update({'Connection': 'keep-alive', 'User-Agent': 'Bale-Bank-Bot/9.4.3'})
    retry_strategy = Retry(
        total=5, backoff_factor=1,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["HEAD", "GET", "OPTIONS", "POST"]
    )
    adapter = HTTPAdapter(max_retries=retry_strategy, pool_connections=20, pool_maxsize=20)
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    return session

requests_session = create_session()
_http_local = threading.local()

def get_http_session():
    """Return one requests.Session per thread; Session objects are not shared."""
    session = getattr(_http_local, 'session', None)
    if session is None:
        session = create_session()
        _http_local.session = session
    return session

# ============================================================
# Connection Pool با مدیریت ایمن
# ============================================================
class SafeConnectionPool:
    def __init__(self, minconn=5, maxconn=30, dsn=None):
        self.dsn = dsn
        self.minconn = minconn
        self.maxconn = maxconn
        self._pool = None
        self._lock = threading.RLock()
        self._direct_connection_ids = set()
        self._direct_semaphore = threading.BoundedSemaphore(5)
        self._create_pool()

    def _create_pool(self):
        try:
            self._pool = psycopg2.pool.ThreadedConnectionPool(
                self.minconn, self.maxconn, self.dsn
            )
            logger.info("✅ Connection pool created")
        except Exception as e:
            logger.error(f"❌ Failed to create pool: {e}")
            self._pool = None

    def getconn(self):
        with self._lock:
            if self._pool is None:
                self._create_pool()
            if self._pool is not None:
                try:
                    return self._pool.getconn()
                except Exception as e:
                    logger.error(f"Pool getconn error: {e}")
                    raise
            if not self._direct_semaphore.acquire(timeout=10):
                raise psycopg2.OperationalError("Direct database connection limit reached")
            try:
                conn = psycopg2.connect(self.dsn)
                self._direct_connection_ids.add(id(conn))
                return conn
            except Exception:
                self._direct_semaphore.release()
                raise

    def putconn(self, conn):
        if conn is None:
            return
        with self._lock:
            if id(conn) in self._direct_connection_ids:
                self._direct_connection_ids.discard(id(conn))
                try:
                    conn.close()
                except Exception:
                    pass
                finally:
                    self._direct_semaphore.release()
                return
            if self._pool is not None:
                try:
                    self._pool.putconn(conn)
                    return
                except Exception as e:
                    logger.error(f"Pool putconn error: {e}, closing directly")
            try:
                conn.close()
            except Exception:
                pass

db_pool = SafeConnectionPool(5, 30, DB_URL)

def get_db_connection():
    return db_pool.getconn()

def return_db_connection(conn):
    db_pool.putconn(conn)

# ============================================================
# کش‌های Thread-Safe با TTL
# ============================================================
class TTLCache:
    def __init__(self, ttl_seconds=10):
        self._cache = {}
        self._timestamps = {}
        self._ttl = ttl_seconds
        self._lock = threading.RLock()

    def get(self, key):
        with self._lock:
            if key in self._cache and time_module.time() - self._timestamps[key] < self._ttl:
                return self._cache[key]
            return None

    def set(self, key, value):
        with self._lock:
            self._cache[key] = value
            self._timestamps[key] = time_module.time()

    def invalidate(self, key=None):
        with self._lock:
            if key is None:
                self._cache.clear()
                self._timestamps.clear()
            elif key in self._cache:
                del self._cache[key]
                del self._timestamps[key]

    def invalidate_all(self):
        self.invalidate(None)

cache_bot_status = TTLCache(ttl_seconds=5)
cache_feature_settings = TTLCache(ttl_seconds=10)
cache_branches = TTLCache(ttl_seconds=300)
cache_today_report = TTLCache(ttl_seconds=30)
cache_top_branches = TTLCache(ttl_seconds=60)
cache_10day_report = TTLCache(ttl_seconds=60)
cache_adaptive = TTLCache(ttl_seconds=60)
cache_forecast_all = TTLCache(ttl_seconds=120)
cache_targets = TTLCache(ttl_seconds=60)
cache_admins = TTLCache(ttl_seconds=300)

def invalidate_branches_cache():
    cache_branches.invalidate('branches')

# ============================================================
# State Management با TTL برای جلوگیری از نشت حافظه
# ============================================================
class TTLUserStates:
    def __init__(self, ttl_seconds=3600):  # 1 ساعت
        self._states = {}
        self._ttl = ttl_seconds
        self._lock = threading.RLock()
        self._cleanup_thread = threading.Thread(target=self._cleanup_loop, daemon=True)
        self._cleanup_thread.start()

    def _cleanup_loop(self):
        while True:
            time_module.sleep(300)  # هر ۵ دقیقه
            self._cleanup_expired()

    def _cleanup_expired(self):
        now = time_module.time()
        with self._lock:
            expired = [k for k, v in self._states.items() if now - v['timestamp'] > self._ttl]
            for k in expired:
                del self._states[k]

    def get(self, chat_id, default=None):
        with self._lock:
            data = self._states.get(chat_id)
            if data and time_module.time() - data['timestamp'] <= self._ttl:
                return data['state']
            return default

    def set(self, chat_id, state):
        with self._lock:
            self._states[chat_id] = {'state': state, 'timestamp': time_module.time()}

    def update(self, chat_id, state_dict):
        with self._lock:
            if chat_id not in self._states:
                self._states[chat_id] = {'state': {}, 'timestamp': time_module.time()}
            if isinstance(state_dict, dict):
                self._states[chat_id]['state'].update(state_dict)
            else:
                self._states[chat_id]['state'] = state_dict
            self._states[chat_id]['timestamp'] = time_module.time()

    def delete(self, chat_id):
        with self._lock:
            if chat_id in self._states:
                del self._states[chat_id]

user_states = TTLUserStates(ttl_seconds=3600)

# ============================================================
# پردازش آپدیت‌ها با قفل
# ============================================================
processed_updates = deque(maxlen=2000)
processed_set = set()
processed_set_lock = threading.Lock()

# ============================================================
# Thread Pool با مدیریت خطا
# ============================================================
executor = ThreadPoolExecutor(max_workers=5)

# ============================================================
# توابع کمکی
# ============================================================
PERSIAN_DIGITS = '۰۱۲۳۴۵۶۷۸۹'
ARABIC_DIGITS = '٠١٢٣٤٥٦٧٨٩'
ENGLISH_DIGITS = '0123456789'
DIGIT_MAP = str.maketrans(PERSIAN_DIGITS + ARABIC_DIGITS, ENGLISH_DIGITS + ENGLISH_DIGITS)

def normalize_digits(text):
    if not text:
        return text
    text = str(text).translate(DIGIT_MAP)
    text = text.replace(',', '').replace('،', '').replace(' ', '')
    return text

def safe_log_value(value, limit=200):
    return re.sub(r'[\r\n\t]+', ' ', str(value))[:limit]

def parse_number(text):
    try:
        if not text or text.strip() == '':
            return None
        text = str(text).strip()
        if text.endswith('-'):
            text = '-' + text[:-1]
        text = normalize_digits(text)
        if not text or text == '-':
            return None
        if len(text) > 20:  # جلوگیری از overflow
            return None
        if '.' in text:
            return int(float(text))
        return int(text)
    except (OverflowError, ValueError, MemoryError):
        logger.error("parse_number error for '%s'", safe_log_value(text))
        return None

def escape_markdown(text):
    if not text:
        return ""
    escape_chars = ['_', '*', '[', ']', '(', ')', '~', '`']
    for char in escape_chars:
        text = text.replace(char, '\\' + char)
    return text

def escape_like_pattern(text):
    """Escape PostgreSQL LIKE wildcards while keeping parameterized queries."""
    return str(text).replace('\\', '\\\\').replace('%', '\\%').replace('_', '\\_')

def split_text_safely(text, max_len=4000):
    if len(text) <= max_len:
        return [text]
    chunks = []
    while text:
        if len(text) <= max_len:
            chunks.append(text)
            break
        cut = max_len
        for sep in ['\n\n', '\n', '. ', '، ', ' ']:
            pos = text.rfind(sep, 0, cut)
            if pos > cut - 50:
                cut = pos + len(sep)
                break
        chunks.append(text[:cut])
        text = text[cut:]
    return chunks

def get_keyboard(role, is_super_admin=False):
    if is_super_admin:
        return get_super_admin_keyboard()
    if role == 'admin':
        return get_admin_keyboard()
    return get_deputy_keyboard()

# ============================================================
# تنظیم فونت فارسی برای نمودارها (با قفل)
# ============================================================
_font_initialized = False
_font_lock = threading.Lock()
_persian_font_property = None

def setup_persian_font_once():
    global _font_initialized, _persian_font_property
    with _font_lock:
        if _font_initialized:
            return
        try:
            font_paths = [
                os.getenv('PERSIAN_FONT_PATH', ''),
                os.path.join(os.path.dirname(os.path.abspath(__file__)), 'fonts', 'Vazirmatn-Regular.ttf'),
                '/usr/share/fonts/truetype/vazirmatn/Vazirmatn-Regular.ttf',
                '/usr/share/fonts/truetype/vazir/Vazir.ttf',
                '/usr/share/fonts/opentype/noto/NotoSansArabic-Regular.ttf',
                '/usr/share/fonts/truetype/noto/NotoNaskhArabic-Regular.ttf',
                '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
                '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
                '/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf',
                '/usr/share/fonts/truetype/ttf-dejavu/DejaVuSans.ttf',
            ]
            plt.rcParams['font.family'] = 'sans-serif'
            plt.rcParams['axes.unicode_minus'] = False
            for path in font_paths:
                if path and os.path.isfile(path):
                    fm.fontManager.addfont(path)
                    prop = fm.FontProperties(fname=path)
                    _persian_font_property = prop
                    plt.rcParams['font.family'] = prop.get_name()
                    logger.info(f"✅ Font loaded: {path}")
                    _font_initialized = True
                    return
            plt.rcParams['font.family'] = ['DejaVu Sans', 'Liberation Sans', 'sans-serif']
            _persian_font_property = fm.FontProperties(family='DejaVu Sans')
            logger.warning("⚠️ No Persian font found, using fallback fonts")
            _font_initialized = True
        except Exception as e:
            logger.error(f"❌ Font setup error: {e}")
            plt.rcParams['font.family'] = 'sans-serif'
            _persian_font_property = fm.FontProperties(family='DejaVu Sans')
            _font_initialized = True

setup_persian_font_once()

def reshape_persian(text):
    if not text:
        return ""
    try:
        value = str(text).replace('\u200e', '').replace('\u200f', '').strip()
        if not re.search(r'[\u0600-\u06ff]', value):
            return value
        reshaped = arabic_reshaper.reshape(value)
        return get_display(reshaped, base_dir='R')
    except Exception:
        return str(text)

# ============================================================
# توابع تاریخ با ZoneInfo
# ============================================================
def get_iran_time():
    return datetime.now(IRAN_TZ)

def get_shamsi_date(days_offset=0):
    now = get_iran_time() + timedelta(days=days_offset)
    shamsi = jdatetime.datetime.fromgregorian(datetime=now)
    return f"{shamsi.year}/{shamsi.month:02d}/{shamsi.day:02d}"

def get_shamsi_date_formatted(shamsi_str):
    if not shamsi_str:
        return "نامعلوم"
    parts = shamsi_str.split('/')
    if len(parts) != 3:
        return shamsi_str
    year, month, day = parts
    month = month.zfill(2)
    day = day.zfill(2)
    months = {
        '01':'فروردین','02':'اردیبهشت','03':'خرداد',
        '04':'تیر','05':'مرداد','06':'شهریور',
        '07':'مهر','08':'آبان','09':'آذر',
        '10':'دی','11':'بهمن','12':'اسفند'
    }
    return f"{int(day)} {months.get(month, '')} {year}"

def safe_format(value, default="0"):
    return value if value is not None else default

def validate_shamsi_date(shamsi_str):
    shamsi_str = normalize_digits(shamsi_str)
    if not re.match(r'^\d{4}/\d{2}/\d{2}$', shamsi_str):
        return False
    try:
        year, month, day = map(int, shamsi_str.split('/'))
        jdatetime.date(year, month, day)
        return True
    except Exception:
        return False

def add_days_to_shamsi(value, days):
    """Date arithmetic through Gregorian dates for all supported jdatetime versions."""
    if isinstance(value, str):
        value = jdatetime.date(*map(int, value.split('/')))
    gregorian = value.togregorian() + timedelta(days=days)
    return jdatetime.date.fromgregorian(date=gregorian)

def is_last_day_of_shamsi_month(shamsi_date_str):
    try:
        parts = shamsi_date_str.split('/')
        year, month, day = map(int, parts)
        try:
            next_day = add_days_to_shamsi(jdatetime.date(year, month, day), 1)
            return next_day.month != month
        except Exception:
            return False
    except Exception:
        return False

def get_shamsi_month_range():
    today = get_iran_time()
    shamsi_today = jdatetime.datetime.fromgregorian(datetime=today)
    first_day = jdatetime.date(shamsi_today.year, shamsi_today.month, 1)
    if shamsi_today.month == 12:
        next_month = jdatetime.date(shamsi_today.year + 1, 1, 1)
    else:
        next_month = jdatetime.date(shamsi_today.year, shamsi_today.month + 1, 1)
    last_day = add_days_to_shamsi(next_month, -1)
    return (
        f"{first_day.year}/{first_day.month:02d}/{first_day.day:02d}",
        f"{last_day.year}/{last_day.month:02d}/{last_day.day:02d}"
    )

# ============================================================
# ایجاد جداول (با ایندکس‌های کامل)
# ============================================================
def create_all_tables_if_not_exists():
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS branches (
                    id SERIAL PRIMARY KEY,
                    name VARCHAR(255) NOT NULL UNIQUE,
                    official_code VARCHAR(20),
                    created_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP
                )
            """)
            cur.execute("ALTER TABLE branches ADD COLUMN IF NOT EXISTS official_code VARCHAR(20)")
            cur.execute("""
                CREATE UNIQUE INDEX IF NOT EXISTS idx_branches_official_code_unique
                ON branches(official_code) WHERE official_code IS NOT NULL
            """)
            branch_code_values = list(OFFICIAL_BRANCH_CODES.items())
            execute_values(cur, """
                UPDATE branches AS b
                SET official_code = v.official_code
                FROM (VALUES %s) AS v(branch_name, official_code)
                WHERE b.name = v.branch_name
                  AND b.official_code IS NULL
                  AND NOT EXISTS (
                      SELECT 1 FROM branches other
                      WHERE other.official_code = v.official_code
                  )
            """, branch_code_values)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS users (
                    id SERIAL PRIMARY KEY,
                    employee_number VARCHAR(20) NOT NULL UNIQUE,
                    telegram_id BIGINT UNIQUE,
                    full_name VARCHAR(255) NOT NULL,
                    role VARCHAR(20) NOT NULL CHECK (role IN ('admin', 'deputy', 'super_admin')),
                    title VARCHAR(255) NOT NULL,
                    branch_id INTEGER REFERENCES branches(id) ON DELETE SET NULL,
                    is_super_admin BOOLEAN DEFAULT FALSE,
                    is_active BOOLEAN NOT NULL DEFAULT TRUE,
                    created_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP
                )
            """)
            # مهاجرت افزایشی نسخه‌های قبلی؛ هیچ کاربر موجودی بازنویسی یا حذف نمی‌شود.
            cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS is_active BOOLEAN")
            cur.execute("UPDATE users SET is_active=TRUE WHERE is_active IS NULL")
            cur.execute("ALTER TABLE users ALTER COLUMN is_active SET DEFAULT TRUE")
            cur.execute("ALTER TABLE users ALTER COLUMN is_active SET NOT NULL")
            cur.execute("""
                CREATE TABLE IF NOT EXISTS collections (
                    id SERIAL PRIMARY KEY,
                    branch_id INTEGER NOT NULL REFERENCES branches(id) ON DELETE CASCADE,
                    deputy_amount BIGINT NOT NULL DEFAULT 0,
                    others_amount BIGINT NOT NULL DEFAULT 0,
                    total_amount BIGINT GENERATED ALWAYS AS (deputy_amount + others_amount) STORED,
                    shamsi_date VARCHAR(10) NOT NULL,
                    recorded_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                    created_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP,
                    CONSTRAINT unique_branch_date UNIQUE (branch_id, shamsi_date)
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS notes (
                    id SERIAL PRIMARY KEY,
                    collection_id INTEGER NOT NULL REFERENCES collections(id) ON DELETE CASCADE,
                    user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    note_text TEXT NOT NULL,
                    created_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS user_activity_log (
                    id SERIAL PRIMARY KEY,
                    user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    action VARCHAR(50) NOT NULL,
                    details TEXT,
                    created_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS settings (
                    key VARCHAR(50) PRIMARY KEY,
                    value TEXT,
                    updated_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS holidays (
                    id SERIAL PRIMARY KEY,
                    shamsi_date VARCHAR(10) NOT NULL UNIQUE,
                    description VARCHAR(255),
                    created_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS problems (
                    id SERIAL PRIMARY KEY,
                    user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    problem_text TEXT NOT NULL,
                    category VARCHAR(50) DEFAULT 'general',
                    status VARCHAR(20) DEFAULT 'pending' CHECK (status IN ('pending', 'resolved', 'rejected')),
                    created_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS scores (
                    id SERIAL PRIMARY KEY,
                    collection_id INTEGER NOT NULL REFERENCES collections(id) ON DELETE CASCADE,
                    score INTEGER NOT NULL DEFAULT 0,
                    created_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP,
                    CONSTRAINT unique_collection_score UNIQUE (collection_id)
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS feature_settings (
                    key VARCHAR(50) PRIMARY KEY,
                    value TEXT,
                    updated_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS actual_stats (
                    id SERIAL PRIMARY KEY,
                    branch_id INTEGER NOT NULL REFERENCES branches(id) ON DELETE CASCADE,
                    shamsi_date VARCHAR(10) NOT NULL,
                    total_actual BIGINT NOT NULL DEFAULT 0,
                    recorded_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                    created_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP,
                    CONSTRAINT unique_branch_actual_date UNIQUE (branch_id, shamsi_date)
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS actual_stats_import_operations (
                    id BIGSERIAL PRIMARY KEY,
                    operation_uuid VARCHAR(36) NOT NULL UNIQUE,
                    shamsi_date VARCHAR(10) NOT NULL,
                    source_file_name VARCHAR(255) NOT NULL,
                    source_sha256 VARCHAR(64) NOT NULL,
                    extracted_count INTEGER NOT NULL,
                    corrected_count INTEGER NOT NULL DEFAULT 0,
                    created_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                    created_at TIMESTAMP WITH TIME ZONE NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    CHECK (shamsi_date ~ '^[0-9]{4}/[0-9]{2}/[0-9]{2}$'),
                    CHECK (source_sha256 ~ '^[0-9a-f]{64}$'),
                    CHECK (extracted_count >= 0),
                    CHECK (corrected_count >= 0)
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS actual_stats_audit_log (
                    id BIGSERIAL PRIMARY KEY,
                    actual_stats_id INTEGER NOT NULL REFERENCES actual_stats(id) ON DELETE RESTRICT,
                    operation_id BIGINT REFERENCES actual_stats_import_operations(id) ON DELETE RESTRICT,
                    branch_id INTEGER NOT NULL REFERENCES branches(id) ON DELETE RESTRICT,
                    shamsi_date VARCHAR(10) NOT NULL,
                    old_total_actual BIGINT,
                    extracted_total_actual BIGINT,
                    new_total_actual BIGINT NOT NULL,
                    source VARCHAR(20) NOT NULL CHECK (source IN ('manual', 'excel')),
                    changed_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                    created_at TIMESTAMP WITH TIME ZONE NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    CHECK (shamsi_date ~ '^[0-9]{4}/[0-9]{2}/[0-9]{2}$')
                )
            """)
            cur.execute("ALTER TABLE actual_stats_audit_log ADD COLUMN IF NOT EXISTS extracted_total_actual BIGINT")
            cur.execute("""
                CREATE TABLE IF NOT EXISTS branch_targets (
                    id SERIAL PRIMARY KEY,
                    branch_id INTEGER NOT NULL REFERENCES branches(id) ON DELETE CASCADE,
                    target_amount BIGINT NOT NULL,
                    target_date VARCHAR(10) NOT NULL,
                    created_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                    created_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP,
                    is_active BOOLEAN DEFAULT TRUE
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS deputy_transfer_operations (
                    id BIGSERIAL PRIMARY KEY,
                    operation_uuid VARCHAR(36) NOT NULL UNIQUE,
                    first_deputy_id INTEGER NOT NULL REFERENCES users(id) ON DELETE RESTRICT,
                    second_deputy_id INTEGER NOT NULL REFERENCES users(id) ON DELETE RESTRICT,
                    first_old_branch_id INTEGER NOT NULL REFERENCES branches(id) ON DELETE RESTRICT,
                    first_new_branch_id INTEGER NOT NULL REFERENCES branches(id) ON DELETE RESTRICT,
                    second_old_branch_id INTEGER NOT NULL REFERENCES branches(id) ON DELETE RESTRICT,
                    second_new_branch_id INTEGER NOT NULL REFERENCES branches(id) ON DELETE RESTRICT,
                    effective_date VARCHAR(10) NOT NULL,
                    performed_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                    reason TEXT,
                    created_at TIMESTAMP WITH TIME ZONE NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    CHECK (effective_date ~ '^[0-9]{4}/[0-9]{2}/[0-9]{2}$'),
                    CHECK (first_deputy_id <> second_deputy_id),
                    CHECK (first_old_branch_id <> first_new_branch_id),
                    CHECK (second_old_branch_id <> second_new_branch_id)
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS deputy_replacement_operations (
                    id BIGSERIAL PRIMARY KEY,
                    operation_uuid VARCHAR(36) NOT NULL UNIQUE,
                    branch_id INTEGER NOT NULL REFERENCES branches(id) ON DELETE RESTRICT,
                    old_deputy_id INTEGER NOT NULL REFERENCES users(id) ON DELETE RESTRICT,
                    new_deputy_id INTEGER NOT NULL REFERENCES users(id) ON DELETE RESTRICT,
                    effective_date VARCHAR(10) NOT NULL,
                    performed_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                    created_at TIMESTAMP WITH TIME ZONE NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    CHECK (effective_date ~ '^[0-9]{4}/[0-9]{2}/[0-9]{2}$'),
                    CHECK (old_deputy_id <> new_deputy_id)
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS deputy_branch_assignments (
                    id BIGSERIAL PRIMARY KEY,
                    deputy_id INTEGER NOT NULL REFERENCES users(id) ON DELETE RESTRICT,
                    branch_id INTEGER NOT NULL REFERENCES branches(id) ON DELETE RESTRICT,
                    effective_from VARCHAR(10) NOT NULL,
                    effective_to VARCHAR(10),
                    operation_id BIGINT REFERENCES deputy_transfer_operations(id) ON DELETE RESTRICT,
                    replacement_operation_id BIGINT,
                    created_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                    source VARCHAR(30) NOT NULL DEFAULT 'manual',
                    created_at TIMESTAMP WITH TIME ZONE NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    CHECK (effective_from ~ '^[0-9]{4}/[0-9]{2}/[0-9]{2}$'),
                    CHECK (effective_to IS NULL OR effective_to ~ '^[0-9]{4}/[0-9]{2}/[0-9]{2}$'),
                    CHECK (effective_to IS NULL OR effective_to >= effective_from)
                )
            """)
            cur.execute("ALTER TABLE deputy_branch_assignments ADD COLUMN IF NOT EXISTS replacement_operation_id BIGINT")
            cur.execute("""
                DO $replacement_fk$
                BEGIN
                    IF NOT EXISTS (
                        SELECT 1 FROM pg_constraint
                        WHERE conname='fk_deputy_assignment_replacement'
                          AND conrelid='deputy_branch_assignments'::regclass
                    ) THEN
                        ALTER TABLE deputy_branch_assignments
                        ADD CONSTRAINT fk_deputy_assignment_replacement
                        FOREIGN KEY (replacement_operation_id)
                        REFERENCES deputy_replacement_operations(id) ON DELETE RESTRICT;
                    END IF;
                END
                $replacement_fk$;
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS message_campaigns (
                    id BIGSERIAL PRIMARY KEY,
                    campaign_uuid VARCHAR(36) NOT NULL UNIQUE,
                    message_type VARCHAR(20) NOT NULL
                        CHECK (message_type IN ('occasion', 'announcement', 'urgent')),
                    message_text TEXT NOT NULL,
                    media_base64 TEXT,
                    media_mime VARCHAR(100),
                    media_name VARCHAR(255),
                    audience_type VARCHAR(20) NOT NULL
                        CHECK (audience_type IN ('all', 'deputies', 'admins', 'branches', 'manual')),
                    audience_config JSONB NOT NULL DEFAULT '{}'::jsonb,
                    personalize BOOLEAN NOT NULL DEFAULT FALSE,
                    status VARCHAR(20) NOT NULL DEFAULT 'draft'
                        CHECK (status IN ('draft', 'scheduled', 'sending', 'completed', 'partial', 'cancelled')),
                    scheduled_at TIMESTAMP WITH TIME ZONE,
                    created_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                    total_recipients INTEGER NOT NULL DEFAULT 0,
                    sent_count INTEGER NOT NULL DEFAULT 0,
                    failed_count INTEGER NOT NULL DEFAULT 0,
                    created_at TIMESTAMP WITH TIME ZONE NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    started_at TIMESTAMP WITH TIME ZONE,
                    completed_at TIMESTAMP WITH TIME ZONE,
                    cancelled_at TIMESTAMP WITH TIME ZONE,
                    CHECK (char_length(message_text) BETWEEN 1 AND 3500)
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS message_deliveries (
                    id BIGSERIAL PRIMARY KEY,
                    campaign_id BIGINT NOT NULL REFERENCES message_campaigns(id) ON DELETE RESTRICT,
                    user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE RESTRICT,
                    telegram_id BIGINT NOT NULL,
                    recipient_name VARCHAR(255) NOT NULL,
                    status VARCHAR(20) NOT NULL DEFAULT 'pending'
                        CHECK (status IN ('pending', 'sending', 'sent', 'failed', 'cancelled')),
                    attempts INTEGER NOT NULL DEFAULT 0,
                    last_error TEXT,
                    sent_at TIMESTAMP WITH TIME ZONE,
                    created_at TIMESTAMP WITH TIME ZONE NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    CONSTRAINT unique_campaign_recipient UNIQUE (campaign_id, user_id)
                )
            """)
            cur.execute("""
                DO $safe_index$
                BEGIN
                    IF NOT EXISTS (SELECT 1 FROM pg_indexes WHERE schemaname='public'
                                   AND indexname='idx_unique_active_target') THEN
                        IF NOT EXISTS (SELECT branch_id FROM branch_targets WHERE is_active=TRUE
                                       GROUP BY branch_id HAVING COUNT(*) > 1) THEN
                            EXECUTE 'CREATE UNIQUE INDEX idx_unique_active_target '
                                    'ON branch_targets(branch_id) WHERE is_active=TRUE';
                        ELSE
                            RAISE WARNING 'Active target duplicates exist; unique index was not created';
                        END IF;
                    END IF;
                END
                $safe_index$;
            """)
            cur.execute("CREATE INDEX IF NOT EXISTS idx_collections_branch_date ON collections(branch_id, shamsi_date);")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_collections_shamsi ON collections(shamsi_date);")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_users_telegram ON users(telegram_id);")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_users_employee ON users(employee_number);")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_users_active_role ON users(is_active, role);")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_branch_targets_branch ON branch_targets(branch_id);")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_branch_targets_active ON branch_targets(is_active);")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_collections_recorded_by ON collections(recorded_by);")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_scores_collection ON scores(collection_id);")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_actual_stats_branch_date ON actual_stats(branch_id, shamsi_date);")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_actual_stats_import_date ON actual_stats_import_operations(shamsi_date, created_at);")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_actual_stats_audit_branch_date ON actual_stats_audit_log(branch_id, shamsi_date, created_at);")
            cur.execute("""
                CREATE UNIQUE INDEX IF NOT EXISTS idx_actual_stats_audit_operation_branch
                ON actual_stats_audit_log(operation_id, branch_id)
                WHERE operation_id IS NOT NULL
            """)
            cur.execute("CREATE INDEX IF NOT EXISTS idx_deputy_assignments_deputy_dates ON deputy_branch_assignments(deputy_id, effective_from, effective_to);")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_deputy_assignments_branch_dates ON deputy_branch_assignments(branch_id, effective_from, effective_to);")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_deputy_replacements_branch_date ON deputy_replacement_operations(branch_id, effective_date);")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_message_campaigns_due ON message_campaigns(status, scheduled_at);")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_message_deliveries_campaign_status ON message_deliveries(campaign_id, status);")
            cur.execute("""
                DO $assignment_indexes$
                BEGIN
                    IF NOT EXISTS (SELECT 1 FROM pg_indexes WHERE schemaname='public' AND indexname='idx_deputy_assignment_one_active')
                       AND NOT EXISTS (SELECT deputy_id FROM deputy_branch_assignments WHERE effective_to IS NULL GROUP BY deputy_id HAVING COUNT(*)>1) THEN
                        CREATE UNIQUE INDEX idx_deputy_assignment_one_active
                        ON deputy_branch_assignments(deputy_id) WHERE effective_to IS NULL;
                    END IF;
                    IF NOT EXISTS (SELECT 1 FROM pg_indexes WHERE schemaname='public' AND indexname='idx_branch_assignment_one_active')
                       AND NOT EXISTS (SELECT branch_id FROM deputy_branch_assignments WHERE effective_to IS NULL GROUP BY branch_id HAVING COUNT(*)>1) THEN
                        CREATE UNIQUE INDEX idx_branch_assignment_one_active
                        ON deputy_branch_assignments(branch_id) WHERE effective_to IS NULL;
                    END IF;
                END
                $assignment_indexes$;
            """)
            # فقط برای معاونانی که هیچ انتصاب فعالی ندارند، وضعیت جاری بدون تغییر
            # داده‌های قبلی به‌عنوان نقطه شروع تاریخچه ثبت می‌شود.
            cur.execute("""
                INSERT INTO deputy_branch_assignments
                    (deputy_id, branch_id, effective_from, source, created_at)
                SELECT u.id, u.branch_id, '0001/01/01', 'initial_backfill', CURRENT_TIMESTAMP
                FROM users u
                WHERE u.role='deputy' AND u.is_active=TRUE AND u.branch_id IS NOT NULL
                  AND NOT EXISTS (
                      SELECT 1 FROM deputy_branch_assignments dba
                      WHERE dba.deputy_id=u.id AND dba.effective_to IS NULL
                  )
                  AND NOT EXISTS (
                      SELECT 1 FROM deputy_branch_assignments dba
                      WHERE dba.branch_id=u.branch_id AND dba.effective_to IS NULL
                  )
                ON CONFLICT DO NOTHING
            """)
            conn.commit()
            logger.info("✅ All tables and indexes created/verified successfully.")
    except Exception as e:
        logger.error(f"❌ Error creating tables: {e}")
        if conn:
            conn.rollback()
    finally:
        if conn:
            return_db_connection(conn)

create_all_tables_if_not_exists()

# ============================================================
# توابع مدیریت تنظیمات
# ============================================================
def get_bot_status_db():
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("SELECT value FROM settings WHERE key = 'bot_status'")
            result = cur.fetchone()
            if result:
                return result[0] == 'active'
            return True
    except Exception as e:
        logger.error(f"get_bot_status_db: {e}")
        if conn:
            conn.rollback()
        return True
    finally:
        if conn:
            return_db_connection(conn)

def get_cached_bot_status():
    cached = cache_bot_status.get('bot_status')
    if cached is not None:
        return cached
    value = get_bot_status_db()
    cache_bot_status.set('bot_status', value)
    return value

def get_bot_status():
    return get_cached_bot_status()

def set_bot_status(status):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO settings (key, value, updated_at)
                VALUES ('bot_status', %s, %s)
                ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value, updated_at = EXCLUDED.updated_at
            """, ('active' if status else 'inactive', get_iran_time()))
            conn.commit()
            cache_bot_status.set('bot_status', status)
            return True
    except Exception as e:
        logger.error(f"set_bot_status: {e}")
        if conn:
            conn.rollback()
        return False
    finally:
        if conn:
            return_db_connection(conn)

def get_feature_setting(key, default='active'):
    cache_key = f'feature_{key}'
    cached = cache_feature_settings.get(cache_key)
    if cached is not None:
        return cached
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("SELECT value FROM feature_settings WHERE key = %s", (key,))
            result = cur.fetchone()
            value = result[0] if result else default
            cache_feature_settings.set(cache_key, value)
            return value
    except Exception as e:
        logger.error(f"get_feature_setting error: {e}")
        if conn:
            conn.rollback()
        return default
    finally:
        if conn:
            return_db_connection(conn)

def set_feature_setting(key, value):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO feature_settings (key, value, updated_at)
                VALUES (%s, %s, %s)
                ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value, updated_at = EXCLUDED.updated_at
            """, (key, value, get_iran_time()))
            conn.commit()
            cache_feature_settings.invalidate(f'feature_{key}')
            return True
    except Exception as e:
        logger.error(f"set_feature_setting error: {e}")
        if conn:
            conn.rollback()
        return False
    finally:
        if conn:
            return_db_connection(conn)

def get_auto_reminder_status():
    return get_feature_setting('auto_reminder', 'active') == 'active'
def set_auto_reminder_status(status):
    return set_feature_setting('auto_reminder', 'active' if status else 'inactive')
def get_auto_report_status():
    return get_feature_setting('auto_report', 'active') == 'active'
def set_auto_report_status(status):
    return set_feature_setting('auto_report', 'active' if status else 'inactive')
def get_auto_alert_status():
    return get_feature_setting('auto_alert', 'active') == 'active'
def set_auto_alert_status(status):
    return set_feature_setting('auto_alert', 'active' if status else 'inactive')
def get_auto_scoring_status():
    return get_feature_setting('auto_scoring', 'active') == 'active'
def set_auto_scoring_status(status):
    return set_feature_setting('auto_scoring', 'active' if status else 'inactive')
def get_weekly_report_status():
    return get_feature_setting('weekly_report', 'active') == 'active'
def set_weekly_report_status(status):
    return set_feature_setting('weekly_report', 'active' if status else 'inactive')
def get_monthly_report_status():
    return get_feature_setting('monthly_report', 'active') == 'active'
def set_monthly_report_status(status):
    return set_feature_setting('monthly_report', 'active' if status else 'inactive')
def get_instant_notification_status():
    return get_feature_setting('instant_notification', 'active') == 'active'
def set_instant_notification_status(status):
    return set_feature_setting('instant_notification', 'active' if status else 'inactive')
def get_adaptive_report_status():
    return get_feature_setting('adaptive_report', 'active') == 'active'
def set_adaptive_report_status(status):
    return set_feature_setting('adaptive_report', 'active' if status else 'inactive')
def get_forecast_report_status():
    return get_feature_setting('forecast_report', 'active') == 'active'
def set_forecast_report_status(status):
    return set_feature_setting('forecast_report', 'active' if status else 'inactive')
def get_chart_report_status():
    return get_feature_setting('chart_report', 'active') == 'active'
def set_chart_report_status(status):
    return set_feature_setting('chart_report', 'active' if status else 'inactive')
def get_actual_stats_status():
    return get_feature_setting('actual_stats', 'active') == 'active'
def set_actual_stats_status(status):
    return set_feature_setting('actual_stats', 'active' if status else 'inactive')

def is_holiday(shamsi_date=None):
    if not shamsi_date:
        shamsi_date = get_shamsi_date()
    shamsi_date = normalize_digits(shamsi_date)
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) FROM holidays WHERE shamsi_date = %s", (shamsi_date,))
            count = cur.fetchone()[0]
            return count > 0
    except Exception as e:
        logger.error(f"is_holiday error: {e}")
        if conn:
            conn.rollback()
        return False
    finally:
        if conn:
            return_db_connection(conn)

def add_holiday(shamsi_date, description=""):
    shamsi_date = normalize_digits(shamsi_date)
    if not validate_shamsi_date(shamsi_date):
        return False, "تاریخ نامعتبر است"
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO holidays (shamsi_date, description, created_at)
                VALUES (%s, %s, %s)
                ON CONFLICT (shamsi_date) DO NOTHING
                RETURNING id
            """, (shamsi_date, description, get_iran_time()))
            result = cur.fetchone()
            conn.commit()
            if result:
                return True, "ثبت شد"
            return False, "این تاریخ قبلاً ثبت شده است"
    except Exception as e:
        logger.error(f"add_holiday error: {e}")
        if conn:
            conn.rollback()
        return False, f"خطا: {e}"
    finally:
        if conn:
            return_db_connection(conn)

def remove_holiday(shamsi_date):
    shamsi_date = normalize_digits(shamsi_date)
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("DELETE FROM holidays WHERE shamsi_date = %s RETURNING id", (shamsi_date,))
            result = cur.fetchone()
            conn.commit()
            return result is not None
    except Exception as e:
        logger.error(f"remove_holiday error: {e}")
        if conn:
            conn.rollback()
        return False
    finally:
        if conn:
            return_db_connection(conn)

def get_all_holidays(limit=30):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT shamsi_date, description, created_at
                FROM holidays
                ORDER BY shamsi_date DESC
                LIMIT %s
            """, (limit,))
            return cur.fetchall()
    except Exception as e:
        logger.error(f"get_all_holidays error: {e}")
        if conn:
            conn.rollback()
        return []
    finally:
        if conn:
            return_db_connection(conn)

# ============================================================
# توابع امتیازدهی (با زمان ۱۶:۳۰)
# ============================================================
def is_on_time(created_at):
    """بررسی اینکه ثبت در زمان ۱۶:۳۰ یا قبل بوده است"""
    iran_time = created_at.astimezone(IRAN_TZ)
    hour = iran_time.hour
    minute = iran_time.minute
    if hour < SCORE_DEADLINE_HOUR:
        return True
    if hour == SCORE_DEADLINE_HOUR and minute <= SCORE_DEADLINE_MINUTE:
        return True
    return False

def calculate_score(collection_time, deputy_amount, others_amount, branch_id, shamsi_date):
    total_amount = deputy_amount + others_amount
    score = 0
    # امتیاز زمان با مهلت ۱۶:۳۰
    if is_on_time(collection_time):
        score += 2
    else:
        score += 1
    # امتیاز نسبت به میانگین ماهانه
    monthly_avg = get_branch_monthly_avg(branch_id, 30)
    if monthly_avg > 0:
        ratio = total_amount / monthly_avg
        if ratio >= 1.5:
            score += 3
        elif ratio >= 1.2:
            score += 2
        elif ratio >= 0.8:
            score += 1
    else:
        if total_amount >= 5_000_000_000:
            score += 3
        elif total_amount >= 2_000_000_000:
            score += 2
        elif total_amount >= 500_000_000:
            score += 1
    if total_amount > 0 and (deputy_amount / total_amount) > 0.5:
        score += 1
    consecutive_days = get_consecutive_days(branch_id, shamsi_date)
    if consecutive_days >= 7:
        score += 2
    elif consecutive_days >= 3:
        score += 1
    return score

def get_consecutive_days(branch_id, shamsi_date):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT shamsi_date
                FROM collections
                WHERE branch_id = %s
                ORDER BY shamsi_date DESC
                LIMIT 30
            """, (branch_id,))
            dates = [row[0] for row in cur.fetchall()]
            if not dates:
                return 0
            target_date = jdatetime.date(*map(int, shamsi_date.split('/')))
            earliest = add_days_to_shamsi(target_date, -45)
            cur.execute("SELECT shamsi_date FROM holidays WHERE shamsi_date BETWEEN %s AND %s",
                        (_shamsi_string(earliest), shamsi_date))
            holidays = {row[0] for row in cur.fetchall()}
            count = 0
            checked_workdays = 0
            offset = 1
            while checked_workdays < 29 and offset <= 45:
                check_date = add_days_to_shamsi(target_date, -offset)
                offset += 1
                check_str = f"{check_date.year}/{check_date.month:02d}/{check_date.day:02d}"
                if check_date.togregorian().weekday() == 4 or check_str in holidays:
                    continue
                checked_workdays += 1
                if check_str in dates:
                    count += 1
                else:
                    break
            return count
    except Exception as e:
        logger.error(f"get_consecutive_days error: {e}")
        if conn:
            conn.rollback()
        return 0
    finally:
        if conn:
            return_db_connection(conn)

def save_score(collection_id, score):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO scores (collection_id, score, created_at)
                VALUES (%s, %s, %s)
                ON CONFLICT (collection_id) DO UPDATE SET score = EXCLUDED.score, updated_at = EXCLUDED.created_at
            """, (collection_id, score, get_iran_time()))
            conn.commit()
            return True
    except Exception as e:
        logger.error(f"save_score error: {e}")
        if conn:
            conn.rollback()
        return False
    finally:
        if conn:
            return_db_connection(conn)

def delete_score(collection_id):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("DELETE FROM scores WHERE collection_id = %s", (collection_id,))
            conn.commit()
            return True
    except Exception as e:
        logger.error(f"delete_score error: {e}")
        if conn:
            conn.rollback()
        return False
    finally:
        if conn:
            return_db_connection(conn)

def get_branch_total_score(branch_id, days=30):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT SUM(s.score)
                FROM scores s
                JOIN collections c ON s.collection_id = c.id
                WHERE c.branch_id = %s
                AND c.shamsi_date >= %s
            """, (branch_id, get_shamsi_date(-days)))
            result = cur.fetchone()[0]
            return result or 0
    except Exception as e:
        logger.error(f"get_branch_total_score: {e}")
        if conn:
            conn.rollback()
        return 0
    finally:
        if conn:
            return_db_connection(conn)

def get_all_branch_scores(days=30):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT b.id, b.name, COALESCE(SUM(s.score), 0) as total_score, COUNT(s.id) as score_count
                FROM branches b
                LEFT JOIN collections c ON b.id = c.branch_id AND c.shamsi_date >= %s
                LEFT JOIN scores s ON c.id = s.collection_id
                GROUP BY b.id, b.name
                ORDER BY total_score DESC
            """, (get_shamsi_date(-days),))
            return cur.fetchall()
    except Exception as e:
        logger.error(f"get_all_branch_scores: {e}")
        if conn:
            conn.rollback()
        return []
    finally:
        if conn:
            return_db_connection(conn)

# ============================================================
# توابع مشکلات
# ============================================================
def save_problem(user_id, problem_text, category="general"):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO problems (user_id, problem_text, category, status, created_at)
                VALUES (%s, %s, %s, %s, %s)
            """, (user_id, problem_text, category, 'pending', get_iran_time()))
            conn.commit()
            return True
    except Exception as e:
        logger.error(f"save_problem error: {e}")
        if conn:
            conn.rollback()
        return False
    finally:
        if conn:
            return_db_connection(conn)

def get_all_problems(status=None, limit=50):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            if status:
                cur.execute("""
                    SELECT p.id, u.full_name, u.employee_number, p.problem_text, p.category, p.status,
                           p.created_at AT TIME ZONE 'Asia/Tehran' as created_at_iran
                    FROM problems p
                    JOIN users u ON p.user_id = u.id
                    WHERE p.status = %s
                    ORDER BY p.created_at DESC
                    LIMIT %s
                """, (status, limit))
            else:
                cur.execute("""
                    SELECT p.id, u.full_name, u.employee_number, p.problem_text, p.category, p.status,
                           p.created_at AT TIME ZONE 'Asia/Tehran' as created_at_iran
                    FROM problems p
                    JOIN users u ON p.user_id = u.id
                    ORDER BY p.created_at DESC
                    LIMIT %s
                """, (limit,))
            return cur.fetchall()
    except Exception as e:
        logger.error(f"get_all_problems: {e}")
        if conn:
            conn.rollback()
        return []
    finally:
        if conn:
            return_db_connection(conn)

def update_problem_status(problem_id, new_status):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                UPDATE problems
                SET status = %s, updated_at = %s
                WHERE id = %s
            """, (new_status, get_iran_time(), problem_id))
            conn.commit()
            return True
    except Exception as e:
        logger.error(f"update_problem_status: {e}")
        if conn:
            conn.rollback()
        return False
    finally:
        if conn:
            return_db_connection(conn)

# ============================================================
# توابع دیتابیس پایه
# ============================================================
def get_all_branches():
    cached = cache_branches.get('branches')
    if cached is not None:
        return cached
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("SELECT id, name FROM branches ORDER BY name")
            result = cur.fetchall()
            cache_branches.set('branches', result)
            return result
    except Exception as e:
        logger.error(f"get_all_branches: {e}")
        if conn:
            conn.rollback()
        return []
    finally:
        if conn:
            return_db_connection(conn)

def get_branches_with_official_codes():
    """Return the authoritative branch-code map used by organization Excel files."""
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT id, name, official_code
                FROM branches
                ORDER BY name
            """)
            return cur.fetchall()
    except Exception as e:
        logger.error("get_branches_with_official_codes: %s", safe_log_value(e))
        if conn:
            conn.rollback()
        return []
    finally:
        if conn:
            return_db_connection(conn)

def find_user_by_employee_number(emp_num, with_status=False):
    emp_num = normalize_digits(emp_num)
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT u.id, u.employee_number, u.full_name, u.role, u.title, u.branch_id, b.name, u.is_super_admin
                FROM users u
                LEFT JOIN branches b ON u.branch_id = b.id
                WHERE u.employee_number = %s AND u.is_active = TRUE
            """, (emp_num,))
            user = cur.fetchone()
            return (user, True) if with_status else user
    except Exception as e:
        logger.exception("find_user_by_employee_number failed: %s", safe_log_value(e))
        if conn:
            conn.rollback()
        return (None, False) if with_status else None
    finally:
        if conn:
            return_db_connection(conn)

def bind_user_telegram_id(user_db_id, chat_id):
    """اتصال اتمیک یک حساب فعال به گفت‌وگوی بله و بازخوانی نتیجه قطعی.

    شماره کارمندی همچنان روش احراز هویت کاربران عادی است. اگر همین گفت‌وگو
    قبلاً به حساب دیگری متصل بوده باشد، فقط شناسه ورود آن حساب آزاد می‌شود؛
    هیچ کاربر، وصول، هدف، یادداشت یا سابقه‌ای حذف یا جابه‌جا نمی‌شود.
    """
    try:
        user_db_id = int(user_db_id)
        chat_id = int(chat_id)
    except (TypeError, ValueError, OverflowError):
        logger.warning("Invalid identifiers supplied for account binding")
        return False, None, "شناسه حساب یا گفت‌وگو معتبر نیست.", {}
    if user_db_id <= 0 or not -(2 ** 63) <= chat_id < 2 ** 63:
        logger.warning("Out-of-range identifiers supplied for account binding")
        return False, None, "شناسه حساب یا گفت‌وگو معتبر نیست.", {}

    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            # یک قفل مشترک کوتاه، ترتیب قفل‌گذاری ورودهای هم‌زمان را قطعی می‌کند.
            cur.execute("SELECT pg_advisory_xact_lock(%s)", (AUTH_BIND_ADVISORY_LOCK_KEY,))
            cur.execute("""
                SELECT u.id, u.employee_number, u.full_name, u.role, u.title,
                       u.branch_id, b.name, u.is_super_admin, u.telegram_id
                FROM users u
                LEFT JOIN branches b ON b.id = u.branch_id
                WHERE u.id = %s AND u.is_active = TRUE
                FOR UPDATE OF u
            """, (user_db_id,))
            target = cur.fetchone()
            if not target:
                conn.rollback()
                return False, None, "حساب کاربری فعال نیست یا دیگر وجود ندارد.", {}

            previous_chat_id = target[8]
            cur.execute("""
                SELECT id, is_active
                FROM users
                WHERE telegram_id = %s AND id <> %s
                FOR UPDATE
            """, (chat_id, user_db_id))
            released_accounts = cur.fetchall()
            now = get_iran_time()
            if released_accounts:
                cur.execute("""
                    UPDATE users
                    SET telegram_id = NULL, updated_at = %s
                    WHERE telegram_id = %s AND id <> %s
                """, (now, chat_id, user_db_id))
                if cur.rowcount != len(released_accounts):
                    raise RuntimeError("تعداد حساب‌های آزادشده با نتیجه قفل‌شده هماهنگ نیست")

            cur.execute("""
                UPDATE users
                SET telegram_id = %s, updated_at = %s
                WHERE id = %s AND is_active = TRUE
            """, (chat_id, now, user_db_id))
            if cur.rowcount != 1:
                raise RuntimeError("اتصال حساب فعال ثبت نشد")

            # پیام ورود موفق فقط پس از بازخوانی نتیجه از خود دیتابیس صادر می‌شود.
            cur.execute("""
                SELECT u.id, u.employee_number, u.full_name, u.role, u.title,
                       u.branch_id, b.name, u.is_super_admin
                FROM users u
                LEFT JOIN branches b ON b.id = u.branch_id
                WHERE u.id = %s AND u.telegram_id = %s AND u.is_active = TRUE
            """, (user_db_id, chat_id))
            verified_user = cur.fetchone()
            if not verified_user:
                raise RuntimeError("بازخوانی اتصال حساب پس از ثبت ناموفق بود")
            conn.commit()
            metadata = {
                "previous_chat_id": previous_chat_id,
                "released_account_ids": [row[0] for row in released_accounts],
                "released_active_count": sum(1 for row in released_accounts if row[1]),
            }
            return True, verified_user, "اتصال حساب با موفقیت ثبت شد.", metadata
    except Exception as e:
        logger.exception("bind_user_telegram_id failed: %s", safe_log_value(e))
        if conn:
            conn.rollback()
        return False, None, "اتصال حساب به بله ثبت نشد؛ لطفاً دوباره تلاش کنید.", {}
    finally:
        if conn:
            return_db_connection(conn)

def update_user_telegram_id(user_db_id, chat_id):
    """پوسته سازگار با کدهای قدیمی؛ موفقیت واقعی را به‌صورت boolean برمی‌گرداند."""
    success, _, _, _ = bind_user_telegram_id(user_db_id, chat_id)
    return success

def find_user_by_telegram_id(chat_id, with_status=False):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT u.id, u.employee_number, u.full_name, u.role, u.title, u.branch_id, b.name, u.is_super_admin
                FROM users u
                LEFT JOIN branches b ON u.branch_id = b.id
                WHERE u.telegram_id = %s AND u.is_active = TRUE
            """, (chat_id,))
            user = cur.fetchone()
            return (user, True) if with_status else user
    except Exception as e:
        logger.exception("find_user_by_telegram_id failed: %s", safe_log_value(e))
        if conn:
            conn.rollback()
        return (None, False) if with_status else None
    finally:
        if conn:
            return_db_connection(conn)

def log_user_activity(user_id, action, details=""):
    safe_action = str(action).replace('\r', ' ').replace('\n', ' ')[:50]
    safe_details = str(details).replace('\r', ' ').replace('\n', ' ')[:4000]
    if action in ["collection_add", "login", "logout"]:
        logger.info("User %s %s", user_id, safe_action)
    else:
        logger.debug("User %s %s", user_id, safe_action)
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO user_activity_log (user_id, action, details, created_at)
                VALUES (%s, %s, %s, %s)
            """, (user_id, safe_action, safe_details, get_iran_time()))
            conn.commit()
    except Exception as e:
        logger.error(f"log_user_activity: {e}")
        if conn:
            conn.rollback()
    finally:
        if conn:
            return_db_connection(conn)

def get_user_activity_log(limit=100):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT l.id, u.full_name, u.employee_number, l.action, l.details,
                       l.created_at AT TIME ZONE 'Asia/Tehran' as created_at_iran
                FROM user_activity_log l
                JOIN users u ON l.user_id = u.id
                ORDER BY l.created_at DESC
                LIMIT %s
            """, (limit,))
            return cur.fetchall()
    except Exception as e:
        logger.error(f"get_user_activity_log: {e}")
        if conn:
            conn.rollback()
        return []
    finally:
        if conn:
            return_db_connection(conn)

def save_note(collection_id, user_id, note_text):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO notes (collection_id, user_id, note_text, created_at)
                VALUES (%s, %s, %s, %s)
            """, (collection_id, user_id, note_text, get_iran_time()))
            conn.commit()
            return True
    except Exception as e:
        logger.error(f"save_note: {e}")
        if conn:
            conn.rollback()
        return False
    finally:
        if conn:
            return_db_connection(conn)

def get_notes_for_collection(collection_id):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT n.id, u.full_name, n.note_text,
                       n.created_at AT TIME ZONE 'Asia/Tehran' as created_at_iran
                FROM notes n
                JOIN users u ON n.user_id = u.id
                WHERE n.collection_id = %s
                ORDER BY n.created_at DESC
            """, (collection_id,))
            return cur.fetchall()
    except Exception as e:
        logger.error(f"get_notes_for_collection: {e}")
        if conn:
            conn.rollback()
        return []
    finally:
        if conn:
            return_db_connection(conn)

def get_all_notes_with_collection(limit=50):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT n.id, b.name, c.shamsi_date, u.full_name, n.note_text,
                       n.created_at AT TIME ZONE 'Asia/Tehran' as created_at_iran
                FROM notes n
                JOIN collections c ON n.collection_id = c.id
                JOIN branches b ON c.branch_id = b.id
                JOIN users u ON n.user_id = u.id
                ORDER BY n.created_at DESC
                LIMIT %s
            """, (limit,))
            return cur.fetchall()
    except Exception as e:
        logger.error(f"get_all_notes_with_collection: {e}")
        if conn:
            conn.rollback()
        return []
    finally:
        if conn:
            return_db_connection(conn)

def check_existing_collection(branch_id, shamsi_date):
    shamsi_date = normalize_digits(shamsi_date)
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT id, deputy_amount, others_amount, created_at
                FROM collections
                WHERE branch_id = %s AND shamsi_date = %s
            """, (branch_id, shamsi_date))
            result = cur.fetchone()
            if result:
                return {
                    'id': result[0],
                    'deputy_amount': result[1],
                    'others_amount': result[2],
                    'created_at': result[3]
                }
            return None
    except Exception as e:
        logger.error(f"check_existing_collection: {e}")
        if conn:
            conn.rollback()
        return None
    finally:
        if conn:
            return_db_connection(conn)

def can_edit_collection(collection_created_at):
    """بررسی اینکه آیا معاون می‌تواند این وصول را ویرایش کند (تا ۱۲ شب همان روز)"""
    now = get_iran_time()
    if collection_created_at.tzinfo is None:
        collection_created_at = collection_created_at.replace(tzinfo=timezone.utc)
    created = collection_created_at.astimezone(IRAN_TZ)
    if created.date() != now.date():
        return False
    if now.hour > EDIT_DEADLINE_HOUR or (now.hour == EDIT_DEADLINE_HOUR and now.minute > EDIT_DEADLINE_MINUTE):
        return False
    return True

# ============================================================
# ذخیره/بروزرسانی وصول با مدیریت امتیاز و محدودیت ویرایش
# ============================================================
def save_or_update_collection_with_note(branch_id, deputy_amount_millions, others_amount_millions, shamsi_date, user_id, note_text=None, update_existing=False):
    conn = None
    created_at_iran = get_iran_time()
    if not (0 <= deputy_amount_millions <= MAX_AMOUNT_MILLIONS and
            0 <= others_amount_millions <= MAX_AMOUNT_MILLIONS):
        return False, None
    deputy_amount = deputy_amount_millions * 1_000_000
    others_amount = others_amount_millions * 1_000_000
    collection_id = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            if update_existing:
                # ابتدا بررسی محدودیت ویرایش
                cur.execute("SELECT created_at FROM collections WHERE branch_id = %s AND shamsi_date = %s", (branch_id, shamsi_date))
                existing = cur.fetchone()
                if not existing:
                    return False, None
                if not can_edit_collection(existing[0]):
                    return False, None  # یا پیام خطا
                cur.execute("""
                    UPDATE collections
                    SET deputy_amount = %s, others_amount = %s, recorded_by = %s, updated_at = %s
                    WHERE branch_id = %s AND shamsi_date = %s
                    RETURNING id
                """, (deputy_amount, others_amount, user_id, created_at_iran, branch_id, shamsi_date))
                result = cur.fetchone()
                if result:
                    collection_id = result[0]
                    cur.execute("DELETE FROM scores WHERE collection_id = %s", (collection_id,))
                else:
                    return False, None
            else:
                cur.execute("""
                    INSERT INTO collections (branch_id, deputy_amount, others_amount, shamsi_date, recorded_by, created_at)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    ON CONFLICT (branch_id, shamsi_date) DO NOTHING
                    RETURNING id
                """, (branch_id, deputy_amount, others_amount, shamsi_date, user_id, created_at_iran))
                result = cur.fetchone()
                collection_id = result[0] if result else None
                if collection_id is None:
                    conn.rollback()
                    return False, None
            if note_text and collection_id:
                cur.execute("""
                    INSERT INTO notes (collection_id, user_id, note_text, created_at)
                    VALUES (%s, %s, %s, %s)
                """, (collection_id, user_id, note_text, created_at_iran))
            conn.commit()
            # پاکسازی کش‌ها
            cache_today_report.invalidate_all()
            cache_top_branches.invalidate('top5')
            cache_10day_report.invalidate('10day')
            cache_adaptive.invalidate('adaptive')
            cache_forecast_all.invalidate('forecast_all')
            invalidate_branches_cache()
            cache_targets.invalidate(f'target_{branch_id}')
            cache_targets.invalidate_all()
            if collection_id and get_instant_notification_status() and not is_holiday(shamsi_date):
                try:
                    executor.submit(
                        send_instant_notification_async,
                        branch_id, deputy_amount_millions, others_amount_millions, shamsi_date, user_id
                    )
                except Exception as e:
                    logger.warning(f"Failed to submit notification task: {e}")
            return True, collection_id
    except Exception as e:
        logger.error(f"save_or_update_collection_with_note: {e}")
        if conn:
            conn.rollback()
        return False, None
    finally:
        if conn:
            return_db_connection(conn)

def send_instant_notification_async(branch_id, deputy_amount, others_amount, shamsi_date, user_id):
    if not get_instant_notification_status():
        return
    if is_holiday(shamsi_date):
        return
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT telegram_id FROM users
                WHERE (role IN ('admin', 'super_admin') OR is_super_admin = TRUE)
                AND telegram_id IS NOT NULL
            """)
            admins = cur.fetchall()
            if not admins:
                return
            cur.execute("SELECT name FROM branches WHERE id = %s", (branch_id,))
            branch_name = cur.fetchone()[0]
            cur.execute("SELECT full_name FROM users WHERE id = %s", (user_id,))
            user_name = cur.fetchone()[0]
            total = deputy_amount + others_amount
            msg = f"🔔 **ثبت وصول جدید**\n━━━━━━━━━━━━━━━━━━\n"
            msg += f"🏢 شعبه: {branch_name}\n"
            msg += f"👤 ثبت‌کننده: {user_name}\n"
            msg += f"📅 تاریخ: {get_shamsi_date_formatted(shamsi_date)}\n"
            msg += f"👤 وصولی معاون: {deputy_amount:,.0f} میلیون ریال\n"
            msg += f"👥 وصولی همکاران: {others_amount:,.0f} میلیون ریال\n"
            msg += f"💰 جمع کل: {total:,.0f} میلیون ریال\n"
            msg += f"⏰ زمان: {get_iran_time().strftime('%H:%M:%S')}"
            for admin in admins:
                chat_id = admin[0]
                if chat_id:
                    send_message(chat_id, msg)
    except Exception as e:
        logger.error(f"send_instant_notification_async error: {e}")
    finally:
        if conn:
            return_db_connection(conn)

# ============================================================
# توابع گزارش‌گیری
# ============================================================
def get_today_province_report(shamsi_date):
    cache_key = f'today_report_{shamsi_date}'
    cached = cache_today_report.get(cache_key)
    if cached is not None:
        return cached
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            # بهینه‌سازی با JOIN به جای subquery
            cur.execute("""
                SELECT
                    b.id,
                    b.name,
                    COALESCE(c.deputy_amount, 0) as deputy_amount,
                    COALESCE(c.others_amount, 0) as others_amount,
                    COALESCE(c.total_amount, 0) as total_amount,
                    bt.target_amount,
                    bt.target_date,
                    bt.created_at as target_created_at,
                    0::BIGINT as collected_since_target
                FROM branches b
                LEFT JOIN collections c ON c.branch_id = b.id AND c.shamsi_date = %s
                LEFT JOIN branch_targets bt ON b.id = bt.branch_id AND bt.is_active = TRUE
                ORDER BY COALESCE(c.total_amount, 0) DESC
            """, (shamsi_date,))
            raw_result = cur.fetchall()
            target_starts = {}
            for row in raw_result:
                if row[5] is not None and row[7] is not None:
                    target_start = jdatetime.datetime.fromgregorian(datetime=row[7])
                    target_starts[row[0]] = f"{target_start.year}/{target_start.month:02d}/{target_start.day:02d}"
            collected_by_branch = {branch_id: 0 for branch_id in target_starts}
            if target_starts:
                cur.execute("""SELECT branch_id, shamsi_date, total_amount
                               FROM collections
                               WHERE branch_id = ANY(%s) AND shamsi_date <= %s""",
                            (list(target_starts), shamsi_date))
                for collection_branch, collection_date, amount in cur.fetchall():
                    if collection_date >= target_starts[collection_branch]:
                        collected_by_branch[collection_branch] += int(amount or 0)
            result = []
            for row in raw_result:
                row = list(row)
                row[8] = collected_by_branch.get(row[0], 0)
                result.append(tuple(row))
            cache_today_report.set(cache_key, result)
            return result
    except Exception as e:
        logger.error(f"get_today_province_report: {e}")
        if conn:
            conn.rollback()
        return []
    finally:
        if conn:
            return_db_connection(conn)

def get_province_10_day_report():
    cached = cache_10day_report.get('10day')
    if cached is not None:
        return cached
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT shamsi_date, SUM(deputy_amount), SUM(others_amount), SUM(total_amount)
                FROM collections
                GROUP BY shamsi_date
                ORDER BY shamsi_date DESC
                LIMIT 10
            """)
            result = cur.fetchall()
            cache_10day_report.set('10day', result)
            return result
    except Exception as e:
        logger.error(f"get_province_10_day_report: {e}")
        if conn:
            conn.rollback()
        return []
    finally:
        if conn:
            return_db_connection(conn)

def get_top_5_branches():
    cached = cache_top_branches.get('top5')
    if cached is not None:
        return cached
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT b.name, SUM(c.total_amount) as total, COUNT(*) as record_count
                FROM collections c
                JOIN branches b ON c.branch_id = b.id
                GROUP BY b.name
                ORDER BY total DESC
                LIMIT 5
            """)
            result = cur.fetchall()
            cache_top_branches.set('top5', result)
            return result
    except Exception as e:
        logger.error(f"get_top_5_branches: {e}")
        if conn:
            conn.rollback()
        return []
    finally:
        if conn:
            return_db_connection(conn)

def get_today_statistics():
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            shamsi_today = get_shamsi_date()
            cur.execute("""
                SELECT COUNT(DISTINCT branch_id), SUM(deputy_amount), SUM(others_amount), SUM(total_amount)
                FROM collections
                WHERE shamsi_date = %s
            """, (shamsi_today,))
            result = cur.fetchone()
            if result and result[0] is not None:
                return result
            return (0, 0, 0, 0)
    except Exception as e:
        logger.error(f"get_today_statistics: {e}")
        if conn:
            conn.rollback()
        return (0, 0, 0, 0)
    finally:
        if conn:
            return_db_connection(conn)

def get_detailed_report(shamsi_date):
    shamsi_date = normalize_digits(shamsi_date)
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT b.name, c.deputy_amount, c.others_amount, c.total_amount, u.full_name
                FROM collections c
                JOIN branches b ON c.branch_id = b.id
                JOIN users u ON c.recorded_by = u.id
                WHERE c.shamsi_date = %s
                ORDER BY c.total_amount DESC
            """, (shamsi_date,))
            return cur.fetchall()
    except Exception as e:
        logger.error(f"get_detailed_report: {e}")
        if conn:
            conn.rollback()
        return []
    finally:
        if conn:
            return_db_connection(conn)

def get_branch_performance(branch_id, days=10):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT shamsi_date, daily_total,
                       AVG(daily_total) OVER (ORDER BY shamsi_date ROWS BETWEEN 2 PRECEDING AND CURRENT ROW) as avg_3day
                FROM (
                    SELECT shamsi_date, SUM(total_amount) as daily_total
                    FROM collections
                    WHERE branch_id = %s
                    GROUP BY shamsi_date
                ) t
                ORDER BY shamsi_date DESC
                LIMIT %s
            """, (branch_id, days))
            return cur.fetchall()
    except Exception as e:
        logger.error(f"get_branch_performance error: {e}")
        if conn:
            conn.rollback()
        return []
    finally:
        if conn:
            return_db_connection(conn)

def get_branch_10_day_report(branch_id):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT shamsi_date, deputy_amount, others_amount, total_amount
                FROM collections
                WHERE branch_id = %s
                ORDER BY shamsi_date DESC
                LIMIT 10
            """, (branch_id,))
            return cur.fetchall()
    except Exception as e:
        logger.error(f"get_branch_10_day_report error: {e}")
        if conn:
            conn.rollback()
        return []
    finally:
        if conn:
            return_db_connection(conn)

def get_daily_comparison():
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    shamsi_date,
                    COUNT(DISTINCT branch_id) as branches_count,
                    SUM(total_amount) as total_collection
                FROM collections
                GROUP BY shamsi_date
                ORDER BY shamsi_date DESC
                LIMIT 7
            """)
            return cur.fetchall()
    except Exception as e:
        logger.error(f"get_daily_comparison: {e}")
        if conn:
            conn.rollback()
        return []
    finally:
        if conn:
            return_db_connection(conn)

def get_deputy_vs_others_ratio():
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    SUM(deputy_amount) as deputy_total,
                    SUM(others_amount) as others_total,
                    ROUND(100.0 * SUM(deputy_amount) / NULLIF(SUM(deputy_amount) + SUM(others_amount), 0), 2) as deputy_percentage
                FROM collections
                WHERE shamsi_date = %s
            """, (get_shamsi_date(),))
            return cur.fetchone()
    except Exception as e:
        logger.error(f"get_deputy_vs_others_ratio: {e}")
        if conn:
            conn.rollback()
        return None
    finally:
        if conn:
            return_db_connection(conn)

def get_report_by_date(shamsi_date):
    shamsi_date = normalize_digits(shamsi_date)
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT b.name, c.deputy_amount, c.others_amount, c.total_amount, u.full_name
                FROM collections c
                JOIN branches b ON c.branch_id = b.id
                JOIN users u ON c.recorded_by = u.id
                WHERE c.shamsi_date = %s
                ORDER BY c.total_amount DESC
            """, (shamsi_date,))
            return cur.fetchall()
    except Exception as e:
        logger.error(f"get_report_by_date: {e}")
        if conn:
            conn.rollback()
        return []
    finally:
        if conn:
            return_db_connection(conn)

def get_branch_report_by_date(branch_id, shamsi_date):
    shamsi_date = normalize_digits(shamsi_date)
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT deputy_amount, others_amount, total_amount
                FROM collections
                WHERE branch_id = %s AND shamsi_date = %s
            """, (branch_id, shamsi_date))
            return cur.fetchone()
    except Exception as e:
        logger.error(f"get_branch_report_by_date: {e}")
        if conn:
            conn.rollback()
        return None
    finally:
        if conn:
            return_db_connection(conn)

def get_branch_full_history(branch_id):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT shamsi_date, deputy_amount, others_amount, total_amount
                FROM collections
                WHERE branch_id = %s
                ORDER BY shamsi_date DESC
            """, (branch_id,))
            return cur.fetchall()
    except Exception as e:
        logger.error(f"get_branch_full_history: {e}")
        if conn:
            conn.rollback()
        return []
    finally:
        if conn:
            return_db_connection(conn)

def get_best_worst_days(limit=5):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT shamsi_date, SUM(total_amount) as total
                FROM collections
                GROUP BY shamsi_date
                ORDER BY total DESC
                LIMIT %s
            """, (limit,))
            best = cur.fetchall()
            cur.execute("""
                SELECT shamsi_date, SUM(total_amount) as total
                FROM collections
                GROUP BY shamsi_date
                ORDER BY total ASC
                LIMIT %s
            """, (limit,))
            worst = cur.fetchall()
            return best, worst
    except Exception as e:
        logger.error(f"get_best_worst_days: {e}")
        if conn:
            conn.rollback()
        return [], []
    finally:
        if conn:
            return_db_connection(conn)

def get_all_users():
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT id, employee_number, full_name, role, title, branch_id, is_super_admin, is_active
                FROM users
                ORDER BY full_name
            """)
            return cur.fetchall()
    except Exception as e:
        logger.error(f"get_all_users: {e}")
        if conn:
            conn.rollback()
        return []
    finally:
        if conn:
            return_db_connection(conn)

def update_user_role(user_id, new_role):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("SELECT role FROM users WHERE id=%s", (user_id,))
            current = cur.fetchone()
            if not current:
                return False
            if current[0] != new_role and ('deputy' in (current[0], new_role)):
                logger.warning("Role conversion to/from deputy blocked to protect assignment history; user_id=%s", user_id)
                return False
            is_super = new_role == 'super_admin'
            cur.execute("""
                UPDATE users
                SET role = %s, is_super_admin = %s, updated_at = %s
                WHERE id = %s
            """, (new_role, is_super, get_iran_time(), user_id))
            conn.commit()
            return True
    except Exception as e:
        logger.error(f"update_user_role: {e}")
        if conn:
            conn.rollback()
        return False
    finally:
        if conn:
            return_db_connection(conn)

def update_user_branch(user_id, branch_id):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("SELECT role, branch_id FROM users WHERE id=%s", (user_id,))
            user_row = cur.fetchone()
            if not user_row:
                return False
            if user_row[0] == 'deputy' and user_row[1] != branch_id:
                logger.warning("Direct deputy branch change blocked; use transfer workflow. user_id=%s", user_id)
                return False
            cur.execute("UPDATE users SET branch_id = %s, updated_at = %s WHERE id = %s", (branch_id, get_iran_time(), user_id))
            conn.commit()
            return True
    except Exception as e:
        logger.error(f"update_user_branch: {e}")
        if conn:
            conn.rollback()
        return False
    finally:
        if conn:
            return_db_connection(conn)

def delete_user(user_id):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    (SELECT COUNT(*) FROM collections WHERE recorded_by=%s),
                    (SELECT COUNT(*) FROM notes WHERE user_id=%s),
                    (SELECT COUNT(*) FROM user_activity_log WHERE user_id=%s),
                    (SELECT COUNT(*) FROM problems WHERE user_id=%s),
                    (SELECT COUNT(*) FROM actual_stats WHERE recorded_by=%s),
                    (SELECT COUNT(*) FROM branch_targets WHERE created_by=%s),
                    (SELECT COUNT(*) FROM deputy_branch_assignments WHERE deputy_id=%s),
                    (SELECT COUNT(*) FROM deputy_transfer_operations WHERE first_deputy_id=%s OR second_deputy_id=%s),
                    (SELECT COUNT(*) FROM deputy_replacement_operations WHERE old_deputy_id=%s OR new_deputy_id=%s),
                    (SELECT COUNT(*) FROM message_deliveries WHERE user_id=%s)
            """, (user_id, user_id, user_id, user_id, user_id, user_id,
                  user_id, user_id, user_id, user_id, user_id, user_id))
            dependencies = cur.fetchone()
            if any(dependencies):
                return False, "این کاربر دارای سابقه عملیاتی است؛ برای حفظ اطلاعات امکان حذف فیزیکی وجود ندارد"
            cur.execute("DELETE FROM users WHERE id = %s", (user_id,))
            if cur.rowcount != 1:
                return False, "کاربر یافت نشد"
            conn.commit()
            return True, "حذف شد"
    except Exception as e:
        logger.error(f"delete_user: {e}")
        if conn:
            conn.rollback()
        return False, "حذف کاربر انجام نشد؛ جزئیات خطا ثبت شد"
    finally:
        if conn:
            return_db_connection(conn)

def get_all_collections(limit=100):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT c.id, b.name, c.shamsi_date, c.deputy_amount, c.others_amount, c.total_amount, u.full_name
                FROM collections c
                JOIN branches b ON c.branch_id = b.id
                JOIN users u ON c.recorded_by = u.id
                ORDER BY c.id DESC
                LIMIT %s
            """, (limit,))
            return cur.fetchall()
    except Exception as e:
        logger.error(f"get_all_collections: {e}")
        if conn:
            conn.rollback()
        return []
    finally:
        if conn:
            return_db_connection(conn)

def delete_collection(collection_id):
    logger.warning("Physical collection deletion blocked to preserve historical data; id=%s", collection_id)
    return False

def update_collection(collection_id, deputy_amount, others_amount):
    if deputy_amount < 0 or others_amount < 0:
        return False
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                UPDATE collections
                SET deputy_amount = %s, others_amount = %s, updated_at = %s
                WHERE id = %s
            """, (deputy_amount, others_amount, get_iran_time(), collection_id))
            if cur.rowcount != 1:
                conn.rollback()
                return False
            cur.execute("DELETE FROM scores WHERE collection_id = %s", (collection_id,))
            conn.commit()
            cache_today_report.invalidate_all()
            cache_top_branches.invalidate('top5')
            cache_10day_report.invalidate('10day')
            cache_adaptive.invalidate('adaptive')
            cache_forecast_all.invalidate('forecast_all')
            invalidate_branches_cache()
            cache_targets.invalidate_all()
            return True
    except Exception as e:
        logger.error(f"update_collection: {e}")
        if conn:
            conn.rollback()
        return False
    finally:
        if conn:
            return_db_connection(conn)

def reset_all_collections():
    """بازنشانی امن خروجی گزارش‌ها؛ هیچ رکوردی از دیتابیس حذف نمی‌شود."""
    cache_today_report.invalidate_all()
    cache_top_branches.invalidate_all()
    cache_10day_report.invalidate_all()
    cache_adaptive.invalidate_all()
    cache_forecast_all.invalidate_all()
    invalidate_branches_cache()
    cache_targets.invalidate_all()
    return True

def get_all_deputies():
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT u.id, u.telegram_id, u.full_name, u.branch_id, b.name
                FROM users u
                LEFT JOIN branches b ON u.branch_id = b.id
                WHERE u.role = 'deputy' AND u.is_active = TRUE
                AND u.telegram_id IS NOT NULL
            """)
            return cur.fetchall()
    except Exception as e:
        logger.error(f"get_all_deputies: {e}")
        if conn:
            conn.rollback()
        return []
    finally:
        if conn:
            return_db_connection(conn)

def get_log_file_path():
    return "bot.log"

def get_branch_weekly_avg(branch_id, days=7):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT AVG(total_amount) as avg_total
                FROM collections
                WHERE branch_id = %s
                AND shamsi_date >= %s
            """, (branch_id, get_shamsi_date(-days)))
            return cur.fetchone()[0] or 0
    except Exception as e:
        logger.error(f"get_branch_weekly_avg: {e}")
        if conn:
            conn.rollback()
        return 0
    finally:
        if conn:
            return_db_connection(conn)

def get_branch_monthly_avg(branch_id, days=30):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT AVG(total_amount) as avg_total
                FROM collections
                WHERE branch_id = %s
                AND shamsi_date >= %s
            """, (branch_id, get_shamsi_date(-days)))
            return cur.fetchone()[0] or 0
    except Exception as e:
        logger.error(f"get_branch_monthly_avg: {e}")
        if conn:
            conn.rollback()
        return 0
    finally:
        if conn:
            return_db_connection(conn)

def get_today_performance_analysis():
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            shamsi_today = get_shamsi_date()
            cur.execute("SELECT SUM(total_amount) FROM collections WHERE shamsi_date = %s", (shamsi_today,))
            today_total = cur.fetchone()[0] or 0
            cur.execute("""
                SELECT b.name, c.total_amount
                FROM collections c
                JOIN branches b ON c.branch_id = b.id
                WHERE c.shamsi_date = %s
                ORDER BY c.total_amount DESC
            """, (shamsi_today,))
            branch_data = cur.fetchall()
            cur.execute("""
                SELECT SUM(deputy_amount), SUM(others_amount)
                FROM collections
                WHERE shamsi_date = %s
            """, (shamsi_today,))
            deputy_others = cur.fetchone()
            deputy_total = deputy_others[0] or 0
            others_total = deputy_others[1] or 0
            cur.execute("""
                SELECT COUNT(DISTINCT branch_id) FROM collections WHERE shamsi_date = %s
            """, (shamsi_today,))
            branches_count = cur.fetchone()[0] or 0
            return {
                "today_total": today_total,
                "branch_data": branch_data,
                "deputy_total": deputy_total,
                "others_total": others_total,
                "branches_count": branches_count
            }
    except Exception as e:
        logger.error(f"get_today_performance_analysis: {e}")
        if conn:
            conn.rollback()
        return None
    finally:
        if conn:
            return_db_connection(conn)

def get_drop_alert_branches():
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            shamsi_today = get_shamsi_date()
            shamsi_week_ago = get_shamsi_date(-7)
            cur.execute("""
                SELECT
                    b.id,
                    b.name,
                    c.total_amount as today_amount,
                    COALESCE((
                        SELECT AVG(c2.total_amount)
                        FROM collections c2
                        WHERE c2.branch_id = b.id
                        AND c2.shamsi_date >= %s
                        AND c2.shamsi_date < %s
                    ), 0) as weekly_avg
                FROM branches b
                LEFT JOIN collections c ON c.branch_id = b.id AND c.shamsi_date = %s
                WHERE c.total_amount IS NOT NULL
                GROUP BY b.id, b.name, c.total_amount
            """, (shamsi_week_ago, shamsi_today, shamsi_today))
            results = []
            for row in cur.fetchall():
                branch_id, name, today, weekly_avg = row
                if weekly_avg > 0 and today < (weekly_avg * 0.6):
                    drop_percent = int(((weekly_avg - today) / weekly_avg) * 100)
                    results.append({
                        "branch_id": branch_id,
                        "name": name,
                        "today": today,
                        "weekly_avg": weekly_avg,
                        "drop_percent": drop_percent
                    })
            return results
    except Exception as e:
        logger.error(f"get_drop_alert_branches: {e}")
        if conn:
            conn.rollback()
        return []
    finally:
        if conn:
            return_db_connection(conn)

def get_branch_trend(branch_id, days=3):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT shamsi_date, total_amount
                FROM collections
                WHERE branch_id = %s
                ORDER BY shamsi_date DESC
                LIMIT %s
            """, (branch_id, days))
            return cur.fetchall()
    except Exception as e:
        logger.error(f"get_branch_trend: {e}")
        if conn:
            conn.rollback()
        return []
    finally:
        if conn:
            return_db_connection(conn)

def get_deputy_performance_report(user_id, days=30):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            shamsi_start = get_shamsi_date(-days)
            cur.execute("""
                SELECT
                    COUNT(*) as total_days,
                    SUM(CASE WHEN (EXTRACT(HOUR FROM created_at AT TIME ZONE 'Asia/Tehran') < 16)
                                OR (EXTRACT(HOUR FROM created_at AT TIME ZONE 'Asia/Tehran') = 16
                                    AND EXTRACT(MINUTE FROM created_at AT TIME ZONE 'Asia/Tehran') <= 30)
                             THEN 1 ELSE 0 END) as on_time_days,
                    AVG(total_amount) as avg_amount,
                    MAX(total_amount) as best_day
                FROM collections
                WHERE recorded_by = %s
                AND shamsi_date >= %s
            """, (user_id, shamsi_start))
            result = cur.fetchone()
            if result:
                total_days = result[0] or 0
                on_time = result[1] or 0
                avg = result[2] or 0
                best = result[3] or 0
                late = total_days - on_time
                return {
                    "total_days": total_days,
                    "on_time": on_time,
                    "late": late,
                    "avg_amount": avg,
                    "best_day": best
                }
            return None
    except Exception as e:
        logger.error(f"get_deputy_performance_report: {e}")
        if conn:
            conn.rollback()
        return None
    finally:
        if conn:
            return_db_connection(conn)

def get_unreported_branches():
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            shamsi_today = get_shamsi_date()
            cur.execute("""
                SELECT DISTINCT b.id, b.name, u.full_name, u.telegram_id
                FROM branches b
                LEFT JOIN users u ON u.branch_id = b.id AND u.role = 'deputy' AND u.is_active = TRUE
                WHERE NOT EXISTS (
                    SELECT 1 FROM collections c
                    WHERE c.branch_id = b.id AND c.shamsi_date = %s
                )
            """, (shamsi_today,))
            return cur.fetchall()
    except Exception as e:
        logger.error(f"get_unreported_branches: {e}")
        if conn:
            conn.rollback()
        return []
    finally:
        if conn:
            return_db_connection(conn)

def get_all_admins():
    cached = cache_admins.get('admins')
    if cached is not None:
        return cached
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT id, telegram_id, full_name, role, is_super_admin
                FROM users
                WHERE (role IN ('admin', 'super_admin') OR is_super_admin = TRUE)
                AND is_active = TRUE
                AND telegram_id IS NOT NULL
            """)
            result = cur.fetchall()
            cache_admins.set('admins', result)
            return result
    except Exception as e:
        logger.error(f"get_all_admins: {e}")
        if conn:
            conn.rollback()
        return []
    finally:
        if conn:
            return_db_connection(conn)

def get_branch_monthly_avg_for_name(branch_name):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT AVG(total_amount)
                FROM collections c
                JOIN branches b ON c.branch_id = b.id
                WHERE b.name = %s
                AND c.shamsi_date >= %s
            """, (branch_name, get_shamsi_date(-30)))
            result = cur.fetchone()[0]
            return result or 0
    except Exception as e:
        logger.error(f"get_branch_monthly_avg_for_name: {e}")
        if conn:
            conn.rollback()
        return 0
    finally:
        if conn:
            return_db_connection(conn)

def generate_management_analysis(analysis):
    lines = []
    today_total = analysis['today_total']
    branch_data = analysis['branch_data']
    deputy_total = analysis['deputy_total']
    others_total = analysis['others_total']
    if branch_data and len(branch_data) >= 4:
        top4_sum = sum([amount for _, amount in branch_data[:4]])
        top4_percent = (top4_sum / today_total * 100) if today_total > 0 else 0
        lines.append(f"📊 {top4_percent:.0f}% وصول استان توسط ۴ شعبه انجام شده است.")
    if branch_data:
        top_branch = branch_data[0]
        lines.append(f"🏆 بیشترین سهم وصول امروز مربوط به {top_branch[0]} است.")
    if deputy_total + others_total > 0:
        dep_percent = (deputy_total / (deputy_total + others_total) * 100) if (deputy_total + others_total) > 0 else 0
        if dep_percent > 50:
            lines.append(f"👤 میانگین وصول معاونان ({dep_percent:.0f}%) از همکاران بیشتر بوده است.")
        else:
            lines.append(f"👥 میانگین وصول همکاران ({100-dep_percent:.0f}%) از معاونان بیشتر بوده است.")
    for branch_name, amount in branch_data[:3]:
        monthly_avg = get_branch_monthly_avg_for_name(branch_name)
        if monthly_avg and monthly_avg > 0:
            growth = ((amount - monthly_avg) / monthly_avg) * 100
            if growth > 10:
                lines.append(f"📈 شعبه {branch_name} نسبت به میانگین ماه، {growth:.0f}% رشد داشته است.")
            elif growth < -10:
                lines.append(f"📉 شعبه {branch_name} نسبت به میانگین ماه، {abs(growth):.0f}% کاهش داشته است.")
    if not lines:
        lines.append("📊 داده‌های کافی برای تحلیل مدیریتی وجود ندارد.")
    return "\n".join(lines)

def get_others_performance_summary():
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    b.id,
                    b.name,
                    COALESCE(SUM(c.others_amount), 0) as total_others,
                    COALESCE(SUM(c.total_amount), 0) as total_branch,
                    COUNT(c.id) as report_days
                FROM branches b
                LEFT JOIN collections c ON b.id = c.branch_id
                GROUP BY b.id, b.name
                ORDER BY total_others DESC
            """)
            return cur.fetchall()
    except Exception as e:
        logger.error(f"get_others_performance_summary error: {e}")
        if conn:
            conn.rollback()
        return []
    finally:
        if conn:
            return_db_connection(conn)

# ============================================================
# توابع گزارش‌های پیشرفته سوپرادمین
# ============================================================
def get_deputy_accuracy_ranking(days=30):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    u.id,
                    u.full_name,
                    COALESCE(STRING_AGG(DISTINCT b.name, '، ' ORDER BY b.name), 'بدون ثبت') as branch_name,
                    COUNT(c.id) as total_days,
                    COALESCE(AVG(
                        CASE
                            WHEN a.total_actual IS NULL THEN NULL
                            WHEN a.total_actual = 0 AND c.total_amount = 0 THEN 100
                            WHEN a.total_actual = 0 AND c.total_amount > 0 THEN 0
                            WHEN a.total_actual < 0 AND c.total_amount >= 0 THEN
                                (LEAST(ABS(a.total_actual), ABS(c.total_amount)) * 100.0) / NULLIF(GREATEST(ABS(a.total_actual), ABS(c.total_amount)), 0)
                            WHEN a.total_actual > 0 AND c.total_amount >= 0 THEN
                                (LEAST(ABS(a.total_actual), ABS(c.total_amount)) * 100.0) / NULLIF(GREATEST(ABS(a.total_actual), ABS(c.total_amount)), 0)
                            ELSE NULL
                        END
                    ), 0) as avg_accuracy
                FROM users u
                LEFT JOIN collections c ON u.id = c.recorded_by AND c.shamsi_date >= %s
                LEFT JOIN branches b ON b.id = c.branch_id
                LEFT JOIN actual_stats a ON c.branch_id = a.branch_id AND c.shamsi_date = a.shamsi_date
                WHERE u.role = 'deputy'
                GROUP BY u.id, u.full_name
                HAVING COUNT(c.id) > 0
                ORDER BY avg_accuracy DESC
            """, (get_shamsi_date(-days),))
            return cur.fetchall()
    except Exception as e:
        logger.error(f"get_deputy_accuracy_ranking: {e}")
        if conn:
            conn.rollback()
        return []
    finally:
        if conn:
            return_db_connection(conn)

def get_branch_accuracy_trend(branch_id, days=30):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    c.shamsi_date,
                    c.total_amount as collected,
                    a.total_actual as actual,
                    CASE
                        WHEN a.total_actual IS NULL THEN NULL
                        WHEN a.total_actual = 0 AND c.total_amount = 0 THEN 100
                        WHEN a.total_actual = 0 AND c.total_amount > 0 THEN 0
                        WHEN a.total_actual < 0 AND c.total_amount >= 0 THEN
                            (LEAST(ABS(a.total_actual), ABS(c.total_amount)) * 100.0) / NULLIF(GREATEST(ABS(a.total_actual), ABS(c.total_amount)), 0)
                        WHEN a.total_actual > 0 AND c.total_amount >= 0 THEN
                            (LEAST(ABS(a.total_actual), ABS(c.total_amount)) * 100.0) / NULLIF(GREATEST(ABS(a.total_actual), ABS(c.total_amount)), 0)
                        ELSE NULL
                    END as accuracy
                FROM collections c
                LEFT JOIN actual_stats a ON c.branch_id = a.branch_id AND c.shamsi_date = a.shamsi_date
                WHERE c.branch_id = %s AND c.shamsi_date >= %s
                ORDER BY c.shamsi_date DESC
            """, (branch_id, get_shamsi_date(-days)))
            return cur.fetchall()
    except Exception as e:
        logger.error(f"get_branch_accuracy_trend: {e}")
        if conn:
            conn.rollback()
        return []
    finally:
        if conn:
            return_db_connection(conn)

def get_best_worst_accuracy_branches(shamsi_date, limit=5):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    b.name,
                    c.total_amount as collected,
                    a.total_actual as actual,
                    CASE
                        WHEN a.total_actual IS NULL THEN NULL
                        WHEN a.total_actual = 0 AND c.total_amount = 0 THEN 100
                        WHEN a.total_actual = 0 AND c.total_amount > 0 THEN 0
                        WHEN a.total_actual < 0 AND c.total_amount >= 0 THEN
                            (LEAST(ABS(a.total_actual), ABS(c.total_amount)) * 100.0) / NULLIF(GREATEST(ABS(a.total_actual), ABS(c.total_amount)), 0)
                        WHEN a.total_actual > 0 AND c.total_amount >= 0 THEN
                            (LEAST(ABS(a.total_actual), ABS(c.total_amount)) * 100.0) / NULLIF(GREATEST(ABS(a.total_actual), ABS(c.total_amount)), 0)
                        ELSE NULL
                    END as accuracy
                FROM branches b
                JOIN collections c ON b.id = c.branch_id AND c.shamsi_date = %s
                LEFT JOIN actual_stats a ON b.id = a.branch_id AND a.shamsi_date = %s
                WHERE a.total_actual IS NOT NULL
                ORDER BY accuracy DESC NULLS LAST
                LIMIT %s
            """, (shamsi_date, shamsi_date, limit))
            best = cur.fetchall()
            cur.execute("""
                SELECT
                    b.name,
                    c.total_amount as collected,
                    a.total_actual as actual,
                    CASE
                        WHEN a.total_actual IS NULL THEN NULL
                        WHEN a.total_actual = 0 AND c.total_amount = 0 THEN 100
                        WHEN a.total_actual = 0 AND c.total_amount > 0 THEN 0
                        WHEN a.total_actual < 0 AND c.total_amount >= 0 THEN
                            (LEAST(ABS(a.total_actual), ABS(c.total_amount)) * 100.0) / NULLIF(GREATEST(ABS(a.total_actual), ABS(c.total_amount)), 0)
                        WHEN a.total_actual > 0 AND c.total_amount >= 0 THEN
                            (LEAST(ABS(a.total_actual), ABS(c.total_amount)) * 100.0) / NULLIF(GREATEST(ABS(a.total_actual), ABS(c.total_amount)), 0)
                        ELSE NULL
                    END as accuracy
                FROM branches b
                JOIN collections c ON b.id = c.branch_id AND c.shamsi_date = %s
                LEFT JOIN actual_stats a ON b.id = a.branch_id AND a.shamsi_date = %s
                WHERE a.total_actual IS NOT NULL
                ORDER BY accuracy ASC NULLS LAST
                LIMIT %s
            """, (shamsi_date, shamsi_date, limit))
            worst = cur.fetchall()
            return best, worst
    except Exception as e:
        logger.error(f"get_best_worst_accuracy_branches: {e}")
        if conn:
            conn.rollback()
        return [], []
    finally:
        if conn:
            return_db_connection(conn)

def get_branch_performance_vs_avg(branch_id, days=30):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            shamsi_start = get_shamsi_date(-days)
            cur.execute("SELECT AVG(total_amount) FROM collections WHERE shamsi_date >= %s", (shamsi_start,))
            avg_province = cur.fetchone()[0] or 0
            cur.execute("""
                SELECT
                    COUNT(*) as days,
                    AVG(total_amount) as avg_amount,
                    SUM(total_amount) as total_amount
                FROM collections
                WHERE branch_id = %s AND shamsi_date >= %s
            """, (branch_id, shamsi_start))
            branch_stats = cur.fetchone()
            if not branch_stats or branch_stats[0] == 0:
                return None
            days, avg_branch, total_branch = branch_stats
            return {
                'days': days,
                'avg_branch': avg_branch,
                'total_branch': total_branch,
                'avg_province': avg_province,
                'diff_avg': avg_branch - avg_province,
                'diff_percent': ((avg_branch - avg_province) / avg_province * 100) if avg_province > 0 else 0
            }
    except Exception as e:
        logger.error(f"get_branch_performance_vs_avg: {e}")
        if conn:
            conn.rollback()
        return None
    finally:
        if conn:
            return_db_connection(conn)

def get_deputy_late_analysis(days=30):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    u.id,
                    u.full_name,
                    COALESCE(STRING_AGG(DISTINCT b.name, '، ' ORDER BY b.name),
                             MAX(current_b.name), 'بدون شعبه') as branch_name,
                    COUNT(c.id) as total_days,
                    COALESCE(SUM(CASE WHEN (EXTRACT(HOUR FROM c.created_at AT TIME ZONE 'Asia/Tehran') > 16)
                                OR (EXTRACT(HOUR FROM c.created_at AT TIME ZONE 'Asia/Tehran') = 16
                                    AND EXTRACT(MINUTE FROM c.created_at AT TIME ZONE 'Asia/Tehran') > 30)
                             THEN 1 ELSE 0 END), 0) as late_days,
                    COALESCE(ROUND(100.0 * SUM(CASE WHEN (EXTRACT(HOUR FROM c.created_at AT TIME ZONE 'Asia/Tehran') > 16)
                                OR (EXTRACT(HOUR FROM c.created_at AT TIME ZONE 'Asia/Tehran') = 16
                                    AND EXTRACT(MINUTE FROM c.created_at AT TIME ZONE 'Asia/Tehran') > 30)
                             THEN 1 ELSE 0 END) / NULLIF(COUNT(c.id), 0), 1), 0) as late_percent
                FROM users u
                LEFT JOIN collections c ON u.id = c.recorded_by AND c.shamsi_date >= %s
                LEFT JOIN branches b ON b.id = c.branch_id
                LEFT JOIN branches current_b ON current_b.id = u.branch_id
                WHERE u.role = 'deputy' AND u.is_active = TRUE
                GROUP BY u.id, u.full_name
                ORDER BY late_percent DESC
            """, (get_shamsi_date(-days),))
            return cur.fetchall()
    except Exception as e:
        logger.error(f"get_deputy_late_analysis: {e}")
        if conn:
            conn.rollback()
        return []
    finally:
        if conn:
            return_db_connection(conn)

# ============================================================
# توابع مدیریت اهداف
# ============================================================
def set_branch_target(branch_id, target_amount, target_date, created_by):
    target_date = normalize_digits(target_date)
    if not validate_shamsi_date(target_date):
        return False, "تاریخ هدف نامعتبر است"
    if target_amount <= 0:
        return False, "مبلغ هدف باید مثبت باشد"
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            # قفل تراکنشی مستقل از وجود ردیف، برای جلوگیری از درج هم‌زمان هدف فعال.
            cur.execute("SELECT pg_advisory_xact_lock(%s)", (int(branch_id),))
            cur.execute("SELECT id FROM branch_targets WHERE branch_id = %s AND is_active = TRUE FOR UPDATE", (branch_id,))
            cur.execute("""
                UPDATE branch_targets
                SET is_active = FALSE, updated_at = %s
                WHERE branch_id = %s AND is_active = TRUE
            """, (get_iran_time(), branch_id))
            cur.execute("""
                INSERT INTO branch_targets (branch_id, target_amount, target_date, created_by, created_at)
                VALUES (%s, %s, %s, %s, %s)
                RETURNING id
            """, (branch_id, target_amount, target_date, created_by, get_iran_time()))
            target_id = cur.fetchone()[0]
            conn.commit()
            cache_targets.invalidate_all()
            return True, target_id
    except Exception as e:
        logger.error(f"set_branch_target error: {e}")
        if conn:
            conn.rollback()
        return False, "ثبت هدف انجام نشد؛ جزئیات خطا در لاگ ثبت شد"
    finally:
        if conn:
            return_db_connection(conn)

def get_active_target(branch_id):
    cached_key = f'target_{branch_id}'
    cached = cache_targets.get(cached_key)
    if cached is not None:
        return cached
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT id, branch_id, target_amount, target_date, created_by, created_at
                FROM branch_targets
                WHERE branch_id = %s AND is_active = TRUE
                ORDER BY created_at DESC, id DESC
                LIMIT 1
            """, (branch_id,))
            result = cur.fetchone()
            if result:
                target = {
                    'id': result[0],
                    'branch_id': result[1],
                    'target_amount': result[2],
                    'target_date': result[3],
                    'created_by': result[4],
                    'created_at': result[5]
                }
                cache_targets.set(cached_key, target)
                return target
            return None
    except Exception as e:
        logger.error(f"get_active_target error: {e}")
        if conn:
            conn.rollback()
        return None
    finally:
        if conn:
            return_db_connection(conn)

def get_branch_collection_since_date(branch_id, start_date, end_date=None):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            if end_date:
                cur.execute("""SELECT COALESCE(SUM(total_amount), 0) FROM collections
                               WHERE branch_id=%s AND shamsi_date BETWEEN %s AND %s""",
                            (branch_id, start_date, end_date))
            else:
                cur.execute("""SELECT COALESCE(SUM(total_amount), 0) FROM collections
                               WHERE branch_id=%s AND shamsi_date >= %s""",
                            (branch_id, start_date))
            return cur.fetchone()[0]
    except Exception as e:
        logger.error(f"get_branch_collection_since_date error: {e}")
        if conn:
            conn.rollback()
        return 0
    finally:
        if conn:
            return_db_connection(conn)

def get_target_progress(branch_id, target_date, target_amount, created_at=None):
    shamsi_today = get_shamsi_date()
    if created_at:
        start_date_obj = jdatetime.datetime.fromgregorian(datetime=created_at)
        start_date = f"{start_date_obj.year}/{start_date_obj.month:02d}/{start_date_obj.day:02d}"
    else:
        start_date = shamsi_today
    collection_end = min(shamsi_today, target_date)
    collected = get_branch_collection_since_date(branch_id, start_date, collection_end)
    progress_percent = (collected / target_amount * 100) if target_amount > 0 else 0
    try:
        target_date_obj = jdatetime.date(*map(int, target_date.split('/')))
        today_obj = jdatetime.date(*map(int, shamsi_today.split('/')))
        days_left = (target_date_obj.togregorian() - today_obj.togregorian()).days
    except Exception:
        days_left = 0
    remaining = target_amount - collected
    if remaining < 0:
        remaining = 0
    return {
        'collected': collected,
        'target_amount': target_amount,
        'progress_percent': progress_percent,
        'remaining': remaining,
        'days_left': days_left
    }

def get_all_active_targets():
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT bt.id, bt.branch_id, b.name, bt.target_amount, bt.target_date,
                       bt.created_at, u.full_name as created_by_name
                FROM branch_targets bt
                JOIN branches b ON bt.branch_id = b.id
                LEFT JOIN users u ON bt.created_by = u.id
                WHERE bt.is_active = TRUE
                ORDER BY b.name
            """)
            return cur.fetchall()
    except Exception as e:
        logger.error(f"get_all_active_targets error: {e}")
        if conn:
            conn.rollback()
        return []
    finally:
        if conn:
            return_db_connection(conn)

def delete_target(target_id):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("SELECT branch_id FROM branch_targets WHERE id = %s", (target_id,))
            result = cur.fetchone()
            if not result:
                return False
            branch_id = result[0]
            cur.execute("""UPDATE branch_targets
                           SET is_active=FALSE, updated_at=%s
                           WHERE id=%s AND is_active=TRUE""", (get_iran_time(), target_id))
            if cur.rowcount != 1:
                conn.rollback()
                return False
            conn.commit()
            cache_targets.invalidate(f'target_{branch_id}')
            cache_targets.invalidate_all()
            return True
    except Exception as e:
        logger.error(f"delete_target error: {e}")
        if conn:
            conn.rollback()
        return False
    finally:
        if conn:
            return_db_connection(conn)

def get_targets_progress_report():
    targets = get_all_active_targets()
    if not targets:
        return None
    report = []
    for target in targets:
        target_id, branch_id, branch_name, target_amount, target_date, created_at, created_by_name = target
        progress = get_target_progress(branch_id, target_date, target_amount, created_at)
        report.append({
            'branch_id': branch_id,
            'branch_name': branch_name,
            'target_amount': target_amount,
            'target_date': target_date,
            'collected': progress['collected'],
            'progress_percent': progress['progress_percent'],
            'remaining': progress['remaining'],
            'days_left': progress['days_left'],
            'created_by': created_by_name
        })
    return report

# ============================================================
# همیار وصول مطالبات - موتور مشترک و کاملاً خواندنی
# ============================================================
def _shamsi_date_obj(value):
    return jdatetime.date(*map(int, str(value).split('/')))

def _shamsi_string(value):
    return f"{value.year:04d}/{value.month:02d}/{value.day:02d}"

def _working_dates(start_date, end_date, holidays=None):
    holidays = set(holidays or [])
    start = _shamsi_date_obj(start_date) if isinstance(start_date, str) else start_date
    end = _shamsi_date_obj(end_date) if isinstance(end_date, str) else end_date
    if end < start:
        return []
    result = []
    current = start
    while current <= end:
        date_text = _shamsi_string(current)
        # Friday is weekday 4 in Python's Gregorian calendar.
        if current.togregorian().weekday() != 4 and date_text not in holidays:
            result.append(date_text)
        current = add_days_to_shamsi(current, 1)
    return result

def _assistant_status(actual_percent, expected_percent, days_left, achieved):
    if achieved:
        return "🏁", "هدف محقق شده"
    gap = actual_percent - expected_percent
    if days_left < 0:
        return "⏳", "مهلت هدف پایان یافته"
    if gap >= 10:
        return "🔵", "جلوتر از برنامه"
    if gap >= -5:
        return "🟢", "روی مسیر"
    if gap >= -15:
        return "🟡", "نیازمند توجه"
    if gap >= -30:
        return "🟠", "عقب از برنامه"
    return "🔴", "بحرانی"

def build_collection_assistant_report(branch_id):
    """Return (ok, message, metrics). It never changes database records."""
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("SELECT name FROM branches WHERE id=%s", (branch_id,))
            branch_row = cur.fetchone()
            if not branch_row:
                return False, "❌ شعبه موردنظر پیدا نشد.", None
            branch_name = branch_row[0]
            cur.execute("""SELECT id,target_amount,target_date,created_at
                           FROM branch_targets
                           WHERE branch_id=%s AND is_active=TRUE
                           ORDER BY created_at DESC LIMIT 1""", (branch_id,))
            target = cur.fetchone()
            if not target:
                return False, f"⚪ در حال حاضر برای شعبه «{branch_name}» هدف فعالی تعریف نشده است.", None
            target_id, target_amount, target_date, target_created_at = target
            start_j = jdatetime.datetime.fromgregorian(datetime=target_created_at).date()
            start_date = _shamsi_string(start_j)
            today = get_shamsi_date()
            cur.execute("SELECT shamsi_date FROM holidays WHERE shamsi_date BETWEEN %s AND %s",
                        (min(start_date, target_date), max(today, target_date)))
            holidays = {row[0] for row in cur.fetchall()}
            cur.execute("""SELECT shamsi_date,deputy_amount,others_amount,total_amount,created_at
                           FROM collections WHERE branch_id=%s AND shamsi_date BETWEEN %s AND %s
                           ORDER BY shamsi_date""", (branch_id, start_date, max(today, target_date)))
            records = cur.fetchall()

        amounts = {row[0]: int(row[3] or 0) for row in records}
        record_by_date = {row[0]: row for row in records}
        analysis_end = min(today, target_date)
        all_workdays = _working_dates(start_date, target_date, holidays)
        candidate_elapsed = _working_dates(start_date, analysis_end, holidays) if analysis_end >= start_date else []
        elapsed_days = [day for day in candidate_elapsed if day < today or day in record_by_date]
        remaining_days = [day for day in all_workdays if day > today or (day == today and day not in record_by_date)]
        total_workdays = max(len(all_workdays), 1)
        elapsed_count = max(len(elapsed_days), 1)
        remaining_count = len(remaining_days)
        collected = sum(amounts.get(day, 0) for day in elapsed_days)
        remaining_amount = max(int(target_amount) - collected, 0)
        actual_percent = collected * 100 / target_amount if target_amount else 0
        expected_percent = min(len(elapsed_days) * 100 / total_workdays, 100)
        expected_amount = int(target_amount * expected_percent / 100)
        current_speed = collected / elapsed_count
        required_speed = remaining_amount / remaining_count if remaining_count else remaining_amount
        speed_gap = ((current_speed - required_speed) * 100 / required_speed) if required_speed else 0
        target_obj = _shamsi_date_obj(target_date)
        today_obj = _shamsi_date_obj(today)
        days_left = (target_obj.togregorian() - today_obj.togregorian()).days
        achieved = collected >= target_amount
        status_icon, status_text = _assistant_status(actual_percent, expected_percent, days_left, achieved)

        today_amount = amounts.get(today, 0)
        minimum_today = max(int(required_speed * .85), 0)
        suitable_today = max(int(required_speed), 0)
        schedule_gap = max(expected_amount - collected, 0)
        recovery_today = min(remaining_amount, int(required_speed + schedule_gap * .25)) if remaining_amount else 0

        last_14 = elapsed_days[-14:]
        recent_days, previous_days = last_14[-7:], last_14[-14:-7]
        recent_avg = sum(amounts.get(day, 0) for day in recent_days) / max(len(recent_days), 1)
        previous_avg = sum(amounts.get(day, 0) for day in previous_days) / max(len(previous_days), 1)
        weekly_change = ((recent_avg - previous_avg) * 100 / previous_avg) if previous_avg else 0
        recent_values = [amounts.get(day, 0) for day in recent_days]
        mean_recent = float(np.mean(recent_values)) if recent_values else 0
        variation = (float(np.std(recent_values)) * 100 / mean_recent) if mean_recent else 0
        stability = "داده کافی نیست" if mean_recent == 0 else "بسیار پایدار" if variation <= 15 else "پایدار" if variation <= 30 else "نسبتاً متغیر" if variation <= 55 else "پرنوسان"
        no_report_days = sum(1 for day in elapsed_days if amounts.get(day, 0) == 0)
        longest_gap = gap = 0
        for day in elapsed_days:
            if amounts.get(day, 0) == 0:
                gap += 1; longest_gap = max(longest_gap, gap)
            else:
                gap = 0

        period_records = [row for row in records if row[0] in elapsed_days]
        deputy_total = sum(int(row[1] or 0) for row in period_records)
        others_total = sum(int(row[2] or 0) for row in period_records)
        participation_total = deputy_total + others_total
        deputy_share = deputy_total * 100 / participation_total if participation_total else 0
        others_share = 100 - deputy_share if participation_total else 0
        positive_records = [row for row in period_records if int(row[3] or 0) > 0]
        best = max(positive_records, key=lambda row: row[3]) if positive_records else None
        worst = min(positive_records, key=lambda row: row[3]) if positive_records else None

        weekday_totals = {}
        for row in positive_records:
            weekday = _shamsi_date_obj(row[0]).togregorian().weekday()
            weekday_totals.setdefault(weekday, []).append(int(row[3]))
        weekday_names = {0:'دوشنبه',1:'سه‌شنبه',2:'چهارشنبه',3:'پنجشنبه',4:'جمعه',5:'شنبه',6:'یکشنبه'}
        weekday_averages = {day: sum(vals)/len(vals) for day,vals in weekday_totals.items() if vals}
        strongest_day = max(weekday_averages, key=weekday_averages.get) if weekday_averages else None
        weakest_day = min(weekday_averages, key=weekday_averages.get) if weekday_averages else None

        forecast_speed = recent_avg if recent_days else current_speed
        forecast_total = collected + forecast_speed * remaining_count
        forecast_percent = forecast_total * 100 / target_amount if target_amount else 0
        consistency_factor = max(0.55, 1 - min(variation, 90) / 200)
        probability = max(2, min(98, (forecast_percent / 100) * 78 * consistency_factor + (actual_percent / max(expected_percent, 1)) * 22))
        if achieved:
            probability = 100
        elif days_left < 0:
            probability = 0

        today_row = record_by_date.get(today)
        if today_row:
            registered_at = today_row[4].astimezone(IRAN_TZ) if today_row[4].tzinfo else today_row[4].replace(tzinfo=timezone.utc).astimezone(IRAN_TZ)
            timing = "به‌موقع" if (registered_at.hour, registered_at.minute) <= (SCORE_DEADLINE_HOUR, SCORE_DEADLINE_MINUTE) else "دیرهنگام"
            today_text = (f"✅ وصول ثبت‌شده امروز: {_fmt_money(today_amount)}\n"
                          f"📌 مقدار مناسب امروز: {_fmt_money(suitable_today)}\n"
                          f"↕️ اختلاف: {_fmt_money(abs(today_amount-suitable_today))} {'بالاتر' if today_amount>=suitable_today else 'پایین‌تر'} از برنامه\n"
                          f"🕒 زمان ثبت: {registered_at.strftime('%H:%M')} — {timing}")
        else:
            deadline = get_iran_time().replace(hour=SCORE_DEADLINE_HOUR, minute=SCORE_DEADLINE_MINUTE, second=0, microsecond=0)
            seconds_left = max(int((deadline-get_iran_time()).total_seconds()),0)
            time_left = f"{seconds_left//3600} ساعت و {(seconds_left%3600)//60} دقیقه" if seconds_left else "مهلت امروز پایان یافته"
            today_text = f"⏳ امروز هنوز وصولی ثبت نشده است.\n📌 مقدار مناسب امروز: {_fmt_money(suitable_today)}\n🕒 زمان باقی‌مانده: {time_left}"

        insights = []
        insights.append(f"سرعت فعلی {abs(speed_gap):.1f}٪ {'بیشتر' if speed_gap>=0 else 'کمتر'} از سرعت موردنیاز است.")
        insights.append(f"روند هفت‌روزه نسبت به هفته قبل {abs(weekly_change):.1f}٪ {'رشد' if weekly_change>=0 else 'افت'} داشته است.")
        if strongest_day is not None:
            insights.append(f"{weekday_names[strongest_day]} قوی‌ترین روز شعبه با میانگین {_fmt_money(weekday_averages[strongest_day])} است.")
        elif no_report_days:
            insights.append(f"در دوره هدف {no_report_days} روز کاری بدون ثبت وجود دارد.")
        else:
            insights.append(f"عملکرد دوره با نوسان {variation:.1f}٪ در وضعیت {stability} قرار دارد.")

        if achieved:
            summary = f"هدف شعبه محقق شده و میزان تحقق به {actual_percent:.1f}٪ رسیده است."
        elif days_left < 0:
            summary = f"مهلت هدف پایان یافته و تحقق نهایی {actual_percent:.1f}٪ بوده است."
        elif actual_percent >= expected_percent - 5:
            summary = f"شعبه روی مسیر هدف قرار دارد و با حفظ میانگین روزانه {_fmt_money(required_speed)} می‌تواند هدف را محقق کند."
        else:
            summary = f"شعبه از برنامه زمانی عقب است و برای رسیدن به هدف به میانگین روزانه {_fmt_money(required_speed)} نیاز دارد."

        msg = (
            f"🤖 **همیار وصول مطالبات**\n🏢 شعبه: {branch_name}\n━━━━━━━━━━━━━━━━━━\n\n"
            f"{status_icon} **وضعیت: {status_text}**\n🎯 هدف: {_fmt_money(target_amount)}\n"
            f"💰 وصول انجام‌شده: {_fmt_money(collected)}\n📊 تحقق: {actual_percent:.1f}٪\n"
            f"📉 باقی‌مانده: {_fmt_money(remaining_amount)}\n📅 روزهای کاری باقی‌مانده: {remaining_count}\n\n"
            f"⚡ **سرعت حرکت**\nمیانگین روزانه فعلی: {_fmt_money(current_speed)}\n"
            f"میانگین روزانه لازم: {_fmt_money(required_speed)}\nاختلاف سرعت: {speed_gap:+.1f}٪\n\n"
            f"📌 **پیشنهاد امروز**\nحداقل قابل‌قبول: {_fmt_money(minimum_today)}\n"
            f"مقدار مناسب: {_fmt_money(suitable_today)}\nمقدار جبرانی: {_fmt_money(recovery_today)}\n\n"
            f"📐 **برنامه زمانی**\nزمان سپری‌شده: {expected_percent:.1f}٪\nپیشرفت واقعی: {actual_percent:.1f}٪\n"
            f"فاصله از برنامه: {actual_percent-expected_percent:+.1f} واحد درصد\n\n"
            f"🔮 **پیش‌بینی موعد هدف**\nوصول پیش‌بینی‌شده: {_fmt_money(forecast_total)}\n"
            f"تحقق پیش‌بینی‌شده: {forecast_percent:.1f}٪\nامتیاز تقریبی احتمال تحقق: {probability:.0f}٪\n\n"
            f"📈 **روند و ثبات**\nمیانگین ۷ روز اخیر: {_fmt_money(recent_avg)}\n"
            f"تغییر با ۷ روز قبل: {weekly_change:+.1f}٪\nثبات: {stability} | نوسان: {variation:.1f}٪\n"
            f"روزهای بدون ثبت: {no_report_days} | طولانی‌ترین وقفه: {longest_gap} روز\n\n"
            f"🤝 **ترکیب مشارکت**\nسهم معاون: {deputy_share:.1f}٪ | سهم همکاران: {others_share:.1f}٪\n\n"
        )
        if best:
            msg += f"🏆 بهترین روز: {get_shamsi_date_formatted(best[0])} — {_fmt_money(best[3])}\n"
        if worst:
            msg += f"📉 ضعیف‌ترین روز ثبت‌شده: {get_shamsi_date_formatted(worst[0])} — {_fmt_money(worst[3])}\n"
        if strongest_day is not None and weakest_day is not None:
            msg += (f"📆 قوی‌ترین روز: {weekday_names[strongest_day]} ({_fmt_money(weekday_averages[strongest_day])})\n"
                    f"📆 ضعیف‌ترین روز: {weekday_names[weakest_day]} ({_fmt_money(weekday_averages[weakest_day])})\n")
        msg += f"\n📍 **عملکرد امروز**\n{today_text}\n\n🧠 **نکات همیار**\n"
        msg += "\n".join(f"{idx}. {text}" for idx,text in enumerate(insights[:3],1))
        msg += f"\n\n📋 **جمع‌بندی:** {summary}"

        cumulative = 0
        actual_path = []
        expected_path = []
        forecast_path = []
        for index, day in enumerate(all_workdays, 1):
            if day <= today:
                cumulative += amounts.get(day, 0)
                actual_path.append(cumulative)
                forecast_path.append(cumulative)
            else:
                actual_path.append(None)
                forecast_path.append(collected + forecast_speed * len([d for d in remaining_days if d <= day]))
            expected_path.append(target_amount * index / total_workdays)
        metrics = {
            'branch_name': branch_name, 'target_id': target_id, 'target_amount': target_amount,
            'labels': all_workdays, 'actual_path': actual_path, 'expected_path': expected_path,
            'forecast_path': forecast_path, 'actual_percent': actual_percent, 'status': status_text,
        }
        return True, msg, metrics
    except Exception:
        if conn: conn.rollback()
        logger.exception("Collection assistant report failed for branch_id=%s", branch_id)
        return False, "❌ تولید گزارش همیار انجام نشد؛ جزئیات خطا ثبت شد.", None
    finally:
        if conn: return_db_connection(conn)

def generate_collection_assistant_chart(metrics):
    if not metrics:
        return None
    ok, _ = get_chart_engine_status()
    if not ok:
        return None
    try:
        labels = [_rtl_plotly_text(label) for label in metrics['labels']]
        fig = go.Figure()
        fig.add_trace(go.Scatter(x=labels,y=metrics['expected_path'],mode='lines',name=_rtl_plotly_text('مسیر موردانتظار'),line=dict(color='#94A3B8',dash='dash',width=3)))
        fig.add_trace(go.Scatter(x=labels,y=metrics['actual_path'],mode='lines+markers',name=_rtl_plotly_text('وصول واقعی'),line=dict(color='#2563EB',width=4)))
        fig.add_trace(go.Scatter(x=labels,y=metrics['forecast_path'],mode='lines',name=_rtl_plotly_text('پیش‌بینی'),line=dict(color='#F59E0B',dash='dot',width=3)))
        fig.update_layout(title=dict(text=f"<span dir='rtl'>مسیر هدف شعبه {html.escape(metrics['branch_name'])}</span>",x=.5),
                          xaxis_title="تاریخ",yaxis_title="مبلغ وصول",font=dict(family="Vazirmatn, Noto Sans Arabic, DejaVu Sans",size=13),
                          paper_bgcolor='white',plot_bgcolor='white',legend=dict(orientation='h',y=1.08),margin=dict(l=80,r=50,t=100,b=90))
        fig.update_xaxes(tickangle=-45,gridcolor='#E5E7EB'); fig.update_yaxes(gridcolor='#E5E7EB')
        return fig.to_image(format='png',width=1400,height=800,scale=1.3,engine='kaleido')
    except Exception:
        logger.exception("Collection assistant chart failed")
        return None

def send_collection_assistant_report(chat_id, branch_id, keyboard, viewer_user_id=None):
    ok, message, metrics = build_collection_assistant_report(branch_id)
    send_message(chat_id, message, keyboard)
    if ok:
        chart = generate_collection_assistant_chart(metrics)
        if chart:
            send_photo(chat_id, chart, f"📈 مسیر هدف — {metrics['branch_name']}", keyboard)
        if viewer_user_id:
            log_user_activity(viewer_user_id, "collection_assistant_view", f"branch_id={branch_id}")
    return ok

def get_assistant_branch_keyboard(branches):
    rows = []
    for index in range(0, len(branches), 2):
        row = [{"text": f"🏢 {branches[index][1]}"}]
        if index + 1 < len(branches):
            row.append({"text": f"🏢 {branches[index + 1][1]}"})
        rows.append(row)
    rows.append([{"text":"🔙 انصراف"}])
    return {"keyboard":rows,"resize_keyboard":True}

# ============================================================
# مرکز گزارش‌های مدیریتی سوپرادمین (تماماً خواندنی)
# ============================================================
MANAGEMENT_REPORT_BUTTONS = {
    "🌅 داشبورد مدیریتی": "executive_dashboard",
    "🚨 شعب نیازمند اقدام": "action_required",
    "🎯 پیشرفت و پیش‌بینی اهداف": "target_outlook",
    "⚖️ مقایسه شعب مشابه": "peer_comparison",
    "📈 روند صعودی و نزولی": "trend",
    "🔎 نوسان‌های غیرعادی": "anomalies",
    "✅ کیفیت ثبت اطلاعات": "data_quality",
    "👤 عملکرد جامع معاونان": "deputy_performance",
    "⏰ انضباط ثبت روزانه": "submission_discipline",
    "🤝 سهم معاون و همکاران": "contribution_mix",
    "🎯 دقت با آمار واقعی": "actual_accuracy",
    "📆 عملکرد روزهای هفته": "weekday_performance",
    "🔄 مقایسه دوره‌ای": "period_comparison",
    "🔮 پیش‌بینی پایان ماه": "month_forecast",
    "🏅 رتبه‌بندی چندمعیاره": "multi_ranking",
    "🩺 سلامت سامانه جامع": "system_health",
    "🧹 کنترل کیفیت داده": "integrity_audit",
    "💾 آمادگی پشتیبان‌گیری": "backup_readiness",
}

VISUAL_REPORT_BUTTONS = {
    "🧪 تست موتور فارسی": "engine_test",
    "🏆 نمودار ۱۰ شعبه برتر": "top_branches",
    "📈 نمودار روند ۳۰ روزه": "province_trend",
    "🤝 نمودار ترکیب مشارکت": "contribution",
    "🎯 نمودار تحقق اهداف": "target_progress",
    "✅ نمودار دقت شعب": "accuracy",
    "⏰ نمودار ثبت دیرهنگام": "late_submissions",
    "📆 نمودار روزهای هفته": "weekday",
    "🔄 نمودار مقایسه دو دوره": "periods",
    "🚨 نمودار هشدار شعب": "risk_map",
}

def get_management_reports_keyboard():
    labels = list(MANAGEMENT_REPORT_BUTTONS)
    rows = [[{"text": labels[i]}, {"text": labels[i + 1]}]
            for i in range(0, len(labels) - 1, 2)]
    if len(labels) % 2:
        rows.append([{"text": labels[-1]}])
    rows.append([{"text": "🔙 بازگشت به پنل سوپرادمین"}])
    return {"keyboard": rows, "resize_keyboard": True}

def get_visual_reports_keyboard():
    labels = list(VISUAL_REPORT_BUTTONS)
    rows = [[{"text": labels[i]}, {"text": labels[i + 1]}]
            for i in range(0, len(labels) - 1, 2)]
    if len(labels) % 2:
        rows.append([{"text": labels[-1]}])
    rows.append([{"text": "🔙 بازگشت به پنل سوپرادمین"}])
    return {"keyboard": rows, "resize_keyboard": True}

def _fmt_money(value):
    return f"{int(value or 0) // 1_000_000:,.0f} میلیون ریال"

def _trend_arrow(percent):
    if percent > 3:
        return "📈"
    if percent < -3:
        return "📉"
    return "➡️"

def _previous_shamsi_month_prefix():
    today = jdatetime.datetime.fromgregorian(datetime=get_iran_time())
    if today.month == 1:
        return f"{today.year - 1}/12"
    return f"{today.year}/{today.month - 1:02d}"

def _report_header(title, period="۳۰ روز اخیر"):
    return f"{title}\n📅 {period} | تولید: {get_iran_time().strftime('%H:%M')}\n━━━━━━━━━━━━━━━━━━\n"

def generate_management_report(report_key):
    """Generate read-only provincial management reports from existing records."""
    conn = None
    today = get_shamsi_date()
    start_7 = get_shamsi_date(-7)
    start_14 = get_shamsi_date(-14)
    start_30 = get_shamsi_date(-30)
    start_60 = get_shamsi_date(-60)
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            if report_key == "executive_dashboard":
                cur.execute("""SELECT COALESCE(SUM(total_amount),0),COUNT(DISTINCT branch_id)
                               FROM collections WHERE shamsi_date=%s""", (today,))
                today_total, reported = cur.fetchone()
                cur.execute("SELECT COUNT(*) FROM branches")
                branch_count = cur.fetchone()[0]
                cur.execute("""SELECT COALESCE(SUM(total_amount),0) FROM collections
                               WHERE shamsi_date=%s""", (get_shamsi_date(-1),))
                yesterday = cur.fetchone()[0]
                change = ((today_total - yesterday) * 100 / yesterday) if yesterday else 0
                cur.execute("""SELECT b.name,COALESCE(c.total_amount,0)
                               FROM branches b LEFT JOIN collections c
                               ON c.branch_id=b.id AND c.shamsi_date=%s
                               ORDER BY COALESCE(c.total_amount,0) DESC LIMIT 5""", (today,))
                top = cur.fetchall()
                msg = _report_header("🌅 داشبورد مدیریتی امروز", get_shamsi_date_formatted(today))
                msg += f"💰 وصول امروز: {_fmt_money(today_total)}\n{_trend_arrow(change)} تغییر با دیروز: {change:+.1f}%\n"
                msg += f"🏢 ثبت‌کننده: {reported} از {branch_count} | ثبت‌نشده: {max(branch_count-reported,0)}\n\n🏆 پنج شعبه اول:\n"
                msg += "\n".join(f"{i}. {name}: {_fmt_money(amount)}" for i, (name, amount) in enumerate(top, 1))
                return msg

            if report_key == "action_required":
                cur.execute("""WITH stats AS (
                    SELECT b.id,b.name,
                      MAX(c.shamsi_date) last_date,
                      AVG(c.total_amount) FILTER (WHERE c.shamsi_date >= %s) avg30,
                      AVG(c.total_amount) FILTER (WHERE c.shamsi_date >= %s) avg7,
                      BOOL_OR(c.shamsi_date=%s) reported_today
                    FROM branches b LEFT JOIN collections c ON c.branch_id=b.id
                    GROUP BY b.id,b.name)
                    SELECT name,last_date,COALESCE(avg7,0),COALESCE(avg30,0),COALESCE(reported_today,false)
                    FROM stats
                    WHERE NOT COALESCE(reported_today,false) OR avg7 < avg30*0.7
                    ORDER BY reported_today, CASE WHEN avg30>0 THEN avg7/avg30 ELSE 0 END""",
                    (start_30, start_7, today))
                rows = cur.fetchall()
                msg = _report_header("🚨 شعب نیازمند اقدام")
                if not rows:
                    return msg + "✅ مورد بحرانی شناسایی نشد."
                for name, last_date, avg7, avg30, reported_today in rows[:20]:
                    reasons = []
                    if not reported_today: reasons.append("ثبت امروز ندارد")
                    if avg30 and avg7 < avg30 * .7: reasons.append(f"افت {100-(avg7*100/avg30):.0f}٪")
                    msg += f"🔴 {name}: {'، '.join(reasons)} | آخرین ثبت: {last_date or 'ندارد'}\n"
                return msg

            if report_key == "target_outlook":
                targets = get_all_active_targets() or []
                msg = _report_header("🎯 پیشرفت و پیش‌بینی اهداف", "اهداف فعال")
                if not targets: return msg + "هدف فعالی ثبت نشده است."
                for _, branch_id, name, amount, target_date, created_at, _ in targets:
                    p = get_target_progress(branch_id, target_date, amount, created_at)
                    elapsed_start = jdatetime.datetime.fromgregorian(datetime=created_at).date()
                    elapsed = max((jdatetime.date(*map(int, today.split('/'))) - elapsed_start).days, 1)
                    daily = p['collected'] / elapsed
                    required = p['remaining'] / max(p['days_left'], 1) if p['days_left'] >= 0 else p['remaining']
                    status = "✅ در مسیر" if daily >= required else "⚠️ عقب از برنامه"
                    msg += f"• {name}: {p['progress_percent']:.1f}٪ | {status}\n  روزانه فعلی: {_fmt_money(daily)}؛ لازم: {_fmt_money(required)}\n"
                return msg

            if report_key == "peer_comparison":
                cur.execute("""WITH s AS (SELECT b.name,COALESCE(AVG(c.total_amount),0) avg_amount,
                    COUNT(c.id) days FROM branches b LEFT JOIN collections c ON c.branch_id=b.id AND c.shamsi_date>=%s
                    GROUP BY b.id,b.name), ranked AS
                    (SELECT *,NTILE(3) OVER(ORDER BY avg_amount) peer_group FROM s)
                    SELECT peer_group,name,avg_amount,days FROM ranked ORDER BY peer_group,avg_amount DESC""", (start_30,))
                rows = cur.fetchall(); msg = _report_header("⚖️ مقایسه شعب مشابه")
                for group in (3,2,1):
                    msg += f"\n{'بزرگ/پربازده' if group==3 else 'متوسط' if group==2 else 'کوچک/کم‌حجم'}:\n"
                    msg += "\n".join(f"• {n}: {_fmt_money(a)} ({d} روز)" for g,n,a,d in rows if g==group)
                return msg

            if report_key == "trend":
                cur.execute("""SELECT b.name,
                    COALESCE(AVG(c.total_amount) FILTER(WHERE c.shamsi_date>=%s),0) recent,
                    COALESCE(AVG(c.total_amount) FILTER(WHERE c.shamsi_date>=%s AND c.shamsi_date<%s),0) previous
                    FROM branches b LEFT JOIN collections c ON c.branch_id=b.id
                    GROUP BY b.id,b.name""", (start_7,start_14,start_7))
                ranked=[]
                for name,recent,previous in cur.fetchall():
                    pct=((recent-previous)*100/previous) if previous else 0
                    ranked.append((abs(pct),pct,name))
                ranked.sort(reverse=True)
                msg=_report_header("📈 روند صعودی و نزولی","۷ روز اخیر در برابر ۷ روز قبل")
                return msg+"\n".join(f"{_trend_arrow(p)} {n}: {p:+.1f}٪" for _,p,n in ranked[:20])

            if report_key == "anomalies":
                cur.execute("""WITH base AS (SELECT c.*,AVG(c.total_amount) OVER(PARTITION BY branch_id
                    ORDER BY shamsi_date ROWS BETWEEN 14 PRECEDING AND 1 PRECEDING) baseline
                    FROM collections c WHERE shamsi_date>=%s)
                    SELECT b.name,base.shamsi_date,base.total_amount,base.baseline
                    FROM base JOIN branches b ON b.id=base.branch_id
                    WHERE baseline>0 AND (total_amount>baseline*2.5 OR total_amount<baseline*.25)
                    ORDER BY shamsi_date DESC LIMIT 25""", (start_60,))
                rows=cur.fetchall(); msg=_report_header("🔎 نوسان‌های غیرعادی","۶۰ روز اخیر")
                return msg+("موردی شناسایی نشد." if not rows else "\n".join(
                    f"• {n} | {d}: {_fmt_money(v)} در برابر میانگین {_fmt_money(a)}" for n,d,v,a in rows))

            if report_key == "data_quality":
                cur.execute("""SELECT b.name,COUNT(DISTINCT c.id),COUNT(n.id),
                    SUM(CASE WHEN (c.created_at AT TIME ZONE 'Asia/Tehran')::time > TIME '16:30' THEN 1 ELSE 0 END)
                    FROM branches b LEFT JOIN collections c ON c.branch_id=b.id AND c.shamsi_date>=%s
                    LEFT JOIN notes n ON n.collection_id=c.id GROUP BY b.id,b.name ORDER BY 4 DESC NULLS LAST""", (start_30,))
                rows=cur.fetchall(); msg=_report_header("✅ کیفیت ثبت اطلاعات")
                return msg+"\n".join(f"• {n}: {cnt} ثبت | {notes} یادداشت | {late or 0} دیرهنگام" for n,cnt,notes,late in rows)

            if report_key == "deputy_performance":
                cur.execute("""SELECT u.full_name,COALESCE(STRING_AGG(DISTINCT b.name,'، ' ORDER BY b.name),MAX(current_b.name)),COUNT(c.id),COALESCE(SUM(c.total_amount),0),
                    COALESCE(AVG(c.total_amount),0),SUM(CASE WHEN (c.created_at AT TIME ZONE 'Asia/Tehran')::time<=TIME '16:30' THEN 1 ELSE 0 END)
                    FROM users u LEFT JOIN collections c ON c.recorded_by=u.id AND c.shamsi_date>=%s
                    LEFT JOIN branches b ON b.id=c.branch_id LEFT JOIN branches current_b ON current_b.id=u.branch_id
                    WHERE u.role='deputy'
                    GROUP BY u.id,u.full_name ORDER BY 4 DESC""", (start_30,))
                rows=cur.fetchall(); msg=_report_header("👤 عملکرد جامع معاونان")
                return msg+"\n".join(f"• {u} ({b or 'بدون شعبه'}): {_fmt_money(total)} | {days} روز | به‌موقع {ontime or 0}" for u,b,days,total,avg,ontime in rows)

            if report_key == "submission_discipline":
                cur.execute("""SELECT b.name,COUNT(c.id),
                    TO_CHAR(AVG((EXTRACT(EPOCH FROM (c.created_at AT TIME ZONE 'Asia/Tehran')::time))), 'FM999999')
                    FROM branches b LEFT JOIN collections c ON c.branch_id=b.id AND c.shamsi_date>=%s
                    GROUP BY b.id,b.name ORDER BY COUNT(c.id) DESC""", (start_30,))
                rows=cur.fetchall(); msg=_report_header("⏰ انضباط ثبت روزانه")
                for name,count,avg_seconds in rows:
                    sec=int(float(avg_seconds or 0)); avg_time=f"{sec//3600:02d}:{(sec%3600)//60:02d}" if count else "—"
                    msg+=f"• {name}: {count} روز | میانگین ساعت ثبت {avg_time}\n"
                return msg

            if report_key == "contribution_mix":
                cur.execute("""SELECT b.name,COALESCE(SUM(c.deputy_amount),0),COALESCE(SUM(c.others_amount),0)
                    FROM branches b LEFT JOIN collections c ON c.branch_id=b.id AND c.shamsi_date>=%s
                    GROUP BY b.id,b.name ORDER BY SUM(c.total_amount) DESC NULLS LAST""", (start_30,))
                rows=cur.fetchall(); msg=_report_header("🤝 سهم معاون و همکاران")
                for name,dep,others in rows:
                    total=dep+others; pct=dep*100/total if total else 0
                    msg+=f"• {name}: معاون {pct:.1f}٪ | همکاران {100-pct:.1f}٪\n"
                return msg

            if report_key == "actual_accuracy":
                cur.execute("""SELECT b.name,COUNT(a.id),AVG(CASE WHEN GREATEST(ABS(a.total_actual),ABS(c.total_amount))=0 THEN 100
                    ELSE LEAST(ABS(a.total_actual),ABS(c.total_amount))*100.0/GREATEST(ABS(a.total_actual),ABS(c.total_amount)) END)
                    FROM branches b LEFT JOIN collections c ON c.branch_id=b.id AND c.shamsi_date>=%s
                    LEFT JOIN actual_stats a ON a.branch_id=c.branch_id AND a.shamsi_date=c.shamsi_date
                    GROUP BY b.id,b.name ORDER BY 3 DESC NULLS LAST""", (start_30,))
                rows=cur.fetchall(); msg=_report_header("🎯 دقت با آمار واقعی")
                return msg+"\n".join(f"• {n}: {float(acc or 0):.1f}٪ ({cnt} تطبیق)" for n,cnt,acc in rows)

            if report_key == "weekday_performance":
                cur.execute("""SELECT EXTRACT(ISODOW FROM created_at AT TIME ZONE 'Asia/Tehran')::int,
                    COUNT(*),AVG(total_amount),SUM(total_amount) FROM collections WHERE shamsi_date>=%s GROUP BY 1 ORDER BY 1""", (start_60,))
                names={6:'شنبه',7:'یکشنبه',1:'دوشنبه',2:'سه‌شنبه',3:'چهارشنبه',4:'پنجشنبه',5:'جمعه'}
                rows=cur.fetchall(); msg=_report_header("📆 عملکرد روزهای هفته","۶۰ روز اخیر")
                return msg+"\n".join(f"• {names.get(day,str(day))}: میانگین {_fmt_money(avg)} | {cnt} ثبت" for day,cnt,avg,total in rows)

            if report_key == "period_comparison":
                cur.execute("""SELECT b.name,
                    COALESCE(SUM(c.total_amount) FILTER(WHERE c.shamsi_date>=%s),0),
                    COALESCE(SUM(c.total_amount) FILTER(WHERE c.shamsi_date>=%s AND c.shamsi_date<%s),0)
                    FROM branches b LEFT JOIN collections c ON c.branch_id=b.id GROUP BY b.id,b.name""", (start_30,start_60,start_30))
                rows=[]
                for n,current,previous in cur.fetchall(): rows.append((((current-previous)*100/previous) if previous else 0,n,current,previous))
                rows.sort(reverse=True); msg=_report_header("🔄 مقایسه دوره‌ای","۳۰ روز اخیر با ۳۰ روز قبل")
                return msg+"\n".join(f"{_trend_arrow(p)} {n}: {p:+.1f}٪" for p,n,c,old in rows)

            if report_key == "month_forecast":
                prefix=today[:7]; cur.execute("SELECT COALESCE(SUM(total_amount),0),COUNT(DISTINCT shamsi_date) FROM collections WHERE shamsi_date LIKE %s", (prefix+'%',))
                total,days=cur.fetchone(); y,m,d=map(int,today.split('/'))
                next_month = jdatetime.date(y + 1, 1, 1) if m == 12 else jdatetime.date(y, m + 1, 1)
                month_days = add_days_to_shamsi(next_month, -1).day
                forecast=(total/days*month_days) if days else 0
                cur.execute("SELECT COALESCE(SUM(target_amount),0) FROM branch_targets WHERE is_active=true")
                targets=cur.fetchone()[0]
                msg=_report_header("🔮 پیش‌بینی پایان ماه",prefix)
                return msg+f"وصول فعلی: {_fmt_money(total)} در {days} روز\nپیش‌بینی پایان ماه: {_fmt_money(forecast)}\nاهداف فعال: {_fmt_money(targets)}\nفاصله پیش‌بینی با اهداف: {_fmt_money(forecast-targets)}"

            if report_key == "multi_ranking":
                cur.execute("""WITH x AS (SELECT b.name,COALESCE(SUM(c.total_amount),0) total,COUNT(c.id) days,
                    COALESCE(SUM(CASE WHEN (c.created_at AT TIME ZONE 'Asia/Tehran')::time<=TIME '16:30' THEN 1 ELSE 0 END),0) ontime
                    FROM branches b LEFT JOIN collections c ON c.branch_id=b.id AND c.shamsi_date>=%s GROUP BY b.id,b.name)
                    SELECT name,total,days,ontime,ROUND((PERCENT_RANK() OVER(ORDER BY total)*60 +
                    PERCENT_RANK() OVER(ORDER BY days)*20 + CASE WHEN days>0 THEN ontime*20.0/days ELSE 0 END)::numeric,1)
                    FROM x ORDER BY 5 DESC""", (start_30,))
                rows=cur.fetchall(); msg=_report_header("🏅 رتبه‌بندی چندمعیاره","وزن وصول ۶۰٪، نظم ۲۰٪، پوشش ثبت ۲۰٪")
                return msg+"\n".join(f"{i}. {n}: امتیاز {score} | {_fmt_money(total)}" for i,(n,total,days,on,score) in enumerate(rows,1))

            if report_key in ("system_health", "integrity_audit", "backup_readiness"):
                if report_key == "system_health":
                    cur.execute("SELECT pg_database_size(current_database()),now(),COUNT(*) FROM pg_stat_activity WHERE datname=current_database()")
                    size,dbtime,connections=cur.fetchone(); cur.execute("SELECT MAX(created_at) FROM user_activity_log")
                    last_activity=cur.fetchone()[0]
                    return _report_header("🩺 سلامت سامانه جامع","وضعیت لحظه‌ای")+f"دیتابیس: {size/1024/1024:.1f} MB\nاتصال‌ها: {connections}\nزمان DB: {dbtime}\nآخرین فعالیت: {last_activity}\nScheduler: فعال در پردازش ربات"
                if report_key == "integrity_audit":
                    checks=[]
                    for label,query in [
                        ("معاون فعال بدون شعبه","SELECT COUNT(*) FROM users WHERE role='deputy' AND is_active=TRUE AND branch_id IS NULL"),
                        ("معاون بدون سابقه انتصاب فعال","""SELECT COUNT(*) FROM users u
                            WHERE u.role='deputy' AND u.is_active=TRUE AND u.branch_id IS NOT NULL
                            AND NOT EXISTS (SELECT 1 FROM deputy_branch_assignments a
                                WHERE a.deputy_id=u.id AND a.effective_to IS NULL)"""),
                        ("ناهماهنگی شعبه فعلی و سابقه انتصاب","""SELECT COUNT(*) FROM users u
                            JOIN deputy_branch_assignments a ON a.deputy_id=u.id AND a.effective_to IS NULL
                            WHERE u.role='deputy' AND u.is_active=TRUE AND u.branch_id IS DISTINCT FROM a.branch_id"""),
                        ("انتصاب فعال تکراری برای معاون","""SELECT COUNT(*) FROM (
                            SELECT deputy_id FROM deputy_branch_assignments WHERE effective_to IS NULL
                            GROUP BY deputy_id HAVING COUNT(*)>1) x"""),
                        ("انتصاب فعال تکراری برای شعبه","""SELECT COUNT(*) FROM (
                            SELECT branch_id FROM deputy_branch_assignments WHERE effective_to IS NULL
                            GROUP BY branch_id HAVING COUNT(*)>1) x"""),
                        ("وصول منفی","SELECT COUNT(*) FROM collections WHERE deputy_amount<0 OR others_amount<0"),
                        ("وصول بدون ثبت‌کننده","SELECT COUNT(*) FROM collections WHERE recorded_by IS NULL"),
                        ("هدف فعال تکراری","SELECT COUNT(*) FROM (SELECT branch_id FROM branch_targets WHERE is_active GROUP BY branch_id HAVING COUNT(*)>1)x"),
                        ("تاریخ با قالب نامعتبر","SELECT COUNT(*) FROM collections WHERE shamsi_date !~ '^[0-9]{4}/[0-9]{2}/[0-9]{2}$'")]:
                        cur.execute(query); checks.append((label,cur.fetchone()[0]))
                    return _report_header("🧹 کنترل کیفیت داده","فقط بررسی؛ بدون تغییر")+"\n".join(f"{'✅' if n==0 else '⚠️'} {label}: {n}" for label,n in checks)
                counts=export_all_data_to_json()
                if counts is None: return _report_header("💾 آمادگی پشتیبان‌گیری")+"❌ خواندن داده‌ها ناموفق بود."
                total=sum(len(v) for v in counts.values())
                return _report_header("💾 آمادگی پشتیبان‌گیری","بررسی خواندن تمام جداول")+f"✅ همه جداول قابل خواندن‌اند.\nتعداد جداول: {len(counts)}\nمجموع ردیف‌ها: {total:,}\nامضای بکاپ: فعال"

            return "گزارش درخواستی شناخته نشد."
    except Exception as e:
        if conn:
            conn.rollback()
        logger.exception("Management report %s failed", report_key)
        return "❌ تولید گزارش انجام نشد؛ جزئیات خطا در لاگ ثبت شد."
    finally:
        if conn:
            return_db_connection(conn)

# ============================================================
# توابع آمار واقعی
# ============================================================
def save_actual_stats(branch_id, shamsi_date, total_actual, user_id):
    conn = None
    try:
        conn = get_db_connection()
        total_actual_rial = total_actual * 1_000_000
        with conn.cursor() as cur:
            cur.execute("""
                SELECT id, total_actual FROM actual_stats
                WHERE branch_id=%s AND shamsi_date=%s
                FOR UPDATE
            """, (branch_id, shamsi_date))
            previous = cur.fetchone()
            cur.execute("""
                INSERT INTO actual_stats (branch_id, shamsi_date, total_actual, recorded_by, created_at)
                VALUES (%s, %s, %s, %s, %s)
                ON CONFLICT (branch_id, shamsi_date) DO UPDATE SET
                    total_actual = EXCLUDED.total_actual,
                    recorded_by = EXCLUDED.recorded_by,
                    updated_at = CURRENT_TIMESTAMP
                RETURNING id
            """, (branch_id, shamsi_date, total_actual_rial, user_id, get_iran_time()))
            actual_stats_id = cur.fetchone()[0]
            cur.execute("""
                INSERT INTO actual_stats_audit_log
                    (actual_stats_id, branch_id, shamsi_date, old_total_actual,
                     new_total_actual, source, changed_by)
                VALUES (%s, %s, %s, %s, %s, 'manual', %s)
            """, (actual_stats_id, branch_id, shamsi_date,
                  previous[1] if previous else None, total_actual_rial, user_id))
            conn.commit()
            return True, "ثبت شد"
    except Exception as e:
        logger.error("save_actual_stats error: %s", safe_log_value(e))
        if conn:
            conn.rollback()
        return False, "ثبت آمار انجام نشد؛ لطفاً دوباره تلاش کنید."
    finally:
        if conn:
            return_db_connection(conn)

def _normalize_excel_header(value):
    if value is None:
        return ''
    text = str(value).translate(DIGIT_MAP)
    text = text.replace('ي', 'ی').replace('ى', 'ی').replace('ك', 'ک')
    text = text.replace('\u200c', '').replace('\u200f', '').replace('\u200e', '')
    return re.sub(r'[\sـ:؛،,\-_]+', '', text).strip().lower()

def _normalize_excel_code(value):
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if not value.is_integer():
            return None
        return str(int(value))
    text = str(value).translate(DIGIT_MAP).strip()
    text = re.sub(r'[\s,،٬\u200c\u200e\u200f]+', '', text)
    if re.fullmatch(r'\d+\.0+', text):
        text = text.split('.', 1)[0]
    return text if re.fullmatch(r'\d+', text) else None

def _parse_excel_actual_amount(value):
    """Return (integer million-rial value, was_blank); blanks intentionally become zero."""
    if value is None or (isinstance(value, str) and not value.strip()):
        return 0, True
    if isinstance(value, bool):
        raise ValueError("boolean amount")
    if isinstance(value, (int, float, Decimal)):
        text = str(value)
    else:
        text = str(value).strip().translate(DIGIT_MAP)
    text = (text.replace('−', '-').replace('–', '-').replace('—', '-')
                .replace('﹣', '-').replace('－', '-').replace('＋', '+'))
    negative_parentheses = text.startswith('(') and text.endswith(')')
    if negative_parentheses:
        text = text[1:-1]
    text = re.sub(r'[\s,،٬\u200c\u200e\u200f]+', '', text)
    if text.endswith('-'):
        text = '-' + text[:-1]
    if negative_parentheses and not text.startswith('-'):
        text = '-' + text
    if text.startswith('+'):
        text = text[1:]
    try:
        amount = Decimal(text)
    except (InvalidOperation, ValueError):
        raise ValueError("invalid numeric amount")
    if not amount.is_finite():
        raise ValueError("amount must be finite")
    # ستون مبلغ در فایل با قالب عدد صحیح نمایش داده می‌شود؛ مقدار خام اعشاری
    # باید دقیقاً مانند Excel به نزدیک‌ترین میلیون ریال گرد شود.
    amount_int = int(amount.quantize(Decimal('1'), rounding=ROUND_HALF_UP))
    if abs(amount_int) > MAX_AMOUNT_MILLIONS:
        raise ValueError("amount outside permitted database range")
    return amount_int, False

def _merged_excel_value(ws, row, column, merged_ranges):
    value = ws.cell(row=row, column=column).value
    if value is not None:
        return value
    for merged in merged_ranges:
        if (merged.min_row <= row <= merged.max_row and
                merged.min_col <= column <= merged.max_col):
            return ws.cell(row=merged.min_row, column=merged.min_col).value
    return None

def _extract_shamsi_date_from_excel(ws):
    candidates = []
    for row_number in range(1, min(ws.max_row or 1, 20) + 1):
        values = []
        for column in range(1, min(ws.max_column or 1, 100) + 1):
            value = ws.cell(row=row_number, column=column).value
            if value is not None:
                values.append(str(value).translate(DIGIT_MAP))
        row_text = ' '.join(values)
        for match in re.findall(r'1[34]\d{2}/\d{1,2}/\d{1,2}', row_text):
            parts = match.split('/')
            normalized = f"{int(parts[0]):04d}/{int(parts[1]):02d}/{int(parts[2]):02d}"
            if validate_shamsi_date(normalized):
                candidates.append((normalized, 'منتهی' in _normalize_excel_header(row_text)))
    preferred = [date for date, is_preferred in candidates if is_preferred]
    if preferred:
        return preferred[0]
    unique_dates = list(dict.fromkeys(date for date, _ in candidates))
    return unique_dates[0] if len(unique_dates) == 1 else None

def parse_actual_stats_excel(file_path, branches):
    """Parse only the official branch-code and day-over-day amount columns."""
    if not OPENPYXL_AVAILABLE:
        return False, "کتابخانه خواندن Excel روی سرور نصب نیست.", None
    configured = {}
    missing_code_names = []
    for branch_id, branch_name, official_code in branches:
        code = _normalize_excel_code(official_code)
        if not code:
            missing_code_names.append(branch_name)
        elif code in configured:
            return False, f"کد رسمی {code} برای بیش از یک شعبه ثبت شده است.", None
        else:
            configured[code] = (branch_id, branch_name)
    if missing_code_names:
        return False, "برای این شعب کد رسمی ثبت نشده است: " + "، ".join(missing_code_names), None
    if not configured:
        return False, "هیچ شعبه دارای کد رسمی در دیتابیس یافت نشد.", None

    # علاوه بر کد اصلی دیتابیس، فقط در این مسیر کدهای جایگزین فایل پذیرفته می‌شوند.
    accepted_codes = {
        code: (branch_id, branch_name, code)
        for code, (branch_id, branch_name) in configured.items()
    }
    for alias_code, canonical_code in ACTUAL_EXCEL_CODE_ALIASES.items():
        alias = _normalize_excel_code(alias_code)
        canonical = _normalize_excel_code(canonical_code)
        if canonical not in configured:
            continue
        branch_id, branch_name = configured[canonical]
        existing = accepted_codes.get(alias)
        if existing and existing[0] != branch_id:
            return False, f"کد جایگزین {alias} با کد اصلی شعبه دیگری تداخل دارد.", None
        accepted_codes[alias] = (branch_id, branch_name, canonical)

    workbook = None
    try:
        workbook = load_workbook(file_path, data_only=True, read_only=False)
        selected = None
        for ws in workbook.worksheets:
            header_rows = min(ws.max_row or 1, 40)
            header_columns = min(ws.max_column or 1, 200)
            merged_ranges = list(ws.merged_cells.ranges)
            normalized_grid = {}
            for row in range(1, header_rows + 1):
                for column in range(1, header_columns + 1):
                    normalized_grid[(row, column)] = _normalize_excel_header(
                        _merged_excel_value(ws, row, column, merged_ranges)
                    )

            code_candidates = []
            for (row, column), header in normalized_grid.items():
                if header == 'کدشعبه':
                    code_candidates.append((row, column))
            amount_candidates = []
            for column in range(1, header_columns + 1):
                column_headers = [normalized_grid[(row, column)] for row in range(1, header_rows + 1)]
                has_section = any('میزانتغییراتروزجاری' in header for header in column_headers)
                has_day_ratio = any('نسبتبهروزقبل' in header for header in column_headers)
                leaf_rows = [row for row in range(1, header_rows + 1)
                             if normalized_grid[(row, column)] == 'مبلغ']
                if has_section and has_day_ratio and leaf_rows:
                    amount_candidates.append((max(leaf_rows), column))
            if code_candidates and len(amount_candidates) == 1:
                code_header_row, code_column = code_candidates[0]
                amount_header_row, amount_column = amount_candidates[0]
                selected = (ws, code_column, amount_column,
                            max(code_header_row, amount_header_row) + 1)
                break

        if not selected:
            return False, ("ستون‌های «کد شعبه» و «میزان تغییرات روز جاری ← نسبت به روز قبل ← مبلغ» "
                           "در فایل پیدا نشدند. قالب فایل را بررسی کنید."), None

        ws, code_column, amount_column, first_data_row = selected
        report_date = _extract_shamsi_date_from_excel(ws)
        found = {}
        duplicates = []
        invalid_rows = []
        ignored_codes = []
        blank_count = 0
        for row_number in range(first_data_row, (ws.max_row or first_data_row) + 1):
            code = _normalize_excel_code(ws.cell(row=row_number, column=code_column).value)
            if not code:
                continue
            if code not in accepted_codes:
                ignored_codes.append(code)
                continue
            branch_id, branch_name, canonical_code = accepted_codes[code]
            if branch_id in found:
                duplicates.append(
                    f"{branch_name} ({found[branch_id]['source_code']} و {code})"
                )
                continue
            try:
                amount, was_blank = _parse_excel_actual_amount(
                    ws.cell(row=row_number, column=amount_column).value
                )
            except ValueError:
                invalid_rows.append(f"ردیف {row_number} (کد {code})")
                continue
            found[branch_id] = {
                'branch_id': branch_id,
                'branch_name': branch_name,
                'official_code': canonical_code,
                'source_code': code,
                'amount_millions': amount,
                'extracted_amount_millions': amount,
                'was_blank': was_blank,
                'corrected': False,
            }
            if was_blank:
                blank_count += 1

        if duplicates:
            return False, "یک شعبه با دو کد در فایل تکرار شده است: " + "، ".join(sorted(set(duplicates))), None
        if invalid_rows:
            return False, "مبلغ نامعتبر در " + "، ".join(invalid_rows[:10]), None
        missing_codes = [
            code for code, (branch_id, _) in configured.items()
            if branch_id not in found
        ]
        if missing_codes:
            missing_names = [f"{configured[code][1]} ({code})" for code in missing_codes]
            return False, ("ردیف این شعب در فایل پیدا نشد؛ برای جلوگیری از ثبت ناقص عملیات متوقف شد: " +
                           "، ".join(missing_names)), None

        rows = sorted(found.values(), key=lambda item: item['branch_name'])
        return True, "", {
            'report_date': report_date,
            'rows': rows,
            'blank_count': blank_count,
            'ignored_code_count': len(set(ignored_codes)),
            'sheet_name': ws.title,
        }
    except Exception as e:
        logger.exception("Actual stats Excel parsing failed")
        return False, f"فایل Excel قابل خواندن نیست ({type(e).__name__}).", None
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass

def get_existing_actual_stats_map(shamsi_date):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("SELECT branch_id, total_actual FROM actual_stats WHERE shamsi_date=%s", (shamsi_date,))
            return {branch_id: total_actual for branch_id, total_actual in cur.fetchall()}
    except Exception as e:
        logger.error("get_existing_actual_stats_map: %s", safe_log_value(e))
        if conn:
            conn.rollback()
        return None
    finally:
        if conn:
            return_db_connection(conn)

def _format_signed_millions(value):
    return f"{value:+,}" if value else "0"

def format_actual_excel_preview(draft):
    shamsi_date = draft.get('report_date') or 'تعیین نشده'
    existing = get_existing_actual_stats_map(shamsi_date) if validate_shamsi_date(shamsi_date) else {}
    if existing is None:
        existing = {}
    lines = [
        "📊 **پیش‌نمایش ثبت آمار واقعی از Excel**",
        "━━━━━━━━━━━━━━━━━━",
        f"📅 تاریخ: {shamsi_date}",
        f"📄 فایل: {escape_markdown(draft.get('file_name', 'نامشخص'))}",
        f"🏢 تعداد شعب: {len(draft.get('rows', []))}",
        "",
    ]
    overwrite_count = 0
    for index, row in enumerate(draft.get('rows', []), 1):
        current_rial = existing.get(row['branch_id'])
        if current_rial is not None:
            overwrite_count += 1
        marker = '✏️' if row.get('corrected') else ('♻️' if current_rial is not None else '▫️')
        source_code = row.get('source_code', row['official_code'])
        code_text = row['official_code']
        if source_code != row['official_code']:
            code_text = f"کد فایل {source_code} ← کد اصلی {row['official_code']}"
        line = (f"{index}. {marker} {row['branch_name']} ({code_text}): "
                f"{_format_signed_millions(row['amount_millions'])} میلیون ریال")
        if current_rial is not None:
            line += f" | قبلی: {_format_signed_millions(int(Decimal(current_rial) / Decimal(1_000_000)))}"
        lines.append(line)
    lines.extend([
        "",
        f"⬜ سلول خالیِ ثبت‌شده به‌صورت صفر: {draft.get('blank_count', 0)}",
        f"➖ کدهای غیرمرتبط/جمع حوزه که نادیده گرفته شدند: {draft.get('ignored_code_count', 0)}",
        f"✏️ اصلاحات دستی در پیش‌نمایش: {sum(1 for row in draft.get('rows', []) if row.get('corrected'))}",
        f"♻️ رکوردهای موجود که پس از تأیید به‌روزرسانی می‌شوند: {overwrite_count}",
        "",
        "تا پیش از زدن «تأیید نهایی»، هیچ داده‌ای در دیتابیس ثبت یا تغییر نمی‌کند.",
    ])
    return '\n'.join(lines)

def save_actual_stats_batch(draft, user_id):
    """Atomically upsert a verified Excel batch and preserve before/after values."""
    conn = None
    try:
        shamsi_date = draft.get('report_date')
        rows = draft.get('rows') or []
        if not validate_shamsi_date(shamsi_date) or not rows:
            return False, "اطلاعات پیش‌نمایش کامل نیست.", None
        branch_ids = [int(row['branch_id']) for row in rows]
        if len(branch_ids) != len(set(branch_ids)):
            return False, "یک شعبه بیش از یک بار در پیش‌نمایش وجود دارد.", None
        for row in rows:
            amount = row.get('amount_millions')
            if not isinstance(amount, int) or abs(amount) > MAX_AMOUNT_MILLIONS:
                return False, "یکی از مبلغ‌ها نامعتبر است.", None

        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("SET LOCAL lock_timeout='10s'")
            cur.execute("SET LOCAL statement_timeout='120s'")
            cur.execute("SELECT pg_advisory_xact_lock(%s)", (int(shamsi_date.replace('/', '')),))
            operation_uuid = str(draft.get('operation_uuid') or uuid.uuid4())
            cur.execute("""
                SELECT shamsi_date, source_sha256
                FROM actual_stats_import_operations
                WHERE operation_uuid=%s
            """, (operation_uuid,))
            already_saved = cur.fetchone()
            if already_saved:
                if (already_saved[0] == shamsi_date and
                        already_saved[1] == draft.get('file_sha256', '')):
                    conn.rollback()
                    return True, "این پیش‌نمایش قبلاً ثبت شده است.", operation_uuid
                raise ValueError("operation UUID collision")
            cur.execute("""
                SELECT id, official_code FROM branches
                WHERE id = ANY(%s)
                FOR UPDATE
            """, (branch_ids,))
            current_branches = {branch_id: _normalize_excel_code(code)
                                for branch_id, code in cur.fetchall()}
            if len(current_branches) != len(branch_ids):
                raise ValueError("branch set changed before confirmation")
            for row in rows:
                if current_branches.get(row['branch_id']) != row['official_code']:
                    raise ValueError("official branch code changed before confirmation")

            corrected_count = sum(1 for row in rows if row.get('corrected'))
            cur.execute("""
                INSERT INTO actual_stats_import_operations
                    (operation_uuid, shamsi_date, source_file_name, source_sha256,
                     extracted_count, corrected_count, created_by)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            """, (operation_uuid, shamsi_date, draft.get('file_name', 'report.xlsx')[:255],
                  draft.get('file_sha256', ''), len(rows), corrected_count, user_id))
            operation_id = cur.fetchone()[0]

            for row in rows:
                branch_id = row['branch_id']
                new_total_rial = row['amount_millions'] * 1_000_000
                cur.execute("""
                    SELECT id, total_actual FROM actual_stats
                    WHERE branch_id=%s AND shamsi_date=%s
                    FOR UPDATE
                """, (branch_id, shamsi_date))
                previous = cur.fetchone()
                cur.execute("""
                    INSERT INTO actual_stats
                        (branch_id, shamsi_date, total_actual, recorded_by, created_at)
                    VALUES (%s, %s, %s, %s, CURRENT_TIMESTAMP)
                    ON CONFLICT (branch_id, shamsi_date) DO UPDATE SET
                        total_actual=EXCLUDED.total_actual,
                        recorded_by=EXCLUDED.recorded_by,
                        updated_at=CURRENT_TIMESTAMP
                    RETURNING id
                """, (branch_id, shamsi_date, new_total_rial, user_id))
                actual_stats_id = cur.fetchone()[0]
                cur.execute("""
                    INSERT INTO actual_stats_audit_log
                        (actual_stats_id, operation_id, branch_id, shamsi_date,
                         old_total_actual, extracted_total_actual, new_total_actual,
                         source, changed_by)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, 'excel', %s)
                """, (actual_stats_id, operation_id, branch_id, shamsi_date,
                      previous[1] if previous else None,
                      row.get('extracted_amount_millions', row['amount_millions']) * 1_000_000,
                      new_total_rial, user_id))
            conn.commit()
            return True, "ثبت کامل انجام شد.", operation_uuid
    except Exception:
        logger.exception("Atomic Excel actual-stats import failed")
        if conn:
            conn.rollback()
        return False, "هیچ داده‌ای ثبت نشد؛ خطای عملیات در لاگ ذخیره شد.", None
    finally:
        if conn:
            return_db_connection(conn)

def get_actual_stats(branch_id, shamsi_date):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT total_actual
                FROM actual_stats
                WHERE branch_id = %s AND shamsi_date = %s
            """, (branch_id, shamsi_date))
            result = cur.fetchone()
            if result:
                return result[0]
            return None
    except Exception as e:
        logger.error(f"get_actual_stats error: {e}")
        if conn:
            conn.rollback()
        return None
    finally:
        if conn:
            return_db_connection(conn)

def get_actual_stats_for_date(shamsi_date):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT b.id, b.name, a.total_actual
                FROM actual_stats a
                JOIN branches b ON a.branch_id = b.id
                WHERE a.shamsi_date = %s
                ORDER BY b.name
            """, (shamsi_date,))
            return cur.fetchall()
    except Exception as e:
        logger.error(f"get_actual_stats_for_date error: {e}")
        if conn:
            conn.rollback()
        return []
    finally:
        if conn:
            return_db_connection(conn)

def compare_collection_with_actual(branch_id, shamsi_date):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT total_amount
                FROM collections
                WHERE branch_id = %s AND shamsi_date = %s
            """, (branch_id, shamsi_date))
            collection = cur.fetchone()
            cur.execute("""
                SELECT total_actual
                FROM actual_stats
                WHERE branch_id = %s AND shamsi_date = %s
            """, (branch_id, shamsi_date))
            actual = cur.fetchone()
            if not actual:
                return None
            claimed = collection[0] if collection else 0
            actual_raw = actual[0]
            return {
                'claimed': claimed,
                'actual': actual_raw,
                'abs_actual': abs(actual_raw),
                'diff_abs': abs(abs(claimed) - abs(actual_raw)),
                'is_claimed_more': abs(claimed) > abs(actual_raw) if actual_raw != 0 else None
            }
    except Exception as e:
        logger.error(f"compare_collection_with_actual error: {e}")
        if conn:
            conn.rollback()
        return None
    finally:
        if conn:
            return_db_connection(conn)

# ============================================================
# توابع گزارش‌های تطبیقی و پیش‌بینی
# ============================================================
def get_adaptive_comparison():
    cached = cache_adaptive.get('adaptive')
    if cached is not None:
        return cached
    shamsi_today = get_shamsi_date()
    shamsi_yesterday = get_shamsi_date(-1)
    shamsi_week_ago = get_shamsi_date(-7)
    shamsi_month_ago = get_shamsi_date(-30)
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("SELECT SUM(total_amount) FROM collections WHERE shamsi_date = %s", (shamsi_today,))
            today_total = cur.fetchone()[0] or 0
            cur.execute("SELECT SUM(total_amount) FROM collections WHERE shamsi_date = %s", (shamsi_yesterday,))
            yesterday_total = cur.fetchone()[0] or 0
            cur.execute("SELECT SUM(total_amount) FROM collections WHERE shamsi_date = %s", (shamsi_week_ago,))
            week_ago_total = cur.fetchone()[0] or 0
            cur.execute("SELECT SUM(total_amount) FROM collections WHERE shamsi_date = %s", (shamsi_month_ago,))
            month_ago_total = cur.fetchone()[0] or 0
            def calc_change(current, previous):
                if previous == 0:
                    if current == 0:
                        return 0
                    else:
                        return 100
                return ((current - previous) / previous) * 100
            result = {
                'today': today_total,
                'yesterday': yesterday_total,
                'week_ago': week_ago_total,
                'month_ago': month_ago_total,
                'change_yesterday': calc_change(today_total, yesterday_total),
                'change_week': calc_change(today_total, week_ago_total),
                'change_month': calc_change(today_total, month_ago_total)
            }
            cache_adaptive.set('adaptive', result)
            return result
    except Exception as e:
        logger.error(f"get_adaptive_comparison error: {e}")
        if conn:
            conn.rollback()
        return None
    finally:
        if conn:
            return_db_connection(conn)

def get_forecast(branch_id=None, days=7):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            if branch_id:
                cur.execute("""
                    SELECT shamsi_date, total_amount
                    FROM collections
                    WHERE branch_id = %s
                    ORDER BY shamsi_date DESC
                    LIMIT 45
                """, (branch_id,))
            else:
                cur.execute("""
                    SELECT shamsi_date, SUM(total_amount) as total
                    FROM collections
                    GROUP BY shamsi_date
                    ORDER BY shamsi_date DESC
                    LIMIT 45
                """)
            data = cur.fetchall()
            if len(data) < 3:
                return None, {'error': 'حداقل ۳ روز داده نیاز است', 'available_days': len(data)}
            dates = []
            amounts = []
            for row in reversed(data):
                shamsi_str = row[0]
                parts = shamsi_str.split('/')
                if len(parts) == 3:
                    try:
                        year, month, day = map(int, parts)
                        greg = jdatetime.date(year, month, day).togregorian()
                        dates.append(greg.toordinal())
                        amounts.append(float(row[1] or 0))
                    except Exception:
                        continue
            if len(dates) < 3:
                return None, {'error': f'تعداد داده‌های معتبر: {len(dates)} (حداقل ۳ روز نیاز است)'}
            x = np.array(dates)
            y = np.array(amounts)
            n = len(x)
            weights = np.exp(np.linspace(0, 1, n))
            weights = weights / weights.sum() * n
            x_mean = np.average(x, weights=weights)
            y_mean = np.average(y, weights=weights)
            cov = np.average((x - x_mean) * (y - y_mean), weights=weights)
            var = np.average((x - x_mean) ** 2, weights=weights)
            if var == 0:
                return None, {'error': 'داده‌ها تغییرات کافی ندارند'}
            slope = cov / var
            intercept = y_mean - slope * x_mean
            y_pred_all = slope * x + intercept
            mse = np.mean((y - y_pred_all) ** 2)
            rmse = np.sqrt(mse)
            ss_tot = np.sum((y - y_mean) ** 2)
            ss_res = np.sum((y - y_pred_all) ** 2)
            r2 = 1 - (ss_res / ss_tot) if ss_tot > 0 else 0
            last_date = dates[-1]
            last_shamsi = data[0][0]
            cur.execute("SELECT shamsi_date FROM holidays WHERE shamsi_date > %s", (last_shamsi,))
            future_holidays = {row[0] for row in cur.fetchall()}
            forecast = []
            calendar_offset = 1
            while len(forecast) < days and calendar_offset <= days * 4:
                future_date = last_date + calendar_offset
                calendar_offset += 1
                future_greg_date = datetime.fromordinal(future_date).date()
                future_shamsi_date = jdatetime.date.fromgregorian(date=future_greg_date)
                shamsi_str = _shamsi_string(future_shamsi_date)
                if future_greg_date.weekday() == 4 or shamsi_str in future_holidays:
                    continue
                predicted = slope * future_date + intercept
                lower = predicted - 1.96 * rmse
                upper = predicted + 1.96 * rmse
                forecast.append({
                    'date': shamsi_str,
                    'predicted': max(0, predicted),
                    'lower': max(0, lower),
                    'upper': max(0, upper)
                })
            trend_analysis = {
                'slope': slope,
                'r2': r2 if not np.isnan(r2) else 0,
                'rmse': rmse,
                'trend': 'صعودی' if slope > 0 else 'نزولی' if slope < 0 else 'ثابت',
                'strength': 'قوی' if r2 > 0.7 else 'متوسط' if r2 > 0.4 else 'ضعیف',
                'avg_amount': np.mean(y),
                'last_amount': y[-1] if len(y) > 0 else 0,
                'data_count': len(dates)
            }
            return forecast, trend_analysis
    except Exception as e:
        logger.error(f"get_forecast error: {e}")
        if conn:
            conn.rollback()
        return None, {'error': str(e)}
    finally:
        if conn:
            return_db_connection(conn)

def get_forecast_for_all_branches(days=7):
    cached = cache_forecast_all.get('forecast_all')
    if cached is not None:
        return cached
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("SELECT id, name FROM branches ORDER BY name")
            branches = cur.fetchall()
            results = {}
            for branch_id, branch_name in branches:
                forecast, trend = get_forecast(branch_id, days)
                if forecast and trend:
                    results[branch_name] = {
                        'forecast': forecast,
                        'trend': trend
                    }
            cache_forecast_all.set('forecast_all', results)
            return results
    except Exception as e:
        logger.error(f"get_forecast_for_all_branches error: {e}")
        if conn:
            conn.rollback()
        return {}
    finally:
        if conn:
            return_db_connection(conn)

# ============================================================
# توابع نمودار با پشتیبانی کامل فارسی
# ============================================================
def _rtl_plotly_text(value):
    """Keep logical Persian order and let Chromium/Kaleido perform shaping."""
    value = str(value or '').replace('\u200e', '').replace('\u200f', '')
    return f"\u202B{value}\u202C" if re.search(r'[\u0600-\u06ff]', value) else value

def _generate_chart_plotly(data, title, x_label, y_label, chart_type, figsize):
    if not PLOTLY_AVAILABLE:
        return None
    labels_raw = [str(item) for item in data.get('labels', [])]
    labels = [_rtl_plotly_text(item) for item in labels_raw]
    values = [float(value or 0) for value in data.get('values', [])]
    values2 = [float(value or 0) for value in data.get('values2', [])]
    series1_name = _rtl_plotly_text(data.get('series1_name', 'معاون'))
    series2_name = _rtl_plotly_text(data.get('series2_name', 'همکاران'))
    scale = float(data.get('display_scale', 1_000_000) or 1)
    suffix = str(data.get('display_suffix', ''))
    formatted = [f"{value / scale:,.0f}{suffix}" for value in values]
    font_family = "Vazirmatn, Noto Sans Arabic, DejaVu Sans, Arial"
    fig = go.Figure()
    if chart_type == 'horizontal':
        fig.add_trace(go.Bar(x=values, y=labels, orientation='h', marker_color='#75C1E3',
                             marker_line_color='#233A8B', marker_line_width=1,
                             text=formatted, textposition='outside', cliponaxis=False))
        fig.update_yaxes(autorange='reversed')
    elif chart_type == 'line':
        fig.add_trace(go.Scatter(x=labels, y=values, mode='lines+markers+text',
                                 line=dict(color='#2455C3', width=3),
                                 marker=dict(size=8), text=formatted, textposition='top center'))
    elif chart_type == 'pie':
        non_zero = [(label, value) for label, value in zip(labels, values) if value > 0]
        if non_zero:
            pie_labels, pie_values = zip(*non_zero)
            fig.add_trace(go.Pie(labels=pie_labels, values=pie_values, textinfo='label+percent', sort=False))
        else:
            fig.add_trace(go.Pie(labels=[_rtl_plotly_text('داده‌ای وجود ندارد')], values=[1], sort=False))
    elif chart_type == 'stacked' and values2:
        fig.add_trace(go.Bar(x=labels, y=values, name=series1_name, marker_color='#2455C3'))
        fig.add_trace(go.Bar(x=labels, y=values2, name=series2_name, marker_color='#F59E0B'))
        fig.update_layout(barmode='stack')
    else:
        fig.add_trace(go.Bar(x=labels, y=values, marker_color='#75C1E3',
                             marker_line_color='#233A8B', marker_line_width=1,
                             text=formatted, textposition='outside', cliponaxis=False))
    fig.update_layout(
        title=dict(text=f"<span dir='rtl'>{html.escape(str(title))}</span>", x=.5, xanchor='center'),
        xaxis_title=f"<span dir='rtl'>{html.escape(str(x_label))}</span>",
        yaxis_title=f"<span dir='rtl'>{html.escape(str(y_label))}</span>",
        font=dict(family=font_family, size=14, color='#111827'),
        paper_bgcolor='white', plot_bgcolor='white',
        margin=dict(l=120 if chart_type == 'horizontal' else 70, r=80, t=90, b=90),
        showlegend=chart_type in ('pie', 'stacked'),
    )
    fig.update_xaxes(showgrid=True, gridcolor='#E5E7EB', zeroline=False, tickangle=-35 if chart_type != 'horizontal' else 0)
    fig.update_yaxes(showgrid=chart_type != 'horizontal', gridcolor='#E5E7EB', zeroline=False)
    width, height = int(figsize[0] * 110), int(figsize[1] * 110)
    return fig.to_image(format='png', width=width, height=height, scale=1.5, engine='kaleido')

def get_chart_engine_status(force_test=False):
    global _chart_engine_error
    if not PLOTLY_AVAILABLE:
        _chart_engine_error = "کتابخانه plotly نصب یا قابل import نیست"
        return False, _chart_engine_error
    if _chart_engine_error is None or force_test:
        try:
            probe = go.Figure(go.Bar(x=[1], y=[1]))
            probe.update_layout(title="آزمون فارسی")
            probe.to_image(format='png', width=200, height=120, engine='kaleido')
            _chart_engine_error = ""
        except Exception as exc:
            _chart_engine_error = f"{type(exc).__name__}: {str(exc)[:500]}"
    return not bool(_chart_engine_error), _chart_engine_error

def generate_chart(data, title, x_label, y_label, chart_type='bar', figsize=(10, 6)):
    global _chart_engine_error
    try:
        engine_ok, engine_error = get_chart_engine_status()
        if engine_ok:
            try:
                plotly_image = _generate_chart_plotly(data, title, x_label, y_label, chart_type, figsize)
                if plotly_image:
                    return plotly_image
            except Exception as plotly_error:
                _chart_engine_error = f"{type(plotly_error).__name__}: {str(plotly_error)[:500]}"
                logger.exception("Plotly chart rendering failed")
        if not ALLOW_LEGACY_CHART_FALLBACK:
            logger.error("Persian chart was not generated because Plotly/Kaleido is unavailable: %s", _chart_engine_error)
            return None
        try:
            plotly_image = _generate_chart_plotly(data, title, x_label, y_label, chart_type, figsize)
            if plotly_image:
                return plotly_image
        except Exception as plotly_error:
            logger.warning("Plotly chart rendering failed; using Matplotlib fallback: %s", plotly_error)
        setup_persian_font_once()
        fig, ax = plt.subplots(figsize=figsize)
        font_prop = _persian_font_property or fm.FontProperties(family='DejaVu Sans')
        title_fa = reshape_persian(title)
        x_label_fa = reshape_persian(x_label)
        y_label_fa = reshape_persian(y_label)
        labels = [reshape_persian(str(lbl)) for lbl in data['labels']]
        values = [float(v) if v is not None else 0 for v in data['values']]
        if chart_type == 'bar':
            ax.bar(labels, values, color='skyblue', edgecolor='navy')
            max_val = max(values) if values else 1
            for i, v in enumerate(values):
                if v > 0:
                    ax.text(i, v + 0.02*max_val, f"{int(v)//1_000_000:,.0f}",
                            ha='center', va='bottom', fontsize=8, fontproperties=font_prop)
        elif chart_type == 'line':
            ax.plot(labels, values, marker='o', linestyle='-', color='blue', linewidth=2, markersize=8)
            max_val = max(values) if values else 1
            for i, v in enumerate(values):
                if v > 0:
                    ax.text(i, v + 0.02*max_val, f"{int(v)//1_000_000:,.0f}",
                            ha='center', va='bottom', fontsize=8, fontproperties=font_prop)
        elif chart_type == 'pie':
            non_zero = [(l, v) for l, v in zip(labels, values) if v > 0]
            if non_zero:
                labels, values = zip(*non_zero)
                _, pie_texts, pie_auto = ax.pie(values, labels=labels, autopct='%1.1f%%', startangle=90)
                for item in list(pie_texts) + list(pie_auto):
                    item.set_fontproperties(font_prop)
            else:
                _, pie_texts = ax.pie([1], labels=[reshape_persian('داده‌ای وجود ندارد')], colors=['lightgray'])
                for item in pie_texts:
                    item.set_fontproperties(font_prop)
        elif chart_type == 'horizontal':
            ax.barh(labels, values, color='skyblue', edgecolor='navy')
            max_val = max(values) if values else 1
            for i, v in enumerate(values):
                if v > 0:
                    ax.text(v + 0.02*max_val, i, f"{int(v)//1_000_000:,.0f}",
                            va='center', fontsize=8, fontproperties=font_prop)
        elif chart_type == 'stacked':
            if 'values2' in data:
                values2 = [float(v) if v is not None else 0 for v in data['values2']]
                ax.bar(labels, values, label=reshape_persian(data.get('series1_name', 'معاون')), color='blue', alpha=0.7)
                ax.bar(labels, values2, label=reshape_persian(data.get('series2_name', 'همکاران')), color='orange', alpha=0.7, bottom=values)
                ax.legend(prop=font_prop)
            else:
                ax.bar(labels, values, color='skyblue')
        ax.set_title(title_fa, fontsize=14, fontweight='bold', fontproperties=font_prop, pad=14)
        ax.set_xlabel(x_label_fa, fontproperties=font_prop)
        ax.set_ylabel(y_label_fa, fontproperties=font_prop)
        for tick in list(ax.get_xticklabels()) + list(ax.get_yticklabels()):
            tick.set_fontproperties(font_prop)
            tick.set_fontsize(10)
        if chart_type != 'horizontal':
            plt.setp(ax.get_xticklabels(), rotation=45, ha='right')
        fig.tight_layout()
        img_bytes = io.BytesIO()
        fig.savefig(img_bytes, format='png', dpi=150, bbox_inches='tight', facecolor='white')
        img_bytes.seek(0)
        return img_bytes.getvalue()
    except Exception as e:
        logger.error(f"generate_chart error: {e}\n{traceback.format_exc()}")
        if 'fig' in locals():
            plt.close(fig)
        return None
    finally:
        if 'fig' in locals():
            plt.close(fig)

def generate_visual_management_report(report_key):
    """Create read-only management charts from existing data."""
    if report_key == 'engine_test':
        ok, _ = get_chart_engine_status(force_test=True)
        if not ok:
            return None
        return generate_chart(
            {'labels':['شعبه مرکزی','میدان معلم','سعدی شمالی'],
             'values':[1_250_000_000,980_000_000,760_000_000]},
            'آزمون نمایش صحیح متن فارسی و عدد ۱۴۰۵',
            'مبلغ وصول', 'نام شعبه', 'horizontal', (10,5)
        )
    conn = None
    today = get_shamsi_date()
    start_7 = get_shamsi_date(-7)
    start_14 = get_shamsi_date(-14)
    start_30 = get_shamsi_date(-30)
    start_60 = get_shamsi_date(-60)
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            if report_key == 'top_branches':
                cur.execute("""SELECT b.name,COALESCE(SUM(c.total_amount),0) total
                    FROM branches b LEFT JOIN collections c ON c.branch_id=b.id AND c.shamsi_date>=%s
                    GROUP BY b.id,b.name ORDER BY total DESC LIMIT 10""", (start_30,))
                rows = cur.fetchall()
                return generate_chart({'labels':[r[0] for r in rows], 'values':[r[1] for r in rows]},
                                      '۱۰ شعبه برتر در ۳۰ روز اخیر', 'مبلغ وصول', 'شعبه', 'horizontal', (12,7))
            if report_key == 'province_trend':
                cur.execute("""SELECT shamsi_date,SUM(total_amount) FROM collections
                    WHERE shamsi_date>=%s GROUP BY shamsi_date ORDER BY shamsi_date""", (start_30,))
                rows=cur.fetchall()
                return generate_chart({'labels':[r[0] for r in rows], 'values':[r[1] for r in rows]},
                                      'روند وصول استان در ۳۰ روز اخیر','تاریخ','مبلغ وصول','line',(13,7))
            if report_key == 'contribution':
                cur.execute("""SELECT b.name,COALESCE(SUM(c.deputy_amount),0),COALESCE(SUM(c.others_amount),0)
                    FROM branches b LEFT JOIN collections c ON c.branch_id=b.id AND c.shamsi_date>=%s
                    GROUP BY b.id,b.name ORDER BY SUM(c.total_amount) DESC NULLS LAST LIMIT 12""", (start_30,))
                rows=cur.fetchall()
                return generate_chart({'labels':[r[0] for r in rows], 'values':[r[1] for r in rows], 'values2':[r[2] for r in rows],
                                       'series1_name':'معاون','series2_name':'همکاران'},
                                      'ترکیب وصول معاون و همکاران','شعبه','مبلغ وصول','stacked',(14,8))
            if report_key == 'target_progress':
                targets=get_all_active_targets() or []; labels=[]; values=[]
                for _,branch_id,name,amount,target_date,created_at,_ in targets:
                    progress=get_target_progress(branch_id,target_date,amount,created_at)
                    labels.append(name); values.append(min(progress['progress_percent'],200))
                return generate_chart({'labels':labels,'values':values,'display_scale':1,'display_suffix':'٪'},
                                      'درصد تحقق اهداف فعال','درصد تحقق','شعبه','horizontal',(12,7))
            if report_key == 'accuracy':
                cur.execute("""SELECT b.name,AVG(CASE WHEN GREATEST(ABS(a.total_actual),ABS(c.total_amount))=0 THEN 100
                    ELSE LEAST(ABS(a.total_actual),ABS(c.total_amount))*100.0/GREATEST(ABS(a.total_actual),ABS(c.total_amount)) END) accuracy
                    FROM branches b JOIN collections c ON c.branch_id=b.id AND c.shamsi_date>=%s
                    JOIN actual_stats a ON a.branch_id=c.branch_id AND a.shamsi_date=c.shamsi_date
                    GROUP BY b.id,b.name ORDER BY accuracy DESC""", (start_30,))
                rows=cur.fetchall()
                return generate_chart({'labels':[r[0] for r in rows], 'values':[r[1] for r in rows], 'display_scale':1,'display_suffix':'٪'},
                                      'میانگین دقت شعب در برابر آمار واقعی','درصد دقت','شعبه','horizontal',(12,7))
            if report_key == 'late_submissions':
                cur.execute("""SELECT b.name,SUM(CASE WHEN (c.created_at AT TIME ZONE 'Asia/Tehran')::time>TIME '16:30' THEN 1 ELSE 0 END) late
                    FROM branches b LEFT JOIN collections c ON c.branch_id=b.id AND c.shamsi_date>=%s
                    GROUP BY b.id,b.name ORDER BY late DESC NULLS LAST""", (start_30,))
                rows=cur.fetchall()
                return generate_chart({'labels':[r[0] for r in rows], 'values':[r[1] or 0 for r in rows], 'display_scale':1},
                                      'تعداد ثبت‌های دیرهنگام در ۳۰ روز اخیر','تعداد ثبت دیرهنگام','شعبه','horizontal',(12,8))
            if report_key == 'weekday':
                cur.execute("SELECT shamsi_date,total_amount FROM collections WHERE shamsi_date>=%s", (start_60,))
                names={6:'شنبه',7:'یکشنبه',1:'دوشنبه',2:'سه‌شنبه',3:'چهارشنبه',4:'پنجشنبه',5:'جمعه'}
                weekday_values = {}
                for date_text, amount in cur.fetchall():
                    iso_day = _shamsi_date_obj(date_text).togregorian().isoweekday()
                    weekday_values.setdefault(iso_day, []).append(float(amount or 0))
                rows = [(day, sum(values) / len(values)) for day, values in sorted(weekday_values.items())]
                return generate_chart({'labels':[names.get(r[0],str(r[0])) for r in rows], 'values':[r[1] for r in rows]},
                                      'میانگین وصول بر اساس روز هفته','روز هفته','میانگین وصول','bar',(11,7))
            if report_key == 'periods':
                cur.execute("""SELECT b.name,
                    COALESCE(SUM(c.total_amount) FILTER(WHERE c.shamsi_date>=%s),0) current_period,
                    COALESCE(SUM(c.total_amount) FILTER(WHERE c.shamsi_date>=%s AND c.shamsi_date<%s),0) previous_period
                    FROM branches b LEFT JOIN collections c ON c.branch_id=b.id GROUP BY b.id,b.name
                    ORDER BY current_period DESC LIMIT 12""", (start_30,start_60,start_30))
                rows=cur.fetchall()
                return generate_chart({'labels':[r[0] for r in rows], 'values':[r[1] for r in rows], 'values2':[r[2] for r in rows],
                                       'series1_name':'۳۰ روز اخیر','series2_name':'۳۰ روز قبل'},
                                      'مقایسه ۳۰ روز اخیر با ۳۰ روز قبل','شعبه','مبلغ وصول','stacked',(14,8))
            if report_key == 'risk_map':
                cur.execute("""SELECT b.name,
                    (CASE WHEN NOT BOOL_OR(c.shamsi_date=%s) THEN 2 ELSE 0 END +
                     CASE WHEN COALESCE(AVG(c.total_amount) FILTER(WHERE c.shamsi_date>=%s),0) <
                                   COALESCE(AVG(c.total_amount) FILTER(WHERE c.shamsi_date>=%s AND c.shamsi_date<%s),0)*.7
                          THEN 2 ELSE 0 END) risk
                    FROM branches b LEFT JOIN collections c ON c.branch_id=b.id GROUP BY b.id,b.name ORDER BY risk DESC,b.name""",
                    (today,start_7,start_14,start_7))
                rows=cur.fetchall()
                return generate_chart({'labels':[r[0] for r in rows], 'values':[r[1] for r in rows], 'display_scale':1},
                                      'شاخص هشدار شعب','امتیاز هشدار','شعبه','horizontal',(12,8))
        return None
    except Exception:
        if conn: conn.rollback()
        logger.exception("Visual management report %s failed", report_key)
        return None
    finally:
        if conn: return_db_connection(conn)

def generate_branch_chart(branch_id, days=10):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT shamsi_date, total_amount
                FROM collections
                WHERE branch_id = %s
                ORDER BY shamsi_date DESC
                LIMIT %s
            """, (branch_id, days))
            data = cur.fetchall()
            if not data:
                return None
            labels = [get_shamsi_date_formatted(row[0]) for row in reversed(data)]
            values = [row[1] for row in reversed(data)]
            return {
                'labels': labels,
                'values': values
            }
    except Exception as e:
        logger.error(f"generate_branch_chart error: {e}")
        if conn:
            conn.rollback()
        return None
    finally:
        if conn:
            return_db_connection(conn)

def generate_province_chart(days=10):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT shamsi_date, SUM(total_amount) as total
                FROM collections
                GROUP BY shamsi_date
                ORDER BY shamsi_date DESC
                LIMIT %s
            """, (days,))
            data = cur.fetchall()
            if not data:
                return None
            labels = [get_shamsi_date_formatted(row[0]) for row in reversed(data)]
            values = [row[1] for row in reversed(data)]
            return {
                'labels': labels,
                'values': values
            }
    except Exception as e:
        logger.error(f"generate_province_chart error: {e}")
        if conn:
            conn.rollback()
        return None
    finally:
        if conn:
            return_db_connection(conn)

def get_analytical_chart_data(chart_type, days=10):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            if chart_type == 'branch_comparison':
                cur.execute("""
                    SELECT b.name, SUM(c.total_amount) as total
                    FROM collections c
                    JOIN branches b ON c.branch_id = b.id
                    WHERE c.shamsi_date >= %s
                    GROUP BY b.name
                    ORDER BY total DESC
                    LIMIT 10
                """, (get_shamsi_date(-days),))
                data = cur.fetchall()
                return {
                    'labels': [row[0] for row in data],
                    'values': [row[1] for row in data]
                }
            elif chart_type == 'deputy_others_ratio':
                cur.execute("""
                    SELECT
                        SUM(deputy_amount) as deputy_total,
                        SUM(others_amount) as others_total
                    FROM collections
                    WHERE shamsi_date >= %s
                """, (get_shamsi_date(-days),))
                row = cur.fetchone()
                return {
                    'labels': ['معاونین', 'همکاران'],
                    'values': [row[0] or 0, row[1] or 0]
                }
            elif chart_type == 'daily_trend':
                cur.execute("""
                    SELECT shamsi_date, SUM(total_amount) as total
                    FROM collections
                    WHERE shamsi_date >= %s
                    GROUP BY shamsi_date
                    ORDER BY shamsi_date DESC
                    LIMIT %s
                """, (get_shamsi_date(-days), days))
                data = cur.fetchall()
                return {
                    'labels': [get_shamsi_date_formatted(row[0]) for row in reversed(data)],
                    'values': [row[1] for row in reversed(data)]
                }
            elif chart_type == 'match_analysis':
                cur.execute("""
                    SELECT
                        b.name,
                        COALESCE(AVG(
                            CASE
                                WHEN a.total_actual IS NULL THEN NULL
                                WHEN a.total_actual = 0 AND c.total_amount = 0 THEN 100
                                WHEN a.total_actual = 0 AND c.total_amount > 0 THEN 0
                                WHEN a.total_actual > 0 AND c.total_amount >= 0 THEN
                                    (LEAST(ABS(a.total_actual), ABS(c.total_amount)) * 100.0) / NULLIF(GREATEST(ABS(a.total_actual), ABS(c.total_amount)), 0)
                                WHEN a.total_actual < 0 AND c.total_amount >= 0 THEN
                                    (LEAST(ABS(a.total_actual), ABS(c.total_amount)) * 100.0) / NULLIF(GREATEST(ABS(a.total_actual), ABS(c.total_amount)), 0)
                                ELSE NULL
                            END
                        ), 0) as accuracy
                    FROM branches b
                    LEFT JOIN collections c ON b.id = c.branch_id
                    LEFT JOIN actual_stats a ON b.id = a.branch_id AND c.shamsi_date = a.shamsi_date
                    WHERE c.shamsi_date >= %s
                    AND a.total_actual IS NOT NULL
                    GROUP BY b.name
                    HAVING COUNT(a.total_actual) > 0
                    ORDER BY accuracy DESC
                """, (get_shamsi_date(-days),))
                data = cur.fetchall()
                return {
                    'labels': [row[0] for row in data],
                    'values': [row[1] for row in data]
                }
            else:
                return None
    except Exception as e:
        logger.error(f"get_analytical_chart_data error: {e}")
        if conn:
            conn.rollback()
        return None
    finally:
        if conn:
            return_db_connection(conn)

# ============================================================
# توابع مدیریت معاونین
# ============================================================
def get_all_deputies_with_details():
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT u.id, u.employee_number, u.full_name, u.title, b.id as branch_id, b.name as branch_name
                FROM users u
                LEFT JOIN branches b ON u.branch_id = b.id
                WHERE u.role = 'deputy' AND u.is_active = TRUE
                ORDER BY b.name, u.full_name
            """)
            return cur.fetchall()
    except Exception as e:
        logger.error(f"get_all_deputies_with_details error: {e}")
        if conn:
            conn.rollback()
        return []
    finally:
        if conn:
            return_db_connection(conn)

def get_branches_with_active_deputy():
    """فهرست شعبی که یک معاون فعال و انتصاب جاری دارند."""
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT b.id, b.name, u.id, u.employee_number, u.full_name
                FROM branches b
                JOIN deputy_branch_assignments a
                  ON a.branch_id=b.id AND a.effective_to IS NULL
                JOIN users u
                  ON u.id=a.deputy_id AND u.branch_id=b.id
                 AND u.role='deputy' AND u.is_active=TRUE
                ORDER BY b.name, u.full_name
            """)
            return cur.fetchall()
    except Exception:
        if conn:
            conn.rollback()
        logger.exception("Loading replaceable deputy branches failed")
        return []
    finally:
        if conn:
            return_db_connection(conn)

def _normalize_deputy_full_name(full_name):
    return re.sub(r'\s+', ' ', str(full_name or '')).strip()

def preview_new_deputy_replacement(branch_id, full_name, employee_number):
    """پیش‌نمایش خواندنی جایگزینی؛ این تابع هیچ داده‌ای تغییر نمی‌دهد."""
    full_name = _normalize_deputy_full_name(full_name)
    employee_number = normalize_digits(employee_number or '').strip()
    effective_date = get_shamsi_date()
    if not 3 <= len(full_name) <= 255:
        return False, "نام و نام خانوادگی باید بین ۳ تا ۲۵۵ نویسه باشد.", None
    if not re.fullmatch(r'[0-9]{1,20}', employee_number):
        return False, "شماره کارمندی باید فقط شامل ۱ تا ۲۰ رقم باشد.", None
    try:
        branch_id = int(branch_id)
    except (TypeError, ValueError):
        return False, "شعبه انتخاب‌شده معتبر نیست.", None

    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT b.id,b.name,u.id,u.employee_number,u.full_name,u.title,
                       a.id,a.effective_from
                FROM branches b
                JOIN deputy_branch_assignments a
                  ON a.branch_id=b.id AND a.effective_to IS NULL
                JOIN users u
                  ON u.id=a.deputy_id AND u.branch_id=b.id
                 AND u.role='deputy' AND u.is_active=TRUE
                WHERE b.id=%s
                FOR SHARE OF b,u,a
            """, (branch_id,))
            current_rows = cur.fetchall()
            if len(current_rows) != 1:
                return False, "این شعبه دقیقاً یک معاون و انتصاب فعالِ هماهنگ ندارد؛ عملیات متوقف شد.", None
            current = current_rows[0]
            if current[7] >= effective_date:
                return False, "انتصاب فعلی از امروز شروع شده است؛ برای جلوگیری از هم‌پوشانی، جایگزینی را در همین روز انجام ندهید.", None
            cur.execute("SELECT id,full_name,role,is_active FROM users WHERE employee_number=%s", (employee_number,))
            existing = cur.fetchone()
            if existing:
                status = "فعال" if existing[3] else "غیرفعال"
                return False, f"این شماره کارمندی قبلاً برای کاربر «{existing[1]}» با وضعیت {status} ثبت شده است.", None
            cur.execute("""SELECT 1 FROM deputy_branch_assignments
                           WHERE branch_id=%s AND effective_from>%s LIMIT 1""",
                        (branch_id, effective_date))
            if cur.fetchone():
                return False, "برای این شعبه انتصاب آینده ثبت شده است؛ جایگزینی خودکار امن نیست.", None
            preview = {
                'branch_id': current[0],
                'branch_name': current[1],
                'old_deputy_id': current[2],
                'old_employee_number': current[3],
                'old_full_name': current[4],
                'old_title': current[5],
                'old_assignment_id': current[6],
                'old_effective_from': current[7],
                'new_full_name': full_name,
                'new_employee_number': employee_number,
                'effective_date': effective_date,
            }
            return True, "پیش‌نمایش معتبر است.", preview
    except Exception:
        if conn:
            conn.rollback()
        logger.exception("New deputy replacement preview failed")
        return False, "بررسی اطلاعات انجام نشد؛ جزئیات خطا ثبت شد.", None
    finally:
        if conn:
            return_db_connection(conn)

def execute_new_deputy_replacement(branch_id, expected_old_deputy_id,
                                   full_name, employee_number, performed_by):
    """جایگزینی اتمیک از امروز؛ هیچ داده مالی، هدف یا گزارش قبلی تغییر نمی‌کند."""
    full_name = _normalize_deputy_full_name(full_name)
    employee_number = normalize_digits(employee_number or '').strip()
    effective_date = get_shamsi_date()
    previous_date = _shamsi_string(add_days_to_shamsi(effective_date, -1))
    operation_uuid = str(uuid.uuid4())
    if not 3 <= len(full_name) <= 255 or not re.fullmatch(r'[0-9]{1,20}', employee_number):
        return False, "نام یا شماره کارمندی معتبر نیست.", None
    try:
        branch_id = int(branch_id)
        expected_old_deputy_id = int(expected_old_deputy_id)
    except (TypeError, ValueError):
        return False, "شناسه شعبه یا معاون قبلی معتبر نیست.", None

    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("SELECT pg_advisory_xact_lock(%s)", (8_300_000_000 + branch_id,))
            cur.execute("SELECT id,name FROM branches WHERE id=%s FOR UPDATE", (branch_id,))
            branch = cur.fetchone()
            if not branch:
                raise ValueError("شعبه انتخاب‌شده وجود ندارد")
            cur.execute("""
                SELECT a.id,a.effective_from,u.id,u.employee_number,u.full_name,u.title
                FROM deputy_branch_assignments a
                JOIN users u ON u.id=a.deputy_id
                WHERE a.branch_id=%s AND a.effective_to IS NULL
                  AND u.branch_id=%s AND u.role='deputy' AND u.is_active=TRUE
                FOR UPDATE OF a,u
            """, (branch_id, branch_id))
            current_rows = cur.fetchall()
            if len(current_rows) != 1:
                raise ValueError("شعبه دقیقاً یک معاون فعال و انتصاب جاری ندارد")
            assignment_id, assignment_start, old_deputy_id, old_employee, old_name, old_title = current_rows[0]
            if old_deputy_id != expected_old_deputy_id:
                raise ValueError("معاون فعلی شعبه پس از پیش‌نمایش تغییر کرده است")
            if assignment_start >= effective_date:
                raise ValueError("انتصاب فعلی از امروز شروع شده و بستن امن آن در همین روز ممکن نیست")
            cur.execute("SELECT id FROM users WHERE employee_number=%s FOR UPDATE", (employee_number,))
            if cur.fetchone():
                raise ValueError("شماره کارمندی در فاصله پیش‌نمایش تا تأیید ثبت شده است")
            cur.execute("""SELECT 1 FROM deputy_branch_assignments
                           WHERE branch_id=%s AND effective_from>%s LIMIT 1""",
                        (branch_id, effective_date))
            if cur.fetchone():
                raise ValueError("برای شعبه انتصاب آینده وجود دارد")

            new_title = f"معاون محترم شعبه {branch[1]}"
            cur.execute("""INSERT INTO users
                           (employee_number,full_name,role,title,branch_id,is_super_admin,is_active,created_at,updated_at)
                           VALUES (%s,%s,'deputy',%s,%s,FALSE,TRUE,%s,%s)
                           RETURNING id""",
                        (employee_number, full_name, new_title, branch_id,
                         get_iran_time(), get_iran_time()))
            new_deputy_id = cur.fetchone()[0]
            cur.execute("""INSERT INTO deputy_replacement_operations
                           (operation_uuid,branch_id,old_deputy_id,new_deputy_id,
                            effective_date,performed_by,created_at)
                           VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
                        (operation_uuid, branch_id, old_deputy_id, new_deputy_id,
                         effective_date, performed_by, get_iran_time()))
            replacement_operation_id = cur.fetchone()[0]
            cur.execute("""UPDATE deputy_branch_assignments SET effective_to=%s
                           WHERE id=%s AND deputy_id=%s AND branch_id=%s AND effective_to IS NULL""",
                        (previous_date, assignment_id, old_deputy_id, branch_id))
            if cur.rowcount != 1:
                raise ValueError("بستن انتصاب معاون قبلی کامل نشد")
            cur.execute("""UPDATE users
                           SET branch_id=NULL,is_active=FALSE,telegram_id=NULL,updated_at=%s
                           WHERE id=%s AND role='deputy' AND is_active=TRUE AND branch_id=%s""",
                        (get_iran_time(), old_deputy_id, branch_id))
            if cur.rowcount != 1:
                raise ValueError("غیرفعال‌سازی دسترسی معاون قبلی کامل نشد")
            cur.execute("""INSERT INTO deputy_branch_assignments
                           (deputy_id,branch_id,effective_from,replacement_operation_id,
                            created_by,source,created_at)
                           VALUES (%s,%s,%s,%s,%s,'replacement',%s)""",
                        (new_deputy_id, branch_id, effective_date,
                         replacement_operation_id, performed_by, get_iran_time()))
            conn.commit()
            cache_admins.invalidate_all()
            invalidate_branches_cache()
            return True, "معاون جدید با موفقیت منصوب شد.", {
                'operation_uuid': operation_uuid,
                'effective_date': effective_date,
                'branch_name': branch[1],
                'old_deputy_id': old_deputy_id,
                'old_full_name': old_name,
                'old_employee_number': old_employee,
                'new_deputy_id': new_deputy_id,
                'new_full_name': full_name,
                'new_employee_number': employee_number,
            }
    except ValueError as exc:
        if conn:
            conn.rollback()
        logger.warning("New deputy replacement validation stopped: %s", safe_log_value(exc))
        return False, safe_log_value(exc), None
    except Exception:
        if conn:
            conn.rollback()
        logger.exception("New deputy replacement transaction failed")
        return False, "خطای فنی رخ داد؛ هیچ تغییر ناقصی ثبت نشد و جزئیات در لاگ ذخیره شد.", None
    finally:
        if conn:
            return_db_connection(conn)

def format_new_deputy_preview(preview):
    return (
        "🔎 **پیش‌نمایش انتصاب معاون جدید**\n━━━━━━━━━━━━━━━━━━\n\n"
        f"🏢 شعبه: {preview['branch_name']}\n"
        f"📅 تاریخ مؤثر: {get_shamsi_date_formatted(preview['effective_date'])}\n\n"
        "👤 **معاون قبلی**\n"
        f"{preview['old_full_name']} ({preview['old_employee_number']})\n"
        "وضعیت پس از تأیید: دسترسی غیرفعال؛ سابقه کاملاً محفوظ\n\n"
        "👤 **معاون جدید**\n"
        f"{preview['new_full_name']} ({preview['new_employee_number']})\n"
        f"سمت: معاون محترم شعبه {preview['branch_name']}\n\n"
        "🛡 **بدون تغییر باقی می‌مانند:**\n"
        "وصول‌ها، مبالغ، اهداف شعبه، آمار واقعی، یادداشت‌ها، امتیازها و تمام سوابق قبلی.\n\n"
        "⏰ تحلیل تأخیر معاون جدید از صفر آغاز می‌شود، چون فقط ثبت‌های انجام‌شده با شناسه کاربری جدید محاسبه خواهند شد.\n\n"
        "⚠️ با تأیید نهایی، عملیات در یک تراکنش ثبت می‌شود و قابل اجرای تکراری نیست."
    )

def preview_deputy_swap(first_deputy_id, second_deputy_id, effective_date):
    """اعتبارسنجی و پیش‌نمایش کاملاً خواندنی؛ هیچ داده‌ای تغییر نمی‌کند."""
    effective_date = normalize_digits(effective_date)
    if first_deputy_id == second_deputy_id or not validate_shamsi_date(effective_date):
        return False, "معاونان یا تاریخ مؤثر معتبر نیستند.", None
    if effective_date > get_shamsi_date():
        return False, "تاریخ مؤثر نمی‌تواند بعد از تاریخ امروز باشد.", None
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""SELECT u.id,u.employee_number,u.full_name,u.title,u.branch_id,b.name
                           FROM users u LEFT JOIN branches b ON b.id=u.branch_id
                           WHERE u.id=ANY(%s) AND u.role='deputy' AND u.is_active=TRUE ORDER BY u.id""",
                        ([first_deputy_id, second_deputy_id],))
            rows = {row[0]: row for row in cur.fetchall()}
            if len(rows) != 2:
                return False, "هر دو کاربر باید معاون فعال و معتبر باشند.", None
            first, second = rows[first_deputy_id], rows[second_deputy_id]
            if not first[4] or not second[4]:
                return False, "برای یکی از معاونان شعبه فعلی تعیین نشده است.", None
            if first[4] == second[4]:
                return False, "دو معاون در حال حاضر به یک شعبه متصل‌اند؛ جابه‌جایی دوطرفه ممکن نیست.", None
            cur.execute("""SELECT deputy_id,branch_id,effective_from,effective_to
                           FROM deputy_branch_assignments
                           WHERE deputy_id=ANY(%s) AND effective_to IS NULL FOR SHARE""",
                        ([first_deputy_id, second_deputy_id],))
            active = {row[0]: row for row in cur.fetchall()}
            if len(active) != 2:
                return False, "تاریخچه فعال یکی از معاونان ناقص است؛ ابتدا باید توسط مدیر سیستم بررسی شود.", None
            if active[first_deputy_id][1] != first[4] or active[second_deputy_id][1] != second[4]:
                return False, "شعبه فعلی کاربر با تاریخچه انتصاب هماهنگ نیست؛ عملیات متوقف شد.", None
            if active[first_deputy_id][2] > effective_date or active[second_deputy_id][2] > effective_date:
                return False, "تاریخ مؤثر قبل از شروع انتصاب فعلی است و با تاریخچه موجود تداخل دارد.", None
            cur.execute("""SELECT deputy_id,branch_id,effective_from,effective_to
                           FROM deputy_branch_assignments
                           WHERE deputy_id=ANY(%s) AND effective_from>%s""",
                        ([first_deputy_id, second_deputy_id], effective_date))
            if cur.fetchone():
                return False, "بعد از تاریخ انتخابی انتصاب دیگری ثبت شده است؛ جابه‌جایی عطف‌به‌ماسبق امن نیست.", None

            audit = {}
            for deputy in (first, second):
                cur.execute("""SELECT c.branch_id,b.name,COUNT(*),COALESCE(SUM(c.total_amount),0)
                               FROM collections c JOIN branches b ON b.id=c.branch_id
                               WHERE c.recorded_by=%s AND c.shamsi_date>=%s
                               GROUP BY c.branch_id,b.name ORDER BY COUNT(*) DESC""",
                            (deputy[0], effective_date))
                audit[deputy[0]] = cur.fetchall()
            preview = {
                'first': first, 'second': second, 'effective_date': effective_date,
                'first_new_branch_id': second[4], 'first_new_branch_name': second[5],
                'second_new_branch_id': first[4], 'second_new_branch_name': first[5],
                'audit': audit,
            }
            return True, "پیش‌نمایش معتبر است.", preview
    except Exception:
        if conn:
            conn.rollback()
        logger.exception("Deputy swap preview failed")
        return False, "بررسی جابه‌جایی انجام نشد؛ جزئیات خطا ثبت شد.", None
    finally:
        if conn:
            return_db_connection(conn)

def execute_deputy_swap(first_deputy_id, second_deputy_id, effective_date,
                        expected_first_branch_id, expected_second_branch_id,
                        performed_by, reason=None):
    """جابه‌جایی اتمیک؛ فقط انتصاب و شعبه فعلی تغییر می‌کند."""
    effective_date = normalize_digits(effective_date)
    previous_date = _shamsi_string(add_days_to_shamsi(effective_date, -1))
    operation_uuid = str(uuid.uuid4())
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            lock_key = min(first_deputy_id, second_deputy_id) * 1_000_000 + max(first_deputy_id, second_deputy_id)
            cur.execute("SELECT pg_advisory_xact_lock(%s)", (lock_key,))
            cur.execute("""SELECT u.id,u.full_name,u.title,u.branch_id,b.name
                           FROM users u JOIN branches b ON b.id=u.branch_id
                           WHERE u.id=ANY(%s) AND u.role='deputy' AND u.is_active=TRUE ORDER BY u.id FOR UPDATE OF u""",
                        ([first_deputy_id, second_deputy_id],))
            users = {row[0]: row for row in cur.fetchall()}
            if len(users) != 2:
                raise ValueError("هر دو معاون معتبر نیستند")
            first, second = users[first_deputy_id], users[second_deputy_id]
            if first[3] != expected_first_branch_id or second[3] != expected_second_branch_id:
                raise ValueError("شعبه یکی از معاونان پس از پیش‌نمایش تغییر کرده است")
            cur.execute("""SELECT id,deputy_id,branch_id,effective_from
                           FROM deputy_branch_assignments
                           WHERE deputy_id=ANY(%s) AND effective_to IS NULL
                           ORDER BY deputy_id FOR UPDATE""",
                        ([first_deputy_id, second_deputy_id],))
            assignments = {row[1]: row for row in cur.fetchall()}
            if len(assignments) != 2:
                raise ValueError("انتصاب فعال معاونان کامل نیست")
            for deputy_id, expected_branch in (
                (first_deputy_id, expected_first_branch_id),
                (second_deputy_id, expected_second_branch_id),
            ):
                assignment = assignments[deputy_id]
                if assignment[2] != expected_branch or assignment[3] > effective_date:
                    raise ValueError("تاریخچه انتصاب با پیش‌نمایش سازگار نیست")
            cur.execute("""SELECT 1 FROM deputy_branch_assignments
                           WHERE deputy_id=ANY(%s) AND effective_from>%s LIMIT 1""",
                        ([first_deputy_id, second_deputy_id], effective_date))
            if cur.fetchone():
                raise ValueError("انتصاب متداخل بعد از تاریخ مؤثر وجود دارد")

            first_title = first[2].replace(first[4], second[4]) if first[4] in first[2] else first[2]
            second_title = second[2].replace(second[4], first[4]) if second[4] in second[2] else second[2]
            cur.execute("""INSERT INTO deputy_transfer_operations
                           (operation_uuid,first_deputy_id,second_deputy_id,
                            first_old_branch_id,first_new_branch_id,
                            second_old_branch_id,second_new_branch_id,
                            effective_date,performed_by,reason,created_at)
                           VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
                        (operation_uuid, first_deputy_id, second_deputy_id,
                         first[3], second[3], second[3], first[3], effective_date,
                         performed_by, reason, get_iran_time()))
            operation_id = cur.fetchone()[0]
            cur.execute("""UPDATE deputy_branch_assignments SET effective_to=%s
                           WHERE id=ANY(%s) AND effective_to IS NULL""",
                        (previous_date, [assignments[first_deputy_id][0], assignments[second_deputy_id][0]]))
            if cur.rowcount != 2:
                raise ValueError("بستن انتصاب‌های قبلی کامل نشد")
            cur.execute("""INSERT INTO deputy_branch_assignments
                           (deputy_id,branch_id,effective_from,operation_id,created_by,source,created_at)
                           VALUES (%s,%s,%s,%s,%s,'swap',%s),(%s,%s,%s,%s,%s,'swap',%s)""",
                        (first_deputy_id, second[3], effective_date, operation_id, performed_by, get_iran_time(),
                         second_deputy_id, first[3], effective_date, operation_id, performed_by, get_iran_time()))
            cur.execute("UPDATE users SET branch_id=%s,title=%s,updated_at=%s WHERE id=%s",
                        (second[3], first_title, get_iran_time(), first_deputy_id))
            cur.execute("UPDATE users SET branch_id=%s,title=%s,updated_at=%s WHERE id=%s",
                        (first[3], second_title, get_iran_time(), second_deputy_id))
            conn.commit()
            cache_admins.invalidate_all()
            invalidate_branches_cache()
            return True, "جابه‌جایی با موفقیت ثبت شد.", {
                'operation_uuid': operation_uuid, 'first_name': first[1], 'second_name': second[1],
                'first_old': first[4], 'first_new': second[4],
                'second_old': second[4], 'second_new': first[4],
                'effective_date': effective_date,
            }
    except ValueError as exc:
        if conn:
            conn.rollback()
        logger.warning("Deputy swap validation stopped: %s", safe_log_value(exc))
        return False, safe_log_value(exc), None
    except Exception:
        if conn:
            conn.rollback()
        logger.exception("Deputy swap transaction failed")
        return False, "خطای فنی رخ داد؛ جزئیات در لاگ ثبت شد", None
    finally:
        if conn:
            return_db_connection(conn)

def get_deputy_assignment_history(limit=30):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""SELECT dba.id,u.full_name,u.employee_number,b.name,
                                  dba.effective_from,dba.effective_to,dba.source,
                                  COALESCE(dto.operation_uuid,dro.operation_uuid)
                           FROM deputy_branch_assignments dba
                           JOIN users u ON u.id=dba.deputy_id
                           JOIN branches b ON b.id=dba.branch_id
                           LEFT JOIN deputy_transfer_operations dto ON dto.id=dba.operation_id
                           LEFT JOIN deputy_replacement_operations dro
                                  ON dro.id=dba.replacement_operation_id
                           ORDER BY dba.created_at DESC,dba.id DESC LIMIT %s""", (limit,))
            return cur.fetchall()
    except Exception:
        if conn:
            conn.rollback()
        logger.exception("Loading deputy assignment history failed")
        return []
    finally:
        if conn:
            return_db_connection(conn)

def format_deputy_swap_preview(preview):
    first, second = preview['first'], preview['second']
    message = (
        "🔎 **پیش‌نمایش جابه‌جایی معاونان**\n━━━━━━━━━━━━━━━━━━\n\n"
        f"📅 تاریخ مؤثر: {get_shamsi_date_formatted(preview['effective_date'])}\n\n"
        f"👤 {first[2]} ({first[1]})\n"
        f"   {first[5]} ← تا روز قبل از تاریخ مؤثر\n"
        f"   {preview['first_new_branch_name']} ← از تاریخ مؤثر\n\n"
        f"👤 {second[2]} ({second[1]})\n"
        f"   {second[5]} ← تا روز قبل از تاریخ مؤثر\n"
        f"   {preview['second_new_branch_name']} ← از تاریخ مؤثر\n\n"
        "🛡 **اطلاعاتی که تغییر نمی‌کنند:**\n"
        "وصول‌ها، مبالغ، تاریخ‌ها، ثبت‌کننده‌ها، یادداشت‌ها، امتیازها، اهداف و آمار واقعی.\n\n"
        "ℹ️ شعبه فعلی دو حساب جابه‌جا می‌شود و اگر نام شعبه در عنوان سازمانی آن‌ها باشد، همان نام نیز اصلاح می‌شود.\n\n"
        "📋 **کنترل ثبت‌های موجود از تاریخ مؤثر:**\n"
    )
    for deputy in (first, second):
        rows = preview['audit'].get(deputy[0], [])
        if not rows:
            message += f"• {deputy[2]}: هیچ وصولی ثبت نکرده است.\n"
            continue
        expected_branch_id = (preview['first_new_branch_id'] if deputy[0] == first[0]
                              else preview['second_new_branch_id'])
        details = "، ".join(
            f"{'✅' if branch_id == expected_branch_id else '⚠️'} {name}: {count} ثبت"
            for branch_id, name, count, _ in rows
        )
        message += f"• {deputy[2]}: {details}\n"
    message += (
        "\n⚠️ این فهرست فقط برای کنترل است؛ حتی در صورت مغایرت، هیچ وصول قبلی خودکار جابه‌جا نمی‌شود.\n"
        "با تأیید نهایی فقط شعبه فعلی و تاریخچه انتصاب دو معاون ثبت خواهد شد."
    )
    return message

def get_deputy_transfer_keyboard():
    return {
        "keyboard": [
            [{"text": "🔄 جابه‌جایی دو معاون"}],
            [{"text": "➕ افزودن معاون جدید"}],
            [{"text": "📜 تاریخچه انتصاب معاونان"}],
            [{"text": "🔙 انصراف"}],
        ],
        "resize_keyboard": True,
    }

def add_deputy(employee_number, full_name, title, branch_id, is_super_admin=False):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO users (employee_number, full_name, role, title, branch_id, is_super_admin, created_at)
                VALUES (%s, %s, 'deputy', %s, %s, %s, %s)
                ON CONFLICT (employee_number) DO NOTHING
                RETURNING id
            """, (employee_number, full_name, title, branch_id, is_super_admin, get_iran_time()))
            result = cur.fetchone()
            if result:
                cur.execute("""INSERT INTO deputy_branch_assignments
                               (deputy_id,branch_id,effective_from,source,created_at)
                               VALUES (%s,%s,%s,'deputy_created',%s)""",
                            (result[0], branch_id, get_shamsi_date(), get_iran_time()))
                conn.commit()
                return result[0]
            conn.rollback()
            return None
    except Exception as e:
        logger.error(f"add_deputy error: {e}")
        if conn:
            conn.rollback()
        return None
    finally:
        if conn:
            return_db_connection(conn)

def update_deputy(user_id, employee_number=None, full_name=None, title=None, branch_id=None):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            if branch_id is not None:
                cur.execute("SELECT branch_id FROM users WHERE id=%s AND role='deputy' AND is_active=TRUE", (user_id,))
                current = cur.fetchone()
                if not current:
                    return False
                if current[0] != branch_id:
                    logger.warning("Direct deputy branch edit blocked; use transfer workflow. user_id=%s", user_id)
                    return False
            updates = []  # ترتیب صریح و ثابت؛ مقادیر همگی پارامتری هستند.
            params = []
            if employee_number is not None:
                updates.append("employee_number = %s")
                params.append(employee_number)
            if full_name is not None:
                updates.append("full_name = %s")
                params.append(full_name)
            if title is not None:
                updates.append("title = %s")
                params.append(title)
            if not updates:
                return False
            updates.append("updated_at = %s")
            params.append(get_iran_time())
            params.append(user_id)
            query = f"UPDATE users SET {', '.join(updates)} WHERE id = %s AND role = 'deputy' AND is_active=TRUE"
            cur.execute(query, params)
            conn.commit()
            return cur.rowcount > 0
    except Exception as e:
        logger.error(f"update_deputy error: {e}")
        if conn:
            conn.rollback()
        return False
    finally:
        if conn:
            return_db_connection(conn)

def delete_deputy(user_id):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("SELECT role FROM users WHERE id = %s", (user_id,))
            result = cur.fetchone()
            if not result or result[0] != 'deputy':
                return False, "کاربر معاون نیست"
            conn.rollback()
            return delete_user(user_id)
    except Exception as e:
        logger.error(f"delete_deputy error: {e}")
        if conn:
            conn.rollback()
        return False, "حذف معاون انجام نشد؛ جزئیات خطا ثبت شد"
    finally:
        if conn:
            return_db_connection(conn)

def get_deputy_by_employee_number(emp_num):
    emp_num = normalize_digits(emp_num)
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT id, employee_number, full_name, title, branch_id
                FROM users
                WHERE employee_number = %s AND role = 'deputy' AND is_active=TRUE
            """, (emp_num,))
            return cur.fetchone()
    except Exception as e:
        logger.error(f"get_deputy_by_employee_number error: {e}")
        if conn:
            conn.rollback()
        return None
    finally:
        if conn:
            return_db_connection(conn)

def get_deputy_match_report(user_id, days=30):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            shamsi_start = get_shamsi_date(-days)
            cur.execute("""
                SELECT
                    c.shamsi_date,
                    c.total_amount as collected,
                    a.total_actual as actual,
                    CASE
                        WHEN a.total_actual IS NULL THEN NULL
                        WHEN a.total_actual = 0 AND c.total_amount = 0 THEN 100
                        WHEN a.total_actual = 0 AND c.total_amount > 0 THEN 0
                        WHEN a.total_actual < 0 AND c.total_amount >= 0 THEN
                            ROUND((LEAST(ABS(a.total_actual), ABS(c.total_amount)) * 100.0) / NULLIF(GREATEST(ABS(a.total_actual), ABS(c.total_amount)), 0), 2)
                        WHEN a.total_actual > 0 AND c.total_amount >= 0 THEN
                            ROUND((LEAST(ABS(a.total_actual), ABS(c.total_amount)) * 100.0) / NULLIF(GREATEST(ABS(a.total_actual), ABS(c.total_amount)), 0), 2)
                        ELSE 0
                    END as match_percent,
                    (ABS(a.total_actual) - ABS(c.total_amount)) as diff
                FROM collections c
                LEFT JOIN actual_stats a ON c.branch_id = a.branch_id AND c.shamsi_date = a.shamsi_date
                WHERE c.recorded_by = %s
                AND c.shamsi_date >= %s
                AND a.total_actual IS NOT NULL
                ORDER BY c.shamsi_date DESC
            """, (user_id, shamsi_start))
            return cur.fetchall()
    except Exception as e:
        logger.error(f"get_deputy_match_report error: {e}")
        if conn:
            conn.rollback()
        return None
    finally:
        if conn:
            return_db_connection(conn)

# ============================================================
# توابع پشتیبان‌گیری و بازیابی
# ============================================================
def export_all_data_to_json():
    conn = None
    data = {}
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            for table in BACKUP_TABLES:
                # ترتیب بر اساس کلید اصلی؛ هیچ فرضی درباره وجود ستون id نمی‌شود.
                cur.execute("""
                    SELECT a.attname
                    FROM pg_index i
                    JOIN pg_attribute a ON a.attrelid = i.indrelid
                                       AND a.attnum = ANY(i.indkey)
                    WHERE i.indrelid = %s::regclass AND i.indisprimary
                    ORDER BY array_position(i.indkey, a.attnum)
                """, (table,))
                primary_keys = [row[0] for row in cur.fetchall()]
                query = sql.SQL("SELECT * FROM {}").format(sql.Identifier(table))
                if primary_keys:
                    query += sql.SQL(" ORDER BY {}").format(
                        sql.SQL(', ').join(sql.Identifier(col) for col in primary_keys)
                    )
                cur.execute(query)
                rows = cur.fetchall()
                colnames = [desc[0] for desc in cur.description]
                data[table] = [dict(zip(colnames, row)) for row in rows]
        return data
    except Exception as e:
        logger.error(f"export_all_data error: {e}")
        return None
    finally:
        if conn:
            return_db_connection(conn)

def import_all_data_from_json(json_data, clear_existing=False, dry_run=False):
    """Restore missing rows only; existing rows are never deleted or overwritten."""
    conn = None
    summary = {'inserted': {}, 'skipped': {}, 'missing_tables': []}
    try:
        if clear_existing:
            logger.warning("clear_existing was requested but is intentionally ignored for data safety")
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("SET LOCAL lock_timeout = '5s'")
            cur.execute("SET LOCAL statement_timeout = '120s'")
            for unexpected in set(json_data) - set(BACKUP_TABLES):
                logger.warning("Ignoring unexpected backup table: %s", unexpected)
            # ترتیب ثابت والد به فرزند، برای رعایت کلیدهای خارجی.
            for table in BACKUP_TABLES:
                rows = json_data.get(table, [])
                if not rows:
                    summary['inserted'][table] = 0
                    summary['skipped'][table] = 0
                    continue
                cur.execute("SELECT to_regclass(%s)", (f'public.{table}',))
                if cur.fetchone()[0] is None:
                    summary['missing_tables'].append(table)
                    continue
                cur.execute("""
                    SELECT column_name, is_generated, is_identity
                    FROM information_schema.columns
                    WHERE table_schema='public' AND table_name=%s
                    ORDER BY ordinal_position
                """, (table,))
                writable = {
                    name for name, generated, identity in cur.fetchall()
                    if generated == 'NEVER' and identity != 'ALWAYS'
                }
                columns = [name for name in rows[0].keys() if name in writable]
                if not columns:
                    summary['skipped'][table] = len(rows)
                    continue

                # برخورد یک شناسه با رکوردی با هویت متفاوت می‌تواند وابستگی‌ها را
                # به شعبه/کاربر اشتباه متصل کند؛ در این حالت کل تراکنش متوقف می‌شود.
                identity_fields = RESTORE_IDENTITY_FIELDS.get(table, ())
                if 'id' in columns and identity_fields:
                    comparable = [field for field in identity_fields if field in columns]
                else:
                    comparable = []
                if comparable:
                    for row in rows:
                        if row.get('id') is None:
                            continue
                        select_columns = sql.SQL(', ').join(sql.Identifier(field) for field in comparable)
                        cur.execute(
                            sql.SQL("SELECT {} FROM {} WHERE id=%s").format(
                                select_columns, sql.Identifier(table)
                            ),
                            (row['id'],)
                        )
                        existing = cur.fetchone()
                        if existing and any(
                            str(existing[index]) != str(row.get(field))
                            for index, field in enumerate(comparable)
                        ):
                            raise ValueError(
                                f"تعارض هویتی در جدول {table} برای id={row['id']}; بازیابی متوقف شد"
                            )
                values = [tuple(
                    json.dumps(row.get(column), ensure_ascii=False)
                    if isinstance(row.get(column), (dict, list)) else row.get(column)
                    for column in columns
                ) for row in rows]
                statement = sql.SQL(
                    "INSERT INTO {} ({}) VALUES %s ON CONFLICT DO NOTHING"
                ).format(
                    sql.Identifier(table),
                    sql.SQL(', ').join(sql.Identifier(column) for column in columns)
                )
                inserted = 0
                for start in range(0, len(values), 200):
                    execute_values(cur, statement.as_string(conn), values[start:start + 200], page_size=200)
                    inserted += max(cur.rowcount, 0)
                summary['inserted'][table] = inserted
                summary['skipped'][table] = len(rows) - inserted

                # همگام‌سازی امن sequence پس از درج شناسه‌های صریح.
                if 'id' in columns:
                    cur.execute("SELECT pg_get_serial_sequence(%s, 'id')", (table,))
                    sequence_name = cur.fetchone()[0]
                    if sequence_name:
                        cur.execute(sql.SQL(
                            "SELECT setval(%s, GREATEST(COALESCE(MAX(id), 1), 1), MAX(id) IS NOT NULL) FROM {}"
                        ).format(sql.Identifier(table)), (sequence_name,))
            if dry_run:
                conn.rollback()
            else:
                conn.commit()
            # پاکسازی کش‌ها
            cache_today_report.invalidate_all()
            cache_top_branches.invalidate_all()
            cache_10day_report.invalidate_all()
            cache_adaptive.invalidate_all()
            cache_forecast_all.invalidate_all()
            cache_targets.invalidate_all()
            invalidate_branches_cache()
            return True, summary
    except Exception as e:
        logger.error(f"import_all_data error: {e}")
        if conn:
            conn.rollback()
        return False, {'error': str(e), **summary}
    finally:
        if conn:
            return_db_connection(conn)

def generate_backup_file():
    size_conn = None
    try:
        size_conn = get_db_connection()
        with size_conn.cursor() as cur:
            cur.execute("SELECT COALESCE(SUM(pg_total_relation_size(to_regclass('public.' || name))),0) FROM unnest(%s::text[]) AS t(name)",
                        (list(BACKUP_TABLES),))
            estimated_size = int(cur.fetchone()[0] or 0)
            if estimated_size > MAX_BACKUP_UNCOMPRESSED_BYTES:
                logger.error("Backup refused: source relations are too large (%s bytes)", estimated_size)
                return None
    finally:
        if size_conn:
            return_db_connection(size_conn)
    data = export_all_data_to_json()
    if data is None:
        return None
    payload = {
        'format': 'zanjan-collection-bot-backup',
        'format_version': BACKUP_FORMAT_VERSION,
        'backup_id': str(uuid.uuid4()),
        'created_at': get_iran_time().isoformat(),
        'tables': data,
        'table_counts': {name: len(rows) for name, rows in data.items()}
    }
    canonical = json.dumps(payload, default=str, ensure_ascii=False,
                           sort_keys=True, separators=(',', ':')).encode('utf-8')
    if len(canonical) > MAX_BACKUP_UNCOMPRESSED_BYTES:
        logger.error("Backup refused: serialized payload is too large (%s bytes)", len(canonical))
        return None
    envelope = {
        'payload': payload,
        'sha256': hashlib.sha256(canonical).hexdigest(),
        'hmac_sha256': hmac.new(BACKUP_SECRET.encode('utf-8'), canonical, hashlib.sha256).hexdigest()
    }
    return gzip.compress(json.dumps(envelope, ensure_ascii=False, default=str).encode('utf-8'), 9)

def restore_from_file(file_bytes, dry_run=True):
    try:
        if not isinstance(file_bytes, (bytes, bytearray)) or len(file_bytes) > MAX_BACKUP_COMPRESSED_BYTES:
            return False, "حجم فایل پشتیبان نامعتبر یا بیش از حد مجاز است.", None
        try:
            with gzip.GzipFile(fileobj=io.BytesIO(file_bytes), mode='rb') as stream:
                raw = stream.read(MAX_BACKUP_UNCOMPRESSED_BYTES + 1)
            if len(raw) > MAX_BACKUP_UNCOMPRESSED_BYTES:
                return False, "حجم بازشده فایل بیش از حد مجاز است.", None
        except (OSError, EOFError):
            raw = bytes(file_bytes)  # سازگاری با بکاپ JSON قدیمی
        document = json.loads(raw.decode('utf-8'))
        if document.get('payload'):
            payload = document['payload']
            canonical = json.dumps(payload, default=str, ensure_ascii=False,
                                   sort_keys=True, separators=(',', ':')).encode('utf-8')
            if not hmac.compare_digest(document.get('sha256', ''), hashlib.sha256(canonical).hexdigest()):
                return False, "فایل پشتیبان تغییر کرده یا خراب است (SHA-256 نامعتبر).", None
            expected_hmac = hmac.new(BACKUP_SECRET.encode('utf-8'), canonical, hashlib.sha256).hexdigest()
            if not hmac.compare_digest(document.get('hmac_sha256', ''), expected_hmac):
                return False, "امضای فایل معتبر نیست؛ BACKUP_SECRET یکسان نیست.", None
            data = payload.get('tables', {})
        else:
            if not ALLOW_UNSIGNED_LEGACY_BACKUP:
                return False, "بکاپ قدیمی فاقد امضای امنیتی است. برای ورود آگاهانه آن، ALLOW_UNSIGNED_LEGACY_BACKUP=true را موقتاً تنظیم کنید.", None
            data = document
        required_tables = ['branches', 'users', 'collections']
        for tbl in required_tables:
            if tbl not in data:
                return False, f"جدول {tbl} در فایل بکاپ وجود ندارد.", None
        success, summary = import_all_data_from_json(data, clear_existing=False, dry_run=dry_run)
        if success:
            action = "آزمون بازیابی موفق بود؛ هیچ داده‌ای تغییر نکرد." if dry_run else "بازیابی افزایشی با موفقیت انجام شد."
            return True, action, summary
        else:
            return False, "خطا در بازیابی داده‌ها.", summary
    except Exception as e:
        logger.error(f"restore_from_file error: {e}")
        return False, "ساختار فایل پشتیبان معتبر نیست؛ جزئیات خطا ثبت شد.", None

# ============================================================
# ارسال پیام و عکس
# ============================================================
def is_super_admin_user(chat_id):
    user = find_user_by_telegram_id(chat_id)
    if user:
        return user[7]
    return False

def send_message(chat_id, text, reply_markup=None, remove_keyboard=False, parse_mode="Markdown", escape_user_text=False):
    bot_status = get_bot_status()
    if not bot_status:
        user = find_user_by_telegram_id(chat_id)
        if not user or not user[7]:
            send_maintenance_message(chat_id)
            return None
    if escape_user_text:
        text = escape_markdown(text)
    if len(text) > 4000:
        chunks = split_text_safely(text, 4000)
        results = []
        for chunk in chunks:
            results.append(send_message_chunk(chat_id, chunk, reply_markup, remove_keyboard, parse_mode, False))
        return results if all(result is not None for result in results) else None
    url = f"{BASE_URL}/sendMessage"
    payload = {"chat_id": chat_id, "text": text}
    if parse_mode:
        payload["parse_mode"] = parse_mode
    if remove_keyboard:
        payload["reply_markup"] = {"remove_keyboard": True}
    elif reply_markup:
        payload["reply_markup"] = reply_markup
    try:
        res = get_http_session().post(url, json=payload, timeout=30)
        if res.status_code == 200:
            return res.json()
        else:
            logger.error(f"sendMessage failed: {res.status_code}")
            if parse_mode:
                payload.pop("parse_mode", None)
                retry_res = get_http_session().post(url, json=payload, timeout=30)
                if retry_res.status_code == 200:
                    return retry_res.json()
            return None
    except requests.exceptions.Timeout:
        logger.error(f"sendMessage timeout for chat_id {chat_id}")
        send_message_chunk(chat_id, "⏳ ارسال گزارش با تأخیر مواجه شد. لطفاً چند لحظه صبر کنید و دوباره تلاش کنید.", remove_keyboard=True)
        return None
    except Exception as e:
        logger.error(f"sendMessage error: {e}")
        return None

def send_message_chunk(chat_id, text, reply_markup=None, remove_keyboard=False, parse_mode="Markdown", escape_user_text=False):
    if escape_user_text:
        text = escape_markdown(text)
    url = f"{BASE_URL}/sendMessage"
    payload = {"chat_id": chat_id, "text": text}
    if parse_mode:
        payload["parse_mode"] = parse_mode
    if remove_keyboard:
        payload["reply_markup"] = {"remove_keyboard": True}
    elif reply_markup:
        payload["reply_markup"] = reply_markup
    try:
        res = get_http_session().post(url, json=payload, timeout=30)
        if res.status_code == 200:
            return res.json()
        if parse_mode:
            payload.pop("parse_mode", None)
            retry_res = get_http_session().post(url, json=payload, timeout=30)
            return retry_res.json() if retry_res.status_code == 200 else None
        return None
    except requests.exceptions.Timeout:
        logger.error(f"send_message_chunk timeout for chat_id {chat_id}")
    except Exception as e:
        logger.error(f"send_message_chunk error: {e}")

def send_photo(chat_id, photo_bytes, caption="", reply_markup=None):
    bot_status = get_bot_status()
    if not bot_status:
        user = find_user_by_telegram_id(chat_id)
        if not user or not user[7]:
            send_maintenance_message(chat_id)
            return None
    if not photo_bytes:
        logger.error("send_photo: photo_bytes is None or empty")
        return None
    url = f"{BASE_URL}/sendPhoto"
    files = {'photo': ('chart.png', photo_bytes, 'image/png')}
    data = {'chat_id': chat_id, 'caption': caption[:1024]}
    if reply_markup:
        data['reply_markup'] = json.dumps(reply_markup)
    try:
        res = get_http_session().post(url, data=data, files=files, timeout=30)
        if res.status_code == 200:
            return res.json()
        else:
            logger.error(f"sendPhoto failed: {res.status_code}")
            return None
    except Exception as e:
        logger.error(f"sendPhoto error: {e}")
        return None

def send_campaign_photo(chat_id, photo_bytes, mime_type, file_name, caption=""):
    """Send an optional campaign image without assuming that it is a chart PNG."""
    if not photo_bytes or not str(mime_type or '').startswith('image/'):
        return None
    files = {'photo': (file_name or 'message-image', photo_bytes, mime_type)}
    data = {'chat_id': chat_id, 'caption': caption[:1024]}
    try:
        response = get_http_session().post(f"{BASE_URL}/sendPhoto", data=data, files=files, timeout=45)
        payload = response.json() if response.status_code == 200 else None
        return payload if payload and payload.get('ok', True) else None
    except Exception:
        logger.exception("Campaign photo send failed for chat_id=%s", chat_id)
        return None

def send_campaign_text(chat_id, text):
    """Campaigns are administrative notices and must not turn into maintenance messages."""
    try:
        response = get_http_session().post(
            f"{BASE_URL}/sendMessage", json={'chat_id': chat_id, 'text': text}, timeout=30
        )
        payload = response.json() if response.status_code == 200 else None
        return payload if payload and payload.get('ok', True) else None
    except Exception:
        logger.exception("Campaign text send failed for chat_id=%s", chat_id)
        return None

# ============================================================
# مرکز پیام‌رسانی سوپرادمین
# ============================================================
def get_messaging_center_keyboard():
    return {
        "keyboard": [
            [{"text": "🎉 پیام مناسبتی"}, {"text": "📢 اطلاعیه"}],
            [{"text": "🚨 پیام فوری"}],
            [{"text": "📝 پیش‌نویس‌ها"}, {"text": "📜 تاریخچه پیام‌ها"}],
            [{"text": "⛔ لغو پیام زمان‌بندی‌شده"}],
            [{"text": "🔁 ارسال مجدد ناموفق‌ها"}],
            [{"text": "🔙 بازگشت به منوی اصلی"}],
        ],
        "resize_keyboard": True,
    }

def get_campaign_audience_keyboard():
    return {
        "keyboard": [
            [{"text": "👥 همه کاربران"}, {"text": "🏢 همه معاونان"}],
            [{"text": "👤 همه ادمین‌ها"}, {"text": "🏬 انتخاب شعب"}],
            [{"text": "☑️ انتخاب دستی کاربران"}],
            [{"text": "🔙 انصراف"}],
        ],
        "resize_keyboard": True,
    }

def get_campaign_recipients(audience_type, audience_config=None):
    config = audience_config or {}
    clauses = ["u.telegram_id IS NOT NULL", "u.is_active=TRUE"]
    params = []
    if audience_type == 'deputies':
        clauses.append("u.role='deputy'")
    elif audience_type == 'admins':
        clauses.append("(u.role IN ('admin','super_admin') OR u.is_super_admin=TRUE)")
    elif audience_type == 'branches':
        branch_ids = [int(value) for value in config.get('branch_ids', [])]
        if not branch_ids:
            return []
        clauses.append("u.branch_id=ANY(%s)")
        params.append(branch_ids)
    elif audience_type == 'manual':
        user_ids = [int(value) for value in config.get('user_ids', [])]
        if not user_ids:
            return []
        clauses.append("u.id=ANY(%s)")
        params.append(user_ids)
    elif audience_type != 'all':
        return []
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute(f"""SELECT u.id,u.telegram_id,u.full_name,u.title,u.role,b.name
                FROM users u LEFT JOIN branches b ON b.id=u.branch_id
                WHERE {' AND '.join(clauses)} ORDER BY u.full_name""", params)
            return cur.fetchall()
    except Exception:
        logger.exception("Campaign recipient resolution failed")
        if conn:
            conn.rollback()
        return []
    finally:
        if conn:
            return_db_connection(conn)

def download_bale_media(file_id, max_bytes=MAX_CAMPAIGN_MEDIA_BYTES):
    try:
        metadata = get_http_session().get(f"{BASE_URL}/getFile", params={"file_id": file_id}, timeout=30)
        if metadata.status_code != 200 or not metadata.json().get('ok'):
            return False, "اطلاعات تصویر دریافت نشد.", None
        file_path = metadata.json().get('result', {}).get('file_path')
        if not file_path:
            return False, "مسیر تصویر دریافت نشد.", None
        response = get_http_session().get(f"https://tapi.bale.ai/file/bot{BOT_TOKEN}/{file_path}", timeout=60, stream=True)
        if response.status_code != 200:
            return False, "دریافت تصویر ناموفق بود.", None
        declared_size = int(response.headers.get('Content-Length') or 0)
        if declared_size > max_bytes:
            response.close()
            return False, "حجم تصویر بیشتر از ۵ مگابایت است.", None
        chunks, total = [], 0
        for chunk in response.iter_content(64 * 1024):
            if not chunk:
                continue
            total += len(chunk)
            if total > max_bytes:
                response.close()
                return False, "حجم تصویر بیشتر از ۵ مگابایت است.", None
            chunks.append(chunk)
        response.close()
        return True, "", b''.join(chunks)
    except Exception:
        logger.exception("Campaign media download failed")
        return False, "دریافت تصویر انجام نشد.", None

def cleanup_stale_actual_excel_files(max_age_seconds=6 * 60 * 60):
    """Delete only abandoned temp files created by this Excel-import feature."""
    temp_dir = tempfile.gettempdir()
    now = time_module.time()
    try:
        for entry in os.scandir(temp_dir):
            if not entry.is_file(follow_symlinks=False):
                continue
            if not entry.name.startswith(ACTUAL_EXCEL_TEMP_PREFIX):
                continue
            try:
                if now - entry.stat(follow_symlinks=False).st_mtime > max_age_seconds:
                    os.unlink(entry.path)
            except OSError:
                logger.warning("Could not remove stale actual-stats temp file")
    except OSError:
        logger.warning("Could not scan temp directory for stale Excel files")

def download_bale_excel_to_temp(file_id):
    """Stream an Excel document to a temporary file without an application size cap."""
    cleanup_stale_actual_excel_files()
    temp_path = None
    response = None
    try:
        metadata = get_http_session().get(
            f"{BASE_URL}/getFile", params={"file_id": file_id}, timeout=30
        )
        if metadata.status_code != 200 or not metadata.json().get('ok'):
            return False, "اطلاعات فایل از بله دریافت نشد.", None, None
        remote_path = metadata.json().get('result', {}).get('file_path')
        if not remote_path:
            return False, "مسیر فایل از بله دریافت نشد.", None, None

        temp_file = tempfile.NamedTemporaryFile(
            mode='wb', prefix=ACTUAL_EXCEL_TEMP_PREFIX, suffix='.xlsx', delete=False
        )
        temp_path = temp_file.name
        digest = hashlib.sha256()
        try:
            response = get_http_session().get(
                f"https://tapi.bale.ai/file/bot{BOT_TOKEN}/{remote_path}",
                timeout=(30, 120), stream=True
            )
            if response.status_code != 200:
                return False, "دریافت فایل Excel ناموفق بود.", temp_path, None
            for chunk in response.iter_content(1024 * 1024):
                if not chunk:
                    continue
                temp_file.write(chunk)
                digest.update(chunk)
            temp_file.flush()
            os.fsync(temp_file.fileno())
        finally:
            temp_file.close()
        if os.path.getsize(temp_path) == 0:
            return False, "فایل دریافت‌شده خالی است.", temp_path, None
        return True, "", temp_path, digest.hexdigest()
    except Exception:
        logger.exception("Excel document download failed")
        return False, "دریافت فایل Excel انجام نشد.", temp_path, None
    finally:
        if response is not None:
            response.close()

def parse_campaign_schedule(value):
    normalized = normalize_digits(value).strip()
    match = re.fullmatch(r'(\d{4})/(\d{2})/(\d{2})\s+(\d{1,2}):(\d{2})', normalized)
    if not match:
        return None
    year, month, day, hour, minute = map(int, match.groups())
    if hour > 23 or minute > 59:
        return None
    try:
        gregorian = jdatetime.datetime(year, month, day, hour, minute).togregorian()
        scheduled = gregorian.replace(tzinfo=IRAN_TZ)
        return scheduled if scheduled > get_iran_time() else None
    except (ValueError, TypeError):
        return None

def format_campaign_preview(draft, recipients, scheduled_at=None):
    type_names = {'occasion': '🎉 مناسبتی', 'announcement': '📢 اطلاعیه', 'urgent': '🚨 فوری'}
    audience_names = {'all': 'همه کاربران', 'deputies': 'همه معاونان', 'admins': 'همه ادمین‌ها',
                      'branches': 'شعب انتخابی', 'manual': 'کاربران انتخابی'}
    names = '، '.join(row[2] for row in recipients[:8])
    if len(recipients) > 8:
        names += f" و {len(recipients)-8} نفر دیگر"
    if scheduled_at:
        local_time = scheduled_at.astimezone(IRAN_TZ)
        shamsi_time = jdatetime.datetime.fromgregorian(datetime=local_time.replace(tzinfo=None))
        when = f"{shamsi_time.year:04d}/{shamsi_time.month:02d}/{shamsi_time.day:02d} {shamsi_time.hour:02d}:{shamsi_time.minute:02d}"
    else:
        when = 'هنوز تعیین نشده'
    return (
        "📋 **پیش‌نمایش پیام**\n━━━━━━━━━━━━━━━━━━\n"
        f"نوع: {type_names.get(draft['message_type'])}\n"
        f"مخاطب: {audience_names.get(draft['audience_type'])}\n"
        f"تعداد دریافت‌کننده: {len(recipients)}\n"
        f"شخصی‌سازی: {'فعال' if draft.get('personalize') else 'غیرفعال'}\n"
        f"تصویر: {'دارد' if draft.get('media_base64') else 'ندارد'}\n"
        f"زمان: {when}\n"
        f"گیرندگان: {names or '—'}\n\n"
        f"متن:\n{draft['message_text']}"
    )

def personalize_campaign_text(text, recipient, enabled):
    if not enabled:
        return text
    _, _, full_name, title, _, branch_name = recipient
    return (text.replace('{name}', full_name or '')
                .replace('{title}', title or '')
                .replace('{branch}', branch_name or ''))

def create_message_campaign(draft, created_by, scheduled_at, initial_status='scheduled'):
    recipients = get_campaign_recipients(draft['audience_type'], draft.get('audience_config'))
    if not recipients:
        return False, "هیچ کاربر متصل به ربات در گروه انتخاب‌شده وجود ندارد.", None
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            campaign_uuid = str(uuid.uuid4())
            cur.execute("""INSERT INTO message_campaigns
                (campaign_uuid,message_type,message_text,media_base64,media_mime,media_name,
                 audience_type,audience_config,personalize,status,scheduled_at,created_by,total_recipients)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s::jsonb,%s,%s,%s,%s,%s)
                RETURNING id""", (
                    campaign_uuid, draft['message_type'], draft['message_text'],
                    draft.get('media_base64'), draft.get('media_mime'), draft.get('media_name'),
                    draft['audience_type'], json.dumps(draft.get('audience_config') or {}),
                    bool(draft.get('personalize')), initial_status, scheduled_at, created_by, len(recipients)
                ))
            campaign_id = cur.fetchone()[0]
            execute_values(cur, """INSERT INTO message_deliveries
                (campaign_id,user_id,telegram_id,recipient_name)
                VALUES %s ON CONFLICT (campaign_id,user_id) DO NOTHING""",
                [(campaign_id, row[0], row[1], row[2]) for row in recipients], page_size=200)
        conn.commit()
        return True, "پیام ثبت شد.", {'id': campaign_id, 'uuid': campaign_uuid, 'count': len(recipients)}
    except Exception:
        logger.exception("Campaign creation failed")
        if conn:
            conn.rollback()
        return False, "ثبت پیام انجام نشد؛ هیچ ارسال ناقصی آغاز نشده است.", None
    finally:
        if conn:
            return_db_connection(conn)

def schedule_draft_campaign(campaign_id, scheduled_at, requested_by):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""UPDATE message_campaigns SET status='scheduled',scheduled_at=%s
                WHERE id=%s AND status='draft' RETURNING campaign_uuid,total_recipients""", (scheduled_at, campaign_id))
            row = cur.fetchone()
            if not row:
                conn.rollback()
                return False, "پیش‌نویس معتبر نیست یا قبلاً زمان‌بندی شده است.", None
        conn.commit()
        log_user_activity(requested_by, 'campaign_schedule_draft', f"campaign={row[0]}")
        return True, "پیش‌نویس در صف ارسال قرار گرفت.", {'uuid': row[0], 'count': row[1]}
    except Exception:
        if conn:
            conn.rollback()
        logger.exception("Draft campaign scheduling failed")
        return False, "زمان‌بندی پیش‌نویس انجام نشد.", None
    finally:
        if conn:
            return_db_connection(conn)

def _campaign_delivery_payload(campaign_id, user_id):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""SELECT d.id,d.telegram_id,u.id,u.telegram_id,u.full_name,u.title,u.role,b.name,
                c.message_text,c.personalize,c.media_base64,c.media_mime,c.media_name
                FROM message_deliveries d JOIN message_campaigns c ON c.id=d.campaign_id
                JOIN users u ON u.id=d.user_id LEFT JOIN branches b ON b.id=u.branch_id
                WHERE d.campaign_id=%s AND d.user_id=%s AND u.is_active=TRUE""", (campaign_id, user_id))
            return cur.fetchone()
    finally:
        if conn:
            return_db_connection(conn)

def process_message_campaign(campaign_id):
    """Idempotent delivery worker; a recipient is claimed before network sending."""
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("SELECT status,scheduled_at FROM message_campaigns WHERE id=%s FOR UPDATE", (campaign_id,))
            row = cur.fetchone()
            if not row or row[0] not in ('scheduled', 'partial') or (row[1] and row[1] > get_iran_time()):
                conn.rollback()
                return
            cur.execute("UPDATE message_campaigns SET status='sending',started_at=COALESCE(started_at,CURRENT_TIMESTAMP) WHERE id=%s", (campaign_id,))
            cur.execute("SELECT user_id FROM message_deliveries WHERE campaign_id=%s AND status IN ('pending','failed') ORDER BY id", (campaign_id,))
            user_ids = [item[0] for item in cur.fetchall()]
        conn.commit()
    except Exception:
        logger.exception("Campaign start failed id=%s", campaign_id)
        if conn:
            conn.rollback()
        return
    finally:
        if conn:
            return_db_connection(conn)

    for user_id in user_ids:
        conn = None
        claim = None
        try:
            conn = get_db_connection()
            with conn.cursor() as cur:
                cur.execute("""UPDATE message_deliveries SET status='sending',attempts=attempts+1,last_error=NULL
                    WHERE campaign_id=%s AND user_id=%s AND status IN ('pending','failed') RETURNING id""", (campaign_id, user_id))
                claim = cur.fetchone()
            conn.commit()
        except Exception:
            if conn:
                conn.rollback()
        finally:
            if conn:
                return_db_connection(conn)
        if not claim:
            continue
        success = False
        error_text = None
        try:
            payload = _campaign_delivery_payload(campaign_id, user_id)
            if not payload:
                raise RuntimeError("اطلاعات گیرنده در دسترس نیست")
            _, chat_id, db_user_id, _, name, title, role, branch, body, personalized, media_b64, mime, media_name = payload
            recipient = (db_user_id, chat_id, name, title, role, branch)
            rendered = personalize_campaign_text(body, recipient, personalized)
            if media_b64:
                media = base64.b64decode(media_b64, validate=True)
                photo_result = send_campaign_photo(chat_id, media, mime, media_name, rendered if len(rendered) <= 1024 else '')
                if not photo_result:
                    raise RuntimeError("ارسال تصویر ناموفق بود")
                result = True if len(rendered) <= 1024 else send_campaign_text(chat_id, rendered)
            else:
                result = send_campaign_text(chat_id, rendered)
            success = bool(result)
            if not success:
                error_text = "پاسخ موفق از سرویس پیام دریافت نشد"
        except Exception as exc:
            error_text = str(exc)[:500]
            logger.exception("Campaign delivery failed campaign=%s user=%s", campaign_id, user_id)
        conn = None
        try:
            conn = get_db_connection()
            with conn.cursor() as cur:
                if success:
                    cur.execute("UPDATE message_deliveries SET status='sent',sent_at=CURRENT_TIMESTAMP,last_error=NULL WHERE campaign_id=%s AND user_id=%s", (campaign_id, user_id))
                else:
                    cur.execute("UPDATE message_deliveries SET status='failed',last_error=%s WHERE campaign_id=%s AND user_id=%s", (error_text, campaign_id, user_id))
            conn.commit()
        finally:
            if conn:
                return_db_connection(conn)
        time.sleep(0.12)

    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""SELECT COUNT(*) FILTER(WHERE status='sent'),COUNT(*) FILTER(WHERE status='failed'),
                COUNT(*) FILTER(WHERE status IN ('pending','sending')) FROM message_deliveries WHERE campaign_id=%s""", (campaign_id,))
            sent, failed, unfinished = cur.fetchone()
            final_status = 'completed' if failed == 0 and unfinished == 0 else 'partial'
            cur.execute("""UPDATE message_campaigns SET status=%s,sent_count=%s,failed_count=%s,
                completed_at=CURRENT_TIMESTAMP WHERE id=%s AND status='sending'""", (final_status, sent, failed, campaign_id))
        conn.commit()
    except Exception:
        logger.exception("Campaign finalization failed id=%s", campaign_id)
        if conn:
            conn.rollback()
    finally:
        if conn:
            return_db_connection(conn)

def process_due_message_campaigns():
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("SELECT id FROM message_campaigns WHERE status='scheduled' AND scheduled_at<=CURRENT_TIMESTAMP ORDER BY scheduled_at LIMIT 10")
            campaign_ids = [row[0] for row in cur.fetchall()]
    except Exception:
        logger.exception("Reading scheduled campaigns failed")
        return
    finally:
        if conn:
            return_db_connection(conn)
    for campaign_id in campaign_ids:
        process_message_campaign(campaign_id)

def get_campaign_history(limit=10):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""SELECT id,campaign_uuid,message_type,status,total_recipients,sent_count,failed_count,
                scheduled_at,created_at,LEFT(message_text,80) FROM message_campaigns ORDER BY id DESC LIMIT %s""", (limit,))
            return cur.fetchall()
    finally:
        if conn:
            return_db_connection(conn)

def cancel_scheduled_campaign(campaign_id, cancelled_by):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""UPDATE message_campaigns SET status='cancelled',cancelled_at=CURRENT_TIMESTAMP
                WHERE id=%s AND status='scheduled' RETURNING campaign_uuid""", (campaign_id,))
            row = cur.fetchone()
            if not row:
                conn.rollback()
                return False, "این پیام زمان‌بندی‌شده نیست یا ارسال آن آغاز شده است."
            cur.execute("UPDATE message_deliveries SET status='cancelled' WHERE campaign_id=%s AND status='pending'", (campaign_id,))
        conn.commit()
        log_user_activity(cancelled_by, 'campaign_cancel', f"campaign={row[0]}")
        return True, f"پیام {row[0]} لغو شد."
    except Exception:
        if conn:
            conn.rollback()
        logger.exception("Campaign cancellation failed")
        return False, "لغو پیام انجام نشد."
    finally:
        if conn:
            return_db_connection(conn)

def retry_failed_campaign(campaign_id, requested_by):
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("SELECT campaign_uuid,status FROM message_campaigns WHERE id=%s FOR UPDATE", (campaign_id,))
            row = cur.fetchone()
            if not row or row[1] != 'partial':
                conn.rollback()
                return False, "این پیام تحویل ناموفق قابل ارسال مجدد ندارد."
            cur.execute("UPDATE message_deliveries SET status='pending',last_error=NULL WHERE campaign_id=%s AND status='failed'", (campaign_id,))
            if cur.rowcount == 0:
                conn.rollback()
                return False, "گیرنده ناموفقی یافت نشد."
            cur.execute("UPDATE message_campaigns SET status='scheduled',scheduled_at=CURRENT_TIMESTAMP,failed_count=0 WHERE id=%s", (campaign_id,))
        conn.commit()
        log_user_activity(requested_by, 'campaign_retry', f"campaign={row[0]}")
        return True, "ارسال مجدد فقط برای گیرندگان ناموفق در صف قرار گرفت."
    except Exception:
        if conn:
            conn.rollback()
        logger.exception("Campaign retry failed")
        return False, "ثبت ارسال مجدد انجام نشد."
    finally:
        if conn:
            return_db_connection(conn)

def send_maintenance_message(chat_id):
    msg = "🔧 با عرض پوزش، ربات در حال بروزرسانی می‌باشد.\nلطفاً بعداً مجدداً تلاش کنید."
    url = f"{BASE_URL}/sendMessage"
    payload = {"chat_id": chat_id, "text": msg, "parse_mode": "Markdown", "reply_markup": {"remove_keyboard": True}}
    try:
        get_http_session().post(url, json=payload, timeout=10)
    except Exception:
        pass

# ============================================================
# کیبوردها
# ============================================================
def get_deputy_keyboard():
    return {
        "keyboard": [
            [{"text": "🤖 همیار وصول مطالبات"}],
            [{"text": "💰 ثبت وصولی روزانه"}, {"text": "📊 گزارش وصولی"}],
            [{"text": "📈 مقایسه عملکرد"}, {"text": "📋 مشاهده ثبت امروز"}],
            [{"text": "📅 گزارش تاریخ خاص"}, {"text": "📊 تاریخچه کامل"}],
            [{"text": "📝 ثبت یادداشت"}, {"text": "📋 مشاهده یادداشت‌ها"}],
            [{"text": "📝 ثبت مشکل"}, {"text": "ℹ️ درباره توسعه‌دهنده"}],
            [{"text": "🔙 خروج"}, {"text": "❓ راهنما"}]
        ],
        "resize_keyboard": True
    }

def get_admin_keyboard():
    return {
        "keyboard": [
            [{"text": "👁 مشاهده وضعیت همیار"}],
            [{"text": "📊 گزارش امروز"}, {"text": "📈 گزارش ۱۰ روز اخیر"}],
            [{"text": "🏆 رتبه‌بندی شعب"}, {"text": "💹 آمار مفصل امروز"}],
            [{"text": "📉 مقایسه روزانه"}, {"text": "🎯 تحلیل مدیریتی"}],
            [{"text": "📅 گزارش تاریخ خاص"}, {"text": "📊 بهترین/بدترین روز"}],
            [{"text": "📊 گزارش روند شعبه"}, {"text": "📋 عملکرد معاونان"}],
            [{"text": "👥 عملکرد همکاران"}, {"text": "📝 مشاهده یادداشت‌ها"}],
            [{"text": "📊 گزارش تطبیقی"}, {"text": "📈 پیش‌بینی عملکرد"}],
            [{"text": "📊 نمودار استان"}, {"text": "📊 نمودار شعبه"}],
            [{"text": "📊 نمودار تحلیلی"}, {"text": "📊 مقایسه انطباق"}],
            [{"text": "📝 ثبت مشکل"}, {"text": "ℹ️ درباره توسعه‌دهنده"}],
            [{"text": "🔙 خروج"}, {"text": "❓ راهنما"}]
        ],
        "resize_keyboard": True
    }

def get_super_admin_keyboard():
    return {
        "keyboard": [
            [{"text": "👁 مشاهده وضعیت همیار"}],
            [{"text": "👥 مدیریت کاربران"}, {"text": "📊 مدیریت گزارش‌ها"}],
            [{"text": "👥 مدیریت معاونین"}, {"text": "📋 مشاهده لاگ‌ها"}],
            [{"text": "📣 مرکز پیام‌رسانی"}],
            [{"text": "🔁 مدیریت جابه‌جایی معاونان"}],
            [{"text": "📊 گزارش امروز"}, {"text": "📈 گزارش ۱۰ روز اخیر"}],
            [{"text": "🏆 رتبه‌بندی شعب"}, {"text": "💹 آمار مفصل امروز"}],
            [{"text": "🎯 تحلیل مدیریتی"}, {"text": "📅 گزارش تاریخ خاص"}],
            [{"text": "📊 بهترین/بدترین روز"}, {"text": "📊 گزارش روند شعبه"}],
            [{"text": "📋 عملکرد معاونان"}, {"text": "👥 عملکرد همکاران"}],
            [{"text": "📝 مشاهده یادداشت‌ها"}, {"text": "📋 لاگ ورود/خروج"}],
            [{"text": "🔧 کنترل خودکار"}, {"text": "📅 مدیریت تعطیلات"}],
            [{"text": "🔄 ریست گزارش‌ها"}],
            [{"text": "⚙️ مدیریت مشکلات"}, {"text": "📊 گزارش هفتگی"}],
            [{"text": "📊 گزارش ماهانه"}, {"text": "📊 گزارش تطبیقی"}],
            [{"text": "📈 پیش‌بینی عملکرد"}, {"text": "📊 نمودار استان"}],
            [{"text": "📊 نمودار شعبه"}, {"text": "📊 نمودار تحلیلی"}],
            [{"text": "📊 مقایسه انطباق"}, {"text": "📊 ثبت آمار واقعی"}],
            [{"text": "📝 ثبت مشکل"}, {"text": "🔧 وضعیت ربات"}],
            [{"text": "ℹ️ درباره توسعه‌دهنده"}, {"text": "🔙 خروج"}],
            [{"text": "❓ راهنما"}],
            [{"text": "🏅 رتبه‌بندی دقت معاونان"}, {"text": "📈 روند دقت شعبه"}],
            [{"text": "📊 بهترین/بدترین دقت روز"}, {"text": "📊 مقایسه عملکرد شعبه با استان"}],
            [{"text": "⏰ تحلیل تاخیر معاونان"}],
            [{"text": "🎯 مدیریت اهداف وصولی"}, {"text": "📊 گزارش پیشرفت اهداف"}],
            [{"text": "🏆 رتبه‌بندی تحقق هدف"}],
            [{"text": "📊 مرکز گزارش‌های مدیریتی"}],
            [{"text": "📈 مرکز گزارش‌های تصویری"}],
            [{"text": "🩺 سلامت دیتابیس"}, {"text": "📦 آمار حجم جداول"}],
            [{"text": "💾 پشتیبان‌گیری از داده‌ها"}, {"text": "📂 بازیابی داده‌ها"}]
        ],
        "resize_keyboard": True
    }

def get_cancel_keyboard():
    return {"keyboard": [[{"text": "🔙 انصراف"}]], "resize_keyboard": True}

def get_actual_stats_method_keyboard():
    return {
        "keyboard": [
            [{"text": "✍️ ثبت دستی آمار واقعی"}],
            [{"text": "📊 ثبت آمار واقعی از Excel"}],
            [{"text": "🔙 انصراف"}],
        ],
        "resize_keyboard": True,
    }

def get_actual_excel_confirm_keyboard():
    return {
        "keyboard": [
            [{"text": "✅ تأیید نهایی و ثبت"}],
            [{"text": "✏️ اصلاح مبلغ یک شعبه"}, {"text": "📅 اصلاح تاریخ"}],
            [{"text": "🔄 ارسال فایل جدید"}],
            [{"text": "🔙 انصراف"}],
        ],
        "resize_keyboard": True,
    }

# ============================================================
# توابع ارسال خودکار
# ============================================================
def send_reminder_to_deputy(chat_id, branch_name):
    msg = f"⏰ یادآوری: شما تا ساعت ۱۵ امروز گزارش وصول شعبه {branch_name} را ثبت نکرده‌اید. لطفاً هرچه سریعتر اقدام فرمایید."
    send_message(chat_id, msg)

def send_reminder_to_admin(chat_id, unreported_list):
    if not unreported_list:
        return
    msg = "📋 **شعب ثبت‌نشده امروز**\n━━━━━━━━━━━━━━━━━━\n"
    for branch in unreported_list:
        msg += f"🏢 {branch[1]} (معاون: {branch[2] or 'نامشخص'})\n"
    send_message(chat_id, msg)

def send_daily_report_to_admins():
    if not get_bot_status() or not get_auto_report_status():
        return
    shamsi_today = get_shamsi_date()
    if is_holiday(shamsi_today):
        return
    analysis = get_today_performance_analysis()
    if not analysis:
        return
    admins = get_all_admins()
    if not admins:
        return
    msg = f"📊 **گزارش پایان روز** - {get_shamsi_date_formatted(shamsi_today)}\n"
    msg += f"━━━━━━━━━━━━━━━━━━\n"
    msg += f"💰 کل وصول استان: {analysis['today_total']//1_000_000:,.0f} میلیون ریال\n"
    msg += f"🏢 تعداد شعب ثبت‌کننده: {analysis['branches_count']}\n"
    msg += f"👤 سهم معاونین: {analysis['deputy_total']//1_000_000:,.0f} میلیون ریال\n"
    msg += f"👥 سهم همکاران: {analysis['others_total']//1_000_000:,.0f} میلیون ریال\n\n"
    if analysis['branch_data']:
        msg += "🏆 **۵ شعبه برتر امروز**\n"
        for i, (name, amount) in enumerate(analysis['branch_data'][:5], 1):
            msg += f"{i}. {name}: {amount//1_000_000:,.0f} میلیون ریال\n"
    msg += "\n📈 **تحلیل مدیریتی**\n"
    msg += generate_management_analysis(analysis)
    for admin in admins:
        admin_id = admin[1]
        if admin_id:
            send_message(admin_id, msg)

def check_and_send_reminders():
    if not get_bot_status() or not get_auto_reminder_status():
        return
    shamsi_today = get_shamsi_date()
    if is_holiday(shamsi_today):
        return
    unreported = get_unreported_branches()
    if unreported:
        for branch in unreported:
            branch_id, name, deputy_name, deputy_chat_id = branch
            if deputy_chat_id:
                send_reminder_to_deputy(deputy_chat_id, name)
        admins = get_all_admins()
        for admin in admins:
            admin_id = admin[1]
            if admin_id:
                send_reminder_to_admin(admin_id, unreported)

def check_and_send_drop_alerts():
    if not get_bot_status() or not get_auto_alert_status():
        return
    shamsi_today = get_shamsi_date()
    if is_holiday(shamsi_today):
        return
    drops = get_drop_alert_branches()
    if drops:
        admins = get_all_admins()
        for admin in admins:
            admin_id = admin[1]
            if not admin_id:
                continue
            msg = "⚠️ **هشدار افت عملکرد**\n━━━━━━━━━━━━━━━━━━\n"
            for drop in drops:
                msg += f"🏢 شعبه {drop['name']}\n"
                msg += f"   امروز: {drop['today']//1_000_000:,.0f} میلیون ریال\n"
                msg += f"   میانگین هفته: {drop['weekly_avg']//1_000_000:,.0f} میلیون ریال\n"
                msg += f"   📉 افت: {drop['drop_percent']}%\n\n"
            send_message(admin_id, msg)

def check_and_auto_score():
    if not get_bot_status() or not get_auto_scoring_status():
        return
    shamsi_today = get_shamsi_date()
    if is_holiday(shamsi_today):
        return
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT c.id, c.branch_id, c.deputy_amount, c.others_amount, c.created_at, c.shamsi_date
                FROM collections c
                LEFT JOIN scores s ON c.id = s.collection_id
                WHERE c.shamsi_date = %s AND s.id IS NULL
            """, (shamsi_today,))
            collections_without_score = cur.fetchall()
            for col in collections_without_score:
                col_id, branch_id, deputy_amount, others_amount, created_at, shamsi_date = col
                score = calculate_score(created_at, deputy_amount, others_amount, branch_id, shamsi_date)
                save_score(col_id, score)
    except Exception as e:
        logger.error(f"check_and_auto_score error: {e}")
        if conn:
            conn.rollback()
    finally:
        if conn:
            return_db_connection(conn)

def generate_weekly_report():
    shamsi_today = get_shamsi_date()
    shamsi_week_ago = get_shamsi_date(-6)
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    b.id,
                    b.name,
                    COUNT(c.id) as report_count,
                    COALESCE(SUM(c.total_amount), 0) as total_amount,
                    COALESCE(AVG(c.total_amount), 0) as avg_amount,
                    COALESCE(COUNT(CASE WHEN (EXTRACT(HOUR FROM c.created_at AT TIME ZONE 'Asia/Tehran') < 16)
                                OR (EXTRACT(HOUR FROM c.created_at AT TIME ZONE 'Asia/Tehran') = 16
                                    AND EXTRACT(MINUTE FROM c.created_at AT TIME ZONE 'Asia/Tehran') <= 30)
                             THEN 1 END), 0) as on_time_count,
                    COALESCE(COUNT(CASE WHEN (EXTRACT(HOUR FROM c.created_at AT TIME ZONE 'Asia/Tehran') > 16)
                                OR (EXTRACT(HOUR FROM c.created_at AT TIME ZONE 'Asia/Tehran') = 16
                                    AND EXTRACT(MINUTE FROM c.created_at AT TIME ZONE 'Asia/Tehran') > 30)
                             THEN 1 END), 0) as late_count
                FROM branches b
                LEFT JOIN collections c ON b.id = c.branch_id AND c.shamsi_date >= %s AND c.shamsi_date <= %s
                GROUP BY b.id, b.name
                ORDER BY total_amount DESC
            """, (shamsi_week_ago, shamsi_today))
            return cur.fetchall()
    except Exception as e:
        logger.error(f"generate_weekly_report: {e}")
        if conn:
            conn.rollback()
        return []
    finally:
        if conn:
            return_db_connection(conn)

def generate_monthly_report():
    shamsi_today = get_shamsi_date()
    shamsi_month_start, shamsi_month_end = get_shamsi_month_range()
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    b.id,
                    b.name,
                    COUNT(c.id) as report_count,
                    COALESCE(SUM(c.total_amount), 0) as total_amount,
                    COALESCE(AVG(c.total_amount), 0) as avg_amount,
                    COALESCE(COUNT(CASE WHEN (EXTRACT(HOUR FROM c.created_at AT TIME ZONE 'Asia/Tehran') < 16)
                                OR (EXTRACT(HOUR FROM c.created_at AT TIME ZONE 'Asia/Tehran') = 16
                                    AND EXTRACT(MINUTE FROM c.created_at AT TIME ZONE 'Asia/Tehran') <= 30)
                             THEN 1 END), 0) as on_time_count,
                    COALESCE(COUNT(CASE WHEN (EXTRACT(HOUR FROM c.created_at AT TIME ZONE 'Asia/Tehran') > 16)
                                OR (EXTRACT(HOUR FROM c.created_at AT TIME ZONE 'Asia/Tehran') = 16
                                    AND EXTRACT(MINUTE FROM c.created_at AT TIME ZONE 'Asia/Tehran') > 30)
                             THEN 1 END), 0) as late_count
                FROM branches b
                LEFT JOIN collections c ON b.id = c.branch_id AND c.shamsi_date >= %s AND c.shamsi_date <= %s
                GROUP BY b.id, b.name
                ORDER BY total_amount DESC
            """, (shamsi_month_start, shamsi_month_end))
            return cur.fetchall()
    except Exception as e:
        logger.error(f"generate_monthly_report: {e}")
        if conn:
            conn.rollback()
        return []
    finally:
        if conn:
            return_db_connection(conn)

def send_weekly_report_to_all():
    if not get_bot_status() or not get_weekly_report_status():
        return
    shamsi_today = get_shamsi_date()
    report_data = generate_weekly_report()
    if not report_data:
        return
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("SELECT telegram_id, full_name, role FROM users WHERE telegram_id IS NOT NULL AND is_active=TRUE")
            users = cur.fetchall()
    finally:
        if conn:
            return_db_connection(conn)
    msg = f"📊 **گزارش هفتگی عملکرد شعب**\n"
    msg += f"📅 بازه: {get_shamsi_date_formatted(get_shamsi_date(-6))} تا {get_shamsi_date_formatted(shamsi_today)}\n"
    msg += f"━━━━━━━━━━━━━━━━━━\n\n"
    total_all = sum(int(row[3] or 0) for row in report_data)
    for idx, row in enumerate(report_data[:10], 1):
        branch_id, name, count, total, avg, on_time, late = row
        msg += f"{idx}. 🏢 {name}\n"
        msg += f"   📊 تعداد گزارش: {count}\n"
        msg += f"   💰 کل وصول: {total//1_000_000:,.0f} میلیون ریال\n"
        msg += f"   📈 میانگین: {avg//1_000_000:,.0f} میلیون ریال\n"
        msg += f"   ⏰ ثبت به‌موقع: {on_time} (قبل ۱۶:۳۰) / دیر: {late}\n\n"
    msg += f"━━━━━━━━━━━━━━━━━━\n"
    msg += f"💰 جمع کل وصول استان: {total_all//1_000_000:,.0f} میلیون ریال"
    for user in users:
        chat_id = user[0]
        if chat_id:
            send_message(chat_id, msg)

def send_monthly_report_to_all(force=False):
    if not get_bot_status() or not get_monthly_report_status():
        return
    shamsi_today = get_shamsi_date()
    if not force and not is_last_day_of_shamsi_month(shamsi_today):
        return
    report_data = generate_monthly_report()
    if not report_data:
        return
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("SELECT telegram_id, full_name, role FROM users WHERE telegram_id IS NOT NULL AND is_active=TRUE")
            users = cur.fetchall()
    finally:
        if conn:
            return_db_connection(conn)
    msg = f"📊 **گزارش ماهانه عملکرد شعب**\n"
    month_start, month_end = get_shamsi_month_range()
    msg += f"📅 بازه: {get_shamsi_date_formatted(month_start)} تا {get_shamsi_date_formatted(month_end)}\n"
    msg += f"━━━━━━━━━━━━━━━━━━\n\n"
    total_all = sum(int(row[3] or 0) for row in report_data)
    for idx, row in enumerate(report_data[:10], 1):
        branch_id, name, count, total, avg, on_time, late = row
        msg += f"{idx}. 🏢 {name}\n"
        msg += f"   📊 تعداد گزارش: {count}\n"
        msg += f"   💰 کل وصول: {total//1_000_000:,.0f} میلیون ریال\n"
        msg += f"   📈 میانگین: {avg//1_000_000:,.0f} میلیون ریال\n"
        msg += f"   ⏰ ثبت به‌موقع: {on_time} (قبل ۱۶:۳۰) / دیر: {late}\n\n"
    msg += f"━━━━━━━━━━━━━━━━━━\n"
    msg += f"💰 جمع کل وصول استان: {total_all//1_000_000:,.0f} میلیون ریال"
    for user in users:
        chat_id = user[0]
        if chat_id:
            send_message(chat_id, msg)

# ============================================================
# توابع کمکی برای Threading (نمودارها)
# ============================================================
def generate_and_send_branch_chart(chat_id, branch_id, branch_name, role, is_super_admin):
    try:
        data = generate_branch_chart(branch_id, 10)
        if data:
            chart_bytes = generate_chart(data, f'روند ۱۰ روز اخیر شعبه {branch_name}', 'تاریخ', 'مبلغ (میلیون ریال)', 'bar')
            if chart_bytes:
                caption = f"📊 **نمودار روند شعبه {branch_name}**\n۱۰ روز اخیر"
                keyboard = get_keyboard(role, is_super_admin)
                send_photo(chat_id, chart_bytes, caption, keyboard)
            else:
                keyboard = get_keyboard(role, is_super_admin)
                send_message(chat_id, "❌ خطا در تولید نمودار.", keyboard)
        else:
            keyboard = get_keyboard(role, is_super_admin)
            send_message(chat_id, f"📭 هیچ داده‌ای برای شعبه {branch_name} یافت نشد.", keyboard)
    except Exception as e:
        logger.error(f"generate_and_send_branch_chart error: {e}")
        send_message(chat_id, "❌ خطا در تولید نمودار. لطفاً دوباره تلاش کنید.", get_cancel_keyboard())

def generate_and_send_province_chart(chat_id, role, is_super_admin):
    try:
        data = generate_province_chart(10)
        if data:
            chart_bytes = generate_chart(data, 'روند ۱۰ روز اخیر استان', 'تاریخ', 'مبلغ (میلیون ریال)', 'line')
            if chart_bytes:
                caption = f"📊 **نمودار روند وصول استان**\n۱۰ روز اخیر"
                keyboard = get_keyboard(role, is_super_admin)
                send_photo(chat_id, chart_bytes, caption, keyboard)
            else:
                keyboard = get_keyboard(role, is_super_admin)
                send_message(chat_id, "❌ خطا در تولید نمودار.", keyboard)
        else:
            keyboard = get_keyboard(role, is_super_admin)
            send_message(chat_id, "📊 داده‌های کافی برای نمودار وجود ندارد.", keyboard)
    except Exception as e:
        logger.error(f"generate_and_send_province_chart error: {e}")
        send_message(chat_id, "❌ خطا در تولید نمودار.", get_cancel_keyboard())

def generate_and_send_analytical_chart(chat_id, chart_key, role, is_super_admin):
    try:
        chart_title_map = {
            "branch_comparison": "مقایسه ۱۰ شعبه برتر (۱۰ روز اخیر)",
            "deputy_others_ratio": "نسبت وصول معاونین و همکاران (۱۰ روز اخیر)",
            "daily_trend": "روند روزانه وصول (۱۰ روز اخیر)",
            "match_analysis": "تحلیل انطباق با آمار واقعی (۱۰ روز اخیر)"
        }
        chart_y_label = {
            "branch_comparison": "مبلغ (میلیون ریال)",
            "deputy_others_ratio": "مبلغ (میلیون ریال)",
            "daily_trend": "مبلغ (میلیون ریال)",
            "match_analysis": "درصد تطابق"
        }
        chart_type = {
            "branch_comparison": "horizontal",
            "deputy_others_ratio": "pie",
            "daily_trend": "line",
            "match_analysis": "bar"
        }
        data = get_analytical_chart_data(chart_key, 10)
        if data and data['values'] and any(v > 0 for v in data['values']):
            chart_bytes = generate_chart(
                data,
                chart_title_map.get(chart_key, "نمودار تحلیلی"),
                "شعبه" if chart_key != "daily_trend" else "تاریخ",
                chart_y_label.get(chart_key, "مبلغ"),
                chart_type.get(chart_key, "bar")
            )
            if chart_bytes:
                caption = f"📊 {chart_title_map.get(chart_key, 'نمودار تحلیلی')}"
                keyboard = get_keyboard(role, is_super_admin)
                send_photo(chat_id, chart_bytes, caption, keyboard)
            else:
                keyboard = get_keyboard(role, is_super_admin)
                send_message(chat_id, "❌ خطا در تولید نمودار.", keyboard)
        else:
            keyboard = get_keyboard(role, is_super_admin)
            send_message(chat_id, "📊 داده‌های کافی برای این نمودار وجود ندارد.", keyboard)
    except Exception as e:
        logger.error(f"generate_and_send_analytical_chart error: {e}")
        send_message(chat_id, "❌ خطا در تولید نمودار.", get_cancel_keyboard())

def generate_and_send_forecast(chat_id, role, is_super_admin):
    try:
        all_forecasts = get_forecast_for_all_branches(7)
        if all_forecasts:
            msg = f"📈 **پیش‌بینی عملکرد شعب (۷ روز آینده)**\n━━━━━━━━━━━━━━━━━━\n\n"
            for branch_name, data in all_forecasts.items():
                trend = data['trend']
                forecast = data['forecast']
                if not forecast:
                    continue
                msg += f"🏢 {branch_name}\n"
                msg += f"   📊 روند: {trend['trend']} (قدرت: {trend['strength']})\n"
                msg += f"   📈 میانگین وصول: {trend['avg_amount']//1_000_000:,.0f} میلیون ریال\n"
                msg += f"   🔮 پیش‌بینی روز بعد: {forecast[0]['predicted']//1_000_000:,.0f} میلیون ریال\n"
                msg += f"   📉 محدوده: {forecast[0]['lower']//1_000_000:,.0f} - {forecast[0]['upper']//1_000_000:,.0f} میلیون ریال\n\n"
            keyboard = get_keyboard(role, is_super_admin)
            send_message(chat_id, msg, keyboard)
        else:
            keyboard = get_keyboard(role, is_super_admin)
            send_message(chat_id, "📈 داده‌های کافی برای پیش‌بینی وجود ندارد.", keyboard)
    except Exception as e:
        logger.error(f"generate_and_send_forecast error: {e}")
        send_message(chat_id, "❌ خطا در تولید پیش‌بینی.", get_cancel_keyboard())

def generate_and_send_management_report(chat_id, report_key, user_db_id):
    try:
        report_text = generate_management_report(report_key)
        send_message(chat_id, report_text, get_management_reports_keyboard())
        log_user_activity(user_db_id, "management_report", report_key)
    except Exception:
        logger.exception("Asynchronous management report failed")
        send_message(chat_id, "❌ تولید گزارش انجام نشد؛ جزئیات خطا ثبت شد.", get_management_reports_keyboard())

def generate_and_send_visual_report(chat_id, report_key, button_text, user_db_id):
    try:
        chart_bytes = generate_visual_management_report(report_key)
        if chart_bytes:
            send_photo(chat_id, chart_bytes, button_text, get_visual_reports_keyboard())
            log_user_activity(user_db_id, "visual_management_report", report_key)
            return
        engine_ok, engine_error = get_chart_engine_status()
        if not engine_ok:
            send_message(chat_id, f"❌ موتور نمودار فارسی اجرا نشد.\nعلت: {engine_error}", get_visual_reports_keyboard())
        else:
            send_message(chat_id, "❌ داده کافی وجود ندارد یا تولید تصویر ناموفق بود.", get_visual_reports_keyboard())
    except Exception:
        logger.exception("Asynchronous visual report failed")
        send_message(chat_id, "❌ تولید نمودار انجام نشد؛ جزئیات خطا ثبت شد.", get_visual_reports_keyboard())

def generate_and_send_backup(chat_id, user_db_id):
    try:
        backup_data = generate_backup_file()
        if not backup_data:
            send_message(chat_id, "❌ تولید پشتیبان انجام نشد یا حجم داده از حد ایمن بیشتر است.", get_super_admin_keyboard())
            return
        backup_name = f"zanjan-backup-{get_iran_time().strftime('%Y%m%d-%H%M%S')}.json.gz"
        files = {'document': (backup_name, backup_data, 'application/gzip')}
        data = {'chat_id': chat_id, 'caption': '📦 فایل پشتیبان امضاشده از تمام داده‌ها'}
        res = get_http_session().post(f"{BASE_URL}/sendDocument", data=data, files=files, timeout=60)
        if res.status_code == 200:
            log_user_activity(user_db_id, "backup", "دریافت فایل پشتیبان")
        else:
            send_message(chat_id, "❌ خطا در ارسال فایل پشتیبان. لطفاً دوباره تلاش کنید.", get_super_admin_keyboard())
    except Exception:
        logger.exception("Asynchronous backup failed")
        send_message(chat_id, "❌ تولید پشتیبان انجام نشد؛ جزئیات در لاگ ثبت شد.", get_super_admin_keyboard())

# ============================================================
# offset management
# ============================================================
OFFSET_FILE = "offset.dat"
_last_offset_save = time_module.time()
_OFFSET_SAVE_INTERVAL = 10
_offset_pending = False
_offset_lock = threading.Lock()
_latest_offset = 0

def save_offset(offset):
    global _offset_pending
    with _offset_lock:
        _offset_pending = True
    if time_module.time() - _last_offset_save > _OFFSET_SAVE_INTERVAL:
        _flush_offset(offset)

def _flush_offset(offset):
    global _last_offset_save, _offset_pending, _latest_offset
    with _offset_lock:
        try:
            temp_path = OFFSET_FILE + '.tmp'
            with open(temp_path, 'w', encoding='ascii') as f:
                f.write(str(max(int(offset), 0)))
                f.flush()
                os.fsync(f.fileno())
            os.replace(temp_path, OFFSET_FILE)
            _latest_offset = offset
            _last_offset_save = time_module.time()
            _offset_pending = False
        except Exception as e:
            logger.error(f"Failed to save offset: {e}")

def load_offset():
    try:
        if os.path.exists(OFFSET_FILE):
            if os.path.getsize(OFFSET_FILE) > 64:
                raise ValueError("offset file is unexpectedly large")
            with open(OFFSET_FILE, 'rb') as f:
                raw = f.read()
            try:
                value = int(raw.decode('ascii').strip())
            except (UnicodeDecodeError, ValueError):
                # مهاجرت محدود از فایل pickle قدیمی که فقط یک عدد صحیح ساده داشته است.
                value = _load_legacy_integer_offset(raw)
            return max(value, 0)
    except Exception as e:
        logger.error(f"Failed to load offset: {e}")
    return 0

def _load_legacy_integer_offset(raw):
    """Parse only the small integer forms produced by previous bot versions."""
    if len(raw) >= 5 and raw[:2] in (b'\x80\x03', b'\x80\x04', b'\x80\x05') and raw[-1:] == b'.':
        opcode = raw[2:3]
        if opcode == b'K' and len(raw) == 5:
            return raw[3]
        if opcode == b'M' and len(raw) == 6:
            return int.from_bytes(raw[3:5], 'little', signed=False)
        if opcode == b'J' and len(raw) == 8:
            return int.from_bytes(raw[3:7], 'little', signed=True)
    raise ValueError("unsupported legacy offset format")

# ============================================================
# Scheduler
# ============================================================
def start_scheduler():
    global _scheduler
    scheduler = BackgroundScheduler(timezone='Asia/Tehran', job_defaults={
        'coalesce': True, 'max_instances': 1, 'misfire_grace_time': 300
    })
    scheduler.add_job(
        check_and_send_reminders,
        CronTrigger(hour=15, minute=0),
        id='reminder_job',
        replace_existing=True
    )
    scheduler.add_job(
        send_daily_report_to_admins,
        CronTrigger(hour=17, minute=30),
        id='daily_report_job',
        replace_existing=True
    )
    scheduler.add_job(
        check_and_send_drop_alerts,
        CronTrigger(hour=18, minute=30),
        id='drop_alert_job',
        replace_existing=True
    )
    scheduler.add_job(
        check_and_auto_score,
        CronTrigger(hour=20, minute=0),
        id='scoring_job',
        replace_existing=True
    )
    scheduler.add_job(
        send_weekly_report_to_all,
        CronTrigger(day_of_week='thu', hour=17, minute=0),
        id='weekly_report_job',
        replace_existing=True
    )
    scheduler.add_job(
        send_monthly_report_to_all,
        CronTrigger(hour=17, minute=0),
        id='monthly_report_job',
        replace_existing=True
    )
    scheduler.add_job(
        process_due_message_campaigns,
        'interval',
        seconds=60,
        id='message_campaign_job',
        replace_existing=True
    )
    scheduler.start()
    _scheduler = scheduler
    logger.info("✅ Scheduler started")

# ============================================================
# Main
# ============================================================
def main():
    global requests_session, processed_set, _latest_offset
    cleanup_stale_actual_excel_files()
    offset = load_offset()
    _latest_offset = offset
    logger.info(f"🤖 Bot started successfully with offset: {offset}")
    threading.Thread(target=run_flask, daemon=True).start()
    logger.info(f"🌐 Flask server started on port {PORT}")
    start_scheduler()
    atexit.register(lambda: _flush_offset(_latest_offset))
    last_offset_save_time = time_module.time()
    while True:
        try:
            url = f"{BASE_URL}/getUpdates"
            params = {"offset": offset, "timeout": 30}
            res = requests_session.get(url, params=params, timeout=45)
            if res.status_code == 200:
                data = res.json()
                if data.get("ok") and data.get("result"):
                    for update in data["result"]:
                        update_id = update["update_id"]
                        with processed_set_lock:
                            if update_id in processed_set:
                                continue
                            processed_set.add(update_id)
                            if len(processed_updates) == processed_updates.maxlen:
                                processed_set.discard(processed_updates[0])
                            processed_updates.append(update_id)
                        try:
                            if "message" in update:
                                handle_message(update["message"])
                        except Exception:
                            with processed_set_lock:
                                processed_set.discard(update_id)
                                try:
                                    processed_updates.remove(update_id)
                                except ValueError:
                                    pass
                            logger.exception("Update processing failed; offset was not advanced: %s", update_id)
                            break
                        offset = update_id + 1
                        _latest_offset = offset
                        if time_module.time() - last_offset_save_time > _OFFSET_SAVE_INTERVAL:
                            _flush_offset(offset)
                            last_offset_save_time = time_module.time()
                else:
                    logger.warning(f"⚠️ API response not ok: {data}")
                    time.sleep(2)
            elif res.status_code == 409:
                logger.error("⚠️ تعارض 409: نمونه دیگری از ربات با همین توکن فعال است؛ این نمونه منتظر می‌ماند.")
                time.sleep(10)
            else:
                logger.error(f"❌ HTTP error: {res.status_code} - {res.text}")
                time.sleep(5)
        except requests.exceptions.Timeout:
            logger.warning("⏳ Timeout in long polling (normal). Reconnecting...")
            time.sleep(2)
        except requests.exceptions.ConnectionError as e:
            logger.error(f"❌ Connection error: {e} – recreating session...")
            requests_session = create_session()
            time.sleep(5)
        except Exception as e:
            logger.error(f"❌ Unexpected error in main loop: {e}")
            time.sleep(5)

# ============================================================
# تابع handle_message (بخش اصلی پردازش پیام)
# ============================================================
def handle_message(message):
    try:
        chat_type = message.get('chat', {}).get('type', 'private')
        if chat_type != 'private':
            text = message.get('text', '').strip()
            if text and text.startswith('/'):
                pass
            else:
                return

        chat_id = message['chat']['id']
        text = message.get('text', '').strip()

        if not text:
            user_state = user_states.get(chat_id, {"state": "LOGGED_IN"})
            current_state = user_state.get("state", "LOGGED_IN") if isinstance(user_state, dict) else "LOGGED_IN"
            # سند بکاپ معمولاً text ندارد؛ باید تا بخش پردازش ریستور ادامه پیدا کند.
            if current_state == "WAITING_FOR_RESTORE_FILE" and message.get('document'):
                pass
            if current_state == "WAITING_FOR_CAMPAIGN_MEDIA" and (message.get('document') or message.get('photo')):
                pass
            if current_state == "WAITING_FOR_ACTUAL_EXCEL_FILE" and message.get('document'):
                pass
            if current_state in ["WAITING_FOR_NOTE", "WAITING_FOR_NOTE_FOR_COLLECTION"]:
                send_message(chat_id, "❌ لطفاً یک متن وارد کنید.")
                return
            elif not ((current_state == "WAITING_FOR_RESTORE_FILE" and message.get('document'))
                      or (current_state == "WAITING_FOR_CAMPAIGN_MEDIA" and (message.get('document') or message.get('photo')))
                      or (current_state == "WAITING_FOR_ACTUAL_EXCEL_FILE" and message.get('document'))):
                return

        if not get_bot_status() and not is_super_admin_user(chat_id):
            send_maintenance_message(chat_id)
            return

        user_state = user_states.get(chat_id, {"state": "LOGGED_OUT"})
        if not isinstance(user_state, dict):
            user_state = {"state": user_state} if user_state else {"state": "LOGGED_OUT"}
        current_state = user_state.get("state", "LOGGED_OUT")

        # ===== State: LOGGED_OUT / WAITING_FOR_EMP_NUM =====
        if current_state == "LOGGED_OUT" or current_state == "WAITING_FOR_EMP_NUM":
            if current_state != "WAITING_FOR_EMP_NUM":
                user_states.set(chat_id, {"state": "WAITING_FOR_EMP_NUM"})
                send_message(chat_id, "👋 سلام! به ربات وصول مطالبات استان زنجان خوش آمدید.\n\n🔐 لطفاً شماره کارمندی خود را ارسال کنید:", remove_keyboard=True)
                return
            normalized_text = normalize_digits(text)
            if not re.match(r'^[0-9]+$', normalized_text):
                send_message(chat_id, "❌ لطفاً شماره کارمندی را فقط با **اعداد** وارد کنید.\nمثال: ۱۲۳۴۵۶")
                return
            emp_user, employee_lookup_ok = find_user_by_employee_number(normalized_text, with_status=True)
            if not employee_lookup_ok:
                send_message(
                    chat_id,
                    "⚠️ بررسی شماره کارمندی موقتاً انجام نشد. لطفاً چند لحظه دیگر همان شماره را دوباره ارسال کنید."
                )
                return
            if emp_user:
                db_id, emp_num, name, role, title, branch_id, branch_name, is_super_admin = emp_user
                if is_super_admin:
                    user_states.update(chat_id, {
                        "state": "WAITING_FOR_SUPER_ADMIN_PASSWORD",
                        "temp_user_data": {
                            "db_id": db_id,
                            "emp_num": emp_num,
                            "name": name,
                            "role": role,
                            "title": title,
                            "branch_id": branch_id,
                            "branch_name": branch_name,
                            "is_super_admin": is_super_admin
                        }
                    })
                    send_message(chat_id, "🔐 شما یک کاربر سوپرادمین هستید. لطفاً رمز عبور خود را وارد کنید:", remove_keyboard=True)
                    return
                else:
                    bind_ok, bound_user, bind_message, bind_meta = bind_user_telegram_id(db_id, chat_id)
                    if not bind_ok:
                        user_states.set(chat_id, {"state": "WAITING_FOR_EMP_NUM"})
                        send_message(
                            chat_id,
                            f"❌ ورود کامل نشد: {bind_message}\n\n"
                            "هیچ دسترسی ناقصی ایجاد نشده است. لطفاً شماره کارمندی را دوباره ارسال کنید.",
                            remove_keyboard=True
                        )
                        return
                    db_id, emp_num, name, role, title, branch_id, branch_name, is_super_admin = bound_user
                    previous_chat_id = bind_meta.get("previous_chat_id")
                    if previous_chat_id is not None and previous_chat_id != chat_id:
                        user_states.delete(previous_chat_id)
                    if bind_meta.get("released_account_ids"):
                        logger.warning(
                            "Bale chat binding reassigned during employee login: target_user=%s released_users=%s active_released=%s",
                            db_id,
                            bind_meta["released_account_ids"],
                            bind_meta.get("released_active_count", 0),
                        )
                    log_user_activity(db_id, "login", f"ورود از chat_id: {chat_id}")
                    user_states.set(chat_id, {
                        "state": "LOGGED_IN",
                        "user_data": {
                            "db_id": db_id,
                            "emp_num": emp_num,
                            "name": name,
                            "role": role,
                            "title": title,
                            "branch_id": branch_id,
                            "branch_name": branch_name,
                            "is_super_admin": is_super_admin
                        }
                    })
                    welcome_msg = (
                        f"✅ هویت شما تایید شد.\n\n"
                        f"👤 {name}\n"
                        f"🏢 {title}\n"
                        f"🏭 واحد: {branch_name or 'بدون شعبه'}\n"
                        f"🔑 شماره کارمندی: {emp_num}\n"
                        f"⏰ زمان ورود: {get_shamsi_date_formatted(get_shamsi_date())} {get_iran_time().strftime('%H:%M:%S')}\n\n"
                        f"خوش آمدید! 👋\n\n"
                        f"📌 **راهنمای ثبت مبلغ:**\n"
                        f"مبالغ را به **میلیون ریال** وارد کنید.\n"
                        f"مثال: ۷۵۷ = ۷۵۷ میلیون ریال"
                    )
                    keyboard = get_keyboard(role, is_super_admin)
                    send_message(chat_id, welcome_msg, keyboard)
            else:
                send_message(chat_id, "❌ شماره کارمندی در سیستم یافت نشد.\nلطفاً شماره کارمندی صحیح خود را بفرستید.")
            return

        # ===== State: WAITING_FOR_SUPER_ADMIN_PASSWORD =====
        if current_state == "WAITING_FOR_SUPER_ADMIN_PASSWORD":
            blocked_until = user_state.get("password_blocked_until", 0)
            if blocked_until > time.time():
                remaining = int((blocked_until - time.time()) / 60) + 1
                send_message(chat_id, f"🔒 به علت تلاش‌های ناموفق، ورود تا حدود {remaining} دقیقه مسدود است.")
                return
            supplied_hash = hashlib.sha256(text.encode()).hexdigest()
            if hmac.compare_digest(supplied_hash, PASSWORD_HASH):
                temp_data = user_state.get("temp_user_data")
                if temp_data:
                    db_id = temp_data["db_id"]
                    bind_ok, bound_user, bind_message, bind_meta = bind_user_telegram_id(db_id, chat_id)
                    if not bind_ok:
                        user_states.set(chat_id, {"state": "LOGGED_OUT"})
                        send_message(
                            chat_id,
                            f"❌ ورود کامل نشد: {bind_message}\n\n"
                            "رمز پذیرفته شد، اما اتصال حساب در دیتابیس قطعی نشد؛ لطفاً ورود را دوباره آغاز کنید.",
                            remove_keyboard=True
                        )
                        return
                    db_id, emp_num, name, role, title, branch_id, branch_name, is_super_admin = bound_user
                    temp_data = {
                        "db_id": db_id,
                        "emp_num": emp_num,
                        "name": name,
                        "role": role,
                        "title": title,
                        "branch_id": branch_id,
                        "branch_name": branch_name,
                        "is_super_admin": is_super_admin,
                    }
                    previous_chat_id = bind_meta.get("previous_chat_id")
                    if previous_chat_id is not None and previous_chat_id != chat_id:
                        user_states.delete(previous_chat_id)
                    if bind_meta.get("released_account_ids"):
                        logger.warning(
                            "Bale chat binding reassigned during super-admin login: target_user=%s released_users=%s active_released=%s",
                            db_id,
                            bind_meta["released_account_ids"],
                            bind_meta.get("released_active_count", 0),
                        )
                    log_user_activity(db_id, "login", "ورود سوپرادمین")
                    user_states.set(chat_id, {
                        "state": "LOGGED_IN",
                        "user_data": temp_data
                    })
                    welcome_msg = (
                        f"✅ هویت سوپرادمین تایید شد.\n\n"
                        f"👤 {temp_data['name']}\n"
                        f"🏢 {temp_data['title']}\n"
                        f"🏭 واحد: {temp_data['branch_name'] or 'بدون شعبه'}\n"
                        f"🔑 شماره کارمندی: {temp_data['emp_num']}\n"
                        f"⏰ زمان ورود: {get_shamsi_date_formatted(get_shamsi_date())} {get_iran_time().strftime('%H:%M:%S')}\n\n"
                        f"شما دسترسی کامل مدیریتی دارید."
                    )
                    send_message(chat_id, welcome_msg, get_super_admin_keyboard())
                else:
                    send_message(chat_id, "❌ خطا در احراز هویت. لطفاً دوباره شماره کارمندی را وارد کنید.")
                    user_states.set(chat_id, {"state": "LOGGED_OUT"})
            else:
                failed_attempts = int(user_state.get("password_failed_attempts", 0)) + 1
                if failed_attempts >= 5:
                    user_states.update(chat_id, {
                        "password_failed_attempts": 0,
                        "password_blocked_until": time.time() + 900
                    })
                    logger.warning("Super-admin login temporarily blocked for chat_id=%s", chat_id)
                    send_message(chat_id, "🔒 پنج تلاش ناموفق ثبت شد؛ ورود برای ۱۵ دقیقه مسدود شد.")
                else:
                    user_states.update(chat_id, {"password_failed_attempts": failed_attempts})
                    send_message(chat_id, f"❌ رمز عبور اشتباه است. {5 - failed_attempts} تلاش باقی مانده است.")
            return

        # نقش و شعبه در هر پیام از دیتابیس تازه‌خوانی می‌شود تا تغییر سطح دسترسی
        # بلافاصله اعمال شود؛ شیوه احراز هویت موجود تغییری نکرده است.
        user, auth_lookup_ok = find_user_by_telegram_id(chat_id, with_status=True)
        if not auth_lookup_ok:
            # خطای موقت دیتابیس نباید به‌عنوان انقضای نشست تعبیر شود یا وضعیت
            # معتبر موجود در حافظه را پاک کند.
            send_message(
                chat_id,
                "⚠️ بررسی اتصال حساب موقتاً انجام نشد. نشست شما حذف نشده است؛ لطفاً چند لحظه دیگر دوباره تلاش کنید."
            )
            return
        if not user:
            user_states.set(chat_id, {"state": "LOGGED_OUT"})
            send_message(
                chat_id,
                "اتصال این گفت‌وگو به حساب کاربری تغییر کرده یا غیرفعال شده است. "
                "لطفاً شماره کارمندی خود را دوباره وارد کنید.",
                remove_keyboard=True
            )
            return
        db_id, emp_num, name, role, title, branch_id, branch_name, is_super_admin = user
        user_data = {
            "db_id": db_id,
            "emp_num": emp_num,
            "name": name,
            "role": role,
            "title": title,
            "branch_id": branch_id,
            "branch_name": branch_name,
            "is_super_admin": is_super_admin
        }
        user_states.update(chat_id, {"user_data": user_data})
        role = user_data["role"]
        branch_id = user_data["branch_id"]
        branch_name = user_data["branch_name"]
        user_db_id = user_data["db_id"]
        is_super_admin = user_data.get("is_super_admin", False)

        # ===== State: انتخاب شعبه برای مشاهده همیار توسط مدیران =====
        if current_state == "WAITING_FOR_ASSISTANT_BRANCH":
            if not (role == 'admin' or is_super_admin):
                user_states.update(chat_id, {"state": "LOGGED_IN"})
                send_message(chat_id, "⛔ دسترسی به این بخش برای شما مجاز نیست.", get_keyboard(role, is_super_admin))
                return
            if text == "🔙 انصراف":
                user_states.update(chat_id, {"state": "LOGGED_IN", "assistant_branches": []})
                send_message(chat_id, "به منوی اصلی بازگشتید.", get_keyboard(role, is_super_admin))
                return
            allowed_branches = user_state.get("assistant_branches") or get_all_branches()
            selected = next(
                ((item[0], item[1]) for item in allowed_branches if text == f"🏢 {item[1]}"),
                None
            )
            if not selected:
                send_message(
                    chat_id,
                    "❌ لطفاً شعبه را فقط از فهرست زیر انتخاب کنید.",
                    get_assistant_branch_keyboard(allowed_branches)
                )
                return
            selected_id, selected_name = selected
            user_states.update(chat_id, {"state": "LOGGED_IN", "assistant_branches": []})
            send_message(chat_id, f"⏳ در حال تهیه گزارش همیار شعبه {selected_name}...", get_keyboard(role, is_super_admin))
            executor.submit(send_collection_assistant_report, chat_id, selected_id,
                            get_keyboard(role, is_super_admin), user_db_id)
            return

        # ===== State: WAITING_FOR_DEPUTY_AMOUNT =====
        if current_state == "WAITING_FOR_DEPUTY_AMOUNT":
            if text == "🔙 انصراف":
                user_states.update(chat_id, {"state": "LOGGED_IN"})
                keyboard = get_keyboard(role, is_super_admin)
                send_message(chat_id, "❌ عملیات لغو شد.\n\nبه منوی اصلی بازگشتید.", keyboard)
                return
            try:
                amount = parse_number(text)
                if amount is None or amount < 0:
                    raise ValueError
                user_states.update(chat_id, {
                    "state": "WAITING_FOR_OTHERS_AMOUNT",
                    "deputy_amount": amount,
                    "edit_mode": user_state.get("edit_mode", False)
                })
                send_message(chat_id, "✏️ اکنون میزان وصولی سایر همکاران شعبه را به **میلیون ریال** وارد کنید:", get_cancel_keyboard())
            except ValueError:
                send_message(chat_id, "❌ خطا: لطفاً مبلغ را به صورت عدد مثبت (میلیون ریال) وارد کنید.\nمثال: ۴۷۰۰ برای ۴.۷ میلیارد ریال")
            return

        # ===== State: WAITING_FOR_OTHERS_AMOUNT =====
        if current_state == "WAITING_FOR_OTHERS_AMOUNT":
            if text == "🔙 انصراف":
                user_states.update(chat_id, {"state": "LOGGED_IN"})
                keyboard = get_keyboard(role, is_super_admin)
                send_message(chat_id, "❌ عملیات لغو شد.\n\nبه منوی اصلی بازگشتید.", keyboard)
                return
            try:
                others_amount = parse_number(text)
                if others_amount is None or others_amount < 0:
                    raise ValueError
                deputy_amount = user_state.get("deputy_amount", 0)
                shamsi_today = get_shamsi_date()
                is_edit = user_state.get("edit_mode", False)
                user_states.update(chat_id, {
                    "state": "WAITING_FOR_NOTE",
                    "collection_data": {
                        "deputy_amount": deputy_amount,
                        "others_amount": others_amount,
                        "shamsi_date": shamsi_today,
                        "is_edit": is_edit
                    }
                })
                send_message(chat_id, "📝 آیا می‌خواهید یادداشتی برای این وصول ثبت کنید؟ (اختیاری)\nلطفاً متن یادداشت را ارسال کنید یا روی «🔙 انصراف» بزنید تا بدون یادداشت ذخیره شود.", get_cancel_keyboard())
            except ValueError:
                send_message(chat_id, "❌ خطا: لطفاً مبلغ را به صورت عدد مثبت (میلیون ریال) وارد کنید.")
            return

        # ===== State: WAITING_FOR_NOTE =====
        if current_state == "WAITING_FOR_NOTE":
            data = user_state.get("collection_data", {})
            if text == "🔙 انصراف":
                success, collection_id = save_or_update_collection_with_note(
                    branch_id=branch_id,
                    deputy_amount_millions=data.get("deputy_amount", 0),
                    others_amount_millions=data.get("others_amount", 0),
                    shamsi_date=data.get("shamsi_date", get_shamsi_date()),
                    user_id=user_db_id,
                    note_text=None,
                    update_existing=data.get("is_edit", False)
                )
                user_states.update(chat_id, {"state": "LOGGED_IN"})
                if success:
                    total = data.get("deputy_amount", 0) + data.get("others_amount", 0)
                    msg = f"✅ ثبت شد.\n💰 جمع کل: {total:,.0f} میلیون ریال"
                    log_user_activity(user_db_id, "collection_add", f"ثبت وصول شعبه {branch_name} - مبلغ: {total} میلیون ریال")
                else:
                    msg = "❌ خطا در ثبت اطلاعات. (ممکن است زمان ویرایش گذشته باشد)"
                keyboard = get_keyboard(role, is_super_admin)
                send_message(chat_id, msg, keyboard)
                return
            else:
                note_text = text
                success, collection_id = save_or_update_collection_with_note(
                    branch_id=branch_id,
                    deputy_amount_millions=data.get("deputy_amount", 0),
                    others_amount_millions=data.get("others_amount", 0),
                    shamsi_date=data.get("shamsi_date", get_shamsi_date()),
                    user_id=user_db_id,
                    note_text=note_text,
                    update_existing=data.get("is_edit", False)
                )
                user_states.update(chat_id, {"state": "LOGGED_IN"})
                if success:
                    total = data.get("deputy_amount", 0) + data.get("others_amount", 0)
                    msg = f"✅ ثبت شد.\n💰 جمع کل: {total:,.0f} میلیون ریال\n📝 یادداشت: {escape_markdown(text)}"
                    log_user_activity(user_db_id, "collection_add_with_note", f"ثبت وصول با یادداشت برای شعبه {branch_name}")
                else:
                    msg = "❌ خطا در ثبت اطلاعات. (ممکن است زمان ویرایش گذشته باشد)"
                keyboard = get_keyboard(role, is_super_admin)
                send_message(chat_id, msg, keyboard)
                return

        # ===== State: WAITING_FOR_EDIT_CONFIRMATION =====
        if current_state == "WAITING_FOR_EDIT_CONFIRMATION":
            if text == "📝 بله، ویرایش شود":
                user_states.update(chat_id, {"state": "WAITING_FOR_DEPUTY_AMOUNT", "edit_mode": True})
                send_message(chat_id, "✏️ لطفاً مبلغ جدید وصولی خود (معاون) را به **میلیون ریال** وارد کنید:", get_cancel_keyboard())
            else:
                user_states.update(chat_id, {"state": "LOGGED_IN"})
                keyboard = get_keyboard(role, is_super_admin)
                send_message(chat_id, "❌ عملیات لغو شد.\n\nبه منوی اصلی بازگشتید.", keyboard)
            return

        # ===== State: WAITING_FOR_BRANCH_DATE (معاون) =====
        if role == 'deputy' and current_state == "WAITING_FOR_BRANCH_DATE":
            if text == "🔙 انصراف":
                user_states.update(chat_id, {"state": "LOGGED_IN"})
                send_message(chat_id, "❌ عملیات لغو شد.", get_deputy_keyboard())
                return
            shamsi_date = normalize_digits(text)
            if validate_shamsi_date(shamsi_date):
                record = get_branch_report_by_date(branch_id, shamsi_date)
                if record:
                    dep, oth, total = record
                    msg = (
                        f"📋 گزارش شعبه {branch_name} برای تاریخ {get_shamsi_date_formatted(shamsi_date)}\n"
                        f"━━━━━━━━━━━━━━━\n"
                        f"👤 وصولی معاون: {dep//1_000_000:,.0f} میلیون ریال\n"
                        f"👥 وصولی همکاران: {oth//1_000_000:,.0f} میلیون ریال\n"
                        f"💰 جمع کل: {total//1_000_000:,.0f} میلیون ریال"
                    )
                    col = check_existing_collection(branch_id, shamsi_date)
                    if col and col.get('id'):
                        notes = get_notes_for_collection(col['id'])
                        if notes:
                            msg += "\n\n📝 **یادداشت‌ها:**\n"
                            for n in notes:
                                note_time = n[3] if hasattr(n[3], 'strftime') else n[3]
                                msg += f"• {n[1]}: {n[2]} ({note_time.strftime('%H:%M') if hasattr(note_time, 'strftime') else note_time})\n"
                    send_message(chat_id, msg, get_deputy_keyboard())
                else:
                    send_message(chat_id, f"📭 هیچ داده‌ای برای تاریخ {get_shamsi_date_formatted(shamsi_date)} یافت نشد.", get_deputy_keyboard())
            else:
                send_message(chat_id, "❌ فرمت تاریخ نامعتبر. لطفاً به صورت YYYY/MM/DD وارد کنید (مثلاً 1403/01/15).")
                return
            user_states.update(chat_id, {"state": "LOGGED_IN"})
            return

        # ===== State: WAITING_FOR_ADMIN_DATE (ادمین/سوپرادمین) =====
        if (role == 'admin' or is_super_admin) and current_state == "WAITING_FOR_ADMIN_DATE":
            if text == "🔙 انصراف":
                user_states.update(chat_id, {"state": "LOGGED_IN"})
                keyboard = get_keyboard(role, is_super_admin)
                send_message(chat_id, "❌ عملیات لغو شد.", keyboard)
                return
            shamsi_date = normalize_digits(text)
            if validate_shamsi_date(shamsi_date):
                report = get_report_by_date(shamsi_date)
                if report:
                    msg = f"📅 گزارش استان برای تاریخ {get_shamsi_date_formatted(shamsi_date)}\n━━━━━━━━━━━━━━━━━━\n\n"
                    total_all = 0
                    for idx, row in enumerate(report, 1):
                        dep = int(safe_format(row[1]))
                        oth = int(safe_format(row[2]))
                        tot = int(safe_format(row[3]))
                        msg += f"{idx}. 🏢 {row[0]}\n"
                        msg += f"   👤 معاون ({row[4]}): {dep//1_000_000:,.0f} میلیون ریال\n"
                        msg += f"   👥 همکاران: {oth//1_000_000:,.0f} میلیون ریال\n"
                        msg += f"   💰 جمع: {tot//1_000_000:,.0f} میلیون ریال\n\n"
                        total_all += tot
                    msg += f"━━━━━━━━━━━━━━━━━━\n💰 جمع کل استان: {total_all//1_000_000:,.0f} میلیون ریال"
                    keyboard = get_keyboard(role, is_super_admin)
                    send_message(chat_id, msg, keyboard)
                else:
                    keyboard = get_keyboard(role, is_super_admin)
                    send_message(chat_id, f"📭 هیچ داده‌ای برای تاریخ {get_shamsi_date_formatted(shamsi_date)} یافت نشد.", keyboard)
            else:
                send_message(chat_id, "❌ فرمت تاریخ نامعتبر. لطفاً به صورت YYYY/MM/DD وارد کنید.")
                return
            user_states.update(chat_id, {"state": "LOGGED_IN"})
            return

        # ===== State: WAITING_FOR_PROBLEM =====
        if current_state == "WAITING_FOR_PROBLEM":
            if text == "🔙 انصراف":
                user_states.update(chat_id, {"state": "LOGGED_IN"})
                keyboard = get_keyboard(role, is_super_admin)
                send_message(chat_id, "❌ عملیات لغو شد.", keyboard)
                return
            if save_problem(user_db_id, text, "general"):
                send_message(chat_id, "✅ مشکل شما با موفقیت ثبت شد. تیم پشتیبانی در اسرع وقت بررسی خواهد کرد.", get_keyboard(role, is_super_admin))
                log_user_activity(user_db_id, "add_problem", f"ثبت مشکل: {text[:50]}...")
            else:
                send_message(chat_id, "❌ خطا در ثبت مشکل. لطفاً مجدداً تلاش کنید.", get_cancel_keyboard())
                return
            user_states.update(chat_id, {"state": "LOGGED_IN"})
            return

        # ===== State: WAITING_FOR_NOTE_FOR_COLLECTION (معاون) =====
        if current_state == "WAITING_FOR_NOTE_FOR_COLLECTION" and role == 'deputy':
            if text == "🔙 انصراف":
                user_states.update(chat_id, {"state": "LOGGED_IN"})
                send_message(chat_id, "❌ عملیات لغو شد.", get_deputy_keyboard())
                return
            try:
                parts = text.split('|', 1)
                if len(parts) == 2:
                    collection_id = int(parts[0].strip())
                    note_text = parts[1].strip()
                    conn = None
                    try:
                        conn = get_db_connection()
                        with conn.cursor() as cur:
                            cur.execute("SELECT branch_id FROM collections WHERE id = %s", (collection_id,))
                            result = cur.fetchone()
                            if not result:
                                send_message(chat_id, "❌ وصول یافت نشد.", get_deputy_keyboard())
                                return
                            if result[0] != branch_id:
                                send_message(chat_id, "❌ شما نمی‌توانید برای این شعبه یادداشت ثبت کنید.", get_deputy_keyboard())
                                return
                    finally:
                        if conn:
                            return_db_connection(conn)
                    if save_note(collection_id, user_db_id, note_text):
                        send_message(chat_id, f"✅ یادداشت برای وصول {collection_id} با موفقیت ثبت شد.", get_deputy_keyboard())
                    else:
                        send_message(chat_id, "❌ خطا در ثبت یادداشت.", get_deputy_keyboard())
                else:
                    send_message(chat_id, "❌ فرمت نامعتبر. لطفاً به شکل `[شناسه] | [متن یادداشت]` وارد کنید.", get_cancel_keyboard())
                    return
            except Exception as e:
                logger.exception("Super-admin operation failed")
                send_message(chat_id, "❌ عملیات انجام نشد؛ جزئیات خطا ثبت شد.", get_cancel_keyboard())
            user_states.update(chat_id, {"state": "LOGGED_IN"})
            return

        # ===== State: WAITING_FOR_MESSAGE_RECIPIENT (سوپرادمین) =====
        if current_state == "WAITING_FOR_MESSAGE_RECIPIENT":
            if text == "🔙 انصراف":
                user_states.update(chat_id, {"state": "LOGGED_IN"})
                send_message(chat_id, "❌ عملیات لغو شد.", get_super_admin_keyboard())
                return
            deputies = user_state.get("deputies", [])
            if not deputies:
                send_message(chat_id, "خطا در دریافت لیست معاونین.", get_super_admin_keyboard())
                return
            recipients = []
            if text == "همه":
                recipients = deputies
            elif text.isdigit():
                idx = int(text)
                if 1 <= idx <= len(deputies):
                    recipients = [deputies[idx-1]]
                else:
                    send_message(chat_id, "❌ شماره نامعتبر.", get_cancel_keyboard())
                    return
            elif ',' in text:
                indices = [int(x.strip()) for x in text.split(',') if x.strip().isdigit()]
                for idx in indices:
                    if 1 <= idx <= len(deputies):
                        recipients.append(deputies[idx-1])
                if not recipients:
                    send_message(chat_id, "❌ هیچ شماره معتبری یافت نشد.", get_cancel_keyboard())
                    return
            else:
                for dep in deputies:
                    if text in dep[2]:
                        recipients.append(dep)
                if not recipients:
                    send_message(chat_id, f"❌ معاونی با نام '{text}' یافت نشد.", get_cancel_keyboard())
                    return
            if not recipients:
                send_message(chat_id, "❌ هیچ مخاطبی انتخاب نشد.", get_cancel_keyboard())
                return
            user_states.update(chat_id, {
                "state": "WAITING_FOR_MESSAGE_TEXT",
                "recipients": recipients,
                "user_data": user_data
            })
            recipient_names = ", ".join([f"{r[2]} ({r[4] or 'بدون شعبه'})" for r in recipients])
            send_message(chat_id, f"📨 مخاطبین انتخاب شدند:\n{recipient_names}\n\n✏️ حالا متن پیام خود را بنویسید:", get_cancel_keyboard())
            return

        # ===== State: WAITING_FOR_MESSAGE_TEXT (سوپرادمین) =====
        if current_state == "WAITING_FOR_MESSAGE_TEXT":
            if text == "🔙 انصراف":
                user_states.update(chat_id, {"state": "LOGGED_IN"})
                send_message(chat_id, "❌ عملیات لغو شد.", get_super_admin_keyboard())
                return
            recipients = user_state.get("recipients", [])
            if not recipients:
                send_message(chat_id, "خطا: مخاطبی انتخاب نشده است.", get_super_admin_keyboard())
                user_states.update(chat_id, {"state": "LOGGED_IN"})
                return
            message_text = escape_markdown(text)
            success_count = 0
            for dep in recipients:
                dep_id, dep_chat_id, dep_name, branch_id, branch_name = dep
                if dep_chat_id:
                    msg = f"📨 **پیام از سوی مدیریت**\n━━━━━━━━━━━━━━━━━━\n\n{message_text}"
                    if send_message(dep_chat_id, msg):
                        success_count += 1
                        log_user_activity(user_db_id, "send_message_to_deputy", f"ارسال پیام به {dep_name} (ID: {dep_id})")
                    else:
                        logger.error(f"Failed to send message to {dep_name} (chat_id: {dep_chat_id})")
                else:
                    logger.warning(f"Deputy {dep_name} has no chat_id")
            final_msg = f"✅ پیام به {success_count} از {len(recipients)} مخاطب ارسال شد."
            if success_count < len(recipients):
                final_msg += f"\n⚠️ {len(recipients) - success_count} مخاطب پیام را دریافت نکردند (احتمالاً ربات را استارت نکرده‌اند)."
            send_message(chat_id, final_msg, get_super_admin_keyboard())
            log_user_activity(user_db_id, "send_message_to_deputies", f"ارسال پیام به {success_count} معاون")
            user_states.update(chat_id, {"state": "LOGGED_IN"})
            return

        # ===== مدیریت اهداف (سوپرادمین) =====
        if is_super_admin:
            if text == "🎯 مدیریت اهداف وصولی":
                keyboard = {
                    "keyboard": [
                        [{"text": "➕ تعیین هدف جدید"}],
                        [{"text": "📋 مشاهده اهداف فعال"}],
                        [{"text": "🗑️ حذف هدف"}],
                        [{"text": "🔙 انصراف"}]
                    ],
                    "resize_keyboard": True
                }
                send_message(chat_id, "🎯 **مدیریت اهداف وصولی شعب**\n\nلطفاً یکی از گزینه‌های زیر را انتخاب کنید:", keyboard)
                return

            if text == "➕ تعیین هدف جدید":
                branches = get_all_branches()
                if not branches:
                    send_message(chat_id, "❌ هیچ شعبه‌ای یافت نشد.", get_super_admin_keyboard())
                    return
                msg = "🏢 **انتخاب شعبه برای تعیین هدف**\n\n"
                for idx, (b_id, b_name) in enumerate(branches, 1):
                    msg += f"{idx}. {b_name}\n"
                msg += "\nلطفاً شماره شعبه مورد نظر را وارد کنید:"
                user_states.update(chat_id, {
                    "state": "WAITING_FOR_TARGET_BRANCH",
                    "branches": branches,
                    "user_data": user_data
                })
                send_message(chat_id, msg, get_cancel_keyboard())
                return

            if text == "📋 مشاهده اهداف فعال":
                targets = get_all_active_targets()
                if not targets:
                    send_message(chat_id, "📭 هیچ هدف فعالی برای شعب تعیین نشده است.", get_super_admin_keyboard())
                    return
                msg = "🎯 **اهداف فعال وصولی شعب**\n━━━━━━━━━━━━━━━━━━\n\n"
                for target in targets:
                    target_id, branch_id, branch_name, target_amount, target_date, created_at, created_by_name = target
                    progress = get_target_progress(branch_id, target_date, target_amount, created_at)
                    days_left_display = progress['days_left']
                    days_text = f"{days_left_display} روز" if days_left_display >= 0 else f"{abs(days_left_display)} روز گذشته"
                    msg += f"🏢 {branch_name}\n"
                    msg += f"   💰 هدف: {target_amount//1_000_000:,.0f} میلیون ریال\n"
                    msg += f"   📅 تاریخ هدف: {get_shamsi_date_formatted(target_date)}\n"
                    msg += f"   📊 پیشرفت: {progress['progress_percent']:.1f}% ({progress['collected']//1_000_000:,.0f} از {target_amount//1_000_000:,.0f} میلیون ریال)\n"
                    msg += f"   📅 زمان باقیمانده: {days_text}\n"
                    msg += f"   📉 فاصله از هدف: {progress['remaining']//1_000_000:,.0f} میلیون ریال\n"
                    msg += f"   👤 ثبت‌کننده: {created_by_name or 'نامشخص'}\n\n"
                send_message(chat_id, msg, get_super_admin_keyboard())
                return

            if text == "🗑️ حذف هدف":
                targets = get_all_active_targets()
                if not targets:
                    send_message(chat_id, "📭 هیچ هدف فعالی برای حذف وجود ندارد.", get_super_admin_keyboard())
                    return
                msg = "🗑️ **انتخاب هدف برای غیرفعال‌سازی**\n\n"
                for idx, target in enumerate(targets, 1):
                    target_id, branch_id, branch_name, target_amount, target_date, created_at, created_by_name = target
                    msg += f"{idx}. {branch_name} - هدف: {target_amount//1_000_000:,.0f} میلیون ریال تا {get_shamsi_date_formatted(target_date)}\n"
                msg += "\nلطفاً شماره هدف مورد نظر را وارد کنید:"
                user_states.update(chat_id, {
                    "state": "WAITING_FOR_TARGET_DELETE",
                    "targets": targets,
                    "user_data": user_data
                })
                send_message(chat_id, msg, get_cancel_keyboard())
                return

            if current_state == "WAITING_FOR_TARGET_BRANCH":
                if text == "🔙 انصراف":
                    user_states.update(chat_id, {"state": "LOGGED_IN"})
                    send_message(chat_id, "❌ عملیات لغو شد.", get_super_admin_keyboard())
                    return
                try:
                    idx = int(text) - 1
                    branches = user_state.get("branches", [])
                    if 0 <= idx < len(branches):
                        branch_id, branch_name = branches[idx]
                        user_states.update(chat_id, {
                            "target_branch_id": branch_id,
                            "target_branch_name": branch_name,
                            "state": "WAITING_FOR_TARGET_AMOUNT"
                        })
                        send_message(chat_id, f"🏢 شعبه {branch_name} انتخاب شد.\n\n✏️ لطفاً **مبلغ هدف** را به **میلیون ریال** وارد کنید:", get_cancel_keyboard())
                    else:
                        send_message(chat_id, "❌ شماره نامعتبر. لطفاً مجدداً تلاش کنید.", get_cancel_keyboard())
                except Exception:
                    send_message(chat_id, "❌ لطفاً یک عدد معتبر وارد کنید.", get_cancel_keyboard())
                return

            if current_state == "WAITING_FOR_TARGET_AMOUNT":
                if text == "🔙 انصراف":
                    user_states.update(chat_id, {"state": "LOGGED_IN"})
                    send_message(chat_id, "❌ عملیات لغو شد.", get_super_admin_keyboard())
                    return
                try:
                    amount = parse_number(text)
                    if amount is None or amount <= 0:
                        raise ValueError
                    user_states.update(chat_id, {
                        "target_amount": amount,
                        "state": "WAITING_FOR_TARGET_DATE"
                    })
                    send_message(chat_id, f"✅ مبلغ هدف {amount:,.0f} میلیون ریال ثبت شد.\n\n📅 لطفاً **تاریخ هدف** را به فرمت YYYY/MM/DD وارد کنید (مثلاً ۱۴۰۴/۰۶/۳۱):", get_cancel_keyboard())
                except Exception:
                    send_message(chat_id, "❌ لطفاً یک عدد مثبت معتبر وارد کنید.", get_cancel_keyboard())
                return

            if current_state == "WAITING_FOR_TARGET_DATE":
                if text == "🔙 انصراف":
                    user_states.update(chat_id, {"state": "LOGGED_IN"})
                    send_message(chat_id, "❌ عملیات لغو شد.", get_super_admin_keyboard())
                    return
                shamsi_date = normalize_digits(text)
                if validate_shamsi_date(shamsi_date):
                    today = get_shamsi_date()
                    try:
                        target_obj = jdatetime.date(*map(int, shamsi_date.split('/')))
                        today_obj = jdatetime.date(*map(int, today.split('/')))
                        if target_obj <= today_obj:
                            send_message(chat_id, "❌ تاریخ هدف باید بزرگتر از تاریخ امروز باشد.", get_cancel_keyboard())
                            return
                    except Exception:
                        send_message(chat_id, "❌ خطا در بررسی تاریخ.", get_cancel_keyboard())
                        return
                    branch_id = user_state.get("target_branch_id")
                    amount = user_state.get("target_amount")
                    success, result = set_branch_target(branch_id, amount * 1_000_000, shamsi_date, user_db_id)
                    if success:
                        branch_name = user_state.get("target_branch_name", "شعبه")
                        send_message(chat_id, f"✅ هدف برای شعبه {branch_name} با موفقیت تعیین شد.\n"
                                            f"💰 مبلغ هدف: {amount:,.0f} میلیون ریال\n"
                                            f"📅 تاریخ هدف: {get_shamsi_date_formatted(shamsi_date)}",
                                            get_super_admin_keyboard())
                        log_user_activity(user_db_id, "set_target", f"تعیین هدف برای شعبه {branch_name}: {amount} میلیون ریال تا {shamsi_date}")
                    else:
                        send_message(chat_id, f"❌ خطا در تعیین هدف: {result}", get_cancel_keyboard())
                        return
                    user_states.update(chat_id, {"state": "LOGGED_IN"})
                else:
                    send_message(chat_id, "❌ فرمت تاریخ نامعتبر. لطفاً به صورت YYYY/MM/DD وارد کنید.", get_cancel_keyboard())
                return

            if current_state == "WAITING_FOR_TARGET_DELETE":
                if text == "🔙 انصراف":
                    user_states.update(chat_id, {"state": "LOGGED_IN"})
                    send_message(chat_id, "❌ عملیات لغو شد.", get_super_admin_keyboard())
                    return
                try:
                    idx = int(text) - 1
                    targets = user_state.get("targets", [])
                    if 0 <= idx < len(targets):
                        target = targets[idx]
                        target_id = target[0]
                        branch_name = target[2]
                        if delete_target(target_id):
                            send_message(chat_id, f"✅ هدف شعبه {branch_name} غیرفعال شد و سابقه آن محفوظ ماند.", get_super_admin_keyboard())
                            log_user_activity(user_db_id, "deactivate_target", f"غیرفعال‌سازی هدف شعبه {branch_name}")
                        else:
                            send_message(chat_id, "❌ خطا در حذف هدف.", get_super_admin_keyboard())
                    else:
                        send_message(chat_id, "❌ شماره نامعتبر.", get_cancel_keyboard())
                        return
                except Exception:
                    send_message(chat_id, "❌ لطفاً یک عدد معتبر وارد کنید.", get_cancel_keyboard())
                    return
                user_states.update(chat_id, {"state": "LOGGED_IN"})
                return

            if text == "📊 گزارش پیشرفت اهداف":
                report = get_targets_progress_report()
                if not report:
                    send_message(chat_id, "📭 هیچ هدف فعالی برای گزارش وجود ندارد.", get_super_admin_keyboard())
                    return
                msg = "📊 **گزارش پیشرفت اهداف وصولی شعب**\n━━━━━━━━━━━━━━━━━━\n\n"
                sorted_report = sorted(report, key=lambda x: x['progress_percent'], reverse=True)
                for idx, item in enumerate(sorted_report, 1):
                    branch_name = item['branch_name']
                    target_amount = item['target_amount']
                    target_date = item['target_date']
                    collected = item['collected']
                    progress = item['progress_percent']
                    remaining = item['remaining']
                    days_left = item['days_left']
                    days_text = f"{days_left} روز" if days_left >= 0 else f"{abs(days_left)} روز گذشته"
                    emoji = "✅" if progress >= 80 else "🟢" if progress >= 50 else "🟡" if progress >= 30 else "🔴"
                    msg += f"{idx}. 🏢 {branch_name}\n"
                    msg += f"   💰 هدف: {target_amount//1_000_000:,.0f} میلیون ریال\n"
                    msg += f"   📅 تاریخ هدف: {get_shamsi_date_formatted(target_date)}\n"
                    msg += f"   📊 پیشرفت: {progress:.1f}% ({collected//1_000_000:,.0f} از {target_amount//1_000_000:,.0f} میلیون ریال) {emoji}\n"
                    msg += f"   📅 زمان باقیمانده: {days_text}\n"
                    msg += f"   📉 فاصله از هدف: {remaining//1_000_000:,.0f} میلیون ریال\n\n"
                send_message(chat_id, msg, get_super_admin_keyboard())
                return

            if text == "🏆 رتبه‌بندی تحقق هدف":
                report = get_targets_progress_report()
                if not report:
                    send_message(chat_id, "📭 هیچ هدف فعالی برای رتبه‌بندی وجود ندارد.", get_super_admin_keyboard())
                    return
                sorted_report = sorted(report, key=lambda x: x['progress_percent'], reverse=True)
                msg = "🏆 **رتبه‌بندی شعب بر اساس تحقق هدف**\n━━━━━━━━━━━━━━━━━━\n\n"
                medals = ["🥇", "🥈", "🥉"]
                for idx, item in enumerate(sorted_report, 1):
                    branch_name = item['branch_name']
                    progress = item['progress_percent']
                    target_amount = item['target_amount']
                    collected = item['collected']
                    days_left = item['days_left']
                    days_text = f"{days_left} روز" if days_left >= 0 else f"{abs(days_left)} روز گذشته"
                    medal = medals[idx-1] if idx <= 3 else f"{idx}."
                    msg += f"{medal} {branch_name}\n"
                    msg += f"   📊 تحقق هدف: {progress:.1f}%\n"
                    msg += f"   💰 جمع وصول: {collected//1_000_000:,.0f} از {target_amount//1_000_000:,.0f} میلیون ریال\n"
                    msg += f"   📅 زمان باقیمانده: {days_text}\n\n"
                send_message(chat_id, msg, get_super_admin_keyboard())
                return

        # ===== دکمه‌های عمومی =====
        if text == "🔙 خروج":
            log_user_activity(user_db_id, "logout", "خروج از سیستم")
            user_states.set(chat_id, {"state": "LOGGED_OUT"})
            send_message(chat_id, "👋 شما از سیستم خارج شدید.\n\nبرای ورود مجدد، شماره کارمندی خود را ارسال کنید.", remove_keyboard=True)
            return

        if text == "❓ راهنما":
            help_text = (
                "📌 **راهنمای ربات وصول مطالبات**\n\n"
                "🔹 **معاونین شعب:**\n"
                "   • ثبت وصولی روزانه (با قابلیت ویرایش تا ۱۲ شب)\n"
                "   • مشاهده گزارش ۱۰ روز اخیر شعبه\n"
                "   • مقایسه عملکرد روزانه شعبه\n"
                "   • مشاهده ثبت امروز\n"
                "   • گزارش یک تاریخ خاص برای شعبه خود\n"
                "   • مشاهده تاریخچه کامل شعبه\n"
                "   • ثبت و مشاهده یادداشت‌ها\n"
                "   • ثبت مشکل\n"
                "   • درباره توسعه‌دهنده\n\n"
                "🔹 **کاربران ارشد (ادمین):**\n"
                "   • گزارش امروز (همه شعب)\n"
                "   • گزارش ۱۰ روز اخیر استان\n"
                "   • رتبه‌بندی شعب برتر\n"
                "   • آمار مفصل امروز\n"
                "   • مقایسه روزانه ۷ روز اخیر\n"
                "   • تحلیل مدیریتی (تحلیل هوشمند داده‌ها)\n"
                "   • گزارش تاریخ خاص برای کل استان\n"
                "   • نمایش بهترین/بدترین روزهای استان\n"
                "   • گزارش روند هر شعبه\n"
                "   • گزارش عملکرد معاونان\n"
                "   • گزارش عملکرد همکاران (مجموع کل دوره)\n"
                "   • گزارش تطبیقی (مقایسه با دوره قبل)\n"
                "   • پیش‌بینی عملکرد (تحلیل روند هوشمند)\n"
                "   • نمودارهای تصویری (استان و شعبه)\n"
                "   • نمودارهای تحلیلی متنوع\n"
                "   • مقایسه انطباق با آمار واقعی (گزارش متنی دقیق)\n"
                "   • مشاهده یادداشت‌ها\n"
                "   • ثبت مشکل\n"
                "   • درباره توسعه‌دهنده\n\n"
                "🔹 **سوپرادمین:**\n"
                "   • مدیریت کاربران و گزارش‌ها\n"
                "   • مدیریت معاونین (افزودن، ویرایش، حذف)\n"
                "   • جابه‌جایی معاونان و انتصاب امن معاون جدید با رسید مستقل\n"
                "   • فعال/غیرفعال کردن ربات\n"
                "   • کنترل اعمال خودکار\n"
                "   • ریست کردن گزارش‌ها\n"
                "   • مرکز پیام‌رسانی: پیام مناسبتی، اطلاعیه و پیام فوری\n"
                "   • مدیریت تعطیلات\n"
                "   • مشاهده لاگ کامل فعالیت‌ها\n"
                "   • مدیریت مشکلات ثبت شده\n"
                "   • ارسال گزارش هفتگی و ماهانه\n"
                "   • فعال/غیرفعال کردن قابلیت‌ها\n"
                "   • ثبت آمار واقعی وصول (هر شعبه یک عدد)\n"
                "   • مشاهده نمودارهای تحلیلی و انطباق\n"
                "   • **مدیریت اهداف وصولی:**\n"
                "      - تعیین هدف برای هر شعبه (مبلغ و تاریخ)\n"
                "      - مشاهده اهداف فعال\n"
                "      - گزارش پیشرفت اهداف\n"
                "      - رتبه‌بندی تحقق هدف\n"
                "   • **گزارش‌های پیشرفته:**\n"
                "      - رتبه‌بندی معاونان بر اساس دقت خوداظهاری\n"
                "      - روند دقت یک شعبه در بازه زمانی\n"
                "      - بهترین/بدترین دقت در یک روز مشخص\n"
                "      - مقایسه عملکرد شعبه با میانگین استان\n"
                "      - تحلیل تاخیر در ثبت وصول معاونان\n"
                "   • **پشتیبان‌گیری و بازیابی:**\n"
                "      - دریافت فایل بکاپ از تمام داده‌ها\n"
                "      - بازیابی کامل سیستم از فایل بکاپ\n\n"
                "💰 **واحد پول:** تمام مبالغ به **میلیون ریال** است.\n"
                "🔸 در هر مرحله می‌توانید با دکمه «انصراف» به منو برگردید.\n"
                "🔸 برای خروج کامل، گزینه «خروج» را انتخاب کنید."
            )
            keyboard = get_keyboard(role, is_super_admin)
            send_message(chat_id, help_text, keyboard)
            return

        if text == "ℹ️ درباره توسعه‌دهنده":
            about_msg = (
                "🤖 **ربات وصول مطالبات استان زنجان**\n"
                "━━━━━━━━━━━━━━━━━━\n\n"
                "این ربات توسط **سید فرهاد سید حسینی**\n"
                "کارشناس حقوقی مدیریت شعب استان زنجان\n\n"
                "با حمایت‌های **آقای هادی بیگدلی**\n"
                "معاونت محترم وقت اعتباری منطقه\n\n"
                "در تابستان سال ۱۴۰۵ توسعه یافته است.\n\n"
                "📅 نسخه: ۹.۴.۳ (ورود پایدار، گزارش‌ها و مدیریت امن معاونان)\n"
                "📧 پشتیبانی: farhad.s.hosseini@gmail.com"
            )
            keyboard = get_keyboard(role, is_super_admin)
            send_message(chat_id, about_msg, keyboard)
            return

        if text == "📝 ثبت مشکل":
            user_states.update(chat_id, {"state": "WAITING_FOR_PROBLEM"})
            send_message(chat_id, "📝 لطفاً مشکل یا پیشنهاد خود را به صورت کامل بنویسید:\n\n(مثال: در ثبت وصول امروز خطایی رخ داد...)", get_cancel_keyboard())
            return

        # ============================================================
        # بخش سوپرادمین
        # ============================================================
        if is_super_admin:
            if text == "📣 مرکز پیام‌رسانی":
                user_states.update(chat_id, {"state": "LOGGED_IN", "campaign_draft": None})
                send_message(chat_id, "📣 **مرکز پیام‌رسانی**\n\nپیام مناسبتی، اطلاعیه یا پیام فوری را ایجاد و مدیریت کنید.", get_messaging_center_keyboard())
                return

            if text == "🔙 بازگشت به منوی اصلی":
                user_states.update(chat_id, {"state": "LOGGED_IN", "campaign_draft": None})
                send_message(chat_id, "به منوی اصلی بازگشتید.", get_super_admin_keyboard())
                return

            campaign_types = {"🎉 پیام مناسبتی": "occasion", "📢 اطلاعیه": "announcement", "🚨 پیام فوری": "urgent"}
            if text in campaign_types:
                user_states.update(chat_id, {
                    "state": "WAITING_FOR_CAMPAIGN_TEXT",
                    "campaign_draft": {"message_type": campaign_types[text]},
                })
                template_keyboard = {"keyboard": [[{"text": "📄 استفاده از قالب آماده"}], [{"text": "🔙 انصراف"}]], "resize_keyboard": True}
                send_message(
                    chat_id,
                    "✍️ متن پیام را وارد کنید.\n\nبرای شخصی‌سازی می‌توانید از `{name}`، `{title}` و `{branch}` استفاده کنید.",
                    template_keyboard
                )
                return

            if current_state == "WAITING_FOR_CAMPAIGN_TEXT":
                if text == "🔙 انصراف":
                    user_states.update(chat_id, {"state": "LOGGED_IN", "campaign_draft": None})
                    send_message(chat_id, "ایجاد پیام لغو شد.", get_messaging_center_keyboard())
                    return
                draft = dict(user_state.get('campaign_draft') or {})
                if text == "📄 استفاده از قالب آماده":
                    templates = {
                        'occasion': "{name} گرامی،\nفرارسیدن این مناسبت را به شما و همکاران محترم شعبه {branch} تبریک عرض می‌کنیم.\nبا آرزوی سلامتی و موفقیت روزافزون",
                        'announcement': "همکار گرامی {name}،\nبه اطلاع می‌رساند:\n\n[متن اطلاعیه]\n\nبا سپاس",
                        'urgent': "🚨 پیام فوری\n\n{name} گرامی، لطفاً در اولین فرصت اطلاعیه زیر را بررسی فرمایید:\n\n[متن پیام فوری]",
                    }
                    text = templates.get(draft.get('message_type'), '')
                if not text or len(text) > MAX_CAMPAIGN_TEXT_LENGTH:
                    send_message(chat_id, f"❌ متن باید بین ۱ تا {MAX_CAMPAIGN_TEXT_LENGTH} نویسه باشد.", get_cancel_keyboard())
                    return
                draft['message_text'] = text
                user_states.update(chat_id, {"state": "WAITING_FOR_CAMPAIGN_MEDIA", "campaign_draft": draft})
                keyboard = {"keyboard": [[{"text": "⏭ بدون تصویر"}], [{"text": "🔙 انصراف"}]], "resize_keyboard": True}
                send_message(chat_id, "🖼 در صورت تمایل یک تصویر حداکثر ۵ مگابایت ارسال کنید؛ یا «بدون تصویر» را بزنید.", keyboard)
                return

            if current_state == "WAITING_FOR_CAMPAIGN_MEDIA":
                if text == "🔙 انصراف":
                    user_states.update(chat_id, {"state": "LOGGED_IN", "campaign_draft": None})
                    send_message(chat_id, "ایجاد پیام لغو شد.", get_messaging_center_keyboard())
                    return
                draft = dict(user_state.get('campaign_draft') or {})
                if text != "⏭ بدون تصویر":
                    document = message.get('document') or {}
                    photos = message.get('photo') or []
                    if document:
                        file_id = document.get('file_id')
                        mime = document.get('mime_type') or ''
                        file_name = document.get('file_name') or 'message-image'
                    elif photos:
                        selected_photo = photos[-1]
                        file_id = selected_photo.get('file_id')
                        mime = 'image/jpeg'
                        file_name = 'message-image.jpg'
                    else:
                        send_message(chat_id, "❌ لطفاً تصویر ارسال کنید یا گزینه «بدون تصویر» را بزنید.", get_cancel_keyboard())
                        return
                    if not mime.startswith('image/'):
                        send_message(chat_id, "❌ فقط فایل تصویری مجاز است.", get_cancel_keyboard())
                        return
                    ok, media_error, media_bytes = download_bale_media(file_id)
                    if not ok:
                        send_message(chat_id, f"❌ {media_error}", get_cancel_keyboard())
                        return
                    draft.update({
                        'media_base64': base64.b64encode(media_bytes).decode('ascii'),
                        'media_mime': mime[:100], 'media_name': file_name[:255],
                    })
                user_states.update(chat_id, {"state": "WAITING_FOR_CAMPAIGN_PERSONALIZE", "campaign_draft": draft})
                keyboard = {"keyboard": [[{"text": "✅ شخصی‌سازی فعال"}, {"text": "❌ بدون شخصی‌سازی"}], [{"text": "🔙 انصراف"}]], "resize_keyboard": True}
                send_message(chat_id, "👤 آیا جایگزینی `{name}`، `{title}` و `{branch}` برای هر گیرنده فعال باشد؟", keyboard)
                return

            if current_state == "WAITING_FOR_CAMPAIGN_PERSONALIZE":
                if text == "🔙 انصراف":
                    user_states.update(chat_id, {"state": "LOGGED_IN", "campaign_draft": None})
                    send_message(chat_id, "ایجاد پیام لغو شد.", get_messaging_center_keyboard())
                    return
                if text not in ("✅ شخصی‌سازی فعال", "❌ بدون شخصی‌سازی"):
                    send_message(chat_id, "یکی از گزینه‌های شخصی‌سازی را انتخاب کنید.")
                    return
                draft = dict(user_state.get('campaign_draft') or {})
                draft['personalize'] = text == "✅ شخصی‌سازی فعال"
                user_states.update(chat_id, {"state": "WAITING_FOR_CAMPAIGN_AUDIENCE", "campaign_draft": draft})
                send_message(chat_id, "🎯 دریافت‌کنندگان پیام را انتخاب کنید.", get_campaign_audience_keyboard())
                return

            if current_state == "WAITING_FOR_CAMPAIGN_AUDIENCE":
                if text == "🔙 انصراف":
                    user_states.update(chat_id, {"state": "LOGGED_IN", "campaign_draft": None})
                    send_message(chat_id, "ایجاد پیام لغو شد.", get_messaging_center_keyboard())
                    return
                audience_map = {"👥 همه کاربران": "all", "🏢 همه معاونان": "deputies", "👤 همه ادمین‌ها": "admins"}
                draft = dict(user_state.get('campaign_draft') or {})
                if text in audience_map:
                    draft.update({'audience_type': audience_map[text], 'audience_config': {}})
                elif text == "🏬 انتخاب شعب":
                    branches = get_all_branches()
                    msg = "شماره یک یا چند شعبه را با ویرگول وارد کنید:\n\n" + "\n".join(f"{i}. {row[1]}" for i, row in enumerate(branches, 1))
                    user_states.update(chat_id, {"state": "WAITING_FOR_CAMPAIGN_BRANCHES", "campaign_draft": draft, "campaign_choices": branches})
                    send_message(chat_id, msg, get_cancel_keyboard())
                    return
                elif text == "☑️ انتخاب دستی کاربران":
                    conn = None
                    try:
                        conn = get_db_connection()
                        with conn.cursor() as cur:
                            cur.execute("SELECT id,full_name,role FROM users WHERE telegram_id IS NOT NULL AND is_active=TRUE ORDER BY full_name")
                            choices = cur.fetchall()
                    finally:
                        if conn:
                            return_db_connection(conn)
                    msg = "شماره کاربران را با ویرگول وارد کنید:\n\n" + "\n".join(f"{i}. {row[1]} ({row[2]})" for i, row in enumerate(choices, 1))
                    user_states.update(chat_id, {"state": "WAITING_FOR_CAMPAIGN_USERS", "campaign_draft": draft, "campaign_choices": choices})
                    send_message(chat_id, msg, get_cancel_keyboard())
                    return
                else:
                    send_message(chat_id, "❌ گروه مخاطبان معتبر نیست.", get_campaign_audience_keyboard())
                    return
                recipients = get_campaign_recipients(draft['audience_type'], draft['audience_config'])
                if not recipients:
                    send_message(chat_id, "❌ در این گروه کاربر متصل به ربات وجود ندارد.", get_campaign_audience_keyboard())
                    return
                user_states.update(chat_id, {"state": "CAMPAIGN_PREVIEW", "campaign_draft": draft, "campaign_recipients": recipients})
                preview_keyboard = {"keyboard": [[{"text": "🧪 ارسال آزمایشی برای خودم"}], [{"text": "💾 ذخیره پیش‌نویس"}, {"text": "➡️ تعیین زمان ارسال"}], [{"text": "🔙 انصراف"}]], "resize_keyboard": True}
                send_message(chat_id, format_campaign_preview(draft, recipients), preview_keyboard)
                return

            if current_state in ("WAITING_FOR_CAMPAIGN_BRANCHES", "WAITING_FOR_CAMPAIGN_USERS"):
                if text == "🔙 انصراف":
                    user_states.update(chat_id, {"state": "LOGGED_IN", "campaign_draft": None})
                    send_message(chat_id, "ایجاد پیام لغو شد.", get_messaging_center_keyboard())
                    return
                choices = user_state.get('campaign_choices') or []
                try:
                    indexes = sorted(set(int(part.strip()) - 1 for part in normalize_digits(text).split(',')))
                except (ValueError, TypeError):
                    indexes = []
                if not indexes or any(index < 0 or index >= len(choices) for index in indexes):
                    send_message(chat_id, "❌ شماره‌ها معتبر نیستند؛ نمونه: ۱,۳,۵", get_cancel_keyboard())
                    return
                draft = dict(user_state.get('campaign_draft') or {})
                if current_state == "WAITING_FOR_CAMPAIGN_BRANCHES":
                    draft.update({'audience_type': 'branches', 'audience_config': {'branch_ids': [choices[i][0] for i in indexes]}})
                else:
                    draft.update({'audience_type': 'manual', 'audience_config': {'user_ids': [choices[i][0] for i in indexes]}})
                recipients = get_campaign_recipients(draft['audience_type'], draft['audience_config'])
                if not recipients:
                    send_message(chat_id, "❌ هیچ‌کدام از انتخاب‌ها به ربات متصل نیستند.", get_cancel_keyboard())
                    return
                user_states.update(chat_id, {"state": "CAMPAIGN_PREVIEW", "campaign_draft": draft, "campaign_recipients": recipients})
                preview_keyboard = {"keyboard": [[{"text": "🧪 ارسال آزمایشی برای خودم"}], [{"text": "💾 ذخیره پیش‌نویس"}, {"text": "➡️ تعیین زمان ارسال"}], [{"text": "🔙 انصراف"}]], "resize_keyboard": True}
                send_message(chat_id, format_campaign_preview(draft, recipients), preview_keyboard)
                return

            if current_state == "CAMPAIGN_PREVIEW":
                if text == "🔙 انصراف":
                    user_states.update(chat_id, {"state": "LOGGED_IN", "campaign_draft": None})
                    send_message(chat_id, "ایجاد پیام لغو شد.", get_messaging_center_keyboard())
                    return
                draft = user_state.get('campaign_draft') or {}
                if text == "💾 ذخیره پیش‌نویس":
                    ok, result_text, result = create_message_campaign(draft, user_db_id, None, initial_status='draft')
                    user_states.update(chat_id, {"state": "LOGGED_IN", "campaign_draft": None, "campaign_recipients": None})
                    if ok:
                        log_user_activity(user_db_id, 'campaign_draft', f"campaign={result['uuid']}")
                        send_message(chat_id, f"✅ پیش‌نویس ذخیره شد.\nشماره: #{result['id']}\nرسید: `{result['uuid']}`", get_messaging_center_keyboard())
                    else:
                        send_message(chat_id, f"❌ {result_text}", get_messaging_center_keyboard())
                    return
                if text == "🧪 ارسال آزمایشی برای خودم":
                    self_rows = get_campaign_recipients('manual', {'user_ids': [user_db_id]})
                    rendered = personalize_campaign_text(draft['message_text'], self_rows[0], draft.get('personalize')) if self_rows else draft['message_text']
                    if draft.get('media_base64'):
                        result = send_campaign_photo(chat_id, base64.b64decode(draft['media_base64']), draft.get('media_mime'), draft.get('media_name'), rendered if len(rendered) <= 1024 else '')
                        if result and len(rendered) > 1024:
                            result = send_message(chat_id, rendered, parse_mode=None)
                    else:
                        result = send_message(chat_id, rendered, parse_mode=None)
                    send_message(chat_id, "✅ ارسال آزمایشی موفق بود." if result else "❌ ارسال آزمایشی ناموفق بود.")
                    return
                if text != "➡️ تعیین زمان ارسال":
                    send_message(chat_id, "ابتدا ارسال آزمایشی یا تعیین زمان را انتخاب کنید.")
                    return
                user_states.update(chat_id, {"state": "WAITING_FOR_CAMPAIGN_TIMING"})
                keyboard = {"keyboard": [[{"text": "⚡ ارسال فوری"}, {"text": "🕒 زمان‌بندی ارسال"}], [{"text": "🔙 انصراف"}]], "resize_keyboard": True}
                send_message(chat_id, "زمان ارسال را مشخص کنید.", keyboard)
                return

            if current_state == "WAITING_FOR_CAMPAIGN_TIMING":
                if text == "🔙 انصراف":
                    user_states.update(chat_id, {"state": "LOGGED_IN", "campaign_draft": None})
                    send_message(chat_id, "ایجاد پیام لغو شد.", get_messaging_center_keyboard())
                    return
                if text == "🕒 زمان‌بندی ارسال":
                    user_states.update(chat_id, {"state": "WAITING_FOR_CAMPAIGN_SCHEDULE"})
                    send_message(chat_id, "تاریخ و ساعت شمسی را وارد کنید.\nنمونه: ۱۴۰۵/۰۶/۰۳ ۰۹:۳۰", get_cancel_keyboard())
                    return
                if text != "⚡ ارسال فوری":
                    send_message(chat_id, "یکی از دو روش ارسال را انتخاب کنید.")
                    return
                scheduled_at = get_iran_time()
                draft = user_state.get('campaign_draft') or {}
                recipients = user_state.get('campaign_recipients') or []
                user_states.update(chat_id, {"state": "WAITING_FOR_CAMPAIGN_CONFIRM", "campaign_scheduled_at": scheduled_at, "campaign_expires_at": time.time() + 600})
                keyboard = {"keyboard": [[{"text": "✅ تأیید نهایی پیام"}], [{"text": "🔙 انصراف"}]], "resize_keyboard": True}
                send_message(chat_id, format_campaign_preview(draft, recipients, scheduled_at), keyboard)
                return

            if current_state == "WAITING_FOR_CAMPAIGN_SCHEDULE":
                if text == "🔙 انصراف":
                    user_states.update(chat_id, {"state": "LOGGED_IN", "campaign_draft": None})
                    send_message(chat_id, "ایجاد پیام لغو شد.", get_messaging_center_keyboard())
                    return
                scheduled_at = parse_campaign_schedule(text)
                if not scheduled_at:
                    send_message(chat_id, "❌ تاریخ باید معتبر و در آینده باشد؛ نمونه: ۱۴۰۵/۰۶/۰۳ ۰۹:۳۰", get_cancel_keyboard())
                    return
                draft = user_state.get('campaign_draft') or {}
                recipients = user_state.get('campaign_recipients') or []
                user_states.update(chat_id, {"state": "WAITING_FOR_CAMPAIGN_CONFIRM", "campaign_scheduled_at": scheduled_at, "campaign_expires_at": time.time() + 600})
                keyboard = {"keyboard": [[{"text": "✅ تأیید نهایی پیام"}], [{"text": "🔙 انصراف"}]], "resize_keyboard": True}
                send_message(chat_id, format_campaign_preview(draft, recipients, scheduled_at), keyboard)
                return

            if current_state == "WAITING_FOR_CAMPAIGN_CONFIRM":
                if text == "🔙 انصراف" or text != "✅ تأیید نهایی پیام":
                    user_states.update(chat_id, {"state": "LOGGED_IN", "campaign_draft": None})
                    send_message(chat_id, "پیام لغو شد و چیزی ارسال نشد.", get_messaging_center_keyboard())
                    return
                if time.time() > user_state.get('campaign_expires_at', 0):
                    user_states.update(chat_id, {"state": "LOGGED_IN", "campaign_draft": None})
                    send_message(chat_id, "⌛ مهلت تأیید تمام شد؛ پیام را دوباره ایجاد کنید.", get_messaging_center_keyboard())
                    return
                draft = user_state.get('campaign_draft') or {}
                scheduled_at = user_state.get('campaign_scheduled_at')
                user_states.update(chat_id, {"state": "LOGGED_IN", "campaign_draft": None, "campaign_recipients": None})
                ok, result_text, result = create_message_campaign(draft, user_db_id, scheduled_at)
                if not ok:
                    send_message(chat_id, f"❌ {result_text}", get_messaging_center_keyboard())
                    return
                log_user_activity(user_db_id, 'campaign_create', f"campaign={result['uuid']} recipients={result['count']}")
                if scheduled_at <= get_iran_time() + timedelta(seconds=5):
                    executor.submit(process_message_campaign, result['id'])
                send_message(chat_id, f"✅ پیام ثبت شد.\n🆔 رسید: `{result['uuid']}`\n👥 مخاطبان: {result['count']}\n\nنتیجه ارسال در تاریخچه قابل مشاهده است.", get_messaging_center_keyboard())
                return

            if text == "📝 پیش‌نویس‌ها":
                rows = [row for row in get_campaign_history(30) if row[3] == 'draft']
                if not rows:
                    send_message(chat_id, "📝 پیش‌نویسی وجود ندارد.", get_messaging_center_keyboard())
                    return
                msg = "📝 **پیش‌نویس‌ها**\n\n" + "\n\n".join(
                    f"#{row[0]} | {row[9]}\nرسید: `{row[1]}`" for row in rows
                )
                msg += "\n\nشماره پیش‌نویسی را که می‌خواهید ارسال کنید وارد نمایید."
                user_states.update(chat_id, {"state": "WAITING_FOR_DRAFT_ID"})
                send_message(chat_id, msg, get_cancel_keyboard())
                return

            if current_state == "WAITING_FOR_DRAFT_ID":
                if text == "🔙 انصراف":
                    user_states.update(chat_id, {"state": "LOGGED_IN"})
                    send_message(chat_id, "عملیات لغو شد.", get_messaging_center_keyboard())
                    return
                try:
                    campaign_id = int(normalize_digits(text).lstrip('#'))
                except ValueError:
                    send_message(chat_id, "❌ شماره پیش‌نویس معتبر نیست.", get_cancel_keyboard())
                    return
                user_states.update(chat_id, {"state": "WAITING_FOR_DRAFT_TIMING", "draft_campaign_id": campaign_id})
                keyboard = {"keyboard": [[{"text": "⚡ ارسال فوری"}, {"text": "🕒 زمان‌بندی ارسال"}], [{"text": "🔙 انصراف"}]], "resize_keyboard": True}
                send_message(chat_id, "زمان ارسال پیش‌نویس را مشخص کنید.", keyboard)
                return

            if current_state == "WAITING_FOR_DRAFT_TIMING":
                if text == "🔙 انصراف":
                    user_states.update(chat_id, {"state": "LOGGED_IN"})
                    send_message(chat_id, "عملیات لغو شد.", get_messaging_center_keyboard())
                    return
                if text == "🕒 زمان‌بندی ارسال":
                    user_states.update(chat_id, {"state": "WAITING_FOR_DRAFT_SCHEDULE"})
                    send_message(chat_id, "تاریخ و ساعت شمسی را وارد کنید؛ نمونه: ۱۴۰۵/۰۶/۰۳ ۰۹:۳۰", get_cancel_keyboard())
                    return
                if text != "⚡ ارسال فوری":
                    send_message(chat_id, "یکی از روش‌های ارسال را انتخاب کنید.")
                    return
                scheduled_at = get_iran_time()
                campaign_id = user_state.get('draft_campaign_id')
                ok, result_text, result = schedule_draft_campaign(campaign_id, scheduled_at, user_db_id)
                user_states.update(chat_id, {"state": "LOGGED_IN", "draft_campaign_id": None})
                if ok:
                    executor.submit(process_message_campaign, campaign_id)
                send_message(chat_id, ("✅ " if ok else "❌ ") + result_text, get_messaging_center_keyboard())
                return

            if current_state == "WAITING_FOR_DRAFT_SCHEDULE":
                if text == "🔙 انصراف":
                    user_states.update(chat_id, {"state": "LOGGED_IN", "draft_campaign_id": None})
                    send_message(chat_id, "عملیات لغو شد.", get_messaging_center_keyboard())
                    return
                scheduled_at = parse_campaign_schedule(text)
                if not scheduled_at:
                    send_message(chat_id, "❌ تاریخ باید معتبر و در آینده باشد.", get_cancel_keyboard())
                    return
                campaign_id = user_state.get('draft_campaign_id')
                ok, result_text, result = schedule_draft_campaign(campaign_id, scheduled_at, user_db_id)
                user_states.update(chat_id, {"state": "LOGGED_IN", "draft_campaign_id": None})
                send_message(chat_id, ("✅ " if ok else "❌ ") + result_text, get_messaging_center_keyboard())
                return

            if text == "📜 تاریخچه پیام‌ها":
                rows = get_campaign_history(12)
                labels = {'occasion': '🎉', 'announcement': '📢', 'urgent': '🚨'}
                status_labels = {'scheduled': 'زمان‌بندی', 'sending': 'درحال ارسال', 'completed': 'کامل', 'partial': 'ناقص', 'cancelled': 'لغوشده', 'draft': 'پیش‌نویس'}
                msg = "📜 **تاریخچه پیام‌ها**\n━━━━━━━━━━━━━━━━━━\n"
                if not rows:
                    msg += "هنوز پیامی ثبت نشده است."
                for row in rows:
                    cid, cuid, kind, status, total, sent, failed, scheduled, created, excerpt = row
                    msg += f"\n#{cid} {labels.get(kind,'📣')} {status_labels.get(status,status)} | ✅{sent} ❌{failed} از {total}\n{excerpt}\nرسید: `{cuid}`\n"
                send_message(chat_id, msg, get_messaging_center_keyboard())
                return

            if text == "⛔ لغو پیام زمان‌بندی‌شده":
                user_states.update(chat_id, {"state": "WAITING_FOR_CAMPAIGN_CANCEL_ID"})
                send_message(chat_id, "شماره پیام زمان‌بندی‌شده را از تاریخچه وارد کنید؛ مثال: ۱۲", get_cancel_keyboard())
                return

            if current_state == "WAITING_FOR_CAMPAIGN_CANCEL_ID":
                if text == "🔙 انصراف":
                    user_states.update(chat_id, {"state": "LOGGED_IN"})
                    send_message(chat_id, "لغو عملیات انجام شد.", get_messaging_center_keyboard())
                    return
                try:
                    campaign_id = int(normalize_digits(text).lstrip('#'))
                except ValueError:
                    send_message(chat_id, "❌ شماره پیام معتبر نیست.", get_cancel_keyboard())
                    return
                ok, result_text = cancel_scheduled_campaign(campaign_id, user_db_id)
                user_states.update(chat_id, {"state": "LOGGED_IN"})
                send_message(chat_id, ("✅ " if ok else "❌ ") + result_text, get_messaging_center_keyboard())
                return

            if text == "🔁 ارسال مجدد ناموفق‌ها":
                user_states.update(chat_id, {"state": "WAITING_FOR_CAMPAIGN_RETRY_ID"})
                send_message(chat_id, "شماره پیام ناقص را از تاریخچه وارد کنید؛ فقط گیرندگان ناموفق دوباره پیام می‌گیرند.", get_cancel_keyboard())
                return

            if current_state == "WAITING_FOR_CAMPAIGN_RETRY_ID":
                if text == "🔙 انصراف":
                    user_states.update(chat_id, {"state": "LOGGED_IN"})
                    send_message(chat_id, "عملیات لغو شد.", get_messaging_center_keyboard())
                    return
                try:
                    campaign_id = int(normalize_digits(text).lstrip('#'))
                except ValueError:
                    send_message(chat_id, "❌ شماره پیام معتبر نیست.", get_cancel_keyboard())
                    return
                ok, result_text = retry_failed_campaign(campaign_id, user_db_id)
                user_states.update(chat_id, {"state": "LOGGED_IN"})
                if ok:
                    executor.submit(process_message_campaign, campaign_id)
                send_message(chat_id, ("✅ " if ok else "❌ ") + result_text, get_messaging_center_keyboard())
                return

            if text == "🔁 مدیریت جابه‌جایی معاونان":
                send_message(
                    chat_id,
                    "🔁 **مدیریت جابه‌جایی معاونان**\n\n"
                    "این بخش جابه‌جایی دو معاون یا انتصاب فردی جدید به‌جای معاون فعلی را "
                    "در یک تراکنش ثبت می‌کند. هیچ سابقه مالی یا هدف شعبه تغییر نمی‌کند.",
                    get_deputy_transfer_keyboard()
                )
                return

            if text == "➕ افزودن معاون جدید":
                branches = get_branches_with_active_deputy()
                if not branches:
                    send_message(
                        chat_id,
                        "❌ هیچ شعبه‌ای با معاون و انتصاب فعالِ هماهنگ برای جایگزینی یافت نشد.",
                        get_deputy_transfer_keyboard()
                    )
                    return
                msg = "🏢 **شعبه موردنظر را انتخاب کنید:**\n\n"
                for index, row in enumerate(branches, 1):
                    msg += f"{index}. {row[1]} — معاون فعلی: {row[4]}\n"
                user_states.update(chat_id, {
                    "state": "WAITING_FOR_NEW_DEPUTY_BRANCH",
                    "replacement_branches": branches,
                })
                send_message(chat_id, msg, get_cancel_keyboard())
                return

            if current_state == "WAITING_FOR_NEW_DEPUTY_BRANCH":
                if text == "🔙 انصراف":
                    user_states.update(chat_id, {"state": "LOGGED_IN"})
                    send_message(chat_id, "عملیات لغو شد.", get_deputy_transfer_keyboard())
                    return
                branches = user_state.get("replacement_branches", [])
                try:
                    index = int(normalize_digits(text)) - 1
                except (TypeError, ValueError):
                    index = -1
                if not 0 <= index < len(branches):
                    send_message(chat_id, "❌ شماره شعبه معتبر نیست.", get_cancel_keyboard())
                    return
                selected_branch = branches[index]
                user_states.update(chat_id, {
                    "state": "WAITING_FOR_NEW_DEPUTY_NAME",
                    "replacement_branch": selected_branch,
                })
                send_message(
                    chat_id,
                    f"✅ شعبه: {selected_branch[1]}\n"
                    f"👤 معاون فعلی: {selected_branch[4]}\n\n"
                    "نام و نام خانوادگی معاون جدید را وارد کنید:",
                    get_cancel_keyboard()
                )
                return

            if current_state == "WAITING_FOR_NEW_DEPUTY_NAME":
                if text == "🔙 انصراف":
                    user_states.update(chat_id, {"state": "LOGGED_IN"})
                    send_message(chat_id, "عملیات لغو شد.", get_deputy_transfer_keyboard())
                    return
                full_name = _normalize_deputy_full_name(text)
                if not 3 <= len(full_name) <= 255:
                    send_message(chat_id, "❌ نام و نام خانوادگی معتبر نیست؛ دوباره وارد کنید.", get_cancel_keyboard())
                    return
                user_states.update(chat_id, {
                    "state": "WAITING_FOR_NEW_DEPUTY_EMPLOYEE_NUMBER",
                    "replacement_new_full_name": full_name,
                })
                send_message(
                    chat_id,
                    f"✅ نام معاون جدید: {full_name}\n\nشماره کارمندی معاون جدید را فقط با اعداد وارد کنید:",
                    get_cancel_keyboard()
                )
                return

            if current_state == "WAITING_FOR_NEW_DEPUTY_EMPLOYEE_NUMBER":
                if text == "🔙 انصراف":
                    user_states.update(chat_id, {"state": "LOGGED_IN"})
                    send_message(chat_id, "عملیات لغو شد.", get_deputy_transfer_keyboard())
                    return
                employee_number = normalize_digits(text).strip()
                branch = user_state.get("replacement_branch")
                full_name = user_state.get("replacement_new_full_name")
                if not branch or not full_name:
                    user_states.update(chat_id, {"state": "LOGGED_IN"})
                    send_message(chat_id, "❌ اطلاعات عملیات منقضی شده است؛ دوباره آغاز کنید.", get_deputy_transfer_keyboard())
                    return
                ok, error, preview = preview_new_deputy_replacement(branch[0], full_name, employee_number)
                if not ok:
                    send_message(chat_id, f"❌ {error}\nشماره کارمندی صحیح را وارد کنید یا عملیات را لغو کنید.", get_cancel_keyboard())
                    return
                user_states.update(chat_id, {
                    "state": "WAITING_FOR_NEW_DEPUTY_CONFIRM",
                    "replacement_preview": preview,
                    "replacement_expires_at": time.time() + 600,
                })
                confirm_keyboard = {
                    "keyboard": [[{"text": "✅ تأیید نهایی انتصاب"}], [{"text": "🔙 انصراف"}]],
                    "resize_keyboard": True,
                }
                send_message(chat_id, format_new_deputy_preview(preview), confirm_keyboard)
                return

            if current_state == "WAITING_FOR_NEW_DEPUTY_CONFIRM":
                if text == "🔙 انصراف" or text != "✅ تأیید نهایی انتصاب":
                    user_states.update(chat_id, {"state": "LOGGED_IN", "replacement_preview": None})
                    send_message(chat_id, "عملیات لغو شد و هیچ تغییری اعمال نشد.", get_deputy_transfer_keyboard())
                    return
                if time.time() > user_state.get("replacement_expires_at", 0):
                    user_states.update(chat_id, {"state": "LOGGED_IN", "replacement_preview": None})
                    send_message(chat_id, "⌛ اعتبار پیش‌نمایش پایان یافته است؛ عملیات را دوباره آغاز کنید.", get_deputy_transfer_keyboard())
                    return
                preview = user_state.get("replacement_preview")
                if not preview:
                    user_states.update(chat_id, {"state": "LOGGED_IN"})
                    send_message(chat_id, "❌ پیش‌نمایش در دسترس نیست؛ هیچ تغییری اعمال نشد.", get_deputy_transfer_keyboard())
                    return
                # بستن state پیش از تراکنش، مانع اجرای دوباره با لمس تکراری دکمه می‌شود.
                user_states.update(chat_id, {"state": "LOGGED_IN", "replacement_preview": None})
                success, result_message, receipt = execute_new_deputy_replacement(
                    preview['branch_id'], preview['old_deputy_id'],
                    preview['new_full_name'], preview['new_employee_number'], user_db_id
                )
                if not success:
                    send_message(
                        chat_id,
                        f"❌ انتصاب انجام نشد: {result_message}\nهیچ تغییر ناقصی ثبت نشده است.",
                        get_deputy_transfer_keyboard()
                    )
                    return
                log_user_activity(
                    user_db_id,
                    "deputy_replacement",
                    f"operation={receipt['operation_uuid']} branch={receipt['branch_name']} "
                    f"old={receipt['old_deputy_id']} new={receipt['new_deputy_id']}"
                )
                receipt_text = (
                    "✅ **رسید انتصاب معاون جدید**\n━━━━━━━━━━━━━━━━━━\n"
                    f"🆔 شناسه عملیات: `{receipt['operation_uuid']}`\n"
                    f"📅 تاریخ مؤثر: {get_shamsi_date_formatted(receipt['effective_date'])}\n"
                    f"🏢 شعبه: {receipt['branch_name']}\n\n"
                    f"👤 معاون قبلی: {receipt['old_full_name']} ({receipt['old_employee_number']})\n"
                    f"👤 معاون جدید: {receipt['new_full_name']} ({receipt['new_employee_number']})\n\n"
                    "🔒 تعداد سوابق مالی، اهداف و گزارش‌های قبلی تغییرکرده: صفر\n"
                    "✅ دسترسی معاون قبلی غیرفعال و سابقه انتصاب او بسته شد.\n"
                    "✅ حساب و انتصاب معاون جدید در همان تراکنش ایجاد شد.\n"
                    "⏰ تحلیل تأخیر معاون جدید از صفر آغاز می‌شود.\n\n"
                    f"🎉 جناب آقای «{receipt['new_full_name']}»، انتصاب شما به‌عنوان معاون "
                    f"شعبه «{receipt['branch_name']}» را تبریک عرض می‌کنیم. برای شما در مسئولیت جدید، "
                    "موفقیت، سربلندی و توفیق روزافزون آرزومندیم."
                )
                send_message(chat_id, receipt_text, get_super_admin_keyboard())
                return

            if text == "🔄 جابه‌جایی دو معاون":
                deputies = [row for row in get_all_deputies_with_details() if row[4] is not None]
                if len(deputies) < 2:
                    send_message(chat_id, "❌ حداقل دو معاون دارای شعبه برای جابه‌جایی لازم است.", get_deputy_transfer_keyboard())
                    return
                msg = "1️⃣ **معاون اول را انتخاب کنید:**\n\n"
                for index, deputy in enumerate(deputies, 1):
                    msg += f"{index}. {deputy[2]} — {deputy[5]}\n"
                user_states.update(chat_id, {
                    "state": "WAITING_FOR_SWAP_FIRST_DEPUTY",
                    "swap_deputies": deputies,
                })
                send_message(chat_id, msg, get_cancel_keyboard())
                return

            if current_state == "WAITING_FOR_SWAP_FIRST_DEPUTY":
                if text == "🔙 انصراف":
                    user_states.update(chat_id, {"state": "LOGGED_IN"})
                    send_message(chat_id, "عملیات لغو شد.", get_super_admin_keyboard())
                    return
                deputies = user_state.get("swap_deputies", [])
                try:
                    index = int(normalize_digits(text)) - 1
                except (TypeError, ValueError):
                    index = -1
                if not 0 <= index < len(deputies):
                    send_message(chat_id, "❌ شماره معاون معتبر نیست.", get_cancel_keyboard())
                    return
                first = deputies[index]
                candidates = [deputy for deputy in deputies if deputy[0] != first[0] and deputy[4] != first[4]]
                msg = f"✅ معاون اول: {first[2]} — {first[5]}\n\n2️⃣ **معاون دوم را انتخاب کنید:**\n\n"
                for candidate_index, deputy in enumerate(candidates, 1):
                    msg += f"{candidate_index}. {deputy[2]} — {deputy[5]}\n"
                user_states.update(chat_id, {
                    "state": "WAITING_FOR_SWAP_SECOND_DEPUTY",
                    "swap_first": first,
                    "swap_candidates": candidates,
                })
                send_message(chat_id, msg, get_cancel_keyboard())
                return

            if current_state == "WAITING_FOR_SWAP_SECOND_DEPUTY":
                if text == "🔙 انصراف":
                    user_states.update(chat_id, {"state": "LOGGED_IN"})
                    send_message(chat_id, "عملیات لغو شد.", get_super_admin_keyboard())
                    return
                candidates = user_state.get("swap_candidates", [])
                try:
                    index = int(normalize_digits(text)) - 1
                except (TypeError, ValueError):
                    index = -1
                if not 0 <= index < len(candidates):
                    send_message(chat_id, "❌ شماره معاون معتبر نیست.", get_cancel_keyboard())
                    return
                second = candidates[index]
                user_states.update(chat_id, {
                    "state": "WAITING_FOR_SWAP_EFFECTIVE_DATE",
                    "swap_second": second,
                })
                send_message(
                    chat_id,
                    f"✅ معاون دوم: {second[2]} — {second[5]}\n\n"
                    "📅 تاریخ مؤثر جابه‌جایی را به فرمت YYYY/MM/DD وارد کنید.\n"
                    "مثال: ۱۴۰۵/۰۵/۲۶",
                    get_cancel_keyboard()
                )
                return

            if current_state == "WAITING_FOR_SWAP_EFFECTIVE_DATE":
                if text == "🔙 انصراف":
                    user_states.update(chat_id, {"state": "LOGGED_IN"})
                    send_message(chat_id, "عملیات لغو شد.", get_super_admin_keyboard())
                    return
                effective_date = normalize_digits(text)
                first = user_state.get("swap_first")
                second = user_state.get("swap_second")
                if not first or not second or not validate_shamsi_date(effective_date):
                    send_message(chat_id, "❌ تاریخ معتبر نیست. نمونه صحیح: ۱۴۰۵/۰۵/۲۶", get_cancel_keyboard())
                    return
                ok, error, _ = preview_deputy_swap(first[0], second[0], effective_date)
                if not ok:
                    send_message(chat_id, f"❌ {error}", get_cancel_keyboard())
                    return
                user_states.update(chat_id, {
                    "state": "WAITING_FOR_SWAP_REASON",
                    "swap_effective_date": effective_date,
                })
                reason_keyboard = {
                    "keyboard": [[{"text": "⏭ بدون توضیح"}], [{"text": "🔙 انصراف"}]],
                    "resize_keyboard": True,
                }
                send_message(chat_id, "📝 علت یا توضیح اداری جابه‌جایی را بنویسید؛ این بخش اختیاری است.", reason_keyboard)
                return

            if current_state == "WAITING_FOR_SWAP_REASON":
                if text == "🔙 انصراف":
                    user_states.update(chat_id, {"state": "LOGGED_IN"})
                    send_message(chat_id, "عملیات لغو شد.", get_super_admin_keyboard())
                    return
                first = user_state.get("swap_first")
                second = user_state.get("swap_second")
                effective_date = user_state.get("swap_effective_date")
                reason = None if text == "⏭ بدون توضیح" else text[:500]
                ok, error, preview = preview_deputy_swap(first[0], second[0], effective_date)
                if not ok:
                    user_states.update(chat_id, {"state": "LOGGED_IN"})
                    send_message(chat_id, f"❌ {error}\nهیچ تغییری اعمال نشد.", get_super_admin_keyboard())
                    return
                user_states.update(chat_id, {
                    "state": "WAITING_FOR_SWAP_CONFIRM",
                    "swap_reason": reason,
                    "swap_preview": preview,
                    "swap_expires_at": time.time() + 600,
                })
                confirm_keyboard = {
                    "keyboard": [[{"text": "✅ تأیید نهایی جابه‌جایی"}], [{"text": "🔙 انصراف"}]],
                    "resize_keyboard": True,
                }
                send_message(chat_id, format_deputy_swap_preview(preview), confirm_keyboard)
                return

            if current_state == "WAITING_FOR_SWAP_CONFIRM":
                if text == "🔙 انصراف" or text != "✅ تأیید نهایی جابه‌جایی":
                    user_states.update(chat_id, {"state": "LOGGED_IN", "swap_preview": None})
                    send_message(chat_id, "عملیات لغو شد و هیچ تغییری اعمال نشد.", get_super_admin_keyboard())
                    return
                if time.time() > user_state.get("swap_expires_at", 0):
                    user_states.update(chat_id, {"state": "LOGGED_IN", "swap_preview": None})
                    send_message(chat_id, "⌛ اعتبار پیش‌نمایش پایان یافته است؛ عملیات را دوباره آغاز کنید.", get_super_admin_keyboard())
                    return
                preview = user_state.get("swap_preview")
                if not preview:
                    user_states.update(chat_id, {"state": "LOGGED_IN"})
                    send_message(chat_id, "❌ اطلاعات پیش‌نمایش در دسترس نیست؛ هیچ تغییری اعمال نشد.", get_super_admin_keyboard())
                    return
                # قبل از اجرای تراکنش state بسته می‌شود تا لمس دوباره دکمه، عملیات را تکرار نکند.
                user_states.update(chat_id, {"state": "LOGGED_IN", "swap_preview": None})
                success, result_message, receipt = execute_deputy_swap(
                    preview['first'][0], preview['second'][0], preview['effective_date'],
                    preview['first'][4], preview['second'][4], user_db_id,
                    user_state.get("swap_reason")
                )
                if not success:
                    send_message(chat_id, f"❌ جابه‌جایی انجام نشد: {result_message}\nهیچ تغییر ناقصی ثبت نشده است.", get_super_admin_keyboard())
                    return
                log_user_activity(
                    user_db_id,
                    "deputy_swap",
                    f"operation={receipt['operation_uuid']} effective={receipt['effective_date']}"
                )
                receipt_text = (
                    "✅ **رسید ثبت جابه‌جایی معاونان**\n━━━━━━━━━━━━━━━━━━\n"
                    f"🆔 شناسه عملیات: `{receipt['operation_uuid']}`\n"
                    f"📅 تاریخ مؤثر: {get_shamsi_date_formatted(receipt['effective_date'])}\n\n"
                    f"👤 {receipt['first_name']}: {receipt['first_old']} ← {receipt['first_new']}\n"
                    f"👤 {receipt['second_name']}: {receipt['second_old']} ← {receipt['second_new']}\n\n"
                    "🔒 تعداد سوابق مالی تغییرکرده: صفر\n"
                    "تاریخچه انتصاب، شعبه فعلی و عنوان وابسته به شعبه در یک تراکنش ثبت شد."
                )
                send_message(chat_id, receipt_text, get_super_admin_keyboard())
                return

            if text == "📜 تاریخچه انتصاب معاونان":
                history = get_deputy_assignment_history(40)
                if not history:
                    send_message(chat_id, "تاریخچه‌ای برای نمایش وجود ندارد.", get_deputy_transfer_keyboard())
                    return
                msg = "📜 **تاریخچه انتصاب معاونان**\n━━━━━━━━━━━━━━━━━━\n"
                for _, name, employee_number, branch, start, end, source, operation_uuid in history:
                    end_text = get_shamsi_date_formatted(end) if end else "اکنون"
                    start_text = "آغاز ثبت سیستم" if start == '0001/01/01' else get_shamsi_date_formatted(start)
                    msg += f"• {name} ({employee_number})\n  {branch}: {start_text} تا {end_text}\n"
                    if operation_uuid:
                        msg += f"  شناسه عملیات: `{operation_uuid}`\n"
                send_message(chat_id, msg, get_deputy_transfer_keyboard())
                return

            if text == "📊 مرکز گزارش‌های مدیریتی":
                send_message(
                    chat_id,
                    "📊 **مرکز گزارش‌های مدیریتی**\n\nتمام گزارش‌های این بخش فقط خواندنی هستند و هیچ داده‌ای را تغییر نمی‌دهند.",
                    get_management_reports_keyboard()
                )
                return

            if text == "📈 مرکز گزارش‌های تصویری":
                engine_ok, engine_error = get_chart_engine_status(force_test=True)
                engine = "✅ Plotly/Kaleido آماده" if engine_ok else "❌ موتور فارسی آماده نیست"
                send_message(
                    chat_id,
                    f"📈 **مرکز گزارش‌های تصویری**\n\nوضعیت: {engine}\n"
                    + (f"علت: {engine_error}\n" if engine_error else "")
                    + "گزارش موردنظر را انتخاب کنید.",
                    get_visual_reports_keyboard()
                )
                return

            if text == "🔙 بازگشت به پنل سوپرادمین":
                send_message(chat_id, "به پنل سوپرادمین بازگشتید.", get_super_admin_keyboard())
                return

            if text in MANAGEMENT_REPORT_BUTTONS:
                report_key = MANAGEMENT_REPORT_BUTTONS[text]
                send_message(chat_id, "⏳ در حال تهیه گزارش...", get_management_reports_keyboard())
                executor.submit(generate_and_send_management_report, chat_id, report_key, user_db_id)
                return

            if text in VISUAL_REPORT_BUTTONS:
                report_key = VISUAL_REPORT_BUTTONS[text]
                send_message(chat_id, "⏳ در حال ساخت نمودار فارسی...", get_visual_reports_keyboard())
                executor.submit(generate_and_send_visual_report, chat_id, report_key, text, user_db_id)
                return

            if text == "🔧 کنترل خودکار":
                reminder_status = "فعال ✅" if get_auto_reminder_status() else "غیرفعال ❌"
                report_status = "فعال ✅" if get_auto_report_status() else "غیرفعال ❌"
                alert_status = "فعال ✅" if get_auto_alert_status() else "غیرفعال ❌"
                scoring_status = "فعال ✅" if get_auto_scoring_status() else "غیرفعال ❌"
                weekly_status = "فعال ✅" if get_weekly_report_status() else "غیرفعال ❌"
                monthly_status = "فعال ✅" if get_monthly_report_status() else "غیرفعال ❌"
                instant_status = "فعال ✅" if get_instant_notification_status() else "غیرفعال ❌"
                adaptive_status = "فعال ✅" if get_adaptive_report_status() else "غیرفعال ❌"
                forecast_status = "فعال ✅" if get_forecast_report_status() else "غیرفعال ❌"
                chart_status = "فعال ✅" if get_chart_report_status() else "غیرفعال ❌"
                actual_status = "فعال ✅" if get_actual_stats_status() else "غیرفعال ❌"
                keyboard = {
                    "keyboard": [
                        [{"text": f"📌 یادآوری: {reminder_status}"}, {"text": f"📌 گزارش روزانه: {report_status}"}],
                        [{"text": f"📌 هشدار افت: {alert_status}"}, {"text": f"📌 امتیازدهی: {scoring_status}"}],
                        [{"text": f"📌 گزارش هفتگی: {weekly_status}"}, {"text": f"📌 گزارش ماهانه: {monthly_status}"}],
                        [{"text": f"🔔 نوتیفیکیشن: {instant_status}"}, {"text": f"📊 گزارش تطبیقی: {adaptive_status}"}],
                        [{"text": f"📈 پیش‌بینی: {forecast_status}"}, {"text": f"📊 نمودار: {chart_status}"}],
                        [{"text": f"📊 آمار واقعی: {actual_status}"}, {"text": "🔙 انصراف"}]
                    ],
                    "resize_keyboard": True
                }
                send_message(chat_id, "⚙️ **کنترل اعمال خودکار ربات**\n\nبرای تغییر وضعیت هر گزینه، روی آن کلیک کنید.", keyboard)
                return

            if text.startswith("📌 یادآوری:"):
                new_status = not get_auto_reminder_status()
                set_auto_reminder_status(new_status)
                send_message(chat_id, f"✅ وضعیت یادآوری خودکار به {'فعال' if new_status else 'غیرفعال'} تغییر یافت.", get_super_admin_keyboard())
                return

            if text.startswith("📌 گزارش روزانه:"):
                new_status = not get_auto_report_status()
                set_auto_report_status(new_status)
                send_message(chat_id, f"✅ وضعیت گزارش روزانه خودکار به {'فعال' if new_status else 'غیرفعال'} تغییر یافت.", get_super_admin_keyboard())
                return

            if text.startswith("📌 هشدار افت:"):
                new_status = not get_auto_alert_status()
                set_auto_alert_status(new_status)
                send_message(chat_id, f"✅ وضعیت هشدار افت عملکرد به {'فعال' if new_status else 'غیرفعال'} تغییر یافت.", get_super_admin_keyboard())
                return

            if text.startswith("📌 امتیازدهی:"):
                new_status = not get_auto_scoring_status()
                set_auto_scoring_status(new_status)
                send_message(chat_id, f"✅ وضعیت امتیازدهی خودکار به {'فعال' if new_status else 'غیرفعال'} تغییر یافت.", get_super_admin_keyboard())
                return

            if text.startswith("📌 گزارش هفتگی:"):
                new_status = not get_weekly_report_status()
                set_weekly_report_status(new_status)
                send_message(chat_id, f"✅ وضعیت گزارش هفتگی به {'فعال' if new_status else 'غیرفعال'} تغییر یافت.", get_super_admin_keyboard())
                return

            if text.startswith("📌 گزارش ماهانه:"):
                new_status = not get_monthly_report_status()
                set_monthly_report_status(new_status)
                send_message(chat_id, f"✅ وضعیت گزارش ماهانه به {'فعال' if new_status else 'غیرفعال'} تغییر یافت.", get_super_admin_keyboard())
                return

            if text.startswith("🔔 نوتیفیکیشن:"):
                new_status = not get_instant_notification_status()
                set_instant_notification_status(new_status)
                send_message(chat_id, f"✅ وضعیت نوتیفیکیشن لحظه‌ای به {'فعال' if new_status else 'غیرفعال'} تغییر یافت.", get_super_admin_keyboard())
                return

            if text.startswith("📊 گزارش تطبیقی:"):
                new_status = not get_adaptive_report_status()
                set_adaptive_report_status(new_status)
                send_message(chat_id, f"✅ وضعیت گزارش تطبیقی به {'فعال' if new_status else 'غیرفعال'} تغییر یافت.", get_super_admin_keyboard())
                return

            if text.startswith("📈 پیش‌بینی:"):
                new_status = not get_forecast_report_status()
                set_forecast_report_status(new_status)
                send_message(chat_id, f"✅ وضعیت پیش‌بینی عملکرد به {'فعال' if new_status else 'غیرفعال'} تغییر یافت.", get_super_admin_keyboard())
                return

            if text.startswith("📊 نمودار:"):
                new_status = not get_chart_report_status()
                set_chart_report_status(new_status)
                send_message(chat_id, f"✅ وضعیت گزارش‌های نموداری به {'فعال' if new_status else 'غیرفعال'} تغییر یافت.", get_super_admin_keyboard())
                return

            if text.startswith("📊 آمار واقعی:"):
                new_status = not get_actual_stats_status()
                set_actual_stats_status(new_status)
                send_message(chat_id, f"✅ وضعیت ثبت آمار واقعی به {'فعال' if new_status else 'غیرفعال'} تغییر یافت.", get_super_admin_keyboard())
                return

            if text == "⚙️ مدیریت مشکلات":
                problems = get_all_problems('pending', 20)
                if problems:
                    msg = "📋 **مشکلات ثبت‌شده (در انتظار بررسی)**\n━━━━━━━━━━━━━━━━━━\n"
                    for p in problems:
                        p_id, name, emp_num, problem_text, category, status, created_at_iran = p
                        shamsi_dt = jdatetime.datetime.fromgregorian(datetime=created_at_iran)
                        shamsi_str = f"{shamsi_dt.year}/{shamsi_dt.month:02d}/{shamsi_dt.day:02d} {shamsi_dt.hour:02d}:{shamsi_dt.minute:02d}"
                        msg += f"🆔 {p_id} | {name} ({emp_num})\n"
                        msg += f"📝 {problem_text[:100]}...\n"
                        msg += f"⏰ {shamsi_str}\n"
                        msg += f"برای بررسی: /resolve_problem {p_id}\n\n"
                    send_message(chat_id, msg, get_super_admin_keyboard())
                else:
                    send_message(chat_id, "✅ هیچ مشکل جدیدی وجود ندارد.", get_super_admin_keyboard())
                return

            if text.startswith("/resolve_problem"):
                parts = text.split()
                if len(parts) == 2:
                    try:
                        problem_id = int(parts[1])
                        if update_problem_status(problem_id, 'resolved'):
                            send_message(chat_id, f"✅ مشکل {problem_id} با موفقیت بررسی و بسته شد.", get_super_admin_keyboard())
                            log_user_activity(user_db_id, "resolve_problem", f"بستن مشکل {problem_id}")
                        else:
                            send_message(chat_id, "❌ خطا در بستن مشکل.", get_super_admin_keyboard())
                    except Exception:
                        send_message(chat_id, "❌ فرمت: /resolve_problem [problem_id]", get_super_admin_keyboard())
                else:
                    send_message(chat_id, "❌ فرمت: /resolve_problem [problem_id]", get_super_admin_keyboard())
                return

            if text == "👥 مدیریت معاونین":
                deputies = get_all_deputies_with_details()
                if deputies:
                    msg = "📋 **لیست معاونین شعب**\n━━━━━━━━━━━━━━━━━━\n"
                    for dep in deputies:
                        dep_id, emp_num, full_name, title, branch_id, branch_name = dep
                        branch_name = branch_name or 'بدون شعبه'
                        msg += f"🆔 {dep_id} | {emp_num} | {full_name}\n"
                        msg += f"   🏢 {branch_name} | {title}\n\n"
                    msg += "برای مدیریت:\n"
                    msg += "▪️ /add_deputy [شماره کارمندی] | [نام کامل] | [شناسه شعبه] | [سمت]\n"
                    msg += "▪️ /edit_deputy [user_id] [field] [value]\n"
                    msg += "▪️ /delete_deputy [user_id]\n"
                    msg += "🔒 تغییر شعبه معاون فقط از بخش «مدیریت جابه‌جایی معاونان» انجام می‌شود."
                    send_message(chat_id, msg, get_super_admin_keyboard())
                else:
                    send_message(chat_id, "هیچ معاونی یافت نشد.", get_super_admin_keyboard())
                return

            if text.startswith("/add_deputy"):
                parts = text.replace("/add_deputy", "", 1).strip().split('|', 3)
                if len(parts) == 4:
                    emp_num = normalize_digits(parts[0].strip())
                    full_name = parts[1].strip()
                    try:
                        branch_id = int(parts[2].strip())
                    except Exception:
                        send_message(chat_id, "❌ شناسه شعبه باید عدد باشد.", get_super_admin_keyboard())
                        return
                    title = parts[3].strip()
                    branches = get_all_branches()
                    branch_ids = [b[0] for b in branches]
                    if branch_id not in branch_ids:
                        send_message(chat_id, "❌ شناسه شعبه نامعتبر.", get_super_admin_keyboard())
                        return
                    existing = get_deputy_by_employee_number(emp_num)
                    if existing:
                        send_message(chat_id, f"❌ شماره کارمندی {emp_num} قبلاً ثبت شده است.", get_super_admin_keyboard())
                        return
                    new_id = add_deputy(emp_num, full_name, title, branch_id, False)
                    if new_id:
                        send_message(chat_id, f"✅ معاون با شماره کارمندی {emp_num} و نام {full_name} اضافه شد.", get_super_admin_keyboard())
                        log_user_activity(user_db_id, "add_deputy", f"افزودن معاون {full_name} با شماره {emp_num}")
                    else:
                        send_message(chat_id, "❌ خطا در افزودن معاون.", get_super_admin_keyboard())
                else:
                    send_message(chat_id, "❌ فرمت: /add_deputy [شماره کارمندی] | [نام کامل] | [شناسه شعبه] | [سمت]", get_super_admin_keyboard())
                return

            if text.startswith("/edit_deputy"):
                parts = text.split(' ', 3)
                if len(parts) >= 4:
                    try:
                        user_id = int(parts[1])
                        field = parts[2]
                        value = parts[3].strip() if len(parts) > 3 else ''
                        if field not in ALLOWED_UPDATE_FIELDS:
                            send_message(chat_id, f"❌ فیلد نامعتبر. فیلدهای قابل ویرایش: {', '.join(ALLOWED_UPDATE_FIELDS)}", get_super_admin_keyboard())
                            return
                        if field == 'employee_number':
                            value = normalize_digits(value)
                            existing = get_deputy_by_employee_number(value)
                            if existing and existing[0] != user_id:
                                send_message(chat_id, f"❌ شماره کارمندی {value} قبلاً ثبت شده است.", get_super_admin_keyboard())
                                return
                            if update_deputy(user_id, employee_number=value):
                                send_message(chat_id, f"✅ شماره کارمندی معاون به {value} تغییر یافت.", get_super_admin_keyboard())
                            else:
                                send_message(chat_id, "❌ خطا در ویرایش.", get_super_admin_keyboard())
                        elif field == 'full_name':
                            if update_deputy(user_id, full_name=value):
                                send_message(chat_id, f"✅ نام معاون به {value} تغییر یافت.", get_super_admin_keyboard())
                            else:
                                send_message(chat_id, "❌ خطا در ویرایش.", get_super_admin_keyboard())
                        elif field == 'title':
                            if update_deputy(user_id, title=value):
                                send_message(chat_id, f"✅ سمت معاون به {value} تغییر یافت.", get_super_admin_keyboard())
                            else:
                                send_message(chat_id, "❌ خطا در ویرایش.", get_super_admin_keyboard())
                        elif field == 'branch_id':
                            try:
                                branch_id = int(value)
                            except Exception:
                                send_message(chat_id, "❌ شناسه شعبه باید عدد باشد.", get_super_admin_keyboard())
                                return
                            branches = get_all_branches()
                            branch_ids = [b[0] for b in branches]
                            if branch_id not in branch_ids:
                                send_message(chat_id, "❌ شناسه شعبه نامعتبر.", get_super_admin_keyboard())
                                return
                            if update_deputy(user_id, branch_id=branch_id):
                                send_message(chat_id, f"✅ شعبه معاون به شناسه {branch_id} تغییر یافت.", get_super_admin_keyboard())
                            else:
                                send_message(chat_id, "❌ خطا در ویرایش.", get_super_admin_keyboard())
                    except Exception as e:
                        logger.exception("Super-admin command failed")
                        send_message(chat_id, "❌ عملیات انجام نشد؛ جزئیات خطا ثبت شد.", get_super_admin_keyboard())
                else:
                    send_message(chat_id, "❌ فرمت: /edit_deputy [user_id] [field] [value]", get_super_admin_keyboard())
                return

            if text.startswith("/delete_deputy"):
                parts = text.split()
                if len(parts) == 2:
                    try:
                        user_id = int(parts[1])
                        success, msg = delete_deputy(user_id)
                        if success:
                            send_message(chat_id, f"✅ {msg}", get_super_admin_keyboard())
                            log_user_activity(user_db_id, "delete_deputy", f"حذف معاون با شناسه {user_id}")
                        else:
                            send_message(chat_id, f"❌ {msg}", get_super_admin_keyboard())
                    except Exception:
                        send_message(chat_id, "❌ فرمت: /delete_deputy [user_id]", get_super_admin_keyboard())
                else:
                    send_message(chat_id, "❌ فرمت: /delete_deputy [user_id]", get_super_admin_keyboard())
                return

            if text == "🔧 وضعیت ربات":
                current_status = get_bot_status()
                status_text = "فعال ✅" if current_status else "غیرفعال ❌"
                keyboard = {
                    "keyboard": [
                        [{"text": "🔛 فعال کردن ربات" if not current_status else "🔛 فعال است"}],
                        [{"text": "🔴 غیرفعال کردن ربات" if current_status else "🔴 غیرفعال است"}],
                        [{"text": "🔙 انصراف"}]
                    ],
                    "resize_keyboard": True
                }
                send_message(chat_id, f"📊 **وضعیت فعلی ربات:** {status_text}", keyboard)
                return

            if text == "🔛 فعال کردن ربات":
                if set_bot_status(True):
                    send_message(chat_id, "✅ ربات با موفقیت **فعال** شد.", get_super_admin_keyboard())
                else:
                    send_message(chat_id, "❌ خطا در فعال‌سازی ربات.", get_super_admin_keyboard())
                return

            if text == "🔴 غیرفعال کردن ربات":
                if set_bot_status(False):
                    send_message(chat_id, "✅ ربات با موفقیت **غیرفعال** شد.", get_super_admin_keyboard())
                else:
                    send_message(chat_id, "❌ خطا در غیرفعال‌سازی ربات.", get_super_admin_keyboard())
                return

            if text == "🔄 ریست گزارش‌ها":
                keyboard = {
                    "keyboard": [
                        [{"text": "✅ بله، ریست کن"}, {"text": "❌ خیر، لغو"}]
                    ],
                    "resize_keyboard": True
                }
                send_message(chat_id, "🔄 این عملیات فقط حافظه موقت گزارش‌ها را پاک و خروجی‌ها را تازه‌سازی می‌کند؛ هیچ وصول، یادداشت یا سابقه‌ای حذف نمی‌شود.\n\nآیا ادامه می‌دهید؟", keyboard)
                user_states.update(chat_id, {"state": "WAITING_FOR_RESET_CONFIRM"})
                return

            if text == "📅 مدیریت تعطیلات":
                keyboard = {
                    "keyboard": [
                        [{"text": "➕ افزودن روز تعطیل"}, {"text": "➖ حذف روز تعطیل"}],
                        [{"text": "📋 مشاهده تعطیلات"}, {"text": "🔙 انصراف"}]
                    ],
                    "resize_keyboard": True
                }
                send_message(chat_id, "📅 **مدیریت تعطیلات**\n\nلطفاً یکی از گزینه‌های زیر را انتخاب کنید:", keyboard)
                return

            if text == "➕ افزودن روز تعطیل":
                user_states.update(chat_id, {"state": "WAITING_FOR_ADD_HOLIDAY"})
                send_message(chat_id, "📅 لطفاً تاریخ مورد نظر را به فرمت **YYYY/MM/DD** وارد کنید (مثلاً ۱۴۰۴/۰۱/۱۵)\nو در صورت تمایل توضیحی وارد کنید:\n\n`تاریخ | توضیح`\nمثال: `۱۴۰۴/۰۱/۱۵ | تعطیلات رسمی`", get_cancel_keyboard())
                return

            if text == "➖ حذف روز تعطیل":
                holidays = get_all_holidays(20)
                if not holidays:
                    send_message(chat_id, "📭 هیچ روز تعطیلی ثبت نشده است.", get_super_admin_keyboard())
                    return
                msg = "📋 **لیست تعطیلات ثبت‌شده**\n━━━━━━━━━━━━━━━━━━\n"
                for i, h in enumerate(holidays, 1):
                    msg += f"{i}. {get_shamsi_date_formatted(h[0])} - {h[1]}\n"
                msg += "\nلطفاً شماره مورد نظر برای حذف را وارد کنید، یا 🔙 انصراف بزنید."
                user_states.update(chat_id, {
                    "state": "WAITING_FOR_REMOVE_HOLIDAY",
                    "holidays_list": holidays
                })
                send_message(chat_id, msg, get_cancel_keyboard())
                return

            if text == "📋 مشاهده تعطیلات":
                holidays = get_all_holidays(30)
                if holidays:
                    msg = "📋 **تعطیلات ثبت‌شده**\n━━━━━━━━━━━━━━━━━━\n"
                    for h in holidays:
                        msg += f"📅 {get_shamsi_date_formatted(h[0])} - {h[1]}\n"
                    send_message(chat_id, msg, get_super_admin_keyboard())
                else:
                    send_message(chat_id, "📭 هیچ روز تعطیلی ثبت نشده است.", get_super_admin_keyboard())
                return

            if text == "📨 ارسال پیام به معاونین":
                deputies = get_all_deputies()
                if not deputies:
                    send_message(chat_id, "هیچ معاونی یافت نشد.", get_super_admin_keyboard())
                    return
                msg = "📨 **ارسال پیام به معاونین**\n\n"
                msg += "لیست معاونین:\n"
                for i, dep in enumerate(deputies, 1):
                    msg += f"{i}. {dep[2]} - {dep[4] or 'بدون شعبه'}\n"
                msg += "\nبرای انتخاب مخاطب، یکی از گزینه‌های زیر را وارد کنید:\n"
                msg += "▪️ `همه` برای ارسال به همه\n"
                msg += "▪️ شماره ردیف (مثلاً `1` یا `1,2,3`)\n"
                msg += "▪️ نام معاون (مثلاً `علی محمدی`)\n"
                msg += "سپس پیام خود را ارسال کنید."
                user_states.update(chat_id, {
                    "state": "WAITING_FOR_MESSAGE_RECIPIENT",
                    "deputies": deputies,
                    "user_data": user_data
                })
                send_message(chat_id, msg, get_cancel_keyboard())
                return

            if text.startswith("/edit_role"):
                parts = text.split()
                if len(parts) == 3:
                    try:
                        user_id = int(parts[1])
                        new_role = parts[2]
                        if new_role in VALID_ROLES:
                            if update_user_role(user_id, new_role):
                                send_message(chat_id, f"✅ نقش کاربر {user_id} به {new_role} تغییر یافت.", get_super_admin_keyboard())
                            else:
                                send_message(chat_id, "❌ خطا در تغییر نقش.", get_super_admin_keyboard())
                        else:
                            send_message(chat_id, f"❌ نقش نامعتبر. فقط {', '.join(VALID_ROLES)} مجاز است.", get_super_admin_keyboard())
                    except Exception:
                        send_message(chat_id, "❌ فرمت: /edit_role [user_id] [role]", get_super_admin_keyboard())
                else:
                    send_message(chat_id, "❌ فرمت: /edit_role [user_id] [role]", get_super_admin_keyboard())
                return

            if text.startswith("/edit_branch"):
                parts = text.split()
                if len(parts) == 3:
                    try:
                        user_id = int(parts[1])
                        branch_id = int(parts[2])
                        if update_user_branch(user_id, branch_id):
                            send_message(chat_id, f"✅ شعبه کاربر {user_id} به {branch_id} تغییر یافت.", get_super_admin_keyboard())
                        else:
                            send_message(chat_id, "❌ خطا در تغییر شعبه.", get_super_admin_keyboard())
                    except Exception:
                        send_message(chat_id, "❌ فرمت: /edit_branch [user_id] [branch_id]", get_super_admin_keyboard())
                else:
                    send_message(chat_id, "❌ فرمت: /edit_branch [user_id] [branch_id]", get_super_admin_keyboard())
                return

            if text.startswith("/delete_user"):
                parts = text.split()
                if len(parts) == 2:
                    try:
                        user_id = int(parts[1])
                        success, msg = delete_user(user_id)
                        if success:
                            send_message(chat_id, f"✅ {msg}", get_super_admin_keyboard())
                        else:
                            send_message(chat_id, f"❌ {msg}", get_super_admin_keyboard())
                    except Exception:
                        send_message(chat_id, "❌ فرمت: /delete_user [user_id]", get_super_admin_keyboard())
                else:
                    send_message(chat_id, "❌ فرمت: /delete_user [user_id]", get_super_admin_keyboard())
                return

            if text.startswith("/delete_collection"):
                parts = text.split()
                if len(parts) == 2:
                    try:
                        col_id = int(parts[1])
                        if delete_collection(col_id):
                            send_message(chat_id, f"✅ گزارش {col_id} حذف شد.", get_super_admin_keyboard())
                        else:
                            send_message(chat_id, "⛔ حذف فیزیکی گزارش برای حفظ سوابق غیرفعال است؛ در صورت نیاز از ویرایش گزارش استفاده کنید.", get_super_admin_keyboard())
                    except Exception:
                        send_message(chat_id, "❌ فرمت: /delete_collection [id]", get_super_admin_keyboard())
                else:
                    send_message(chat_id, "❌ فرمت: /delete_collection [id]", get_super_admin_keyboard())
                return

            if text.startswith("/edit_collection"):
                parts = text.split()
                if len(parts) == 4:
                    try:
                        col_id = int(parts[1])
                        deputy = int(parts[2]) * 1_000_000
                        others = int(parts[3]) * 1_000_000
                        if update_collection(col_id, deputy, others):
                            send_message(chat_id, f"✅ گزارش {col_id} به‌روزرسانی شد.", get_super_admin_keyboard())
                        else:
                            send_message(chat_id, "❌ خطا در ویرایش گزارش.", get_super_admin_keyboard())
                    except Exception:
                        send_message(chat_id, "❌ فرمت: /edit_collection [id] [deputy_amount_millions] [others_amount_millions]", get_super_admin_keyboard())
                else:
                    send_message(chat_id, "❌ فرمت: /edit_collection [id] [deputy_amount_millions] [others_amount_millions]", get_super_admin_keyboard())
                return

            if text == "📊 گزارش هفتگی":
                send_message(chat_id, "🔄 گزارش هفتگی در پس‌زمینه تولید و برای کاربران ارسال می‌شود.", get_super_admin_keyboard())
                executor.submit(send_weekly_report_to_all)
                return

            if text == "📊 گزارش ماهانه":
                send_message(chat_id, "🔄 گزارش ماه جاری در پس‌زمینه تولید و برای کاربران ارسال می‌شود.", get_super_admin_keyboard())
                executor.submit(send_monthly_report_to_all, True)
                return

            if text == "👥 مدیریت کاربران":
                users = get_all_users()
                if users:
                    msg = "📋 **لیست کاربران**\n━━━━━━━━━━━━━━━━━━\n"
                    for u in users:
                        access_status = "فعال" if u[7] else "غیرفعال (سابقه محفوظ)"
                        msg += f"🆔 {u[0]} | {u[1]} | {u[2]} | نقش: {u[3]} | شعبه: {u[5]} | {access_status}\n"
                    msg += "\nبرای مدیریت، از گزینه‌های زیر استفاده کنید:\n"
                    msg += "▪️ /edit_role [user_id] [admin|deputy|super_admin]\n"
                    msg += "▪️ /edit_branch [user_id] [branch_id]\n"
                    msg += "▪️ /delete_user [user_id]"
                    send_message(chat_id, msg, get_super_admin_keyboard())
                else:
                    send_message(chat_id, "هیچ کاربری یافت نشد.", get_super_admin_keyboard())
                return

            if text == "📊 مدیریت گزارش‌ها":
                collections = get_all_collections(20)
                if collections:
                    msg = "📊 **۲۰ گزارش اخیر**\n━━━━━━━━━━━━━━━━━━\n"
                    for c in collections:
                        msg += f"🆔 {c[0]} | {c[1]} | {c[2]} | {c[5]//1_000_000:,.0f} میلیون ریال | ثبت: {c[6]}\n"
                    msg += "\nبرای حذف: /delete_collection [id]\n"
                    msg += "برای ویرایش: /edit_collection [id] [deputy_amount] [others_amount] (به میلیون ریال)"
                    send_message(chat_id, msg, get_super_admin_keyboard())
                else:
                    send_message(chat_id, "هیچ گزارشی یافت نشد.", get_super_admin_keyboard())
                return

            if text == "📋 مشاهده لاگ‌ها":
                log_file = get_log_file_path()
                if os.path.exists(log_file):
                    try:
                        with open(log_file, 'r', encoding='utf-8') as f:
                            lines = f.readlines()[-50:]
                            log_text = "".join(lines)
                            if len(log_text) > 4000:
                                log_text = log_text[-4000:]
                            send_message(chat_id, f"📋 **آخرین لاگ‌ها**\n```\n{log_text}\n```", get_super_admin_keyboard())
                    except Exception as e:
                        logger.exception("Reading log failed")
                        send_message(chat_id, "❌ خواندن لاگ انجام نشد.", get_super_admin_keyboard())
                else:
                    send_message(chat_id, "فایل لاگ وجود ندارد.", get_super_admin_keyboard())
                return

            if text == "📋 لاگ ورود/خروج":
                logs = get_user_activity_log(50)
                if logs:
                    msg = "📋 **لاگ فعالیت کاربران**\n━━━━━━━━━━━━━━━━━━\n"
                    for log in logs:
                        created_at_iran = log[5]
                        shamsi_dt = jdatetime.datetime.fromgregorian(datetime=created_at_iran)
                        shamsi_str = f"{shamsi_dt.year}/{shamsi_dt.month:02d}/{shamsi_dt.day:02d} {shamsi_dt.hour:02d}:{shamsi_dt.minute:02d}"
                        msg += f"👤 {log[1]} ({log[2]}) | {log[3]}\n"
                        msg += f"📝 {log[4]}\n"
                        msg += f"⏰ {shamsi_str}\n\n"
                    send_message(chat_id, msg, get_super_admin_keyboard())
                else:
                    send_message(chat_id, "هیچ فعالیتی ثبت نشده است.", get_super_admin_keyboard())
                return

            if text == "📝 مشاهده یادداشت‌ها":
                notes = get_all_notes_with_collection(30)
                if notes:
                    msg = "📝 **یادداشت‌های اخیر**\n━━━━━━━━━━━━━━━━━━\n"
                    for note in notes:
                        note_time = note[5]
                        msg += f"🏢 {note[1]} | 📅 {note[2]}\n"
                        msg += f"👤 {note[3]}: {note[4]}\n"
                        msg += f"⏰ {note_time.strftime('%H:%M') if hasattr(note_time, 'strftime') else note_time}\n\n"
                    send_message(chat_id, msg, get_super_admin_keyboard())
                else:
                    send_message(chat_id, "هیچ یادداشتی وجود ندارد.", get_super_admin_keyboard())
                return

            if text == "📊 ثبت آمار واقعی":
                if not get_actual_stats_status():
                    send_message(chat_id, "🔴 ثبت آمار واقعی در حال حاضر غیرفعال است.", get_super_admin_keyboard())
                    return
                user_states.update(chat_id, {"state": "WAITING_FOR_ACTUAL_METHOD"})
                send_message(
                    chat_id,
                    "📊 **ثبت آمار واقعی**\n\nروش ثبت را انتخاب کنید:",
                    get_actual_stats_method_keyboard(),
                )
                return

            if text == "✍️ ثبت دستی آمار واقعی":
                if not get_actual_stats_status():
                    send_message(chat_id, "🔴 ثبت آمار واقعی در حال حاضر غیرفعال است.", get_super_admin_keyboard())
                    return
                user_states.update(chat_id, {"state": "WAITING_FOR_ACTUAL_DATE"})
                send_message(chat_id, "📅 لطفاً **تاریخ** مورد نظر برای ثبت آمار واقعی را به فرمت YYYY/MM/DD وارد کنید:\n\n(مثلاً 1403/01/15)", get_cancel_keyboard())
                return

            if text == "📊 ثبت آمار واقعی از Excel":
                if not get_actual_stats_status():
                    send_message(chat_id, "🔴 ثبت آمار واقعی در حال حاضر غیرفعال است.", get_super_admin_keyboard())
                    return
                if not OPENPYXL_AVAILABLE:
                    send_message(
                        chat_id,
                        "❌ امکان خواندن Excel روی سرور نصب نیست. ابتدا requirements.txt را نصب/به‌روزرسانی کنید.",
                        get_super_admin_keyboard(),
                    )
                    return
                coded_branches = get_branches_with_official_codes()
                missing_codes = [name for _, name, code in coded_branches if not code]
                if not coded_branches or missing_codes:
                    details = "، ".join(missing_codes) if missing_codes else "دیتابیس در دسترس نیست"
                    send_message(chat_id, f"❌ کد رسمی شعب کامل نیست: {details}", get_super_admin_keyboard())
                    return
                user_states.update(chat_id, {"state": "WAITING_FOR_ACTUAL_EXCEL_FILE"})
                send_message(
                    chat_id,
                    "📎 فایل اصلی **Excel با پسوند .xlsx** را ارسال کنید.\n\n"
                    "ربات کد شعبه و ستون «نسبت به روز قبل ← مبلغ» را می‌خواند. "
                    "سلول مبلغ خالی صفر در نظر گرفته می‌شود و قبل از ثبت، پیش‌نمایش کامل نمایش داده خواهد شد.",
                    get_cancel_keyboard(),
                )
                return

            # ===== گزارش‌های جدید سوپرادمین =====
            if text == "🏅 رتبه‌بندی دقت معاونان":
                ranking = get_deputy_accuracy_ranking(30)
                if not ranking:
                    send_message(chat_id, "📭 داده‌های کافی برای رتبه‌بندی دقت وجود ندارد.", get_super_admin_keyboard())
                    return
                msg = "🏅 **رتبه‌بندی معاونان بر اساس دقت خوداظهاری (۳۰ روز اخیر)**\n━━━━━━━━━━━━━━━━━━\n\n"
                for idx, row in enumerate(ranking, 1):
                    dep_id, name, branch, days, avg_acc = row
                    medal = "🥇" if idx == 1 else "🥈" if idx == 2 else "🥉" if idx == 3 else f"{idx}."
                    msg += f"{medal} {name} - {branch}\n"
                    msg += f"   📊 تعداد روزهای ثبت: {days}\n"
                    msg += f"   🎯 میانگین دقت: {avg_acc:.1f}%\n\n"
                send_message(chat_id, msg, get_super_admin_keyboard())
                return

            if text == "📈 روند دقت شعبه":
                user_states.update(chat_id, {"state": "WAITING_FOR_BRANCH_ACCURACY", "accuracy_context": "trend"})
                send_message(chat_id, "🏢 لطفاً **نام شعبه** مورد نظر را برای مشاهده روند دقت وارد کنید:", get_cancel_keyboard())
                return

            if text == "📊 بهترین/بدترین دقت روز":
                user_states.update(chat_id, {"state": "WAITING_FOR_BRANCH_ACCURACY", "accuracy_context": "best_worst"})
                send_message(chat_id, "📅 لطفاً **تاریخ** مورد نظر را به فرمت YYYY/MM/DD وارد کنید:", get_cancel_keyboard())
                return

            if text == "📊 مقایسه عملکرد شعبه با استان":
                user_states.update(chat_id, {"state": "WAITING_FOR_BRANCH_ACCURACY", "accuracy_context": "compare_avg"})
                send_message(chat_id, "🏢 لطفاً **نام شعبه** مورد نظر را برای مقایسه با میانگین استان وارد کنید:", get_cancel_keyboard())
                return

            if text == "⏰ تحلیل تاخیر معاونان":
                late_data = get_deputy_late_analysis(30)
                if not late_data:
                    send_message(chat_id, "📭 داده‌های کافی برای تحلیل تاخیر وجود ندارد.", get_super_admin_keyboard())
                    return
                msg = "⏰ **تحلیل تاخیر در ثبت وصول معاونان (۳۰ روز اخیر)**\n━━━━━━━━━━━━━━━━━━\n\n"
                for idx, row in enumerate(late_data, 1):
                    dep_id, name, branch, total_days, late_days, late_percent = row
                    emoji = "🔴" if late_percent > 30 else "🟡" if late_percent > 15 else "🟢"
                    msg += f"{idx}. {name} - {branch}\n"
                    msg += f"   📅 تعداد روزهای ثبت: {total_days}\n"
                    msg += f"   ⏰ تاخیر: {late_days} روز ({late_percent:.1f}%) {emoji}\n\n"
                send_message(chat_id, msg, get_super_admin_keyboard())
                return

            # ===== مدیریت Stateهای دقت =====
            if current_state == "WAITING_FOR_BRANCH_ACCURACY":
                if text == "🔙 انصراف":
                    user_states.update(chat_id, {"state": "LOGGED_IN"})
                    send_message(chat_id, "❌ عملیات لغو شد.", get_super_admin_keyboard())
                    return
                context = user_state.get("accuracy_context")
                if context == "trend":
                    conn = None
                    try:
                        conn = get_db_connection()
                        with conn.cursor() as cur:
                            cur.execute("SELECT id FROM branches WHERE name ILIKE %s ESCAPE '\\' LIMIT 1", (f"%{escape_like_pattern(text)}%",))
                            result = cur.fetchone()
                            if not result:
                                send_message(chat_id, f"❌ شعبه‌ای با نام {text} یافت نشد.", get_cancel_keyboard())
                                return
                            branch_id = result[0]
                            trend = get_branch_accuracy_trend(branch_id, 30)
                            if not trend:
                                send_message(chat_id, f"📭 داده‌های کافی برای روند دقت شعبه {text} وجود ندارد.", get_super_admin_keyboard())
                                user_states.update(chat_id, {"state": "LOGGED_IN"})
                                return
                            msg = f"📈 **روند دقت خوداظهاری شعبه {text} (۳۰ روز اخیر)**\n━━━━━━━━━━━━━━━━━━\n\n"
                            for row in trend:
                                date, collected, actual, accuracy = row
                                if accuracy is None:
                                    continue
                                emoji = "✅" if accuracy >= 95 else "🟢" if accuracy >= 80 else "🟡" if accuracy >= 50 else "🔴"
                                msg += f"📅 {get_shamsi_date_formatted(date)}\n"
                                msg += f"   ادعا: {collected//1_000_000:,.0f} میلیون ریال | واقعی: {abs(actual)//1_000_000 if actual else 0:,.0f} میلیون ریال\n"
                                msg += f"   🎯 دقت: {accuracy:.1f}% {emoji}\n\n"
                            send_message(chat_id, msg, get_super_admin_keyboard())
                    except Exception as e:
                        logger.exception("Branch report failed")
                        send_message(chat_id, "❌ عملیات انجام نشد؛ جزئیات خطا ثبت شد.", get_cancel_keyboard())
                    finally:
                        if conn:
                            return_db_connection(conn)
                        user_states.update(chat_id, {"state": "LOGGED_IN"})
                elif context == "best_worst":
                    shamsi_date = normalize_digits(text)
                    if not validate_shamsi_date(shamsi_date):
                        send_message(chat_id, "❌ فرمت تاریخ نامعتبر. لطفاً به صورت YYYY/MM/DD وارد کنید.", get_cancel_keyboard())
                        return
                    best, worst = get_best_worst_accuracy_branches(shamsi_date, 5)
                    if not best and not worst:
                        send_message(chat_id, f"📭 هیچ داده‌ای برای تاریخ {get_shamsi_date_formatted(shamsi_date)} یافت نشد.", get_super_admin_keyboard())
                        user_states.update(chat_id, {"state": "LOGGED_IN"})
                        return
                    msg = f"📊 **بهترین و بدترین دقت خوداظهاری - {get_shamsi_date_formatted(shamsi_date)}**\n━━━━━━━━━━━━━━━━━━\n\n"
                    msg += "✅ **بهترین دقت:**\n"
                    if best:
                        for row in best:
                            name, collected, actual, acc = row
                            msg += f"🏢 {name}: {acc:.1f}% (ادعا: {collected//1_000_000:,.0f} | واقعی: {abs(actual)//1_000_000 if actual else 0:,.0f})\n"
                    else:
                        msg += "هیچ داده‌ای موجود نیست.\n"
                    msg += "\n🔴 **بدترین دقت:**\n"
                    if worst:
                        for row in worst:
                            name, collected, actual, acc = row
                            msg += f"🏢 {name}: {acc:.1f}% (ادعا: {collected//1_000_000:,.0f} | واقعی: {abs(actual)//1_000_000 if actual else 0:,.0f})\n"
                    else:
                        msg += "هیچ داده‌ای موجود نیست."
                    send_message(chat_id, msg, get_super_admin_keyboard())
                    user_states.update(chat_id, {"state": "LOGGED_IN"})
                elif context == "compare_avg":
                    conn = None
                    try:
                        conn = get_db_connection()
                        with conn.cursor() as cur:
                            cur.execute("SELECT id FROM branches WHERE name ILIKE %s ESCAPE '\\' LIMIT 1", (f"%{escape_like_pattern(text)}%",))
                            result = cur.fetchone()
                            if not result:
                                send_message(chat_id, f"❌ شعبه‌ای با نام {text} یافت نشد.", get_cancel_keyboard())
                                return
                            branch_id = result[0]
                            comp = get_branch_performance_vs_avg(branch_id, 30)
                            if not comp:
                                send_message(chat_id, f"📭 داده‌های کافی برای مقایسه شعبه {text} با استان وجود ندارد.", get_super_admin_keyboard())
                                user_states.update(chat_id, {"state": "LOGGED_IN"})
                                return
                            msg = f"📊 **مقایسه عملکرد شعبه {text} با میانگین استان (۳۰ روز اخیر)**\n━━━━━━━━━━━━━━━━━━\n\n"
                            msg += f"🏢 شعبه {text}\n"
                            msg += f"   📅 تعداد روزهای ثبت: {comp['days']}\n"
                            msg += f"   💰 میانگین روزانه: {comp['avg_branch']//1_000_000:,.0f} میلیون ریال\n"
                            msg += f"   💰 کل وصول: {comp['total_branch']//1_000_000:,.0f} میلیون ریال\n\n"
                            msg += f"📊 میانگین استان: {comp['avg_province']//1_000_000:,.0f} میلیون ریال\n"
                            diff = comp['diff_percent']
                            if diff > 0:
                                msg += f"📈 عملکرد شعبه {diff:.1f}% **بالاتر** از میانگین استان است."
                            elif diff < 0:
                                msg += f"📉 عملکرد شعبه {abs(diff):.1f}% **پایین‌تر** از میانگین استان است."
                            else:
                                msg += f"➡️ عملکرد شعبه برابر با میانگین استان است."
                            send_message(chat_id, msg, get_super_admin_keyboard())
                    except Exception as e:
                        logger.exception("Branch comparison failed")
                        send_message(chat_id, "❌ عملیات انجام نشد؛ جزئیات خطا ثبت شد.", get_cancel_keyboard())
                    finally:
                        if conn:
                            return_db_connection(conn)
                        user_states.update(chat_id, {"state": "LOGGED_IN"})
                else:
                    user_states.update(chat_id, {"state": "LOGGED_IN"})
                    send_message(chat_id, "❌ خطا در تشخیص درخواست.", get_super_admin_keyboard())
                return

            # ===== پشتیبان‌گیری و بازیابی =====
            if text == "🩺 سلامت دیتابیس":
                conn = None
                try:
                    conn = get_db_connection()
                    with conn.cursor() as cur:
                        cur.execute("SELECT current_database(), current_user, version(), pg_database_size(current_database()), now()")
                        db_name, db_user, db_version, db_size, db_now = cur.fetchone()
                        cur.execute("""
                            SELECT COUNT(*) FILTER (WHERE state = 'active'), COUNT(*)
                            FROM pg_stat_activity WHERE datname = current_database()
                        """)
                        active_connections, total_connections = cur.fetchone()
                    msg = (f"🩺 **سلامت دیتابیس**\n\n"
                           f"✅ اتصال و اجرای کوئری موفق\n"
                           f"🗄 نام: {db_name}\n👤 کاربر: {db_user}\n"
                           f"💾 حجم: {db_size / 1024 / 1024:.2f} MB\n"
                           f"🔌 اتصال فعال/کل: {active_connections}/{total_connections}\n"
                           f"🕒 زمان سرور: {db_now}\n"
                           f"🐘 نسخه: {db_version.split(',')[0]}")
                    send_message(chat_id, msg, get_super_admin_keyboard())
                    log_user_activity(user_db_id, "database_health", "بررسی سلامت دیتابیس")
                except Exception as e:
                    if conn:
                        conn.rollback()
                    logger.error(f"Database health error: {e}")
                    send_message(chat_id, "❌ بررسی سلامت دیتابیس انجام نشد؛ جزئیات در لاگ ثبت شد.", get_super_admin_keyboard())
                finally:
                    if conn:
                        return_db_connection(conn)
                return

            if text == "📦 آمار حجم جداول":
                conn = None
                try:
                    conn = get_db_connection()
                    with conn.cursor() as cur:
                        cur.execute("""
                            SELECT relname, n_live_tup,
                                   pg_total_relation_size(relid) AS total_bytes
                            FROM pg_stat_user_tables
                            ORDER BY total_bytes DESC, relname
                        """)
                        rows = cur.fetchall()
                    msg = "📦 **آمار جداول**\n━━━━━━━━━━━━━━━━━━\n"
                    for table_name, estimated_rows, total_bytes in rows:
                        msg += f"• {table_name}: حدود {estimated_rows:,} ردیف | {total_bytes / 1024:.1f} KB\n"
                    send_message(chat_id, msg, get_super_admin_keyboard())
                except Exception as e:
                    if conn:
                        conn.rollback()
                    logger.exception("Table statistics failed")
                    send_message(chat_id, "❌ دریافت آمار جداول انجام نشد.", get_super_admin_keyboard())
                finally:
                    if conn:
                        return_db_connection(conn)
                return

            if text == "💾 پشتیبان‌گیری از داده‌ها":
                send_message(chat_id, "⏳ در حال تولید فایل پشتیبان... لطفاً صبر کنید.", get_super_admin_keyboard())
                executor.submit(generate_and_send_backup, chat_id, user_db_id)
                return

            if text == "📂 بازیابی داده‌ها":
                user_states.update(chat_id, {"state": "WAITING_FOR_RESTORE_FILE"})
                send_message(chat_id, "📂 فایل پشتیبان را به صورت **سند** ارسال کنید.\n\nابتدا آزمون تراکنشی انجام می‌شود و هیچ داده‌ای تغییر نمی‌کند. پس از نمایش نتیجه، تأیید جداگانه لازم است. ریستور فقط ردیف‌های غایب را اضافه می‌کند و داده‌های موجود را حذف یا بازنویسی نمی‌کند.", get_cancel_keyboard())
                return

            if current_state == "WAITING_FOR_RESTORE_FILE":
                if text == "🔙 انصراف":
                    user_states.update(chat_id, {"state": "LOGGED_IN"})
                    send_message(chat_id, "❌ عملیات لغو شد.", get_super_admin_keyboard())
                    return
                if message.get('document'):
                    file_id = None
                    if message.get('document'):
                        file_id = message['document']['file_id']
                    if file_id:
                        try:
                            file_url_res = get_http_session().get(f"{BASE_URL}/getFile", params={"file_id": file_id}, timeout=30)
                            if file_url_res.status_code == 200 and file_url_res.json().get('ok'):
                                file_path = file_url_res.json().get('result', {}).get('file_path')
                                if file_path:
                                    file_data_res = get_http_session().get(
                                        f"https://tapi.bale.ai/file/bot{BOT_TOKEN}/{file_path}",
                                        timeout=60,
                                        stream=True
                                    )
                                    if file_data_res.status_code == 200:
                                        declared_size = int(file_data_res.headers.get('Content-Length') or 0)
                                        if declared_size > MAX_BACKUP_COMPRESSED_BYTES:
                                            file_data_res.close()
                                            send_message(chat_id, "❌ حجم فایل بیش از حد مجاز است.", get_cancel_keyboard())
                                            return
                                        chunks = []
                                        downloaded = 0
                                        for chunk in file_data_res.iter_content(chunk_size=64 * 1024):
                                            if not chunk:
                                                continue
                                            downloaded += len(chunk)
                                            if downloaded > MAX_BACKUP_COMPRESSED_BYTES:
                                                file_data_res.close()
                                                send_message(chat_id, "❌ حجم فایل بیش از حد مجاز است.", get_cancel_keyboard())
                                                return
                                            chunks.append(chunk)
                                        file_data_res.close()
                                        file_bytes = b''.join(chunks)
                                        success, msg, summary = restore_from_file(file_bytes, dry_run=True)
                                        if success:
                                            inserted = sum(summary.get('inserted', {}).values())
                                            skipped = sum(summary.get('skipped', {}).values())
                                            user_states.update(chat_id, {
                                                "state": "WAITING_FOR_RESTORE_CONFIRM",
                                                "restore_file_bytes": file_bytes,
                                                "restore_expires_at": time.time() + 300,
                                                "user_data": user_data
                                            })
                                            keyboard = {"keyboard": [[{"text": "✅ تأیید بازیابی افزایشی"}], [{"text": "🔙 انصراف"}]], "resize_keyboard": True}
                                            send_message(chat_id, f"✅ {msg}\n\nردیف‌های قابل افزودن: {inserted}\nردیف‌های موجود/ردشده: {skipped}\n\nبرای اجرای واقعی تا ۵ دقیقه آینده تأیید کنید.", keyboard)
                                            return
                                        else:
                                            send_message(chat_id, f"❌ {msg}", get_cancel_keyboard())
                                    else:
                                        send_message(chat_id, "❌ خطا در دریافت فایل از سرور.", get_cancel_keyboard())
                                else:
                                    send_message(chat_id, "❌ مسیر فایل دریافت نشد.", get_cancel_keyboard())
                            else:
                                send_message(chat_id, "❌ خطا در دریافت اطلاعات فایل.", get_cancel_keyboard())
                        except Exception as e:
                            logger.error(f"Restore error: {e}")
                            send_message(chat_id, "❌ بازیابی انجام نشد؛ جزئیات خطا ثبت شد.", get_cancel_keyboard())
                    else:
                        send_message(chat_id, "❌ فایل معتبری یافت نشد.", get_cancel_keyboard())
                else:
                    send_message(chat_id, "❌ لطفاً فایل پشتیبان را به صورت سند ارسال کنید.", get_cancel_keyboard())
                user_states.update(chat_id, {"state": "LOGGED_IN"})
                return

            if current_state == "WAITING_FOR_RESTORE_CONFIRM":
                if text == "🔙 انصراف":
                    user_states.update(chat_id, {"state": "LOGGED_IN", "restore_file_bytes": None})
                    send_message(chat_id, "❌ بازیابی لغو شد و هیچ تغییری اعمال نشد.", get_super_admin_keyboard())
                    return
                if text != "✅ تأیید بازیابی افزایشی":
                    send_message(chat_id, "برای ادامه دکمه تأیید را بزنید یا انصراف دهید.", get_cancel_keyboard())
                    return
                if time.time() > user_state.get("restore_expires_at", 0):
                    user_states.update(chat_id, {"state": "LOGGED_IN", "restore_file_bytes": None})
                    send_message(chat_id, "⌛ مهلت تأیید پایان یافت؛ فایل را دوباره ارسال کنید.", get_super_admin_keyboard())
                    return
                success, msg, summary = restore_from_file(user_state.get("restore_file_bytes", b""), dry_run=False)
                user_states.update(chat_id, {"state": "LOGGED_IN", "restore_file_bytes": None})
                if success:
                    inserted = sum(summary.get('inserted', {}).values())
                    skipped = sum(summary.get('skipped', {}).values())
                    send_message(chat_id, f"✅ {msg}\nافزوده‌شده: {inserted}\nبدون تغییر/ردشده: {skipped}", get_super_admin_keyboard())
                    log_user_activity(user_db_id, "restore_additive", f"بازیابی افزایشی؛ {inserted} ردیف افزوده شد")
                else:
                    send_message(chat_id, f"❌ {msg}", get_super_admin_keyboard())
                return

        # ============================================================
        # ادامه منوی ادمین
        # ============================================================
        if role == 'admin' or is_super_admin:
            if text == "👁 مشاهده وضعیت همیار":
                branches = get_all_branches()
                if not branches:
                    send_message(chat_id, "❌ فهرست شعب در دسترس نیست.", get_keyboard(role, is_super_admin))
                    return
                user_states.update(chat_id, {
                    "state": "WAITING_FOR_ASSISTANT_BRANCH",
                    "assistant_branches": branches
                })
                send_message(
                    chat_id,
                    "🏢 شعبه موردنظر را برای مشاهده همان گزارش همیار معاون انتخاب کنید:",
                    get_assistant_branch_keyboard(branches)
                )
                return

            if text == "🎯 تحلیل مدیریتی":
                if is_holiday():
                    send_message(chat_id, "📅 امروز تعطیل است، گزارشی ثبت نشده است.", get_keyboard(role, is_super_admin))
                    return
                analysis = get_today_performance_analysis()
                if analysis:
                    msg = f"📈 **تحلیل مدیریتی امروز** - {get_shamsi_date_formatted(get_shamsi_date())}\n"
                    msg += f"━━━━━━━━━━━━━━━━━━\n\n"
                    msg += generate_management_analysis(analysis)
                    msg += f"\n\n💰 کل وصول: {analysis['today_total']//1_000_000:,.0f} میلیون ریال"
                    keyboard = get_keyboard(role, is_super_admin)
                    send_message(chat_id, msg, keyboard)
                else:
                    keyboard = get_keyboard(role, is_super_admin)
                    send_message(chat_id, "📊 داده‌های کافی برای تحلیل وجود ندارد.", keyboard)
                return

            if text == "📊 گزارش روند شعبه":
                user_states.update(chat_id, {"state": "WAITING_FOR_BRANCH_TREND"})
                send_message(chat_id, "🏢 لطفاً **نام شعبه** مورد نظر را برای مشاهده روند وارد کنید:", get_cancel_keyboard())
                return

            if text == "📋 عملکرد معاونان":
                deputies = get_all_deputies()
                if not deputies:
                    keyboard = get_keyboard(role, is_super_admin)
                    send_message(chat_id, "هیچ معاونی یافت نشد.", keyboard)
                    return
                msg = "📋 **گزارش عملکرد معاونان (۳۰ روز اخیر)**\n━━━━━━━━━━━━━━━━━━\n\n"
                for dep in deputies:
                    dep_id, dep_chat_id, dep_name, branch_id, branch_name = dep
                    perf = get_deputy_performance_report(dep_id, 30)
                    if perf:
                        msg += f"👤 {dep_name} - شعبه فعلی: {branch_name or 'بدون شعبه'}\n"
                        msg += f"   📅 ثبت به‌موقع: {perf['on_time']} روز\n"
                        msg += f"   📅 تاخیر: {perf['late']} روز\n"
                        msg += f"   💰 میانگین وصول: {perf['avg_amount']//1_000_000:,.0f} میلیون ریال\n"
                        msg += f"   🏆 بهترین روز: {perf['best_day']//1_000_000:,.0f} میلیون ریال\n"
                        match_data = get_deputy_match_report(dep_id, 30)
                        if match_data:
                            total_match = 0
                            count = 0
                            for row in match_data:
                                total_match += row[3]
                                count += 1
                            if count > 0:
                                avg_match = total_match / count
                                msg += f"   📊 میانگین انطباق با آمار واقعی: {avg_match:.1f}%\n"
                        msg += "\n"
                keyboard = get_keyboard(role, is_super_admin)
                send_message(chat_id, msg, keyboard)
                return

            if text == "👥 عملکرد همکاران":
                report = get_others_performance_summary()
                if report:
                    msg = f"📊 **عملکرد کلی همکاران (کل دوره)**\n━━━━━━━━━━━━━━━━━━\n\n"
                    total_others_all = 0
                    for idx, row in enumerate(report, 1):
                        branch_name = row[1]
                        total_others = int(row[2])
                        total_branch = int(row[3])
                        report_days = row[4]
                        msg += f"{idx}. 🏢 {branch_name}\n"
                        msg += f"   👥 کل وصولی همکاران: {total_others//1_000_000:,.0f} میلیون ریال\n"
                        msg += f"   📈 کل وصول شعبه: {total_branch//1_000_000:,.0f} میلیون ریال\n"
                        if total_branch > 0:
                            percent = (total_others / total_branch) * 100
                            msg += f"   📊 سهم همکاران: {percent:.1f}%\n"
                        else:
                            msg += f"   📊 سهم همکاران: ۰%\n"
                        msg += f"   📅 تعداد روزهای ثبت: {report_days}\n\n"
                        total_others_all += total_others
                    msg += f"━━━━━━━━━━━━━━━━━━\n"
                    msg += f"💰 کل وصولی همکاران استان: {total_others_all//1_000_000:,.0f} میلیون ریال"
                    keyboard = get_keyboard(role, is_super_admin)
                    send_message(chat_id, msg, keyboard)
                else:
                    keyboard = get_keyboard(role, is_super_admin)
                    send_message(chat_id, "📭 هیچ داده‌ای برای عملکرد همکاران یافت نشد.", keyboard)
                return

            if text == "📊 گزارش تطبیقی":
                if not get_adaptive_report_status():
                    send_message(chat_id, "🔴 گزارش تطبیقی در حال حاضر غیرفعال است.", get_keyboard(role, is_super_admin))
                    return
                if is_holiday():
                    send_message(chat_id, "📅 امروز تعطیل است، گزارشی ثبت نشده است.", get_keyboard(role, is_super_admin))
                    return
                comparison = get_adaptive_comparison()
                if comparison:
                    msg = f"📊 **گزارش تطبیقی** - {get_shamsi_date_formatted(get_shamsi_date())}\n"
                    msg += f"━━━━━━━━━━━━━━━━━━\n\n"
                    msg += f"💰 امروز: {comparison['today']//1_000_000:,.0f} میلیون ریال\n"
                    msg += f"📅 دیروز: {comparison['yesterday']//1_000_000:,.0f} میلیون ریال\n"
                    msg += f"📊 تغییر: {comparison['change_yesterday']:+.1f}%\n\n"
                    msg += f"📅 هفته قبل: {comparison['week_ago']//1_000_000:,.0f} میلیون ریال\n"
                    msg += f"📊 تغییر: {comparison['change_week']:+.1f}%\n\n"
                    msg += f"📅 ماه قبل: {comparison['month_ago']//1_000_000:,.0f} میلیون ریال\n"
                    msg += f"📊 تغییر: {comparison['change_month']:+.1f}%"
                    keyboard = get_keyboard(role, is_super_admin)
                    send_message(chat_id, msg, keyboard)
                else:
                    keyboard = get_keyboard(role, is_super_admin)
                    send_message(chat_id, "📊 داده‌های کافی برای گزارش تطبیقی وجود ندارد.", keyboard)
                return

            if text == "📈 پیش‌بینی عملکرد":
                if not get_forecast_report_status():
                    send_message(chat_id, "🔴 پیش‌بینی عملکرد در حال حاضر غیرفعال است.", get_keyboard(role, is_super_admin))
                    return
                if is_holiday():
                    send_message(chat_id, "📅 امروز تعطیل است، داده‌های کافی برای پیش‌بینی وجود ندارد.", get_keyboard(role, is_super_admin))
                    return
                send_message(chat_id, "⏳ در حال تحلیل داده‌ها و پیش‌بینی عملکرد هر شعبه... لطفاً چند لحظه صبر کنید.")
                executor.submit(generate_and_send_forecast, chat_id, role, is_super_admin)
                return

            if text == "📊 نمودار استان":
                if not get_chart_report_status():
                    send_message(chat_id, "🔴 گزارش‌های نموداری در حال حاضر غیرفعال است.", get_keyboard(role, is_super_admin))
                    return
                send_message(chat_id, "⏳ در حال تولید نمودار استان... لطفاً چند لحظه صبر کنید.")
                executor.submit(generate_and_send_province_chart, chat_id, role, is_super_admin)
                return

            if text == "📊 نمودار شعبه":
                if not get_chart_report_status():
                    send_message(chat_id, "🔴 گزارش‌های نموداری در حال حاضر غیرفعال است.", get_keyboard(role, is_super_admin))
                    return
                user_states.update(chat_id, {"state": "WAITING_FOR_BRANCH_CHART"})
                send_message(chat_id, "🏢 لطفاً **نام شعبه** مورد نظر را برای نمایش نمودار وارد کنید:", get_cancel_keyboard())
                return

            if text == "📊 نمودار تحلیلی":
                if not get_chart_report_status():
                    send_message(chat_id, "🔴 گزارش‌های نموداری در حال حاضر غیرفعال است.", get_keyboard(role, is_super_admin))
                    return
                keyboard = {
                    "keyboard": [
                        [{"text": "📊 مقایسه شعب برتر"}, {"text": "📊 نسبت معاون/همکار"}],
                        [{"text": "📈 روند روزانه"}, {"text": "📊 تحلیل انطباق"}],
                        [{"text": "🔙 انصراف"}]
                    ],
                    "resize_keyboard": True
                }
                user_states.update(chat_id, {"state": "WAITING_FOR_ANALYTICAL_CHART"})
                send_message(chat_id, "📊 **نمودارهای تحلیلی**\n\nلطفاً نوع نمودار مورد نظر را انتخاب کنید:", keyboard)
                return

            if text == "📊 مقایسه انطباق":
                if not get_actual_stats_status():
                    send_message(chat_id, "🔴 ثبت آمار واقعی در حال حاضر غیرفعال است.", get_keyboard(role, is_super_admin))
                    return
                user_states.update(chat_id, {"state": "WAITING_FOR_MATCH_DATE"})
                send_message(chat_id, "📅 لطفاً **تاریخ** مورد نظر برای گزارش مقایسه را به فرمت YYYY/MM/DD وارد کنید:", get_cancel_keyboard())
                return

            if text == "📝 مشاهده یادداشت‌ها":
                notes = get_all_notes_with_collection(30)
                if notes:
                    msg = "📝 **یادداشت‌های اخیر**\n━━━━━━━━━━━━━━━━━━\n"
                    for note in notes:
                        note_time = note[5]
                        msg += f"🏢 {note[1]} | 📅 {note[2]}\n"
                        msg += f"👤 {note[3]}: {note[4]}\n"
                        msg += f"⏰ {note_time.strftime('%H:%M') if hasattr(note_time, 'strftime') else note_time}\n\n"
                    keyboard = get_keyboard(role, is_super_admin)
                    send_message(chat_id, msg, keyboard)
                else:
                    keyboard = get_keyboard(role, is_super_admin)
                    send_message(chat_id, "هیچ یادداشتی وجود ندارد.", keyboard)
                return

            # ============================================================
            # گزارش امروز (اصلاح‌شده با نمایش هدف و N+1 رفع شده)
            # ============================================================
            if text == "📊 گزارش امروز":
                shamsi_today = get_shamsi_date()
                if is_holiday(shamsi_today):
                    send_message(chat_id, f"📅 امروز {get_shamsi_date_formatted(shamsi_today)} تعطیل است و گزارشی ثبت نشده است.", get_keyboard(role, is_super_admin))
                    return
                report = get_today_province_report(shamsi_today)
                stats = get_today_statistics()
                if report:
                    msg = f"📊 گزارش وصول امروز\n📅 تاریخ: {get_shamsi_date_formatted(shamsi_today)}\n━━━━━━━━━━━━━━━━━━\n\n"
                    total_province = 0
                    for idx, row in enumerate(report, 1):
                        branch_id = row[0]
                        branch_name = row[1]
                        dep = int(safe_format(row[2]))
                        oth = int(safe_format(row[3]))
                        tot = int(safe_format(row[4]))
                        target_amount = row[5]
                        target_date = row[6]
                        target_created_at = row[7]
                        collected_since_target = row[8]
                        
                        msg += f"{idx}. 🏢 {branch_name}\n"
                        msg += f"   👤 معاون: {dep//1_000_000:,.0f} میلیون ریال\n"
                        msg += f"   👥 همکاران: {oth//1_000_000:,.0f} میلیون ریال\n"
                        msg += f"   💰 جمع: {tot//1_000_000:,.0f} میلیون ریال\n"
                        
                        if target_amount is not None and target_date is not None:
                            progress_percent = (collected_since_target / target_amount * 100) if target_amount > 0 else 0
                            try:
                                target_date_obj = jdatetime.date(*map(int, target_date.split('/')))
                                today_obj = jdatetime.date(*map(int, shamsi_today.split('/')))
                                days_left = (target_date_obj - today_obj).days
                            except Exception:
                                days_left = 0
                            remaining = target_amount - collected_since_target
                            if remaining < 0:
                                remaining = 0
                            days_text = f"{days_left} روز" if days_left >= 0 else f"{abs(days_left)} روز گذشته"
                            msg += f"   🎯 هدف تا {get_shamsi_date_formatted(target_date)}: وصول {target_amount//1_000_000:,.0f} میلیون ریال\n"
                            msg += f"   📊 پیشرفت: {progress_percent:.1f}% ({collected_since_target//1_000_000:,.0f} از {target_amount//1_000_000:,.0f} میلیون ریال)\n"
                            msg += f"   📅 زمان باقیمانده تا پایان فرصت: {days_text}\n"
                            msg += f"   📉 فاصله از هدف: {remaining//1_000_000:,.0f} میلیون ریال\n"
                        else:
                            msg += f"   🎯 هدفی برای این شعبه تعریف نشده است\n"
                        msg += "\n"
                        total_province += tot
                    msg += f"━━━━━━━━━━━━━━━━━━\n"
                    if stats:
                        s0 = int(safe_format(stats[0]))
                        s1 = int(safe_format(stats[1]))
                        s2 = int(safe_format(stats[2]))
                        msg += f"📈 خلاصه:\n"
                        msg += f"   تعداد شعب ثبت شده: {s0}\n"
                        msg += f"   کل وصولی معاونین: {s1//1_000_000:,.0f} میلیون ریال\n"
                        msg += f"   کل وصولی همکاران: {s2//1_000_000:,.0f} میلیون ریال\n"
                        msg += f"   💰 جمع کل استان: {total_province//1_000_000:,.0f} میلیون ریال"
                    keyboard = get_keyboard(role, is_super_admin)
                    send_message(chat_id, msg, keyboard)
                else:
                    keyboard = get_keyboard(role, is_super_admin)
                    send_message(chat_id, f"📊 امروز ({shamsi_today}) هنوز هیچ شعبه‌ای اطلاعات ثبت نکرده است.", keyboard)
                return

            if text == "📈 گزارش ۱۰ روز اخیر":
                report = get_province_10_day_report()
                if report:
                    msg = f"📈 گزارش ۱۰ روز اخیر استان زنجان\n━━━━━━━━━━━━━━━━━━\n\n"
                    total_all = 0
                    for row in report:
                        r1 = int(safe_format(row[1]))
                        r2 = int(safe_format(row[2]))
                        r3 = int(safe_format(row[3]))
                        msg += f"📅 {get_shamsi_date_formatted(row[0])}\n"
                        msg += f"   👤 معاونین: {r1//1_000_000:,.0f} میلیون ریال\n"
                        msg += f"   👥 سایر همکاران: {r2//1_000_000:,.0f} میلیون ریال\n"
                        msg += f"   💰 جمع: {r3//1_000_000:,.0f} میلیون ریال\n\n"
                        total_all += r3
                    msg += f"━━━━━━━━━━━━━━━━━━\n"
                    msg += f"📊 کل ۱۰ روز: {total_all//1_000_000:,.0f} میلیون ریال"
                    keyboard = get_keyboard(role, is_super_admin)
                    send_message(chat_id, msg, keyboard)
                else:
                    keyboard = get_keyboard(role, is_super_admin)
                    send_message(chat_id, "📈 دیتابیس خالی است.", keyboard)
                return

            if text == "🏆 رتبه‌بندی شعب":
                report = get_top_5_branches()
                if report:
                    msg = f"🏆 ۵ شعبه برتر استان زنجان\n━━━━━━━━━━━━━━━━━━\n\n"
                    medals = ["🥇", "🥈", "🥉", "4️⃣", "5️⃣"]
                    for idx, row in enumerate(report):
                        tot = int(safe_format(row[1]))
                        cnt = int(safe_format(row[2]))
                        msg += f"{medals[idx]} {row[0]}\n"
                        msg += f"    💰 کل وصولی: {tot//1_000_000:,.0f} میلیون ریال\n"
                        msg += f"    📊 تعداد ثبت: {cnt} روز\n\n"
                    keyboard = get_keyboard(role, is_super_admin)
                    send_message(chat_id, msg, keyboard)
                else:
                    keyboard = get_keyboard(role, is_super_admin)
                    send_message(chat_id, "🏆 داده کافی برای رتبه‌بندی وجود ندارد.", keyboard)
                return

            if text == "💹 آمار مفصل امروز":
                shamsi_today = get_shamsi_date()
                if is_holiday(shamsi_today):
                    send_message(chat_id, f"📅 امروز {get_shamsi_date_formatted(shamsi_today)} تعطیل است، گزارشی ثبت نشده است.", get_keyboard(role, is_super_admin))
                    return
                report = get_detailed_report(shamsi_today)
                if report:
                    msg = f"💹 آمار مفصل امروز\n━━━━━━━━━━━━━━━━━━\n\n"
                    for idx, row in enumerate(report, 1):
                        dep = int(safe_format(row[1]))
                        oth = int(safe_format(row[2]))
                        tot = int(safe_format(row[3]))
                        msg += f"{idx}. 🏢 {row[0]}\n"
                        msg += f"   👤 معاون ({row[4]}): {dep//1_000_000:,.0f} میلیون ریال\n"
                        msg += f"   👥 سایرین: {oth//1_000_000:,.0f} میلیون ریال\n"
                        msg += f"   💰 جمع: {tot//1_000_000:,.0f} میلیون ریال\n\n"
                    keyboard = get_keyboard(role, is_super_admin)
                    send_message(chat_id, msg, keyboard)
                else:
                    keyboard = get_keyboard(role, is_super_admin)
                    send_message(chat_id, "💹 برای امروز اطلاعاتی وجود ندارد.", keyboard)
                return

            if text == "📉 مقایسه روزانه":
                comparison = get_daily_comparison()
                if comparison:
                    msg = f"📉 مقایسه روزانه (۷ روز اخیر)\n━━━━━━━━━━━━━━━━━━\n\n"
                    for row in comparison:
                        br = int(safe_format(row[1]))
                        tot = int(safe_format(row[2]))
                        msg += f"📅 {get_shamsi_date_formatted(row[0])}\n"
                        msg += f"    🏢 شعب ثبت‌کننده: {br}\n"
                        msg += f"    💰 کل وصولی: {tot//1_000_000:,.0f} میلیون ریال\n\n"
                    keyboard = get_keyboard(role, is_super_admin)
                    send_message(chat_id, msg, keyboard)
                else:
                    keyboard = get_keyboard(role, is_super_admin)
                    send_message(chat_id, "📉 داده کافی وجود ندارد.", keyboard)
                return

            if text == "📅 گزارش تاریخ خاص":
                user_states.update(chat_id, {"state": "WAITING_FOR_ADMIN_DATE"})
                send_message(chat_id, "📅 لطفاً تاریخ مورد نظر را به فرمت **YYYY/MM/DD** وارد کنید (مثلاً ۱۴۰۳/۰۱/۱۵):", get_cancel_keyboard())
                return

            if text == "📊 بهترین/بدترین روز":
                best, worst = get_best_worst_days(5)
                msg = "📊 **بهترین روزهای استان**\n━━━━━━━━━━━━━━━━━━\n"
                if best:
                    for i, row in enumerate(best, 1):
                        msg += f"{i}. 📅 {get_shamsi_date_formatted(row[0])} -> {int(row[1])//1_000_000:,.0f} میلیون ریال\n"
                else:
                    msg += "هیچ داده‌ای موجود نیست.\n"
                msg += "\n📊 **بدترین روزهای استان**\n━━━━━━━━━━━━━━━━━━\n"
                if worst:
                    for i, row in enumerate(worst, 1):
                        msg += f"{i}. 📅 {get_shamsi_date_formatted(row[0])} -> {int(row[1])//1_000_000:,.0f} میلیون ریال\n"
                else:
                    msg += "هیچ داده‌ای موجود نیست."
                keyboard = get_keyboard(role, is_super_admin)
                send_message(chat_id, msg, keyboard)
                return

            # ===== Stateهای ادمین =====
            if current_state == "WAITING_FOR_BRANCH_TREND" and (role == 'admin' or is_super_admin):
                if text == "🔙 انصراف":
                    user_states.update(chat_id, {"state": "LOGGED_IN"})
                    keyboard = get_keyboard(role, is_super_admin)
                    send_message(chat_id, "❌ عملیات لغو شد.", keyboard)
                    return
                conn = None
                try:
                    conn = get_db_connection()
                    with conn.cursor() as cur:
                        cur.execute("SELECT id FROM branches WHERE name ILIKE %s ESCAPE '\\' LIMIT 1", (f"%{escape_like_pattern(text)}%",))
                        result = cur.fetchone()
                        if result:
                            branch_id = result[0]
                            trend = get_branch_trend(branch_id, 5)
                            if trend:
                                msg = f"📊 **روند ۵ روز اخیر شعبه {text}**\n━━━━━━━━━━━━━━━━━━\n"
                                for i in range(len(trend)):
                                    date, amount = trend[i]
                                    if i == 0:
                                        trend_symbol = "📊"
                                    else:
                                        prev_amount = trend[i-1][1]
                                        if amount > prev_amount:
                                            trend_symbol = "📈"
                                        elif amount < prev_amount:
                                            trend_symbol = "📉"
                                        else:
                                            trend_symbol = "➡️"
                                    msg += f"{trend_symbol} 📅 {get_shamsi_date_formatted(date)}: {amount//1_000_000:,.0f} میلیون ریال\n"
                                keyboard = get_keyboard(role, is_super_admin)
                                send_message(chat_id, msg, keyboard)
                            else:
                                keyboard = get_keyboard(role, is_super_admin)
                                send_message(chat_id, f"📭 هیچ داده‌ای برای شعبه {text} یافت نشد.", keyboard)
                        else:
                            send_message(chat_id, f"❌ شعبه‌ای با نام {text} یافت نشد. لطفاً نام دقیق شعبه را وارد کنید.", get_cancel_keyboard())
                            return
                except Exception as e:
                    logger.exception("Accuracy trend failed")
                    send_message(chat_id, "❌ عملیات انجام نشد؛ جزئیات خطا ثبت شد.", get_cancel_keyboard())
                finally:
                    if conn:
                        return_db_connection(conn)
                user_states.update(chat_id, {"state": "LOGGED_IN"})
                return

            if current_state == "WAITING_FOR_BRANCH_CHART" and (role == 'admin' or is_super_admin):
                if text == "🔙 انصراف":
                    user_states.update(chat_id, {"state": "LOGGED_IN"})
                    keyboard = get_keyboard(role, is_super_admin)
                    send_message(chat_id, "❌ عملیات لغو شد.", keyboard)
                    return
                conn = None
                try:
                    conn = get_db_connection()
                    with conn.cursor() as cur:
                        cur.execute("SELECT id FROM branches WHERE name ILIKE %s ESCAPE '\\' LIMIT 1", (f"%{escape_like_pattern(text)}%",))
                        result = cur.fetchone()
                        if result:
                            branch_id = result[0]
                            send_message(chat_id, "⏳ در حال تولید نمودار... لطفاً چند لحظه صبر کنید.")
                            executor.submit(generate_and_send_branch_chart, chat_id, branch_id, text, role, is_super_admin)
                            user_states.update(chat_id, {"state": "LOGGED_IN"})
                            return
                        else:
                            send_message(chat_id, f"❌ شعبه‌ای با نام {text} یافت نشد.", get_cancel_keyboard())
                            return
                except Exception as e:
                    logger.exception("Accuracy comparison failed")
                    send_message(chat_id, "❌ عملیات انجام نشد؛ جزئیات خطا ثبت شد.", get_cancel_keyboard())
                finally:
                    if conn:
                        return_db_connection(conn)
                user_states.update(chat_id, {"state": "LOGGED_IN"})
                return

            if current_state == "WAITING_FOR_ANALYTICAL_CHART" and (role == 'admin' or is_super_admin):
                if text == "🔙 انصراف":
                    user_states.update(chat_id, {"state": "LOGGED_IN"})
                    keyboard = get_keyboard(role, is_super_admin)
                    send_message(chat_id, "❌ عملیات لغو شد.", keyboard)
                    return
                chart_type_map = {
                    "📊 مقایسه شعب برتر": "branch_comparison",
                    "📊 نسبت معاون/همکار": "deputy_others_ratio",
                    "📈 روند روزانه": "daily_trend",
                    "📊 تحلیل انطباق": "match_analysis"
                }
                chart_key = chart_type_map.get(text)
                if chart_key:
                    send_message(chat_id, "⏳ در حال تولید نمودار تحلیلی... لطفاً چند لحظه صبر کنید.")
                    executor.submit(generate_and_send_analytical_chart, chat_id, chart_key, role, is_super_admin)
                else:
                    send_message(chat_id, "❌ گزینه نامعتبر.", get_cancel_keyboard())
                user_states.update(chat_id, {"state": "LOGGED_IN"})
                return

            if current_state == "WAITING_FOR_MATCH_DATE" and (role == 'admin' or is_super_admin):
                if text == "🔙 انصراف":
                    user_states.update(chat_id, {"state": "LOGGED_IN"})
                    keyboard = get_keyboard(role, is_super_admin)
                    send_message(chat_id, "❌ عملیات لغو شد.", keyboard)
                    return
                shamsi_date = normalize_digits(text)
                if validate_shamsi_date(shamsi_date):
                    actual_data = get_actual_stats_for_date(shamsi_date)
                    if not actual_data:
                        keyboard = get_keyboard(role, is_super_admin)
                        send_message(chat_id, f"📭 هیچ آمار واقعی برای تاریخ {get_shamsi_date_formatted(shamsi_date)} ثبت نشده است.", keyboard)
                        user_states.update(chat_id, {"state": "LOGGED_IN"})
                        return

                    comparison_data = []
                    for item in actual_data:
                        branch_id, branch_name, total_act_rial = item
                        comp = compare_collection_with_actual(branch_id, shamsi_date)
                        if comp:
                            comparison_data.append({
                                'branch_name': branch_name,
                                'claimed': comp['claimed'],
                                'actual': comp['actual'],
                                'abs_actual': comp['abs_actual'],
                                'diff_abs': comp['diff_abs'],
                                'is_claimed_more': comp['is_claimed_more']
                            })

                    if not comparison_data:
                        keyboard = get_keyboard(role, is_super_admin)
                        send_message(chat_id, f"📭 هیچ وصولی برای تاریخ {get_shamsi_date_formatted(shamsi_date)} ثبت نشده است.", keyboard)
                        user_states.update(chat_id, {"state": "LOGGED_IN"})
                        return

                    msg = f"📊 **گزارش مقایسه خوداظهاری معاونین شعب - {get_shamsi_date_formatted(shamsi_date)}**\n"
                    msg += f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"

                    negative_actual_branches = []
                    positive_actual_branches = []
                    negative_accuracies = []

                    for item in comparison_data:
                        branch_name = item['branch_name']
                        claimed = item['claimed']
                        actual = item['actual']
                        abs_actual = item['abs_actual']
                        diff_abs = item['diff_abs']
                        is_claimed_more = item['is_claimed_more']

                        if actual < 0:
                            if claimed == 0 and abs_actual == 0:
                                accuracy = 100.0
                            elif claimed == 0 and abs_actual > 0:
                                accuracy = 0.0
                            else:
                                accuracy = (1 - (diff_abs / max(claimed, abs_actual))) * 100
                        else:
                            if claimed == 0 and actual == 0:
                                accuracy = 100.0
                            elif claimed == 0 or actual == 0:
                                accuracy = 0.0
                            else:
                                accuracy = (min(claimed, actual) / max(claimed, actual)) * 100

                        if accuracy >= 95:
                            emoji = "✅"
                        elif accuracy >= 80:
                            emoji = "🟢"
                        elif accuracy >= 50:
                            emoji = "🟡"
                        else:
                            emoji = "🔴"

                        diff_text = "بیشتر" if is_claimed_more else "کمتر" if is_claimed_more is not None else ""

                        msg += f"🏢 **{branch_name}**\n"
                        msg += f"📝 ادعای وصول معاون: {claimed//1_000_000:,.0f} میلیون ریال\n"
                        if actual < 0:
                            msg += f"📉 کاهش واقعی مطالبات: {abs_actual//1_000_000:,.0f} میلیون ریال\n"
                        else:
                            msg += f"📈 افزایش واقعی مطالبات: {actual//1_000_000:,.0f} میلیون ریال\n"
                        if actual < 0:
                            msg += f"↕️ اختلاف: {diff_abs//1_000_000:,.0f} میلیون ریال (ادعا {diff_text} از واقعیت بوده)\n"
                        msg += f"🎯 دقت خوداظهاری: {accuracy:.1f}% {emoji}\n\n"

                        if actual < 0:
                            negative_actual_branches.append(branch_name)
                            negative_accuracies.append(accuracy)
                        else:
                            positive_actual_branches.append(branch_name)

                    msg += f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                    msg += f"📊 **خلاصه کلی**\n"
                    msg += f"🏢 تعداد شعب: {len(comparison_data)}\n"

                    if negative_actual_branches:
                        avg_neg_accuracy = sum(negative_accuracies) / len(negative_accuracies)
                        msg += f"📉 شعب با کاهش واقعی مطالبات: {len(negative_actual_branches)} شعبه — میانگین دقت خوداظهاری: {avg_neg_accuracy:.1f}%\n"
                    else:
                        msg += f"📉 شعب با کاهش واقعی مطالبات: ۰ شعبه\n"

                    if positive_actual_branches:
                        msg += f"📈 شعب با افزایش واقعی مطالبات: {len(positive_actual_branches)} شعبه\n"
                    else:
                        msg += f"📈 شعب با افزایش واقعی مطالبات: ۰ شعبه\n"

                    keyboard = get_keyboard(role, is_super_admin)
                    send_message(chat_id, msg, keyboard)
                    user_states.update(chat_id, {"state": "LOGGED_IN"})
                else:
                    send_message(chat_id, "❌ فرمت تاریخ نامعتبر. لطفاً به صورت YYYY/MM/DD وارد کنید.")
                return

            # ===== Stateهای سوپرادمین =====
            if current_state == "WAITING_FOR_ACTUAL_METHOD" and is_super_admin:
                if text == "🔙 انصراف":
                    user_states.update(chat_id, {"state": "LOGGED_IN"})
                    send_message(chat_id, "❌ عملیات لغو شد.", get_super_admin_keyboard())
                else:
                    send_message(chat_id, "لطفاً یکی از دو روش ثبت را انتخاب کنید.", get_actual_stats_method_keyboard())
                return

            if current_state == "WAITING_FOR_ACTUAL_EXCEL_FILE" and is_super_admin:
                if text == "🔙 انصراف":
                    user_states.update(chat_id, {"state": "LOGGED_IN"})
                    send_message(chat_id, "❌ عملیات لغو شد.", get_super_admin_keyboard())
                    return
                document = message.get('document')
                if not document:
                    send_message(chat_id, "❌ لطفاً فایل Excel با پسوند .xlsx ارسال کنید.", get_cancel_keyboard())
                    return
                raw_file_name = str(document.get('file_name') or 'report.xlsx')
                file_name = re.sub(r'[\x00-\x1f\x7f]+', '', os.path.basename(raw_file_name))[:255]
                if not file_name.lower().endswith('.xlsx'):
                    send_message(chat_id, "❌ فقط فایل اصلی Excel با پسوند .xlsx پذیرفته می‌شود.", get_cancel_keyboard())
                    return
                file_id = document.get('file_id')
                if not file_id:
                    send_message(chat_id, "❌ شناسه فایل دریافت نشد؛ فایل را دوباره ارسال کنید.", get_cancel_keyboard())
                    return

                send_message(chat_id, "⏳ فایل در حال خواندن و کنترل است؛ لطفاً کمی صبر کنید.", get_cancel_keyboard())
                temp_path = None
                try:
                    downloaded, download_message, temp_path, file_sha256 = download_bale_excel_to_temp(file_id)
                    if not downloaded:
                        send_message(chat_id, f"❌ {download_message}", get_cancel_keyboard())
                        return
                    branches = get_branches_with_official_codes()
                    parsed, parse_message, result = parse_actual_stats_excel(temp_path, branches)
                    if not parsed:
                        send_message(chat_id, f"❌ {parse_message}\n\nهیچ داده‌ای ثبت نشد.", get_cancel_keyboard())
                        return
                    try:
                        os.unlink(temp_path)
                        temp_path = None
                    except OSError:
                        logger.exception("Excel was parsed but immediate temp deletion failed")
                        send_message(
                            chat_id,
                            "❌ حذف فایل موقت تأیید نشد؛ برای جلوگیری از اشغال حافظه، عملیات ثبت متوقف شد.",
                            get_cancel_keyboard(),
                        )
                        return
                    draft = result
                    draft.update({
                        'file_name': file_name,
                        'file_sha256': file_sha256,
                        'operation_uuid': str(uuid.uuid4()),
                        'expires_at': time_module.time() + ACTUAL_EXCEL_CONFIRM_TTL_SECONDS,
                    })
                    if not draft.get('report_date'):
                        user_states.update(chat_id, {
                            "state": "WAITING_FOR_ACTUAL_EXCEL_DATE",
                            "actual_excel_draft": draft,
                        })
                        send_message(
                            chat_id,
                            "✅ اطلاعات شعب استخراج شد، اما تاریخ گزارش با اطمینان تشخیص داده نشد.\n"
                            "لطفاً تاریخ را به فرمت YYYY/MM/DD وارد کنید:",
                            get_cancel_keyboard(),
                        )
                        return
                    user_states.update(chat_id, {
                        "state": "WAITING_FOR_ACTUAL_EXCEL_CONFIRM",
                        "actual_excel_draft": draft,
                    })
                    send_message(chat_id, format_actual_excel_preview(draft), get_actual_excel_confirm_keyboard())
                finally:
                    # فایل در هیچ حالتی پس از پردازش روی سرور باقی نمی‌ماند.
                    if temp_path and os.path.isfile(temp_path):
                        try:
                            os.unlink(temp_path)
                        except OSError:
                            logger.exception("Could not delete processed Excel temp file")
                return

            if current_state == "WAITING_FOR_ACTUAL_EXCEL_DATE" and is_super_admin:
                if text == "🔙 انصراف":
                    user_states.update(chat_id, {"state": "LOGGED_IN", "actual_excel_draft": None})
                    send_message(chat_id, "❌ عملیات لغو شد.", get_super_admin_keyboard())
                    return
                shamsi_date = normalize_digits(text)
                if not validate_shamsi_date(shamsi_date):
                    send_message(chat_id, "❌ تاریخ نامعتبر است؛ به فرمت YYYY/MM/DD وارد کنید.", get_cancel_keyboard())
                    return
                draft = user_state.get('actual_excel_draft')
                if not isinstance(draft, dict):
                    user_states.update(chat_id, {"state": "LOGGED_IN"})
                    send_message(chat_id, "❌ پیش‌نمایش منقضی شده است؛ فایل را دوباره ارسال کنید.", get_super_admin_keyboard())
                    return
                draft['report_date'] = shamsi_date
                draft['expires_at'] = time_module.time() + ACTUAL_EXCEL_CONFIRM_TTL_SECONDS
                user_states.update(chat_id, {
                    "state": "WAITING_FOR_ACTUAL_EXCEL_CONFIRM",
                    "actual_excel_draft": draft,
                })
                send_message(chat_id, format_actual_excel_preview(draft), get_actual_excel_confirm_keyboard())
                return

            if current_state == "WAITING_FOR_ACTUAL_EXCEL_CONFIRM" and is_super_admin:
                draft = user_state.get('actual_excel_draft')
                if not isinstance(draft, dict) or time_module.time() > draft.get('expires_at', 0):
                    user_states.update(chat_id, {"state": "LOGGED_IN", "actual_excel_draft": None})
                    send_message(chat_id, "⌛ پیش‌نمایش منقضی شد؛ برای امنیت، فایل را دوباره ارسال کنید.", get_super_admin_keyboard())
                    return
                if text == "🔙 انصراف":
                    user_states.update(chat_id, {"state": "LOGGED_IN", "actual_excel_draft": None})
                    send_message(chat_id, "❌ عملیات لغو شد و هیچ داده‌ای ثبت نشد.", get_super_admin_keyboard())
                    return
                if text == "🔄 ارسال فایل جدید":
                    user_states.update(chat_id, {"state": "WAITING_FOR_ACTUAL_EXCEL_FILE", "actual_excel_draft": None})
                    send_message(chat_id, "📎 فایل جدید .xlsx را ارسال کنید.", get_cancel_keyboard())
                    return
                if text == "📅 اصلاح تاریخ":
                    user_states.update(chat_id, {
                        "state": "WAITING_FOR_ACTUAL_EXCEL_DATE",
                        "actual_excel_draft": draft,
                    })
                    send_message(chat_id, "📅 تاریخ صحیح را به فرمت YYYY/MM/DD وارد کنید:", get_cancel_keyboard())
                    return
                if text == "✏️ اصلاح مبلغ یک شعبه":
                    rows = draft.get('rows', [])
                    branch_list = '\n'.join(
                        f"{index}. {row['branch_name']} ({row['official_code']})"
                        for index, row in enumerate(rows, 1)
                    )
                    user_states.update(chat_id, {
                        "state": "WAITING_FOR_ACTUAL_EXCEL_EDIT_BRANCH",
                        "actual_excel_draft": draft,
                    })
                    send_message(
                        chat_id,
                        "✏️ شماره شعبه‌ای را که باید اصلاح شود وارد کنید:\n\n" + branch_list,
                        get_cancel_keyboard(),
                    )
                    return
                if text == "✅ تأیید نهایی و ثبت":
                    user_states.update(chat_id, {
                        "state": "ACTUAL_EXCEL_COMMITTING",
                        "actual_excel_draft": draft,
                    })
                    success, save_message, receipt_uuid = save_actual_stats_batch(draft, user_db_id)
                    if success:
                        log_user_activity(
                            user_db_id,
                            "import_actual_stats_excel",
                            f"ثبت Excel آمار واقعی تاریخ {draft['report_date']}، "
                            f"{len(draft.get('rows', []))} شعبه، رسید {receipt_uuid}",
                        )
                        user_states.update(chat_id, {"state": "LOGGED_IN", "actual_excel_draft": None})
                        send_message(
                            chat_id,
                            "✅ **ثبت نهایی با موفقیت انجام شد.**\n\n"
                            f"📅 تاریخ: {draft['report_date']}\n"
                            f"🏢 تعداد شعب: {len(draft.get('rows', []))}\n"
                            f"🧾 شماره رسید: `{receipt_uuid}`\n\n"
                            "فایل Excel پیش‌تر از حافظه موقت سرور حذف شده است.",
                            get_super_admin_keyboard(),
                        )
                    else:
                        user_states.update(chat_id, {
                            "state": "WAITING_FOR_ACTUAL_EXCEL_CONFIRM",
                            "actual_excel_draft": draft,
                        })
                        send_message(chat_id, f"❌ {save_message}", get_actual_excel_confirm_keyboard())
                    return
                send_message(chat_id, "لطفاً یکی از گزینه‌های پیش‌نمایش را انتخاب کنید.", get_actual_excel_confirm_keyboard())
                return

            if current_state == "WAITING_FOR_ACTUAL_EXCEL_EDIT_BRANCH" and is_super_admin:
                if text == "🔙 انصراف":
                    user_states.update(chat_id, {"state": "LOGGED_IN", "actual_excel_draft": None})
                    send_message(chat_id, "❌ عملیات لغو شد و هیچ داده‌ای ثبت نشد.", get_super_admin_keyboard())
                    return
                draft = user_state.get('actual_excel_draft')
                selection = parse_number(text)
                rows = draft.get('rows', []) if isinstance(draft, dict) else []
                if selection is None or selection < 1 or selection > len(rows):
                    send_message(chat_id, "❌ شماره شعبه معتبر نیست.", get_cancel_keyboard())
                    return
                edit_index = selection - 1
                user_states.update(chat_id, {
                    "state": "WAITING_FOR_ACTUAL_EXCEL_EDIT_VALUE",
                    "actual_excel_draft": draft,
                    "actual_excel_edit_index": edit_index,
                })
                send_message(
                    chat_id,
                    f"مبلغ صحیح شعبه **{rows[edit_index]['branch_name']}** را به میلیون ریال وارد کنید.\n"
                    "عدد منفی یا مثبت پذیرفته می‌شود؛ مثال: `4700-`",
                    get_cancel_keyboard(),
                )
                return

            if current_state == "WAITING_FOR_ACTUAL_EXCEL_EDIT_VALUE" and is_super_admin:
                if text == "🔙 انصراف":
                    user_states.update(chat_id, {"state": "LOGGED_IN", "actual_excel_draft": None})
                    send_message(chat_id, "❌ عملیات لغو شد و هیچ داده‌ای ثبت نشد.", get_super_admin_keyboard())
                    return
                draft = user_state.get('actual_excel_draft')
                edit_index = user_state.get('actual_excel_edit_index')
                corrected_value = parse_number(text)
                rows = draft.get('rows', []) if isinstance(draft, dict) else []
                if (corrected_value is None or abs(corrected_value) > MAX_AMOUNT_MILLIONS or
                        not isinstance(edit_index, int) or edit_index < 0 or edit_index >= len(rows)):
                    send_message(chat_id, "❌ مبلغ معتبر نیست؛ یک عدد صحیح وارد کنید.", get_cancel_keyboard())
                    return
                if rows[edit_index].get('was_blank'):
                    draft['blank_count'] = max(0, draft.get('blank_count', 0) - 1)
                rows[edit_index]['amount_millions'] = corrected_value
                rows[edit_index]['was_blank'] = False
                rows[edit_index]['corrected'] = True
                draft['expires_at'] = time_module.time() + ACTUAL_EXCEL_CONFIRM_TTL_SECONDS
                user_states.update(chat_id, {
                    "state": "WAITING_FOR_ACTUAL_EXCEL_CONFIRM",
                    "actual_excel_draft": draft,
                    "actual_excel_edit_index": None,
                })
                send_message(chat_id, format_actual_excel_preview(draft), get_actual_excel_confirm_keyboard())
                return

            if current_state == "ACTUAL_EXCEL_COMMITTING" and is_super_admin:
                send_message(chat_id, "⏳ ثبت نهایی در حال انجام است؛ لطفاً منتظر بمانید.")
                return

            if current_state == "WAITING_FOR_ACTUAL_DATE" and is_super_admin:
                if text == "🔙 انصراف":
                    user_states.update(chat_id, {"state": "LOGGED_IN"})
                    send_message(chat_id, "❌ عملیات لغو شد.", get_super_admin_keyboard())
                    return
                shamsi_date = normalize_digits(text)
                if validate_shamsi_date(shamsi_date):
                    user_states.update(chat_id, {"actual_date": shamsi_date, "state": "WAITING_FOR_ACTUAL_BRANCH"})
                    branches = get_all_branches()
                    if not branches:
                        send_message(chat_id, "❌ هیچ شعبه‌ای یافت نشد.", get_super_admin_keyboard())
                        return
                    user_states.update(chat_id, {"actual_branches": branches, "actual_branch_index": 0})
                    branch = branches[0]
                    msg = f"📊 **ثبت آمار واقعی برای تاریخ {get_shamsi_date_formatted(shamsi_date)}**\n"
                    msg += f"━━━━━━━━━━━━━━━━━━\n"
                    msg += f"🏢 شعبه: {branch[1]}\n\n"
                    msg += "📝 لطفاً **کل مبلغ وصول واقعی** را به **میلیون ریال** وارد کنید.\n"
                    msg += "(برای کاهش از علامت منفی استفاده کنید، برای افزایش مثبت)\n"
                    msg += "مثال: 4700- برای کاهش ۴.۷ میلیاردی"
                    send_message(chat_id, msg, get_cancel_keyboard())
                else:
                    send_message(chat_id, "❌ فرمت تاریخ نامعتبر. لطفاً به صورت YYYY/MM/DD وارد کنید.")
                return

            if current_state == "WAITING_FOR_ACTUAL_BRANCH" and is_super_admin:
                if text == "🔙 انصراف":
                    user_states.update(chat_id, {"state": "LOGGED_IN"})
                    send_message(chat_id, "❌ عملیات لغو شد.", get_super_admin_keyboard())
                    return
                try:
                    total_value = parse_number(text)
                    if total_value is None:
                        raise ValueError
                    shamsi_date = user_state.get("actual_date")
                    branches = user_state.get("actual_branches", [])
                    index = user_state.get("actual_branch_index", 0)
                    if index < len(branches):
                        branch_id = branches[index][0]
                        success, message = save_actual_stats(branch_id, shamsi_date, total_value, user_db_id)
                        if success:
                            log_user_activity(user_db_id, "add_actual_stats", f"ثبت آمار واقعی برای شعبه {branches[index][1]} تاریخ {shamsi_date}: {total_value} میلیون ریال")
                        else:
                            send_message(chat_id, f"❌ خطا در ثبت آمار واقعی: {message}", get_cancel_keyboard())
                            return
                        index += 1
                        if index < len(branches):
                            user_states.update(chat_id, {"actual_branch_index": index})
                            branch = branches[index]
                            msg = f"📊 **ثبت آمار واقعی برای تاریخ {get_shamsi_date_formatted(shamsi_date)}**\n"
                            msg += f"━━━━━━━━━━━━━━━━━━\n"
                            msg += f"🏢 شعبه: {branch[1]}\n\n"
                            msg += "📝 لطفاً **کل مبلغ وصول واقعی** را به میلیون ریال وارد کنید."
                            send_message(chat_id, msg, get_cancel_keyboard())
                        else:
                            send_message(chat_id, "✅ ثبت آمار واقعی برای همه شعب با موفقیت انجام شد.", get_super_admin_keyboard())
                            user_states.update(chat_id, {"state": "LOGGED_IN"})
                    else:
                        send_message(chat_id, "✅ ثبت آمار واقعی کامل شد.", get_super_admin_keyboard())
                        user_states.update(chat_id, {"state": "LOGGED_IN"})
                except ValueError:
                    send_message(chat_id, "❌ لطفاً یک عدد معتبر وارد کنید.")
                return

            if current_state == "WAITING_FOR_ADD_HOLIDAY" and is_super_admin:
                if text == "🔙 انصراف":
                    user_states.update(chat_id, {"state": "LOGGED_IN"})
                    send_message(chat_id, "❌ عملیات لغو شد.", get_super_admin_keyboard())
                    return
                parts = text.split('|', 1)
                shamsi_date = normalize_digits(parts[0].strip())
                description = parts[1].strip() if len(parts) > 1 else "تعطیل"
                if validate_shamsi_date(shamsi_date):
                    success, msg = add_holiday(shamsi_date, description)
                    if success:
                        send_message(chat_id, f"✅ روز {get_shamsi_date_formatted(shamsi_date)} با موفقیت به عنوان تعطیل ثبت شد.\nتوضیح: {description}", get_super_admin_keyboard())
                        log_user_activity(user_db_id, "add_holiday", f"افزودن تعطیل: {shamsi_date} - {description}")
                    else:
                        send_message(chat_id, f"❌ {msg}", get_cancel_keyboard())
                        return
                else:
                    send_message(chat_id, "❌ فرمت تاریخ نامعتبر. لطفاً به صورت YYYY/MM/DD وارد کنید.", get_cancel_keyboard())
                    return
                user_states.update(chat_id, {"state": "LOGGED_IN"})
                return

            if current_state == "WAITING_FOR_REMOVE_HOLIDAY" and is_super_admin:
                if text == "🔙 انصراف":
                    user_states.update(chat_id, {"state": "LOGGED_IN"})
                    send_message(chat_id, "❌ عملیات لغو شد.", get_super_admin_keyboard())
                    return
                try:
                    index = int(text) - 1
                    holidays = user_state.get("holidays_list", [])
                    if 0 <= index < len(holidays):
                        shamsi_date = holidays[index][0]
                        if remove_holiday(shamsi_date):
                            send_message(chat_id, f"✅ روز {get_shamsi_date_formatted(shamsi_date)} از تعطیلات حذف شد.", get_super_admin_keyboard())
                            log_user_activity(user_db_id, "remove_holiday", f"حذف تعطیل: {shamsi_date}")
                        else:
                            send_message(chat_id, "❌ خطا در حذف تعطیل.", get_super_admin_keyboard())
                    else:
                        send_message(chat_id, "❌ شماره نامعتبر.", get_cancel_keyboard())
                        user_states.update(chat_id, {"state": "LOGGED_IN"})
                        return
                except Exception:
                    send_message(chat_id, "❌ لطفاً یک عدد معتبر وارد کنید.", get_cancel_keyboard())
                    user_states.update(chat_id, {"state": "LOGGED_IN"})
                    return
                user_states.update(chat_id, {"state": "LOGGED_IN"})
                return

            if current_state == "WAITING_FOR_RESET_CONFIRM" and is_super_admin:
                if text == "✅ بله، ریست کن":
                    if reset_all_collections():
                        send_message(chat_id, "✅ حافظه موقت گزارش‌ها پاک شد؛ همه اطلاعات ثبت‌شده محفوظ هستند.", get_super_admin_keyboard())
                        log_user_activity(user_db_id, "refresh_report_cache", "تازه‌سازی امن حافظه گزارش‌ها")
                    else:
                        send_message(chat_id, "❌ خطا در ریست گزارش‌ها.", get_super_admin_keyboard())
                else:
                    send_message(chat_id, "❌ عملیات ریست لغو شد.", get_super_admin_keyboard())
                user_states.update(chat_id, {"state": "LOGGED_IN"})
                return

            # ===== منوی عمومی ادمین =====
            keyboard = get_keyboard(role, is_super_admin)
            send_message(chat_id, "لطفاً یک گزینه از منو انتخاب کنید:", keyboard)
            return

        # ============================================================
        # منوی معاون
        # ============================================================
        if role == 'deputy':
            if text == "🤖 همیار وصول مطالبات":
                if not branch_id:
                    send_message(chat_id, "❌ برای حساب شما شعبه‌ای تعیین نشده است.", get_deputy_keyboard())
                    return
                send_message(chat_id, "⏳ در حال تهیه گزارش اختصاصی همیار وصول مطالبات...", get_deputy_keyboard())
                executor.submit(send_collection_assistant_report, chat_id, branch_id,
                                get_deputy_keyboard(), user_db_id)
                return

            if text == "💰 ثبت وصولی روزانه":
                shamsi_today = get_shamsi_date()
                if is_holiday(shamsi_today):
                    send_message(chat_id, f"📅 امروز {get_shamsi_date_formatted(shamsi_today)} تعطیل است، نیازی به ثبت وصول نیست.", get_deputy_keyboard())
                    return
                existing = check_existing_collection(branch_id, shamsi_today)
                if existing:
                    if can_edit_collection(existing['created_at']):
                        user_states.update(chat_id, {"state": "WAITING_FOR_EDIT_CONFIRMATION"})
                        confirm_keyboard = {
                            "keyboard": [[{"text": "📝 بله، ویرایش شود"}, {"text": "❌ خیر، لغو شود"}]],
                            "resize_keyboard": True
                        }
                        msg = (
                            f"⚠️ اطلاعات امروز قبلاً ثبت شده است.\n\n"
                            f"📋 وضعیت فعلی:\n"
                            f"━━━━━━━━━━━━━━━\n"
                            f"🏢 شعبه: {branch_name}\n"
                            f"📅 تاریخ: {get_shamsi_date_formatted(shamsi_today)}\n"
                            f"👤 وصولی معاون: {int(safe_format(existing['deputy_amount']))//1_000_000:,.0f} میلیون ریال\n"
                            f"👥 وصولی همکاران: {int(safe_format(existing['others_amount']))//1_000_000:,.0f} میلیون ریال\n"
                            f"💰 جمع کل: {(int(safe_format(existing['deputy_amount'])) + int(safe_format(existing['others_amount'])))//1_000_000:,.0f} میلیون ریال\n"
                            f"━━━━━━━━━━━━━━━\n\n"
                            f"❓ آیا مایل به ویرایش هستید؟ (فقط تا ۱۲ شب امکان ویرایش دارید)"
                        )
                        send_message(chat_id, msg, confirm_keyboard)
                    else:
                        msg = (
                            f"⚠️ اطلاعات امروز قبلاً ثبت شده است و زمان ویرایش (تا ۱۲ شب) گذشته است.\n\n"
                            f"📋 وضعیت فعلی:\n"
                            f"━━━━━━━━━━━━━━━\n"
                            f"🏢 شعبه: {branch_name}\n"
                            f"📅 تاریخ: {get_shamsi_date_formatted(shamsi_today)}\n"
                            f"👤 وصولی معاون: {int(safe_format(existing['deputy_amount']))//1_000_000:,.0f} میلیون ریال\n"
                            f"👥 وصولی همکاران: {int(safe_format(existing['others_amount']))//1_000_000:,.0f} میلیون ریال\n"
                            f"💰 جمع کل: {(int(safe_format(existing['deputy_amount'])) + int(safe_format(existing['others_amount'])))//1_000_000:,.0f} میلیون ریال"
                        )
                        send_message(chat_id, msg, get_deputy_keyboard())
                else:
                    user_states.update(chat_id, {"state": "WAITING_FOR_DEPUTY_AMOUNT", "edit_mode": False})
                    send_message(chat_id, "📝 لطفاً میزان وصولی خود (معاون) را به **میلیون ریال** وارد کنید:", get_cancel_keyboard())
                return

            if text == "📊 گزارش وصولی":
                report = get_branch_10_day_report(branch_id)
                if report:
                    msg = f"📊 گزارش وصول شعبه {branch_name}\n(۱۰ روز اخیر)\n━━━━━━━━━━━━━━━━━━\n\n"
                    total_sum = 0
                    for i, row in enumerate(report, 1):
                        dep = int(safe_format(row[1]))
                        oth = int(safe_format(row[2]))
                        tot = int(safe_format(row[3]))
                        msg += f"{i}. 📅 {get_shamsi_date_formatted(row[0])}\n"
                        msg += f"   👤 معاون: {dep//1_000_000:,.0f} میلیون ریال\n"
                        msg += f"   👥 همکاران: {oth//1_000_000:,.0f} میلیون ریال\n"
                        msg += f"   💰 جمع: {tot//1_000_000:,.0f} میلیون ریال\n\n"
                        total_sum += tot
                    msg += f"━━━━━━━━━━━━━━━━━━\n"
                    msg += f"📈 جمع ۱۰ روز: {total_sum//1_000_000:,.0f} میلیون ریال\n"
                    msg += f"📊 میانگین روزانه: {total_sum//len(report)//1_000_000:,.0f} میلیون ریال"
                    send_message(chat_id, msg, get_deputy_keyboard())
                else:
                    send_message(chat_id, "📊 هیچ سابقه وصولی برای شعبه شما یافت نشد.", get_deputy_keyboard())
                return

            if text == "📈 مقایسه عملکرد":
                perf = get_branch_performance(branch_id, 7)
                if perf:
                    msg = f"📈 تحلیل عملکرد شعبه {branch_name}\n(۷ روز اخیر)\n━━━━━━━━━━━━━━━━━━\n\n"
                    for i in range(len(perf)):
                        row = perf[i]
                        date = row[0]
                        daily = int(safe_format(row[1]))
                        avg = int(safe_format(row[2]))
                        if i == 0:
                            trend = "📊"
                        else:
                            prev_amount = int(safe_format(perf[i-1][1]))
                            if daily > prev_amount:
                                trend = "📈"
                            elif daily < prev_amount:
                                trend = "📉"
                            else:
                                trend = "➡️"
                        msg += f"{trend} 📅 {get_shamsi_date_formatted(date)}\n"
                        msg += f"   جمع روزانه: {daily//1_000_000:,.0f} میلیون ریال\n"
                        msg += f"   میانگین متحرک: {avg//1_000_000:,.0f} میلیون ریال\n\n"
                    send_message(chat_id, msg, get_deputy_keyboard())
                else:
                    send_message(chat_id, "📈 داده کافی برای تحلیل وجود ندارد.", get_deputy_keyboard())
                return

            if text == "📋 مشاهده ثبت امروز":
                shamsi_today = get_shamsi_date()
                if is_holiday(shamsi_today):
                    send_message(chat_id, f"📅 امروز {get_shamsi_date_formatted(shamsi_today)} تعطیل است، ثبت وصولی وجود ندارد.", get_deputy_keyboard())
                    return
                existing = check_existing_collection(branch_id, shamsi_today)
                if existing:
                    msg = (
                        f"📋 ثبت امروز شعبه {branch_name}\n"
                        f"📅 تاریخ: {get_shamsi_date_formatted(shamsi_today)}\n"
                        f"━━━━━━━━━━━━━━━\n"
                        f"👤 وصولی معاون: {int(safe_format(existing['deputy_amount']))//1_000_000:,.0f} میلیون ریال\n"
                        f"👥 وصولی همکاران: {int(safe_format(existing['others_amount']))//1_000_000:,.0f} میلیون ریال\n"
                        f"💰 جمع کل: {(int(safe_format(existing['deputy_amount'])) + int(safe_format(existing['others_amount'])))//1_000_000:,.0f} میلیون ریال"
                    )
                    send_message(chat_id, msg, get_deputy_keyboard())
                else:
                    send_message(chat_id, f"📭 امروز ({shamsi_today}) هنوز ثبت نشده است.", get_deputy_keyboard())
                return

            if text == "📅 گزارش تاریخ خاص":
                user_states.update(chat_id, {"state": "WAITING_FOR_BRANCH_DATE"})
                send_message(chat_id, "📅 لطفاً تاریخ مورد نظر را به فرمت **YYYY/MM/DD** وارد کنید (مثلاً ۱۴۰۳/۰۱/۱۵):", get_cancel_keyboard())
                return

            if text == "📊 تاریخچه کامل":
                history = get_branch_full_history(branch_id)
                if history:
                    msg = f"📊 تاریخچه کامل شعبه {branch_name}\n━━━━━━━━━━━━━━━━━━\n\n"
                    total_all = 0
                    for i, row in enumerate(history, 1):
                        dep = int(safe_format(row[1]))
                        oth = int(safe_format(row[2]))
                        tot = int(safe_format(row[3]))
                        msg += f"{i}. 📅 {get_shamsi_date_formatted(row[0])}\n"
                        msg += f"   👤 معاون: {dep//1_000_000:,.0f} میلیون ریال\n"
                        msg += f"   👥 همکاران: {oth//1_000_000:,.0f} میلیون ریال\n"
                        msg += f"   💰 جمع: {tot//1_000_000:,.0f} میلیون ریال\n\n"
                        total_all += tot
                    msg += f"━━━━━━━━━━━━━━━━━━\n"
                    msg += f"📈 جمع کل از ابتدا: {total_all//1_000_000:,.0f} میلیون ریال"
                    send_message(chat_id, msg, get_deputy_keyboard())
                else:
                    send_message(chat_id, "📭 هیچ سابقه‌ای برای شعبه شما وجود ندارد.", get_deputy_keyboard())
                return

            if text == "📝 ثبت یادداشت":
                user_states.update(chat_id, {"state": "WAITING_FOR_NOTE_FOR_COLLECTION"})
                send_message(chat_id, "📝 لطفاً **شناسه وصول** (ID) که در گزارش‌ها مشاهده می‌کنید و متن یادداشت را به این فرمت وارد کنید:\n\n`[شناسه] | [متن یادداشت]`\n\nمثال: `42 | وصول از پرونده شماره ۱۲۳۴۵`", get_cancel_keyboard())
                return

            if text == "📋 مشاهده یادداشت‌ها":
                conn = None
                try:
                    conn = get_db_connection()
                    with conn.cursor() as cur:
                        cur.execute("""
                            SELECT n.id, b.name, c.shamsi_date, n.note_text,
                                   n.created_at AT TIME ZONE 'Asia/Tehran' as created_at_iran
                            FROM notes n
                            JOIN collections c ON n.collection_id = c.id
                            JOIN branches b ON c.branch_id = b.id
                            WHERE n.user_id = %s
                            ORDER BY n.created_at DESC
                            LIMIT 20
                        """, (user_db_id,))
                        notes = cur.fetchall()
                        if notes:
                            msg = "📝 **یادداشت‌های شما**\n━━━━━━━━━━━━━━━━━━\n"
                            for note in notes:
                                note_time = note[4]
                                msg += f"🏢 {note[1]} | 📅 {note[2]}\n"
                                msg += f"📝 {note[3]}\n"
                                msg += f"⏰ {note_time.strftime('%H:%M') if hasattr(note_time, 'strftime') else note_time}\n\n"
                            send_message(chat_id, msg, get_deputy_keyboard())
                        else:
                            send_message(chat_id, "شما هیچ یادداشتی ثبت نکرده‌اید.", get_deputy_keyboard())
                except Exception as e:
                    logger.exception("Deputy command failed")
                    send_message(chat_id, "❌ عملیات انجام نشد؛ جزئیات خطا ثبت شد.", get_deputy_keyboard())
                finally:
                    if conn:
                        return_db_connection(conn)
                return

            send_message(chat_id, "لطفاً یک گزینه از منو انتخاب کنید:", get_deputy_keyboard())
            return

        send_message(chat_id, "نقش شما نامعتبر است. لطفاً با پشتیبان تماس بگیرید.")

    except Exception as e:
        logger.error(f"❌ handle_message error: {e}", exc_info=True)
        try:
            send_message(message['chat']['id'], "❌ خطایی رخ داد. لطفاً مجدداً تلاش کنید.")
        except Exception:
            pass
        raise

if __name__ == "__main__":
    main()
